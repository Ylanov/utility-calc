import io
from typing import Dict, List, Tuple
from decimal import Decimal
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from app.models import User, MeterReading, BillingPeriod
from app.auth import get_password_hash
from app.services.debt_import import find_user_fuzzy, normalize_name

ZERO = Decimal("0.00")


async def import_users_from_excel(file_content: bytes, db: AsyncSession) -> dict:
    try:
        # Загружаем Excel-файл из байтов в режиме только для чтения (оптимизация памяти)
        workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        worksheet = workbook.active

        # ИСПРАВЛЕНИЕ: Загружаем ТОЛЬКО активных пользователей (исключаем удаленных - Soft Delete)
        users_result = await db.execute(select(User).where(User.is_deleted == False))
        existing_users: Dict[str, User] = {user.username: user for user in users_result.scalars().all()}

        added_count = 0
        updated_count = 0
        skipped_count = 0
        errors: List[str] = []
        new_users: List[User] = []
        password_cache: Dict[str, str] = {}

        # Читаем строки, пропуская заголовок (min_row=2)
        rows = worksheet.iter_rows(min_row=2, values_only=True)

        for row_index, row in enumerate(rows, start=2):
            try:
                # Пропускаем пустые строки
                if not row or not row[0]:
                    skipped_count += 1
                    continue

                username = str(row[0]).strip()
                if not username:
                    skipped_count += 1
                    continue

                # Если пароль не указан, используем логин в качестве пароля
                password = str(row[1]).strip() if len(row) > 1 and row[1] else username
                dormitory = str(row[2]).strip() if len(row) > 2 and row[2] else None

                try:
                    apartment_area = Decimal(str(row[3])) if len(row) > 3 and row[3] else ZERO
                    residents_count = int(row[4]) if len(row) > 4 and row[4] else 1
                    total_room_residents = int(row[5]) if len(row) > 5 and row[5] else residents_count
                except Exception:
                    apartment_area = ZERO
                    residents_count = 1
                    total_room_residents = 1
                    errors.append(f"Строка {row_index}: Ошибка числовых значений. Установлены значения по умолчанию.")

                workplace = str(row[6]).strip() if len(row) > 6 and row[6] else None

                # Кэширование хэшей паролей (ускоряет работу, если у многих одинаковый дефолтный пароль)
                if password in password_cache:
                    hashed_password = password_cache[password]
                else:
                    hashed_password = get_password_hash(password)
                    password_cache[password] = hashed_password

                # Если пользователь уже существует — обновляем его данные
                if username in existing_users:
                    user = existing_users[username]
                    user.dormitory = dormitory
                    user.apartment_area = apartment_area
                    user.residents_count = residents_count
                    user.total_room_residents = total_room_residents
                    user.workplace = workplace

                    if password:
                        user.hashed_password = hashed_password

                    # Обновляем счетчик только если это не дубль внутри текущего загружаемого файла
                    if user not in new_users:
                        updated_count += 1

                # Если пользователя нет — создаем нового
                else:
                    new_user = User(
                        username=username,
                        hashed_password=hashed_password,
                        role="user",
                        dormitory=dormitory,
                        apartment_area=apartment_area,
                        residents_count=residents_count,
                        total_room_residents=total_room_residents,
                        workplace=workplace,
                        is_deleted=False
                    )
                    new_users.append(new_user)
                    existing_users[username] = new_user
                    added_count += 1

            except Exception as error:
                skipped_count += 1
                errors.append(f"Строка {row_index}: {str(error)}")

        # Массовое добавление новых пользователей
        if new_users:
            db.add_all(new_users)

        await db.commit()
        workbook.close()

        return {
            "status": "success",
            "message": "Импорт завершен",
            "added": added_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "errors": errors
        }

    except Exception as error:
        await db.rollback()
        return {
            "status": "error",
            "message": f"Ошибка импорта: {str(error)}",
            "added": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [str(error)]
        }


async def generate_billing_report_xlsx(db: AsyncSession, period_id: int) -> Tuple[io.BytesIO, str]:
    period_result = await db.execute(select(BillingPeriod).where(BillingPeriod.id == period_id))
    period = period_result.scalars().first()

    if not period:
        raise ValueError("Период не найден")

    # ВАЖНО: Здесь мы СПЕЦИАЛЬНО не фильтруем по User.is_deleted == False.
    # Потому что если жилец съехал (и был мягко удален), его исторические
    # начисления за закрытые периоды всё равно ДОЛЖНЫ быть в отчетах бухгалтерии.
    statement = (
        select(User, MeterReading)
        .join(MeterReading, User.id == MeterReading.user_id)
        .where(MeterReading.period_id == period_id, MeterReading.is_approved.is_(True))
        .order_by(User.dormitory, User.username)
    )

    result = await db.execute(statement)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Сводная ведомость"

    headers = [
        "Общежитие/Комната", "ФИО (Логин)", "Площадь", "Жильцов",
        "ГВС (руб)", "ХВС (руб)", "Водоотв. (руб)", "Электроэнергия (руб)",
        "Содержание (руб)", "Наем (руб)", "ТКО (руб)", "Отопление + ОДН (руб)",
        "Счет 209 (Комм.)", "Счет 205 (Найм)", "ИТОГО (руб)"
    ]
    worksheet.append(headers)

    total_sum = ZERO
    total_209_sum = ZERO
    total_205_sum = ZERO

    for user, reading in result:
        total_cost = Decimal(reading.total_cost or 0)
        t_209 = Decimal(reading.total_209 or 0)
        t_205 = Decimal(reading.total_205 or 0)

        total_sum += total_cost
        total_209_sum += t_209
        total_205_sum += t_205

        # Если пользователь удален, помечаем это в отчете (опционально, для удобства бухгалтера)
        username_display = user.username
        if user.is_deleted:
            # Убираем системный суффикс "_deleted_ID" для красивого отображения
            username_display = username_display.split("_deleted_")[0] + " (Выселен)"

        worksheet.append([
            user.dormitory,
            username_display,
            user.apartment_area,
            f"{user.residents_count}/{user.total_room_residents}",
            reading.cost_hot_water,
            reading.cost_cold_water,
            reading.cost_sewage,
            reading.cost_electricity,
            reading.cost_maintenance,
            reading.cost_social_rent,
            reading.cost_waste,
            reading.cost_fixed_part,
            t_209,
            t_205,
            total_cost
        ])

    worksheet.append([""] * 11 + ["ИТОГО:", total_209_sum, total_205_sum, total_sum])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Формируем безопасное имя файла
    filename = f"Report_{period.name}".replace(" ", "_") + ".xlsx"

    return output, filename

async def import_readings_from_excel(file_content: bytes, db: AsyncSession, period_id: int) -> dict:
    """Импорт показаний (черновиков) счетчиков из Excel"""
    try:
        workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        worksheet = workbook.active

        # Загружаем активных пользователей для поиска
        users_result = await db.execute(select(User).where(User.is_deleted == False))
        users_map = {normalize_name(u.username): u.id for u in users_result.scalars().all()}

        added_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        # Читаем со второй строки (пропускаем заголовки)
        # Ожидаемый формат: [ФИО, ГВС, ХВС, Электричество]
        rows = worksheet.iter_rows(min_row=2, values_only=True)
        for row_index, row in enumerate(rows, start=2):
            if not row or not row[0]:
                skipped_count += 1
                continue

            fio_raw = str(row[0]).strip()
            user_id = find_user_fuzzy(fio_raw, users_map)

            if not user_id:
                errors.append(f"Строка {row_index}: Жилец '{fio_raw}' не найден")
                skipped_count += 1
                continue

            try:
                # Читаем объемы. Если пусто - ставим 0
                hot = Decimal(str(row[1]).replace(',', '.').replace(' ', '')) if len(row) > 1 and row[1] is not None else Decimal("0.00")
                cold = Decimal(str(row[2]).replace(',', '.').replace(' ', '')) if len(row) > 2 and row[2] is not None else Decimal("0.00")
                elect = Decimal(str(row[3]).replace(',', '.').replace(' ', '')) if len(row) > 3 and row[3] is not None else Decimal("0.00")
            except Exception:
                errors.append(f"Строка {row_index}: Ошибка в числах для '{fio_raw}'")
                skipped_count += 1
                continue

            # Ищем, есть ли уже ЧЕРНОВИК в этом периоде у юзера
            exist_res = await db.execute(
                select(MeterReading).where(
                    MeterReading.user_id == user_id,
                    MeterReading.period_id == period_id,
                    MeterReading.is_approved == False
                )
            )
            draft = exist_res.scalars().first()

            if draft:
                # Обновляем черновик
                draft.hot_water = hot
                draft.cold_water = cold
                draft.electricity = elect
                updated_count += 1
            else:
                # Создаем новый черновик
                new_draft = MeterReading(
                    user_id=user_id,
                    period_id=period_id,
                    hot_water=hot,
                    cold_water=cold,
                    electricity=elect,
                    is_approved=False,
                    anomaly_flags="IMPORTED_DRAFT" # Пометка, что загружено из админки
                )
                db.add(new_draft)
                added_count += 1

        await db.commit()
        workbook.close()

        return {
            "status": "success",
            "added": added_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "errors": errors
        }

    except Exception as error:
        await db.rollback()
        return {"status": "error", "message": str(error), "errors": [str(error)]}