# Отчётность по долгам: список должников, зеркало по квартирам, KPI, dormitories, unassigned, экспорт.
# Механически выделено из монолитного routers/financier.py (распил на
# пакет financier/): код перенесён дословно, поведение/пути/тексты 1:1.

import io
from app.core.time_utils import utcnow
from typing import Optional
from fastapi import Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, or_, desc, asc
from app.core.database import get_db
from app.modules.utility.models import User, MeterReading, BillingPeriod, Room, DebtImportLog
from app.core.dependencies import get_current_user
from app.modules.utility.schemas import PaginatedResponse, UserDebtResponse
from app.modules.utility.services.user_service import countable_resident_condition
from app.modules.utility.services.search_utils import like_contains

from ._shared import (
    router,
    _resolve_view_period,
    _require_finance,
)


@router.get(
    "/users-status",
    response_model=PaginatedResponse[UserDebtResponse],
    summary="Список пользователей с долгами"
)
async def get_users_with_debts(
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=500),
        search: str | None = Query(None),
        # Новые фильтры/сортировка для вкладки «Долги 1С»
        only_debtors: bool = Query(False, description="Только с положительным долгом 209 или 205"),
        only_overpaid: bool = Query(False, description="Только с положительной переплатой"),
        has_data: bool = Query(False, description="Скрыть жильцов без данных из 1С (все 8 финансовых полей = 0)"),
        dormitory: Optional[str] = Query(None, description="Фильтр по названию общежития"),
        min_debt: Optional[float] = Query(None, ge=0, description="Минимальный суммарный долг (209+205)"),
        sort_by: str = Query("room", pattern="^(room|username|debt|overpay|total)$"),
        sort_dir: str = Query("asc", pattern="^(asc|desc)$"),
        period_id: Optional[int] = Query(None, description="Период просмотра; по умолчанию активный → последний импорт → свежий"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db)
):
    if current_user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    offset = (page - 1) * limit

    active_period = await _resolve_view_period(db, period_id)
    period_id = active_period.id if active_period else None

    d209 = func.coalesce(func.sum(MeterReading.debt_209), 0).label("debt_209")
    o209 = func.coalesce(func.sum(MeterReading.overpayment_209), 0).label("overpayment_209")
    d205 = func.coalesce(func.sum(MeterReading.debt_205), 0).label("debt_205")
    o205 = func.coalesce(func.sum(MeterReading.overpayment_205), 0).label("overpayment_205")
    # Bug V: обороты для UI «движение средств».
    od209 = func.coalesce(func.sum(MeterReading.obor_debit_209), 0).label("obor_debit_209")
    oc209 = func.coalesce(func.sum(MeterReading.obor_credit_209), 0).label("obor_credit_209")
    od205 = func.coalesce(func.sum(MeterReading.obor_debit_205), 0).label("obor_debit_205")
    oc205 = func.coalesce(func.sum(MeterReading.obor_credit_205), 0).label("obor_credit_205")
    total = func.coalesce(func.sum(MeterReading.total_cost), 0).label("current_total_cost")

    stmt = select(
        User, Room, d209, o209, d205, o205, total,
        od209, oc209, od205, oc205,
    ).outerjoin(
        Room, User.room_id == Room.id
    ).outerjoin(
        MeterReading,
        (User.id == MeterReading.user_id) &
        (MeterReading.period_id == period_id)
    ).where(
        User.is_deleted.is_(False),
        User.role == "user",
        # «свои дома» (безкомнатные без долга) в список «Долги 1С» не берём
        countable_resident_condition(),
    )

    search_condition = None
    if search:
        search_value = like_contains(search.lower())
        search_condition = or_(
            func.lower(User.username).like(search_value),
            func.lower(Room.dormitory_name).like(search_value),
            func.lower(Room.room_number).like(search_value)
        )
        stmt = stmt.where(search_condition)

    if dormitory:
        stmt = stmt.where(Room.dormitory_name == dormitory)

    stmt = stmt.group_by(User.id, Room.id)

    # Фильтры по агрегированным значениям — HAVING
    if only_debtors:
        stmt = stmt.having((d209 + d205) > 0)
    if only_overpaid:
        stmt = stmt.having((o209 + o205) > 0)
    if min_debt is not None:
        stmt = stmt.having((d209 + d205) >= min_debt)
    if has_data:
        # Bug AB: «не показывать пустых» — хотя бы одно из 8 финансовых
        # полей (сальдо + обороты по 209 и 205) > 0.
        stmt = stmt.having(
            (d209 + o209 + d205 + o205 + od209 + oc209 + od205 + oc205) > 0
        )

    # Сортировка: столбец + направление
    sort_map = {
        "room": (Room.dormitory_name, Room.room_number),
        "username": (User.username,),
        "debt": ((d209 + d205).label("__debt_sum"),),
        "overpay": ((o209 + o205).label("__over_sum"),),
        "total": (total,),
    }
    cols = sort_map[sort_by]
    direction = desc if sort_dir == "desc" else asc
    order_cols = [direction(c).nulls_last() for c in cols]
    # Стабилизатор — вторичная сортировка по username
    if sort_by != "username":
        order_cols.append(asc(User.username))
    stmt = stmt.order_by(*order_cols).limit(limit).offset(offset)

    # count
    count_stmt = select(func.count(User.id)).outerjoin(Room, User.room_id == Room.id).where(
        User.is_deleted.is_(False), User.role == "user",
        countable_resident_condition(),
    )
    if search_condition is not None:
        count_stmt = count_stmt.where(search_condition)
    if dormitory:
        count_stmt = count_stmt.where(Room.dormitory_name == dormitory)

    # Для only_debtors/only_overpaid/min_debt/has_data count тоже надо
    # пересчитать через HAVING — делаем через subquery вместо дублирования.
    if only_debtors or only_overpaid or min_debt is not None or has_data:
        inner = stmt.with_only_columns(User.id).limit(None).offset(None).order_by(None).subquery()
        count_stmt = select(func.count()).select_from(inner)

    total_res = await db.execute(count_stmt)
    total_items = total_res.scalar_one()

    result = await db.execute(stmt)
    rows = result.all()

    # Покрытие импортами 1С активного периода: кто из жильцов попал в ПОСЛЕДНИЙ
    # импорт счёта (applied_state ключуется по str(user_id)). seen_2xx:
    #   None  — импорта этого счёта в периоде не было (нечему быть «не найденным»);
    #   set   — множество user_id, попавших в последний импорт счёта.
    # Жилец с долгом >0 всегда был в импорте (его touch'нули), поэтому «не найден»
    # покажется только тем, у кого реально нет данных по счёту, а не «долг 0».
    async def _seen_ids(acct: str):
        if not period_id:
            return None
        last_imp = (await db.execute(
            select(DebtImportLog)
            .where(
                DebtImportLog.period_id == period_id,
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
            )
            .order_by(desc(DebtImportLog.started_at))
            .limit(1)
        )).scalars().first()
        if not last_imp:
            return None
        st = last_imp.applied_state or {}
        return {int(k) for k in st.keys() if str(k).isdigit()}

    seen_209 = await _seen_ids("209")
    seen_205 = await _seen_ids("205")

    items = []
    for row in rows:
        user_obj, room_obj = row[0], row[1]
        items.append({
            "id": user_obj.id,
            "username": user_obj.username,
            "room": room_obj,
            "debt_209": row[2],
            "overpayment_209": row[3],
            "debt_205": row[4],
            "overpayment_205": row[5],
            "current_total_cost": row[6],
            # Bug V: движение средств — обороты периода.
            "obor_debit_209": row[7],
            "obor_credit_209": row[8],
            "obor_debit_205": row[9],
            "obor_credit_205": row[10],
            # Покрытие импортом (None если импорта счёта не было).
            "seen_209": (user_obj.id in seen_209) if seen_209 is not None else None,
            "seen_205": (user_obj.id in seen_205) if seen_205 is not None else None,
        })

    return {"total": total_items, "page": page, "size": limit, "items": items}


# =========================================================================
# ЗЕРКАЛО ПО КВАРТИРАМ: та же отчётность, но агрегация по ПОМЕЩЕНИЮ, без ФИО.
# Долг квартиры = сумма долгов всех жильцов комнаты (по room_id). Адрес вместо
# ФИО; кто живёт — в детализации /rooms/{id}/residents-finance.
# =========================================================================
@router.get("/rooms-status", summary="Список квартир с долгами (агрегация по помещению)")
async def get_rooms_with_debts(
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=500),
        search: str | None = Query(None),
        only_debtors: bool = Query(False),
        only_overpaid: bool = Query(False),
        has_data: bool = Query(False),
        dormitory: Optional[str] = Query(None),
        place_type: Optional[str] = Query(None, pattern="^(dormitory|house)$"),
        min_debt: Optional[float] = Query(None, ge=0),
        sort_by: str = Query("room", pattern="^(room|debt|overpay|total)$"),
        sort_dir: str = Query("asc", pattern="^(asc|desc)$"),
        period_id: Optional[int] = Query(None, description="Период просмотра; по умолчанию активный → последний импорт → свежий"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db),
):
    if current_user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    offset = (page - 1) * limit
    active = await _resolve_view_period(db, period_id)
    period_id = active.id if active else None

    d209 = func.coalesce(func.sum(MeterReading.debt_209), 0).label("debt_209")
    o209 = func.coalesce(func.sum(MeterReading.overpayment_209), 0).label("overpayment_209")
    d205 = func.coalesce(func.sum(MeterReading.debt_205), 0).label("debt_205")
    o205 = func.coalesce(func.sum(MeterReading.overpayment_205), 0).label("overpayment_205")
    od209 = func.coalesce(func.sum(MeterReading.obor_debit_209), 0).label("obor_debit_209")
    oc209 = func.coalesce(func.sum(MeterReading.obor_credit_209), 0).label("obor_credit_209")
    od205 = func.coalesce(func.sum(MeterReading.obor_debit_205), 0).label("obor_debit_205")
    oc205 = func.coalesce(func.sum(MeterReading.obor_credit_205), 0).label("obor_credit_205")
    total = func.coalesce(func.sum(MeterReading.total_cost), 0).label("current_total_cost")
    residents = func.count(func.distinct(MeterReading.user_id)).label("residents_count")

    stmt = select(
        Room, d209, o209, d205, o205, total, od209, oc209, od205, oc205, residents,
    ).outerjoin(
        MeterReading,
        (Room.id == MeterReading.room_id) & (MeterReading.period_id == period_id),
    )
    search_condition = None
    if search:
        sv = like_contains(search.lower())
        search_condition = or_(
            func.lower(Room.dormitory_name).like(sv),
            func.lower(Room.room_number).like(sv),
            func.lower(Room.street).like(sv),
            func.lower(Room.house_number).like(sv),
            func.lower(Room.apartment_number).like(sv),
        )
        stmt = stmt.where(search_condition)
    if dormitory:
        stmt = stmt.where(Room.dormitory_name == dormitory)
    if place_type:
        stmt = stmt.where(Room.place_type == place_type)
    stmt = stmt.group_by(Room.id)
    if only_debtors:
        stmt = stmt.having((d209 + d205) > 0)
    if only_overpaid:
        stmt = stmt.having((o209 + o205) > 0)
    if min_debt is not None:
        stmt = stmt.having((d209 + d205) >= min_debt)
    if has_data:
        stmt = stmt.having(
            (d209 + o209 + d205 + o205 + od209 + oc209 + od205 + oc205) > 0
        )

    sort_map = {
        "room": (Room.dormitory_name, Room.room_number, Room.street, Room.house_number),
        "debt": ((d209 + d205).label("__d"),),
        "overpay": ((o209 + o205).label("__o"),),
        "total": (total,),
    }
    direction = desc if sort_dir == "desc" else asc
    order_cols = [direction(c).nulls_last() for c in sort_map[sort_by]]
    order_cols.append(asc(Room.id))
    stmt = stmt.order_by(*order_cols).limit(limit).offset(offset)

    count_stmt = select(func.count(Room.id))
    if search_condition is not None:
        count_stmt = count_stmt.where(search_condition)
    if dormitory:
        count_stmt = count_stmt.where(Room.dormitory_name == dormitory)
    if place_type:
        count_stmt = count_stmt.where(Room.place_type == place_type)
    if only_debtors or only_overpaid or min_debt is not None or has_data:
        inner = stmt.with_only_columns(Room.id).limit(None).offset(None).order_by(None).subquery()
        count_stmt = select(func.count()).select_from(inner)
    total_items = (await db.execute(count_stmt)).scalar_one()

    rows = (await db.execute(stmt)).all()
    items = []
    for row in rows:
        room = row[0]
        items.append({
            "room_id": room.id,
            "address": room.format_address,
            "place_type": room.place_type,
            "dormitory_name": room.dormitory_name,
            "room_number": room.room_number,
            "residents_count": int(row[10] or 0),
            "debt_209": row[1], "overpayment_209": row[2],
            "debt_205": row[3], "overpayment_205": row[4],
            "current_total_cost": row[5],
            "obor_debit_209": row[6], "obor_credit_209": row[7],
            "obor_debit_205": row[8], "obor_credit_205": row[9],
        })
    return {"total": total_items, "page": page, "size": limit, "items": items}


@router.get("/rooms/{room_id}/residents-finance",
            summary="Финансы жильцов конкретной квартиры (раскрытие строки)")
async def room_residents_finance(
        room_id: int,
        period_id: Optional[int] = Query(None, description="Период просмотра; по умолчанию активный → последний импорт → свежий"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db),
):
    if current_user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    active = await _resolve_view_period(db, period_id)
    period_id = active.id if active else None

    d209 = func.coalesce(func.sum(MeterReading.debt_209), 0)
    o209 = func.coalesce(func.sum(MeterReading.overpayment_209), 0)
    d205 = func.coalesce(func.sum(MeterReading.debt_205), 0)
    o205 = func.coalesce(func.sum(MeterReading.overpayment_205), 0)
    total = func.coalesce(func.sum(MeterReading.total_cost), 0)

    stmt = select(User, d209, o209, d205, o205, total).outerjoin(
        MeterReading,
        (User.id == MeterReading.user_id) & (MeterReading.period_id == period_id),
    ).where(
        User.room_id == room_id, User.is_deleted.is_(False),
    ).group_by(User.id).order_by(User.username)
    rows = (await db.execute(stmt)).all()
    return {
        "room_id": room_id,
        "residents": [{
            "user_id": r[0].id,
            "username": r[0].username,
            "full_name": getattr(r[0], "full_name", None),
            "debt_209": r[1], "overpayment_209": r[2],
            "debt_205": r[3], "overpayment_205": r[4],
            "current_total_cost": r[5],
        } for r in rows],
    }


# =========================================================================
# NEW: KPI / STATS / EXPORT / HISTORY / UNDO / RECONCILE
# =========================================================================

@router.get("/debts/stats", summary="KPI по долгам (выбранный/активный период)")
async def debts_stats(
    period_id: Optional[int] = Query(None, description="Период просмотра; по умолчанию активный → последний импорт → свежий"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сводка долгов для шапки вкладки «Долги 1С»."""
    _require_finance(current_user)

    active_period = await _resolve_view_period(db, period_id)
    period_id = active_period.id if active_period else None

    # Агрегация по readings активного периода. ВАЖНО: join User + фильтр
    # is_deleted/role — чтобы KPI считал ТОЛЬКО активных жильцов и совпадал со
    # списком «Долги 1С» (users-status фильтрует так же). Иначе долги, оставшиеся
    # на user_id удалённых/выехавших жильцов после импорта 1С, раздувают счётчик
    # («Должников: 800», а в списке 1).
    _active_user = [User.is_deleted.is_(False), User.role == "user"]
    agg_q = (
        select(
            func.coalesce(func.sum(MeterReading.debt_209), 0),
            func.coalesce(func.sum(MeterReading.overpayment_209), 0),
            func.coalesce(func.sum(MeterReading.debt_205), 0),
            func.coalesce(func.sum(MeterReading.overpayment_205), 0),
            func.count(MeterReading.id),
        )
        .join(User, User.id == MeterReading.user_id)
        .where(MeterReading.period_id == period_id, *_active_user)
    )
    agg = (await db.execute(agg_q)).one()
    total_debt_209, total_over_209, total_debt_205, total_over_205, readings_count = agg

    # Должников: активных жильцов где debt_209+205 > 0 в активном периоде
    debtors_q = (
        select(func.count(func.distinct(MeterReading.user_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            *_active_user,
            (MeterReading.debt_209 > 0) | (MeterReading.debt_205 > 0),
        )
    )
    debtors_count = (await db.execute(debtors_q)).scalar_one()

    # Переплатчиков
    overpayers_q = (
        select(func.count(func.distinct(MeterReading.user_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            *_active_user,
            (MeterReading.overpayment_209 > 0) | (MeterReading.overpayment_205 > 0),
        )
    )
    overpayers_count = (await db.execute(overpayers_q)).scalar_one()

    # --- Учёт по КВАРТИРАМ (помещениям), а не по жильцам ---
    # Квартир с долгом: distinct room_id где сумма debt_209+205 > 0 (по активным
    # жильцам — join User, чтобы не считать квартиры по долгам выехавших).
    rooms_debt_q = (
        select(func.count(func.distinct(MeterReading.room_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            MeterReading.room_id.isnot(None),
            *_active_user,
            (MeterReading.debt_209 > 0) | (MeterReading.debt_205 > 0),
        )
    )
    rooms_with_debt_count = (await db.execute(rooms_debt_q)).scalar_one()

    # Квартир с переплатой
    rooms_over_q = (
        select(func.count(func.distinct(MeterReading.room_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            MeterReading.room_id.isnot(None),
            *_active_user,
            (MeterReading.overpayment_209 > 0) | (MeterReading.overpayment_205 > 0),
        )
    )
    rooms_overpaying_count = (await db.execute(rooms_over_q)).scalar_one()

    # Всего квартир с данными в периоде (для шапки в режиме «Квартиры»)
    total_rooms_q = (
        select(func.count(func.distinct(MeterReading.room_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            MeterReading.room_id.isnot(None),
            *_active_user,
        )
    )
    total_rooms = (await db.execute(total_rooms_q)).scalar_one()

    # Всего учитываемых жильцов (с комнатой ИЛИ с долгом). Без «своих домов» —
    # чтобы число сходилось с долговой популяцией (деньги/должники reading-based).
    total_users_q = select(func.count(User.id)).where(
        User.is_deleted.is_(False), User.role == "user",
        countable_resident_condition(),
    )
    total_users = (await db.execute(total_users_q)).scalar_one()

    total_debt = float(total_debt_209 or 0) + float(total_debt_205 or 0)
    total_over = float(total_over_209 or 0) + float(total_over_205 or 0)
    avg_debt = (total_debt / debtors_count) if debtors_count else 0.0
    avg_debt_room = (total_debt / rooms_with_debt_count) if rooms_with_debt_count else 0.0

    # Последний импорт
    last_log = (await db.execute(
        select(DebtImportLog).order_by(desc(DebtImportLog.started_at)).limit(1)
    )).scalars().first()

    # Распределение по общежитиям: ТОП-10 по долгу
    by_dorm_q = (
        select(
            Room.dormitory_name,
            func.sum(MeterReading.debt_209 + MeterReading.debt_205).label("total_debt"),
            func.count(func.distinct(MeterReading.user_id)).label("debtors"),
        )
        .select_from(MeterReading)
        .join(Room, Room.id == MeterReading.room_id)
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            *_active_user,
            (MeterReading.debt_209 > 0) | (MeterReading.debt_205 > 0),
        )
        .group_by(Room.dormitory_name)
        .order_by(desc("total_debt"))
        .limit(10)
    )
    # У домов dormitory_name = NULL (адрес в street/house/apartment), поэтому
    # группа без названия — это дома. Подписываем её «🏠 Дома», а не «—».
    by_dorm = [
        {"name": r[0] or "🏠 Дома", "total_debt": float(r[1] or 0), "debtors": int(r[2] or 0)}
        for r in (await db.execute(by_dorm_q)).all()
    ]

    return {
        "period_name": active_period.name if active_period else None,
        "period_id": period_id,
        "total_users": total_users,
        "debtors_count": debtors_count,
        "overpayers_count": overpayers_count,
        "total_debt_209": float(total_debt_209 or 0),
        "total_debt_205": float(total_debt_205 or 0),
        "total_debt": round(total_debt, 2),
        "total_overpay_209": float(total_over_209 or 0),
        "total_overpay_205": float(total_over_205 or 0),
        "total_overpay": round(total_over, 2),
        "avg_debt_per_debtor": round(avg_debt, 2),
        # Учёт по квартирам (помещениям)
        "rooms_with_debt_count": int(rooms_with_debt_count or 0),
        "rooms_overpaying_count": int(rooms_overpaying_count or 0),
        "total_rooms": int(total_rooms or 0),
        "avg_debt_per_room": round(avg_debt_room, 2),
        "readings_count": int(readings_count or 0),
        "last_import": {
            "id": last_log.id,
            "account_type": last_log.account_type,
            "status": last_log.status,
            "started_at": last_log.started_at.isoformat() if last_log.started_at else None,
            "started_by": last_log.started_by_username,
            "updated": last_log.updated,
            "created": last_log.created,
            "not_found_count": last_log.not_found_count,
        } if last_log else None,
        "by_dormitory": by_dorm,
    }


@router.get("/debts/dormitories", summary="Список общежитий для фильтра")
async def debts_dormitories(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    rows = (await db.execute(
        select(Room.dormitory_name).distinct().order_by(Room.dormitory_name)
    )).scalars().all()
    return [r for r in rows if r]


@router.get("/debts/unassigned", summary="Неразнесённые долги (ФИО не сопоставлены с жильцом)")
async def debts_unassigned(
    period_id: Optional[int] = Query(None, description="Период; по умолчанию активный → последний импорт → свежий"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сумма и список долгов из 1С, которые НЕ привязались ни к одному жильцу
    (ФИО нет в базе / не сопоставлено, либо у жильца нет комнаты — долг хранится
    в показании, а оно требует комнату). Берём not_found из ПОСЛЕДНЕГО импорта
    каждого счёта (209/205) за период и сводим по ФИО. Деньги не теряются из
    вида: разнесутся, когда заведёшь жильца с комнатой и сделаешь переимпорт."""
    _require_finance(current_user)
    period = await _resolve_view_period(db, period_id)
    if not period:
        return {"period_name": None, "period_id": None, "total_debt": 0.0,
                "total_overpayment": 0.0, "count": 0, "items": []}

    merged: dict[str, dict] = {}
    for acct in ("209", "205"):
        log = (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.period_id == period.id,
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()
        if not log or not log.not_found_users:
            continue
        for item in log.not_found_users:
            if isinstance(item, dict):
                fio = (item.get("fio") or "").strip()
                debt = float(item.get("debt") or 0)
                over = float(item.get("overpayment") or 0)
            else:
                fio, debt, over = str(item).strip(), 0.0, 0.0
            if not fio:
                continue
            key = " ".join(fio.lower().split())
            slot = merged.get(key)
            if slot is None:
                slot = merged[key] = {"fio": fio, "debt": 0.0, "overpayment": 0.0, "accounts": []}
            slot["debt"] += debt
            slot["overpayment"] += over
            if acct not in slot["accounts"]:
                slot["accounts"].append(acct)

    items = sorted(merged.values(), key=lambda x: -x["debt"])
    return {
        "period_name": period.name,
        "period_id": period.id,
        "total_debt": round(sum(i["debt"] for i in items), 2),
        "total_overpayment": round(sum(i["overpayment"] for i in items), 2),
        "count": len(items),
        "items": items,
    }


@router.get("/debts/export", summary="Excel-выгрузка текущего списка долгов")
async def debts_export(
    search: str | None = Query(None),
    only_debtors: bool = Query(False),
    only_overpaid: bool = Query(False),
    dormitory: Optional[str] = Query(None),
    min_debt: Optional[float] = Query(None, ge=0),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Excel-файл с теми же фильтрами, что и в таблице UI.
    Без пагинации — выгружает все подходящие записи.
    """
    _require_finance(current_user)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    period_id = active_period.id if active_period else None

    d209 = func.coalesce(func.sum(MeterReading.debt_209), 0).label("debt_209")
    o209 = func.coalesce(func.sum(MeterReading.overpayment_209), 0).label("overpayment_209")
    d205 = func.coalesce(func.sum(MeterReading.debt_205), 0).label("debt_205")
    o205 = func.coalesce(func.sum(MeterReading.overpayment_205), 0).label("overpayment_205")
    tot = func.coalesce(func.sum(MeterReading.total_cost), 0).label("total_cost")

    stmt = select(User, Room, d209, o209, d205, o205, tot).outerjoin(
        Room, User.room_id == Room.id
    ).outerjoin(
        MeterReading,
        (User.id == MeterReading.user_id) & (MeterReading.period_id == period_id)
    ).where(User.is_deleted.is_(False), User.role == "user")

    if search:
        sv = like_contains(search.lower())
        stmt = stmt.where(or_(
            func.lower(User.username).like(sv),
            func.lower(Room.dormitory_name).like(sv),
            func.lower(Room.room_number).like(sv),
        ))
    if dormitory:
        stmt = stmt.where(Room.dormitory_name == dormitory)
    stmt = stmt.group_by(User.id, Room.id)
    if only_debtors:
        stmt = stmt.having((d209 + d205) > 0)
    if only_overpaid:
        stmt = stmt.having((o209 + o205) > 0)
    if min_debt is not None:
        stmt = stmt.having((d209 + d205) >= min_debt)

    stmt = stmt.order_by(Room.dormitory_name.asc().nulls_last(),
                         Room.room_number.asc().nulls_last(),
                         User.username.asc())
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Долги 1С"
    headers = ["ID", "ФИО", "Общежитие", "Комната", "Долг 209", "Перепл. 209",
               "Долг 205", "Перепл. 205", "Итого начислено"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E9D5FF")
    for i, r in enumerate(rows, 2):
        u, room = r[0], r[1]
        ws.cell(row=i, column=1, value=u.id)
        ws.cell(row=i, column=2, value=u.username)
        # housing_001/E2-A: для дома колонка "Общежитие" заполняется
        # улицей+номером дома, "Комната" — номером квартиры. Для общаги
        # сохраняем старое поведение (dormitory_name + room_number).
        if room and room.place_type == "house":
            _addr = ", ".join(filter(None, [
                f"ул. {room.street}" if room.street else None,
                f"д. {room.house_number}" if room.house_number else None,
            ])) or ""
            ws.cell(row=i, column=3, value=_addr)
            ws.cell(row=i, column=4, value=(f"кв. {room.apartment_number}" if room.apartment_number else ""))
        else:
            ws.cell(row=i, column=3, value=(room.dormitory_name if room else ""))
            ws.cell(row=i, column=4, value=(room.room_number if room else ""))
        ws.cell(row=i, column=5, value=float(r[2] or 0))
        ws.cell(row=i, column=6, value=float(r[3] or 0))
        ws.cell(row=i, column=7, value=float(r[4] or 0))
        ws.cell(row=i, column=8, value=float(r[5] or 0))
        ws.cell(row=i, column=9, value=float(r[6] or 0))
    for col, w in [("A", 6), ("B", 30), ("C", 22), ("D", 10),
                   ("E", 12), ("F", 12), ("G", 12), ("H", 12), ("I", 14)]:
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"debts_{utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
