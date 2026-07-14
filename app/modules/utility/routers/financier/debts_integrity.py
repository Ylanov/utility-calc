# Диагностика: resident-coverage, parser-diagnose, probe-update, integrity check/fix, zombie/orphan readings.
# Механически выделено из монолитного routers/financier.py (распил на
# пакет financier/): код перенесён дословно, поведение/пути/тексты 1:1.

from decimal import Decimal
from typing import Optional
from fastapi import Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, or_, desc
from app.core.database import get_db
from app.modules.utility.models import User, MeterReading, BillingPeriod, Room, DebtImportLog
from app.core.dependencies import get_current_user

from ._shared import (
    router,
    logger,
    _require_finance,
)


@router.get(
    "/debts/check-resident-coverage/{user_id}",
    summary="Найти жильца в архивах последних импортов 1С (диагностика)",
)
async def debts_check_resident_coverage(
    user_id: int,
    last_n: int = 10,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Для конкретного жильца перебирает последние N импортов 1С,
    парсит архивные xlsx, ищет ФИО (точное совпадение + substring).
    Полезно для диагностики «почему у Миронова нет долгов»:
      - если в архивах есть с цифрами → fuzzy-привязка ошиблась, нужен reassign
      - если есть с нулями → нормально, нет долга
      - если нет вообще → жильца не передавали из 1С
    """
    if current_user.role not in ("admin", "financier"):
        raise HTTPException(403, "Недостаточно прав")

    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Жилец не найден")

    fio_db = (user.full_name or user.username or "").strip()
    if not fio_db:
        raise HTTPException(400, "У жильца нет ФИО — нечего искать")

    # Нормализация для substring-сравнения (нижний регистр, без точек,
    # без двойных пробелов).
    import re as _re
    def _norm(s: str) -> str:
        s = (s or "").lower().replace(".", " ").replace(",", " ")
        s = _re.sub(r"\s+", " ", s).strip()
        return s

    fio_db_norm = _norm(fio_db)
    # Также берём фамилию + первую букву имени для substring-поиска.
    parts = fio_db_norm.split()
    surname = parts[0] if parts else ""

    logs = (await db.execute(
        select(DebtImportLog)
        .where(
            DebtImportLog.archive_path.is_not(None),
            DebtImportLog.status.in_(["completed", "reverted"]),
        )
        .order_by(desc(DebtImportLog.id))
        .limit(last_n)
    )).scalars().all()

    import openpyxl as _opx
    import os as _os
    from decimal import Decimal as _D
    results = []
    for log in logs:
        item = {
            "log_id": log.id,
            "account_type": log.account_type,
            "started_at": log.started_at.isoformat() if log.started_at else None,
            "status": log.status,
            "matches": [],
            "error": None,
        }
        try:
            if not _os.path.exists(log.archive_path):
                item["error"] = "archive_missing"
                results.append(item)
                continue
            wb = _opx.load_workbook(filename=log.archive_path, read_only=True, data_only=True)
            ws = wb.active
            # ФИО в ОСВ 1С может быть в любой строковой колонке (зависит
            # от шаблона выгрузки). Ищем substring фамилии во ВСЕХ
            # колонках строки.
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not row:
                    continue
                # Ищем колонку с ФИО (substring "фамилия" в строковом значении).
                fio_cell = None
                fio_col_idx = None
                for col_idx, cell_val in enumerate(row):
                    if cell_val is None or not isinstance(cell_val, str):
                        continue
                    cell_norm = _norm(cell_val)
                    if not cell_norm or not surname:
                        continue
                    if surname not in cell_norm:
                        continue
                    # Sanity: ячейка должна выглядеть как ФИО (несколько слов
                    # с заглавных букв), а не как «Договор...» или «Сальдо...».
                    # Фильтруем явные ключевые слова ОСВ.
                    if any(kw in cell_norm for kw in [
                        "договор", "сальдо", "оборот", "итого", "период",
                        "квартир", "общежит", "счёт", "счет", "помещен",
                    ]):
                        continue
                    fio_cell = cell_val
                    fio_col_idx = col_idx
                    break
                if fio_cell is None:
                    continue
                fio_cell_norm = _norm(str(fio_cell))
                # Собираем числовые значения из строки (после колонки ФИО),
                # чтобы показать сальдо.
                numeric_cols = []
                for col_val in row[fio_col_idx + 1:]:
                    if col_val is None or col_val == "":
                        continue
                    try:
                        d = _D(str(col_val).replace(",", "."))
                        if d != 0:
                            numeric_cols.append(float(d))
                    except Exception:
                        pass
                exact = fio_cell_norm == fio_db_norm
                item["matches"].append({
                    "row_excel": row_idx,
                    "col_excel": fio_col_idx + 1,  # 1-based для удобства админа
                    "fio_in_excel": str(fio_cell).strip(),
                    "exact_match": exact,
                    "numeric_values": numeric_cols[:6],
                })
            wb.close()
        except Exception as exc:
            item["error"] = f"parse_failed: {exc}"
        results.append(item)

    return {
        "user_id": user_id,
        "fio_db": fio_db,
        "imports_checked": len(logs),
        "results": results,
    }


@router.get(
    "/debts/import-history/{log_id}/parser-diagnose",
    summary="Диагностика парсера: какие колонки нашёл, что извлёк",
)
async def debts_parser_diagnose(
    log_id: int,
    fio_search: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Прогоняет логику парсера заголовков на архивном файле этого импорта
    и возвращает: какие колонки нашёл (debt_first/last, overpay_first/last),
    где была account-total row (209.34 / 205.X), какие numeric_positions
    в ней, sample 3 строк жильцов с распарсенными debt/over.

    Этот endpoint **не** делает импорт — только показывает что парсер
    видит. Помогает диагностировать «почему у Бендаса всё ещё 2385.07».
    """
    if current_user.role not in ("admin", "financier"):
        raise HTTPException(403, "Недостаточно прав")

    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог не найден")
    if not log.archive_path:
        raise HTTPException(400, "У этого импорта нет архива (старый импорт без archive_path)")

    import os as _os
    if not _os.path.exists(log.archive_path):
        raise HTTPException(404, f"Архив не найден на диске: {log.archive_path}")

    # Используем тот же парсер что и в основном импорте — копируем сюда
    # ключевые шаги.
    import openpyxl as _opx
    from app.modules.utility.services.debt_import import clean_decimal, pick_saldo_pair
    try:
        ws = _opx.load_workbook(filename=log.archive_path, read_only=True, data_only=True).active
    except Exception as e:
        raise HTTPException(400, f"openpyxl не открыл файл: {e}")

    section_markers: dict = {}
    debit_cols: list = []
    credit_cols: list = []
    account_total = None  # {row_idx, label_col, label, numeric_positions, all_values}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if not row:
            continue
        # Section markers + debit/credit positions
        for col_idx, cell in enumerate(row):
            if cell is None or not isinstance(cell, str):
                continue
            cell_norm = cell.strip().lower()
            if not cell_norm:
                continue
            if "сальдо" in cell_norm and "начал" in cell_norm and "начало" not in section_markers:
                section_markers["начало"] = col_idx
            elif "оборот" in cell_norm and ("период" in cell_norm or len(cell_norm) < 30) and "обороты" not in section_markers:
                section_markers["обороты"] = col_idx
            elif "сальдо" in cell_norm and "конец" in cell_norm and "конец" not in section_markers:
                section_markers["конец"] = col_idx
            elif cell_norm == "дебет":
                debit_cols.append(col_idx)
            elif cell_norm == "кредит":
                credit_cols.append(col_idx)
        # Account total row
        if account_total is None:
            for col_label in range(min(3, len(row))):
                cell = row[col_label]
                if cell is None:
                    continue
                s = str(cell).strip()
                if s.startswith("209.") or s.startswith("205.") or s == "209" or s == "205":
                    numeric_positions = []
                    all_values = {}
                    for col_idx in range(col_label + 1, len(row)):
                        c = row[col_idx]
                        if c is None or c == "":
                            continue
                        try:
                            d = clean_decimal(c)
                            if d != 0:
                                numeric_positions.append(col_idx)
                                all_values[col_idx] = float(d)
                        except Exception:
                            pass
                    account_total = {
                        "row_idx": row_idx,
                        "label_col": col_label,
                        "label": s,
                        "numeric_positions": numeric_positions,
                        "all_values": all_values,
                    }
                    break

    debit_cols = sorted(set(debit_cols))
    credit_cols = sorted(set(credit_cols))

    # Какие колонки выберет парсер
    chosen = {
        "debt_col_first": None,
        "debt_col_last": None,
        "overpay_col_first": None,
        "overpay_col_last": None,
        "obor_debit_col": None,
        "obor_credit_col": None,
        "strategy": None,
    }
    if account_total and len(account_total["numeric_positions"]) >= 4:
        np_list = account_total["numeric_positions"]
        chosen["debt_col_first"] = np_list[0]
        chosen["overpay_col_first"] = np_list[1] if len(np_list) > 1 else np_list[0]
        if len(np_list) >= 6:
            chosen["obor_debit_col"] = np_list[2]
            chosen["obor_credit_col"] = np_list[3]
            chosen["debt_col_last"] = np_list[4]
            chosen["overpay_col_last"] = np_list[5]
        elif len(np_list) == 5:
            chosen["obor_debit_col"] = np_list[2]
            chosen["debt_col_last"] = np_list[3]
            chosen["overpay_col_last"] = np_list[4]
        elif len(np_list) == 4:
            chosen["debt_col_last"] = np_list[2]
            chosen["overpay_col_last"] = np_list[3]
        chosen["strategy"] = "0_account_total_row"

    # Sample жильцов: для каждого извлекаем debt/over через pick_saldo_pair.
    # Если fio_search задан — ищем только этого жильца (substring match).
    # Иначе — первые 3 жильца как preview.
    samples = []
    search_norm = (fio_search or "").strip().lower()
    if chosen["debt_col_last"] is not None and chosen["overpay_col_last"] is not None:
        count = 0
        max_count = 50 if search_norm else 3
        for row in ws.iter_rows(min_row=10, max_row=2000, values_only=True):
            if count >= max_count or not row:
                continue
            # Ищем ФИО в первых 5 колонках
            fio = None
            fio_col = None
            for col_idx in range(min(5, len(row))):
                cell = row[col_idx]
                if not isinstance(cell, str):
                    continue
                s = str(cell).strip()
                if " " in s and len(s.split()) >= 2 and any('А' <= c <= 'я' for c in s):
                    # Sanity: не "Договор", не "Сальдо", не "Контрагенты"
                    s_low = s.lower()
                    if any(kw in s_low for kw in ["договор", "сальдо", "оборот", "итого", "контрагент", "счёт", "счет", "помещен", "период"]):
                        continue
                    fio = s
                    fio_col = col_idx
                    break
            if not fio:
                continue
            # Если задан поиск — фильтруем по substring.
            if search_norm and search_norm not in fio.lower():
                continue
            try:
                debt, over = pick_saldo_pair(
                    row,
                    end_debit_col=chosen["debt_col_last"],
                    end_credit_col=chosen["overpay_col_last"],
                    start_debit_col=chosen["debt_col_first"],
                    start_credit_col=chosen["overpay_col_first"],
                    obor_debit_col=chosen["obor_debit_col"],
                    obor_credit_col=chosen["obor_credit_col"],
                )
            except Exception:
                debt, over = 0, 0

            # Raw values в каждой ключевой колонке — для понимания структуры.
            def _raw(col):
                if col is None or col >= len(row):
                    return None
                v = row[col]
                if v is None or v == "":
                    return None
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    return float(clean_decimal(v))
                except Exception:
                    return str(v)

            sample = {
                "fio": fio,
                "fio_col": fio_col,
                "debt_extracted": float(debt),
                "overpayment_extracted": float(over),
                "raw_values": {
                    f"col{chosen['debt_col_first']}_start_debit": _raw(chosen["debt_col_first"]),
                    f"col{chosen['overpay_col_first']}_start_credit": _raw(chosen["overpay_col_first"]),
                    **({f"col{chosen['obor_debit_col']}_obor_debit": _raw(chosen["obor_debit_col"])} if chosen.get("obor_debit_col") is not None else {}),
                    **({f"col{chosen['obor_credit_col']}_obor_credit": _raw(chosen["obor_credit_col"])} if chosen.get("obor_credit_col") is not None else {}),
                    f"col{chosen['debt_col_last']}_end_debit": _raw(chosen["debt_col_last"]),
                    f"col{chosen['overpay_col_last']}_end_credit": _raw(chosen["overpay_col_last"]),
                },
            }

            # Если поиск конкретного жильца — добавляем сравнение с БД.
            if search_norm:
                # Ищем жильца в БД через нормализацию.
                from app.modules.utility.services.debt_import import normalize_name
                from rapidfuzz import process, fuzz
                norm = normalize_name(fio)
                # Загружаем кэш жильцов.
                users_all = (await db.execute(
                    select(User).where(User.is_deleted.is_(False))
                )).scalars().all()
                user_map = {normalize_name(u.username): u for u in users_all}
                matched_user = user_map.get(norm)
                fuzzy_match_info = None
                if not matched_user:
                    # Fuzzy
                    match = process.extractOne(
                        norm, list(user_map.keys()),
                        scorer=fuzz.token_sort_ratio,
                    )
                    if match:
                        best_key, score, _ = match
                        if score >= 80:
                            matched_user = user_map[best_key]
                            fuzzy_match_info = {"key": best_key, "score": score}
                        else:
                            fuzzy_match_info = {"key": best_key, "score": score, "too_low": True}

                sample["db_lookup"] = None
                if matched_user:
                    # Загружаем текущие debt/over из активного period.
                    active_period = (await db.execute(
                        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
                    )).scalars().first()
                    db_reading = None
                    if active_period:
                        db_reading = (await db.execute(
                            select(MeterReading).where(
                                MeterReading.user_id == matched_user.id,
                                MeterReading.period_id == active_period.id,
                            ).limit(1)
                        )).scalars().first()
                    is_account_209 = log.account_type == "209"
                    sample["db_lookup"] = {
                        "matched_user_id": matched_user.id,
                        "matched_username": matched_user.username,
                        "matched_full_name": matched_user.full_name,
                        "fuzzy": fuzzy_match_info,
                        "db_debt": float(db_reading.debt_209 if is_account_209 else db_reading.debt_205) if db_reading else None,
                        "db_overpayment": float(db_reading.overpayment_209 if is_account_209 else db_reading.overpayment_205) if db_reading else None,
                        "expected_debt": float(debt),
                        "expected_overpayment": float(over),
                        "mismatch": (db_reading is None) or (
                            abs(float(debt) - float(db_reading.debt_209 if is_account_209 else db_reading.debt_205 or 0)) > 0.01
                        ),
                    }
                else:
                    sample["db_lookup"] = {
                        "matched_user_id": None,
                        "fuzzy": fuzzy_match_info,
                        "expected_debt": float(debt),
                        "expected_overpayment": float(over),
                        "mismatch": True,
                        "reason": "user_not_found_in_db",
                    }
            samples.append(sample)
            count += 1

    ws.parent.close()

    return {
        "log_id": log_id,
        "archive_path": log.archive_path,
        "section_markers": section_markers,
        "debit_cols_in_header": debit_cols,
        "credit_cols_in_header": credit_cols,
        "account_total": account_total,
        "chosen": chosen,
        "samples": samples,
    }


@router.post(
    "/debts/probe-update/{user_id}",
    summary="Bug AF: проверить, что UPDATE на reading жильца реально доходит до БД",
)
async def debts_probe_update(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bug AF probe: пробует UPDATE двумя стратегиями (только по `id` —
    как сейчас в импорте; и по `(id, created_at)` — кандидат на фикс
    партиционирования) и показывает rowcount/значение после каждой.

    БЕЗОПАСНО: в конце делает rollback — БД не меняется. Используется
    как «истина в последней инстанции» — если UPDATE по `id` возвращает
    rowcount=0, это партиционная проблема (composite PK + RANGE BY
    created_at). Если rowcount=1, но value не меняется — что-то другое.
    """
    _require_finance(current_user)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")

    reading = (await db.execute(
        select(MeterReading).where(
            MeterReading.user_id == user_id,
            MeterReading.period_id == active_period.id,
        ).limit(1)
    )).scalars().first()
    if not reading:
        raise HTTPException(404, f"Reading для user_id={user_id} в активном периоде не найден")

    from sqlalchemy import update as _sa_update, text as _sa_text

    reading_id = reading.id
    created_at = reading.created_at
    debt_before = float(reading.debt_209 or 0)

    # Стратегия A: UPDATE по id (как сейчас в импорте).
    res_a = await db.execute(
        _sa_update(MeterReading)
        .where(MeterReading.id == reading_id)
        .values(debt_209=0)
    )
    rowcount_a = res_a.rowcount or 0
    val_a = (await db.execute(
        _sa_text("SELECT debt_209 FROM readings WHERE id = :rid AND created_at = :ca"),
        {"rid": reading_id, "ca": created_at},
    )).scalar()

    # Откатываем A перед B — чистый эксперимент.
    await db.rollback()

    # Стратегия B: UPDATE по (id, created_at).
    res_b = await db.execute(
        _sa_update(MeterReading)
        .where(MeterReading.id == reading_id)
        .where(MeterReading.created_at == created_at)
        .values(debt_209=0)
    )
    rowcount_b = res_b.rowcount or 0
    val_b = (await db.execute(
        _sa_text("SELECT debt_209 FROM readings WHERE id = :rid AND created_at = :ca"),
        {"rid": reading_id, "ca": created_at},
    )).scalar()

    # Стратегия C: raw SQL — на случай если ORM что-то ломает.
    await db.rollback()
    res_c = await db.execute(
        _sa_text(
            "UPDATE readings SET debt_209 = 0 "
            "WHERE id = :rid AND created_at = :ca"
        ),
        {"rid": reading_id, "ca": created_at},
    )
    rowcount_c = res_c.rowcount or 0
    val_c = (await db.execute(
        _sa_text("SELECT debt_209 FROM readings WHERE id = :rid AND created_at = :ca"),
        {"rid": reading_id, "ca": created_at},
    )).scalar()

    # Финальный rollback — никаких изменений в БД.
    await db.rollback()

    return {
        "user_id": user_id,
        "reading_id": reading_id,
        "created_at": created_at.isoformat() if created_at else None,
        "debt_209_before": debt_before,
        "strategies": {
            "A_orm_by_id_only": {
                "rowcount": rowcount_a,
                "value_after_in_db": float(val_a) if val_a is not None else None,
                "worked": (rowcount_a == 1 and val_a == 0),
            },
            "B_orm_by_id_and_created_at": {
                "rowcount": rowcount_b,
                "value_after_in_db": float(val_b) if val_b is not None else None,
                "worked": (rowcount_b == 1 and val_b == 0),
            },
            "C_raw_sql_by_id_and_created_at": {
                "rowcount": rowcount_c,
                "value_after_in_db": float(val_c) if val_c is not None else None,
                "worked": (rowcount_c == 1 and val_c == 0),
            },
        },
        "diagnosis": (
            "partitioning_blocks_update" if not (rowcount_a == 1) and rowcount_b == 1
            else "orm_issue" if rowcount_b != 1 and rowcount_c == 1
            else "all_work_check_other_writer" if rowcount_a == 1
            else "all_fail_deeper_problem"
        ),
        "note": "Все стратегии в конце откатываются — БД не изменена.",
    }


@router.get(
    "/debts/integrity-check",
    summary="Анализатор: сравнить applied_state свежего импорта с БД (Этап 2)",
)
async def debts_integrity_check(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Этап 2: проверка целостности долгов в активном периоде.

    Сравнивает что **должно** быть (по applied_state последних 209 и 205
    импортов) с тем, что **есть** в `readings.debt_*`. Три категории
    проблем:

      1) **drift** — applied_state[u] и reading[u] оба есть, но debt
         различается > 1₽. Симптом: что-то перезаписало после импорта
         (manual_receipt, recalc, ручная правка).
      2) **missing_in_db** — applied_state ожидает долг у юзера, а
         reading'а у него вообще нет. Симптом: импорт не дошёл, или
         reading удалён вручную.
      3) **extra_in_db** — reading с долгом есть, в applied_state юзера
         нет. Симптом: zombie от старого Bug AG (см. /debts/zombie-readings).

    Read-only. Auto-fix не делает (на каждую категорию — свой инструмент:
    drift → reparse, missing → reparse, extra → cleanup-zombie-readings).
    """
    _require_finance(current_user)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")
    period_id = active_period.id

    async def _latest_applied(acct: str):
        log = (await db.execute(
            select(DebtImportLog)
            .where(
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
                DebtImportLog.applied_state.is_not(None),
            )
            .order_by(DebtImportLog.id.desc()).limit(1)
        )).scalars().first()
        return log

    log_209 = await _latest_applied("209")
    log_205 = await _latest_applied("205")
    state_209 = (log_209.applied_state or {}) if log_209 else {}
    state_205 = (log_205.applied_state or {}) if log_205 else {}

    # Объединяем expected_state по юзерам: для каждого user_id — ожидаемые debt_209/205.
    expected: dict[int, dict] = {}
    for uid_str, vals in state_209.items():
        try:
            uid = int(uid_str)
        except Exception:
            continue
        expected.setdefault(uid, {"username": None, "room_label": None})
        expected[uid]["debt_209"] = float(vals.get("debt_209", "0") or 0)
        expected[uid]["overpayment_209"] = float(vals.get("overpayment_209", "0") or 0)
        expected[uid]["username"] = vals.get("username")
        expected[uid]["room_label"] = vals.get("room_label")
    for uid_str, vals in state_205.items():
        try:
            uid = int(uid_str)
        except Exception:
            continue
        expected.setdefault(uid, {"username": None, "room_label": None})
        expected[uid]["debt_205"] = float(vals.get("debt_205", "0") or 0)
        expected[uid]["overpayment_205"] = float(vals.get("overpayment_205", "0") or 0)
        if not expected[uid].get("username"):
            expected[uid]["username"] = vals.get("username")
        if not expected[uid].get("room_label"):
            expected[uid]["room_label"] = vals.get("room_label")

    # Все reading'и активного периода (одной выборкой).
    readings = (await db.execute(
        select(MeterReading).where(MeterReading.period_id == period_id)
    )).scalars().all()
    readings_by_user: dict[int, "MeterReading"] = {}
    for r in readings:
        if r.user_id is not None:
            readings_by_user[r.user_id] = r

    drift = []
    missing_in_db = []
    extra_in_db = []

    THR = 1.0  # порог расхождения в рублях

    # 1+2: сверяем expected → reality
    for uid, exp in expected.items():
        r = readings_by_user.get(uid)
        exp_d209 = exp.get("debt_209", 0.0)
        exp_o209 = exp.get("overpayment_209", 0.0)
        exp_d205 = exp.get("debt_205", 0.0)
        exp_o205 = exp.get("overpayment_205", 0.0)
        if r is None:
            # missing — только если ожидалось ненулевое сальдо
            if max(exp_d209, exp_o209, exp_d205, exp_o205) > THR:
                missing_in_db.append({
                    "user_id": uid,
                    "username": exp.get("username"),
                    "room_label": exp.get("room_label"),
                    "expected": {
                        "debt_209": exp_d209, "overpayment_209": exp_o209,
                        "debt_205": exp_d205, "overpayment_205": exp_o205,
                    },
                })
            continue

        actual_d209 = float(r.debt_209 or 0)
        actual_o209 = float(r.overpayment_209 or 0)
        actual_d205 = float(r.debt_205 or 0)
        actual_o205 = float(r.overpayment_205 or 0)

        diff_d209 = actual_d209 - exp_d209
        diff_o209 = actual_o209 - exp_o209
        diff_d205 = actual_d205 - exp_d205
        diff_o205 = actual_o205 - exp_o205
        max_abs_diff = max(abs(diff_d209), abs(diff_o209), abs(diff_d205), abs(diff_o205))
        if max_abs_diff > THR:
            drift.append({
                "user_id": uid,
                "reading_id": r.id,
                "username": exp.get("username"),
                "room_label": exp.get("room_label"),
                "expected": {
                    "debt_209": exp_d209, "overpayment_209": exp_o209,
                    "debt_205": exp_d205, "overpayment_205": exp_o205,
                },
                "actual": {
                    "debt_209": actual_d209, "overpayment_209": actual_o209,
                    "debt_205": actual_d205, "overpayment_205": actual_o205,
                },
                "max_abs_diff": max_abs_diff,
            })

    # 3: reading'и, которых нет в expected (zombie)
    known = set(expected.keys())
    user_ids_for_rooms = set()
    for r in readings:
        if r.user_id is None:
            continue
        if r.user_id in known:
            continue
        has_money = (
            float(r.debt_209 or 0) > THR or float(r.debt_205 or 0) > THR
            or float(r.overpayment_209 or 0) > THR or float(r.overpayment_205 or 0) > THR
        )
        if not has_money:
            continue
        user_ids_for_rooms.add(r.user_id)

    # Загружаем username/комнаты для zombie batch'ем
    extra_users_map = {}
    extra_rooms_map = {}
    if user_ids_for_rooms:
        u_rows = (await db.execute(
            select(User).where(User.id.in_(user_ids_for_rooms))
        )).scalars().all()
        extra_users_map = {u.id: u for u in u_rows}
        rids = {r.room_id for r in readings if r.user_id in user_ids_for_rooms and r.room_id}
        if rids:
            r_rows = (await db.execute(
                select(Room).where(Room.id.in_(rids))
            )).scalars().all()
            extra_rooms_map = {rm.id: rm for rm in r_rows}

    for r in readings:
        if r.user_id is None or r.user_id in known:
            continue
        has_money = (
            float(r.debt_209 or 0) > THR or float(r.debt_205 or 0) > THR
            or float(r.overpayment_209 or 0) > THR or float(r.overpayment_205 or 0) > THR
        )
        if not has_money:
            continue
        u = extra_users_map.get(r.user_id)
        rm = extra_rooms_map.get(r.room_id) if r.room_id else None
        extra_in_db.append({
            "user_id": r.user_id,
            "reading_id": r.id,
            "username": u.username if u else None,
            "room_label": (rm.format_address if rm else None),
            "actual": {
                "debt_209": float(r.debt_209 or 0),
                "overpayment_209": float(r.overpayment_209 or 0),
                "debt_205": float(r.debt_205 or 0),
                "overpayment_205": float(r.overpayment_205 or 0),
            },
        })

    drift.sort(key=lambda x: -x["max_abs_diff"])
    extra_in_db.sort(
        key=lambda x: -(
            x["actual"]["debt_209"] + x["actual"]["debt_205"]
            + x["actual"]["overpayment_209"] + x["actual"]["overpayment_205"]
        )
    )

    return {
        "period_id": period_id,
        "threshold_rub": THR,
        "latest_209_log_id": log_209.id if log_209 else None,
        "latest_205_log_id": log_205.id if log_205 else None,
        "summary": {
            "drift_count": len(drift),
            "missing_in_db_count": len(missing_in_db),
            "extra_in_db_count": len(extra_in_db),
            "expected_users": len(expected),
            "actual_readings": len(readings_by_user),
        },
        "drift": drift[:200],
        "missing_in_db": missing_in_db[:200],
        "extra_in_db": extra_in_db[:200],
    }


@router.post(
    "/debts/integrity-fix",
    summary="Авто-фикс расхождений integrity-check (Bug AK)",
)
async def debts_integrity_fix(
    category: str = Query("all", pattern="^(all|drift|missing|user)$"),
    user_id: Optional[int] = None,
    confirm: str = Query(..., pattern="^YES$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Применяет ожидаемые значения из applied_state свежих 209/205-импортов
    в БД. Покрывает категории:

      - **drift**: UPDATE существующих reading'ов до значений из applied_state
        (когда импорт правильно посчитал, но что-то после перезаписало).
      - **missing**: INSERT недостающих reading'ов из applied_state
        (когда жилец есть в файле, а в БД его reading нет).
      - **all**: drift + missing вместе.
      - **user**: фикс только для конкретного user_id (точечно).

    Extra/Zombie фиксится отдельным endpoint'ом /debts/cleanup-zombie-readings —
    у них нет «ожидаемого значения», только зануление.

    Требует ?confirm=YES.
    """
    _require_finance(current_user)

    # Реюзаем диагностику чтобы не дублировать логику расчёта расхождений.
    data = await debts_integrity_check(current_user=current_user, db=db)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")

    from sqlalchemy import update as _sa_update

    fixed_drift = 0
    fixed_missing = 0
    errors = []

    drift_items = data.get("drift", [])
    missing_items = data.get("missing_in_db", [])

    # Фильтр по user_id если category=user
    if category == "user":
        if not user_id:
            raise HTTPException(400, "category=user требует user_id")
        drift_items = [d for d in drift_items if d.get("user_id") == user_id]
        missing_items = [m for m in missing_items if m.get("user_id") == user_id]

    # 1) drift — UPDATE существующих reading'ов
    if category in ("all", "drift", "user"):
        for item in drift_items:
            try:
                res = await db.execute(
                    _sa_update(MeterReading)
                    .where(MeterReading.id == item["reading_id"])
                    .values(
                        debt_209=Decimal(str(item["expected"]["debt_209"])),
                        overpayment_209=Decimal(str(item["expected"]["overpayment_209"])),
                        debt_205=Decimal(str(item["expected"]["debt_205"])),
                        overpayment_205=Decimal(str(item["expected"]["overpayment_205"])),
                    )
                )
                if res.rowcount:
                    fixed_drift += 1
            except Exception as e:
                errors.append({
                    "kind": "drift",
                    "user_id": item.get("user_id"),
                    "error": str(e)[:200],
                })

    # 2) missing — INSERT недостающих reading'ов из applied_state
    if category in ("all", "missing", "user"):
        for item in missing_items:
            try:
                user = await db.get(User, item["user_id"])
                if not user:
                    errors.append({
                        "kind": "missing",
                        "user_id": item.get("user_id"),
                        "error": "user не найден в БД",
                    })
                    continue
                new_reading = MeterReading(
                    user_id=item["user_id"],
                    room_id=user.room_id,
                    period_id=active_period.id,
                    is_approved=False,
                    debt_209=Decimal(str(item["expected"]["debt_209"])),
                    overpayment_209=Decimal(str(item["expected"]["overpayment_209"])),
                    debt_205=Decimal(str(item["expected"]["debt_205"])),
                    overpayment_205=Decimal(str(item["expected"]["overpayment_205"])),
                    obor_debit_209=Decimal("0"), obor_credit_209=Decimal("0"),
                    obor_debit_205=Decimal("0"), obor_credit_205=Decimal("0"),
                )
                db.add(new_reading)
                fixed_missing += 1
            except Exception as e:
                errors.append({
                    "kind": "missing",
                    "user_id": item.get("user_id"),
                    "error": str(e)[:200],
                })

    await db.commit()
    logger.info(
        "[INTEGRITY-FIX] category=%s drift=%d missing=%d errors=%d (by %s)",
        category, fixed_drift, fixed_missing, len(errors), current_user.username,
    )

    return {
        "status": "ok",
        "category": category,
        "fixed_drift": fixed_drift,
        "fixed_missing": fixed_missing,
        "errors": errors[:50],
    }


@router.get(
    "/debts/zombie-readings",
    summary="Reading'и с долгом, которых нет в свежем импорте (Этап 3)",
)
async def debts_zombie_readings(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bug AG cleanup: после переключения импорта на per-user-key (Bug AG)
    в БД могут остаться reading'и с ненулевыми debt_*/overpayment_*, которых
    в свежем impart'е 1С уже нет (т.е. в файле от 1С этого жильца не передавали,
    значит долг должен быть 0). Раньше эти суммы суммировались в общий
    reading комнаты — после Bug AG они становятся «висяком» на чужом юзере.

    Логика: смотрим последние completed-импорты 209 и 205, собираем все
    user_id из их applied_state. Все reading'и активного периода с
    долгом/переплатой, чей user_id НЕ упомянут ни в одном из этих логов —
    кандидаты на zombie.

    Read-only. POST /debts/cleanup-zombie-readings занулит их (с
    подтверждением).
    """
    _require_finance(current_user)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")
    period_id = active_period.id

    async def _latest_applied(acct: str):
        log = (await db.execute(
            select(DebtImportLog)
            .where(
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
                DebtImportLog.applied_state.is_not(None),
            )
            .order_by(DebtImportLog.id.desc()).limit(1)
        )).scalars().first()
        return log

    log_209 = await _latest_applied("209")
    log_205 = await _latest_applied("205")
    state_209 = (log_209.applied_state or {}) if log_209 else {}
    state_205 = (log_205.applied_state or {}) if log_205 else {}
    known_user_ids = set(state_209.keys()) | set(state_205.keys())

    if not known_user_ids:
        return {
            "period_id": period_id,
            "count": 0,
            "zombies": [],
            "note": "Нет свежих импортов с applied_state — нечего сравнивать.",
        }

    readings = (await db.execute(
        select(MeterReading)
        .where(
            MeterReading.period_id == period_id,
            or_(
                MeterReading.debt_209 > 0,
                MeterReading.debt_205 > 0,
                MeterReading.overpayment_209 > 0,
                MeterReading.overpayment_205 > 0,
            ),
        )
    )).scalars().all()

    user_ids_in_db = {r.user_id for r in readings if r.user_id}
    if not user_ids_in_db:
        return {"period_id": period_id, "count": 0, "zombies": []}

    users = (await db.execute(
        select(User).where(User.id.in_(user_ids_in_db))
    )).scalars().all()
    users_map = {u.id: u for u in users}

    room_ids = {r.room_id for r in readings if r.room_id}
    rooms_map = {}
    if room_ids:
        rooms = (await db.execute(
            select(Room).where(Room.id.in_(room_ids))
        )).scalars().all()
        rooms_map = {r.id: r for r in rooms}

    zombies = []
    for r in readings:
        if not r.user_id:
            continue
        if str(r.user_id) in known_user_ids:
            continue  # есть в свежем импорте — не зомби
        user = users_map.get(r.user_id)
        room = rooms_map.get(r.room_id) if r.room_id else None
        zombies.append({
            "reading_id": r.id,
            "user_id": r.user_id,
            "username": user.username if user else None,
            "room_id": r.room_id,
            "room_label": (
                room.format_address if room else None
            ),
            "debt_209": float(r.debt_209 or 0),
            "overpayment_209": float(r.overpayment_209 or 0),
            "debt_205": float(r.debt_205 or 0),
            "overpayment_205": float(r.overpayment_205 or 0),
            "total_to_clean": (
                float(r.debt_209 or 0) + float(r.debt_205 or 0)
                + float(r.overpayment_209 or 0) + float(r.overpayment_205 or 0)
            ),
        })

    zombies.sort(key=lambda x: -x["total_to_clean"])
    return {
        "period_id": period_id,
        "latest_209_log_id": log_209.id if log_209 else None,
        "latest_205_log_id": log_205.id if log_205 else None,
        "count": len(zombies),
        "zombies": zombies,
    }


@router.post(
    "/debts/cleanup-zombie-readings",
    summary="Занулить debt_*/overpayment_* у zombie-reading'ов (Этап 3)",
)
async def debts_cleanup_zombie_readings(
    confirm: str = Query(..., pattern="^YES$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Зануляет debt_209/205 и overpayment_209/205 у всех reading'ов,
    которые попали в список /debts/zombie-readings.

    Reading'и НЕ удаляются (audit/история сохраняется) — только зануляются
    финансовые поля. После этого дашборд показывает 0₽ у соответствующих
    жильцов.

    Требует ?confirm=YES.
    """
    _require_finance(current_user)

    # Реюзаем логику /debts/zombie-readings — чтобы не дублировать.
    result = await debts_zombie_readings(current_user=current_user, db=db)
    zombies = result.get("zombies", [])
    if not zombies:
        return {"status": "ok", "cleaned": 0, "note": "Zombie-reading'ов нет"}

    reading_ids = [z["reading_id"] for z in zombies]
    from sqlalchemy import update as _sa_update
    total = 0
    for rid in reading_ids:
        res = await db.execute(
            _sa_update(MeterReading)
            .where(MeterReading.id == rid)
            .values(
                debt_209=Decimal("0.00"),
                overpayment_209=Decimal("0.00"),
                debt_205=Decimal("0.00"),
                overpayment_205=Decimal("0.00"),
            )
        )
        total += res.rowcount or 0

    await db.commit()
    logger.info(
        "[ZOMBIE-CLEANUP] %s reading-ов занулено (запросил %s)",
        total, current_user.username,
    )
    return {
        "status": "ok",
        "cleaned": total,
        "requested": len(reading_ids),
        "zombies": zombies[:50],
    }


@router.get(
    "/debts/orphan-readings",
    summary="Жильцы с >1 reading в активном периоде (диагностика для Bug AF)",
)
async def debts_orphan_readings(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bug AF: после переездов / auto-Vacant может оказаться, что у одного
    user_id в активном периоде существует несколько MeterReading с разными
    room_id. Дашборд агрегирует SUM(debt_*) по user_id, а импорт 1С
    обновляет reading по room_id. Если осиротевший reading с прошлой
    комнатой не зачищен — его debt_209 продолжает суммироваться в общий,
    и reparse его не чинит (импорт обновляет только текущий room_id).

    Read-only endpoint: возвращает список жильцов с >1 reading + раскладку
    каждого reading'а (room_id, debt, current/orphan). По нему уже решаем,
    что чистить — отдельный POST-endpoint.
    """
    _require_finance(current_user)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")
    period_id = active_period.id

    dup_q = (
        select(MeterReading.user_id, func.count(MeterReading.id).label("cnt"))
        .where(
            MeterReading.period_id == period_id,
            MeterReading.user_id.is_not(None),
        )
        .group_by(MeterReading.user_id)
        .having(func.count(MeterReading.id) > 1)
    )
    dup_rows = (await db.execute(dup_q)).all()
    user_ids = [r.user_id for r in dup_rows]
    if not user_ids:
        return {"period_id": period_id, "count": 0, "users": []}

    readings = (await db.execute(
        select(MeterReading)
        .where(
            MeterReading.period_id == period_id,
            MeterReading.user_id.in_(user_ids),
        )
        .order_by(MeterReading.user_id, MeterReading.created_at)
    )).scalars().all()

    users = (await db.execute(
        select(User).where(User.id.in_(user_ids))
    )).scalars().all()
    users_map = {u.id: u for u in users}

    room_ids = {r.room_id for r in readings if r.room_id}
    room_ids.update(u.room_id for u in users if u.room_id)
    rooms_map = {}
    if room_ids:
        rooms = (await db.execute(
            select(Room).where(Room.id.in_(room_ids))
        )).scalars().all()
        rooms_map = {r.id: r for r in rooms}

    def _room_label(rid):
        if rid is None:
            return None
        r = rooms_map.get(rid)
        return r.format_address if r else f"id={rid}"

    by_user: dict[int, list] = {}
    for r in readings:
        by_user.setdefault(r.user_id, []).append(r)

    items = []
    for uid, urs in by_user.items():
        user = users_map.get(uid)
        cur_room_id = user.room_id if user else None
        orphan_debt = sum(
            float(r.debt_209 or 0) + float(r.debt_205 or 0)
            for r in urs if r.room_id != cur_room_id
        )
        items.append({
            "user_id": uid,
            "username": user.username if user else None,
            "current_room_id": cur_room_id,
            "current_room_label": _room_label(cur_room_id),
            "total_debt_209": sum(float(r.debt_209 or 0) for r in urs),
            "total_debt_205": sum(float(r.debt_205 or 0) for r in urs),
            "orphan_debt_sum": orphan_debt,
            "readings": [
                {
                    "id": r.id,
                    "room_id": r.room_id,
                    "room_label": _room_label(r.room_id) or "(нет комнаты)",
                    "is_current_room": (r.room_id == cur_room_id),
                    "debt_209": float(r.debt_209 or 0),
                    "overpayment_209": float(r.overpayment_209 or 0),
                    "debt_205": float(r.debt_205 or 0),
                    "overpayment_205": float(r.overpayment_205 or 0),
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                    "is_approved": r.is_approved,
                }
                for r in urs
            ],
        })

    # Сортировка: сначала те, у кого больше всего «осиротевших» денег.
    items.sort(key=lambda x: -x["orphan_debt_sum"])

    return {
        "period_id": period_id,
        "count": len(items),
        "users": items,
    }
