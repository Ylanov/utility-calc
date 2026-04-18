# app/modules/utility/routers/admin_initial_readings.py

"""
Загрузка начальных показаний счётчиков.

ПРОБЛЕМА: При запуске проекта все счётчики начинаются с 0.
Когда жилец передаёт первые реальные показания (например ГВС=150.5),
система считает потребление как 150.5 - 0 = 150.5 м³ → огромный счёт.

РЕШЕНИЕ: Администратор загружает начальные показания (baseline) для каждой комнаты.
Они сохраняются как утверждённая запись MeterReading с флагом INITIAL_SETUP.
После этого первый расчёт жильца будет: 151.2 - 150.5 = 0.7 м³ → нормальный счёт.
"""

import io
import asyncio
import logging
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core.database import get_db
from app.modules.utility.models import User, Room, MeterReading, BillingPeriod
from app.core.dependencies import RoleChecker

router = APIRouter(prefix="/api/rooms", tags=["Initial Readings"])
logger = logging.getLogger(__name__)

ZERO = Decimal("0.000")
allow_management = RoleChecker(["accountant", "admin"])


# =====================================================================
# УСТАНОВКА НАЧАЛЬНЫХ ПОКАЗАНИЙ ДЛЯ ОДНОЙ КОМНАТЫ
# =====================================================================
@router.post("/{room_id}/initial-readings", summary="Установить начальные показания счётчиков")
async def set_initial_readings(
    room_id: int,
    hot_water: Decimal = Query(..., ge=0, description="Текущее показание ГВС"),
    cold_water: Decimal = Query(..., ge=0, description="Текущее показание ХВС"),
    electricity: Decimal = Query(..., ge=0, description="Текущее показание электричества"),
    current_user: User = Depends(allow_management),
    db: AsyncSession = Depends(get_db)
):
    """
    Устанавливает начальные (базовые) показания счётчиков для комнаты.

    Создаёт утверждённую запись MeterReading с флагом INITIAL_SETUP.
    Обновляет кэш комнаты (last_hot_water и т.д.).

    Если начальные показания уже были установлены — перезаписывает их.
    """
    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(status_code=404, detail="Комната не найдена")

    # Ищем существующую начальную запись для этой комнаты
    existing = (await db.execute(
        select(MeterReading).where(
            MeterReading.room_id == room_id,
            MeterReading.anomaly_flags == "INITIAL_SETUP"
        )
    )).scalars().first()

    if existing:
        # Обновляем существующую
        existing.hot_water = hot_water
        existing.cold_water = cold_water
        existing.electricity = electricity
        db.add(existing)
    else:
        # Создаём новую
        # Находим любого жильца комнаты для user_id (или NULL)
        user_result = await db.execute(
            select(User.id).where(
                User.room_id == room_id,
                User.is_deleted.is_(False)
            ).limit(1)
        )
        user_id = user_result.scalar_one_or_none()

        reading = MeterReading(
            room_id=room_id,
            user_id=user_id,
            period_id=None,  # Начальные показания не привязаны к периоду
            hot_water=hot_water,
            cold_water=cold_water,
            electricity=electricity,
            is_approved=True,  # Сразу утверждено — это baseline
            anomaly_flags="INITIAL_SETUP",
            anomaly_score=0,
            total_cost=ZERO,
            total_209=ZERO,
            total_205=ZERO,
        )
        db.add(reading)

    # Обновляем кэш комнаты
    room.last_hot_water = hot_water
    room.last_cold_water = cold_water
    room.last_electricity = electricity
    db.add(room)

    await db.commit()

    logger.info(
        f"Initial readings set for room {room_id}: "
        f"hot={hot_water}, cold={cold_water}, elect={electricity} "
        f"by {current_user.username}"
    )

    return {
        "status": "success",
        "room_id": room_id,
        "hot_water": str(hot_water),
        "cold_water": str(cold_water),
        "electricity": str(electricity),
    }


# =====================================================================
# ПОЛУЧЕНИЕ ТЕКУЩИХ БАЗОВЫХ ПОКАЗАНИЙ КОМНАТЫ
# =====================================================================
@router.get("/{room_id}/current-readings", summary="Текущие показания счётчиков комнаты")
async def get_current_readings(
    room_id: int,
    current_user: User = Depends(allow_management),
    db: AsyncSession = Depends(get_db)
):
    """Возвращает последние известные показания комнаты (из кэша или из истории)."""
    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(status_code=404, detail="Комната не найдена")

    return {
        "room_id": room_id,
        "dormitory_name": room.dormitory_name,
        "room_number": room.room_number,
        "hot_water": str(room.last_hot_water or ZERO),
        "cold_water": str(room.last_cold_water or ZERO),
        "electricity": str(room.last_electricity or ZERO),
    }


# =====================================================================
# МАССОВАЯ ЗАГРУЗКА НАЧАЛЬНЫХ ПОКАЗАНИЙ ИЗ EXCEL
# =====================================================================
@router.post("/import-initial-readings", summary="Массовая загрузка начальных показаний из Excel")
async def import_initial_readings(
    file: UploadFile = File(...),
    current_user: User = Depends(allow_management),
    db: AsyncSession = Depends(get_db)
):
    """
    Загружает начальные показания счётчиков из Excel-файла.

    Формат файла (колонки):
    1. Общежитие (название)
    2. Номер комнаты
    3. ГВС (текущее показание)
    4. ХВС (текущее показание)
    5. Электричество (текущее показание)

    Система находит комнату по общежитию + номеру и устанавливает показания.
    Если комната не найдена — строка пропускается с ошибкой.
    """
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Поддерживаются только файлы Excel (.xlsx)")

    content = await file.read()

    try:
        wb = await asyncio.to_thread(
            lambda: load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        )
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать файл. Проверьте формат.")

    ws = wb.active

    # ============================================================
    # Шаг 1: Сначала пробегаем по Excel и собираем все уникальные
    # (dormitory, room_number) — это позволит одним запросом достать
    # только нужные комнаты, не загружая всю таблицу (~10к записей) в RAM.
    # ============================================================
    raw_rows = []
    room_keys = set()
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        dormitory = str(row[0]).strip() if row[0] else None
        room_number = str(row[1]).strip() if len(row) > 1 and row[1] else None
        if dormitory and room_number:
            room_keys.add((dormitory, room_number))
        raw_rows.append((row_idx, row, dormitory, room_number))

    # ============================================================
    # Шаг 2: batch-запрос только тех комнат, что есть в файле.
    # WHERE (dormitory_name, room_number) IN (...) — один проход по индексу.
    # ============================================================
    rooms_map: dict[tuple[str, str], Room] = {}
    if room_keys:
        rooms_list = list(room_keys)
        # tuple_() для PostgreSQL IN по составному ключу
        from sqlalchemy import tuple_ as _tuple
        query = select(Room).where(
            _tuple(Room.dormitory_name, Room.room_number).in_(rooms_list)
        )
        rooms_result = await db.execute(query)
        for r in rooms_result.scalars():
            rooms_map[(r.dormitory_name, r.room_number)] = r

    # ============================================================
    # Шаг 3: batch-запрос существующих INITIAL_SETUP readings
    # для всех комнат сразу — один запрос вместо N.
    # ============================================================
    existing_readings: dict[int, MeterReading] = {}
    room_ids = [r.id for r in rooms_map.values()]
    if room_ids:
        existing_result = await db.execute(
            select(MeterReading).where(
                MeterReading.room_id.in_(room_ids),
                MeterReading.anomaly_flags == "INITIAL_SETUP"
            )
        )
        for mr in existing_result.scalars():
            existing_readings[mr.room_id] = mr

    # ============================================================
    # Шаг 4: batch-запрос первого юзера для каждой комнаты.
    # DISTINCT ON: первый не-удалённый юзер на комнату.
    # ============================================================
    users_map: dict[int, int] = {}  # room_id -> user_id
    if room_ids:
        users_result = await db.execute(
            select(User.id, User.room_id)
            .where(User.room_id.in_(room_ids), User.is_deleted.is_(False))
            .order_by(User.room_id, User.id)
        )
        for uid, rid in users_result.all():
            users_map.setdefault(rid, uid)

    updated = 0
    skipped = 0
    errors = []

    # ============================================================
    # Шаг 5: основной проход без БД-запросов — только ORM-операции.
    # ============================================================
    for row_idx, row, dormitory, room_number in raw_rows:
        try:
            if not dormitory or not room_number:
                errors.append(f"Строка {row_idx}: не указано общежитие или номер комнаты")
                skipped += 1
                continue

            room = rooms_map.get((dormitory, room_number))
            if not room:
                errors.append(f"Строка {row_idx}: комната '{dormitory}, {room_number}' не найдена в базе")
                skipped += 1
                continue

            hot = Decimal(str(row[2]).replace(',', '.')) if len(row) > 2 and row[2] else ZERO
            cold = Decimal(str(row[3]).replace(',', '.')) if len(row) > 3 and row[3] else ZERO
            elect = Decimal(str(row[4]).replace(',', '.')) if len(row) > 4 and row[4] else ZERO

            if hot < 0 or cold < 0 or elect < 0:
                errors.append(f"Строка {row_idx}: отрицательные показания")
                skipped += 1
                continue

            existing = existing_readings.get(room.id)
            if existing:
                existing.hot_water = hot
                existing.cold_water = cold
                existing.electricity = elect
            else:
                db.add(MeterReading(
                    room_id=room.id,
                    user_id=users_map.get(room.id),
                    period_id=None,
                    hot_water=hot, cold_water=cold, electricity=elect,
                    is_approved=True, anomaly_flags="INITIAL_SETUP", anomaly_score=0,
                    total_cost=ZERO, total_209=ZERO, total_205=ZERO,
                ))

            # Обновляем кэш комнаты
            room.last_hot_water = hot
            room.last_cold_water = cold
            room.last_electricity = elect

            updated += 1

        except Exception as e:
            errors.append(f"Строка {row_idx}: ошибка обработки — {str(e)}")
            skipped += 1

    await db.commit()

    logger.info(f"Initial readings import: updated={updated}, skipped={skipped} by {current_user.username}")

    return {
        "status": "success",
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],  # Ограничиваем количество ошибок в ответе
    }


# =====================================================================
# СКАЧАТЬ ШАБЛОН EXCEL ДЛЯ ЗАГРУЗКИ
# =====================================================================
@router.get("/initial-readings/template", summary="Скачать шаблон Excel для начальных показаний")
async def download_initial_readings_template(
    current_user: User = Depends(allow_management),
    db: AsyncSession = Depends(get_db)
):
    """
    Генерирует Excel-шаблон, предзаполненный списком всех комнат.
    Администратор заполняет колонки показаний и загружает обратно.
    """
    rooms_result = await db.execute(
        select(Room).order_by(Room.dormitory_name, Room.room_number)
    )
    rooms = rooms_result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Начальные показания"

    headers = ["Общежитие", "Номер комнаты", "ГВС (м³)", "ХВС (м³)", "Электричество (кВт·ч)"]
    ws.append(headers)

    # Стиль заголовков
    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 22

    # Заполняем комнатами (показания пустые — для заполнения вручную)
    for room in rooms:
        current_hot = float(room.last_hot_water or 0)
        current_cold = float(room.last_cold_water or 0)
        current_elect = float(room.last_electricity or 0)

        # Если показания уже есть — подставляем их, иначе пусто
        ws.append([
            room.dormitory_name,
            room.room_number,
            current_hot if current_hot > 0 else None,
            current_cold if current_cold > 0 else None,
            current_elect if current_elect > 0 else None,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Initial_Readings_Template.xlsx"}
    )