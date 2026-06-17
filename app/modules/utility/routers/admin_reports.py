# app/modules/utility/routers/admin_reports.py

import io
import os
import shutil
import tempfile
import uuid
import asyncio
import logging
from starlette.background import BackgroundTask
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from celery.result import AsyncResult
from openpyxl import Workbook
from decimal import Decimal
from urllib.parse import quote

from app.core.database import get_db
from app.modules.utility.models import User, MeterReading, Tariff, BillingPeriod, Adjustment, Room, RentalContract
from app.core.dependencies import get_current_user
from app.modules.utility.services.pdf_generator import generate_receipt_pdf
from app.modules.utility.services.s3_client import s3_service
from app.modules.utility.services.period_helpers import period_chron_key
from app.modules.utility.tasks import generate_receipt_task, start_bulk_receipt_generation

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Admin Reports"])
ZERO = Decimal("0.00")


@router.get("/api/admin/receipts/{reading_id}")
async def get_receipt_pdf(
        reading_id: int,
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db)
):
    if current_user.role not in ("accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    stmt = (
        select(MeterReading)
        .options(selectinload(MeterReading.user).selectinload(User.room), selectinload(MeterReading.period))
        .where(MeterReading.id == reading_id)
    )
    reading = (await db.execute(stmt)).scalars().first()

    if not reading or not reading.user or not reading.period or not reading.user.room:
        raise HTTPException(404, "Данные не найдены или жилец не привязан к помещению")

    user, room = reading.user, reading.user.room

    # Тариф через единый сервис: Room.tariff_id → User.tariff_id → default.
    # tariff_cache использует in-memory кеш и учитывает приоритет комнатной привязки.
    from app.modules.utility.services.tariff_cache import tariff_cache
    tariff = tariff_cache.get_effective_tariff(user=user, room=room)
    if not tariff:
        tariff = (await db.execute(select(Tariff).where(Tariff.is_active))).scalars().first()
        if not tariff:
            raise HTTPException(404, "Активный тариф не найден")

    prev = (await db.execute(
        select(MeterReading)
        .where(MeterReading.room_id == room.id, MeterReading.is_approved.is_(True),
               MeterReading.created_at < reading.created_at)
        .order_by(MeterReading.created_at.desc()).limit(1)
    )).scalars().first()

    adjustments = (await db.execute(
        select(Adjustment).where(Adjustment.user_id == user.id, Adjustment.period_id == reading.period_id)
    )).scalars().all()

    # tempfile.TemporaryDirectory вместо хардкода "/tmp" — Sonar
    # python:S5443. Изолированная dir с правами 700, удаление при выходе
    # автоматическое (включая случай exception). Возврат JSON, не
    # FileResponse, поэтому файл уже не нужен после return — БЕЗОПАСНО
    # очистить здесь же.
    try:
        with tempfile.TemporaryDirectory(prefix="utility_pdf_") as output_dir:
            pdf_path = await asyncio.to_thread(
                generate_receipt_pdf,
                user=user, room=room, reading=reading, period=reading.period,
                tariff=tariff, prev_reading=prev, adjustments=adjustments, output_dir=output_dir
            )
            s3_key = f"receipts/{reading.period.id}/admin_view_{user.id}_{uuid.uuid4().hex[:8]}.pdf"

            if await asyncio.to_thread(s3_service.upload_file, pdf_path, s3_key):
                # TemporaryDirectory сам удалит локальную копию, явный os.remove не нужен.
                download_url = await asyncio.to_thread(s3_service.get_presigned_url, s3_key, 300)
                return {"status": "success", "url": download_url}
            else:
                # S3 недоступен — копируем PDF в статику. Используем copy
                # (не move): TemporaryDirectory корректно удалит исходник.
                static_dir = "/app/static/generated_files"
                await asyncio.to_thread(os.makedirs, static_dir, exist_ok=True)
                filename = os.path.basename(pdf_path)
                static_path = os.path.join(static_dir, filename)
                await asyncio.to_thread(shutil.copy, pdf_path, static_path)
                return {"status": "success", "url": f"/generated_files/{filename}"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Ошибка генерации PDF: {e}")


@router.get(
    "/api/admin/receipts/{reading_id}/download",
    summary="Скачать PDF квитанции (streaming, без S3)",
)
async def stream_admin_receipt(
        reading_id: int,
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db),
):
    """
    Стримит PDF квитанции конкретного жильца напрямую через FastAPI с правильными
    заголовками Content-Disposition.

    Этот эндпоинт заменяет старый двухшаговый процесс (GET JSON с url → window.open),
    в котором происходил редирект на portal.html при сбое S3/протухшем токене.
    Вызывается фронтендом через api.download(...) — всё идёт одним авторизованным
    запросом, без посредников.
    """
    if current_user.role not in ("accountant", "admin", "financier"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    stmt = (
        select(MeterReading)
        .options(
            selectinload(MeterReading.user).selectinload(User.room),
            selectinload(MeterReading.period),
        )
        .where(MeterReading.id == reading_id)
    )
    reading = (await db.execute(stmt)).scalars().first()

    if not reading or not reading.user or not reading.period or not reading.user.room:
        raise HTTPException(404, "Данные не найдены или жилец не привязан к помещению")

    user, room = reading.user, reading.user.room

    # Тариф через единый сервис (Room.tariff_id → User.tariff_id → default).
    from app.modules.utility.services.tariff_cache import tariff_cache
    tariff = tariff_cache.get_effective_tariff(user=user, room=room)
    if not tariff:
        tariff = (await db.execute(select(Tariff).where(Tariff.is_active))).scalars().first()
        if not tariff:
            raise HTTPException(404, "Активный тариф не найден")

    prev = (await db.execute(
        select(MeterReading).where(
            MeterReading.room_id == room.id,
            MeterReading.is_approved.is_(True),
            MeterReading.created_at < reading.created_at,
        ).order_by(MeterReading.created_at.desc()).limit(1)
    )).scalars().first()

    adjustments = (await db.execute(
        select(Adjustment).where(
            Adjustment.user_id == user.id,
            Adjustment.period_id == reading.period_id,
        )
    )).scalars().all()

    # mkdtemp + BackgroundTask вместо "/tmp": Sonar python:S5443 +
    # попутно фикс утечки PDF в /tmp (раньше после стрима файл не
    # удалялся). FileResponse читает файл АСИНХРОННО уже после возврата
    # из эндпоинта, поэтому with TemporaryDirectory здесь не подходит —
    # директория удалилась бы до того, как клиент успел скачать файл.
    output_dir = tempfile.mkdtemp(prefix="utility_pdf_")
    try:
        pdf_path = await asyncio.to_thread(
            generate_receipt_pdf,
            user=user, room=room, reading=reading, period=reading.period,
            tariff=tariff, prev_reading=prev, adjustments=adjustments, output_dir=output_dir,
        )
    except Exception as e:
        shutil.rmtree(output_dir, ignore_errors=True)
        logger.error(f"PDF generation failed for reading_id={reading_id}: {e}", exc_info=True)
        raise HTTPException(500, "Ошибка генерации квитанции. Попробуйте позже.")

    if not os.path.exists(pdf_path):
        shutil.rmtree(output_dir, ignore_errors=True)
        raise HTTPException(500, "Не удалось получить файл квитанции на сервере")

    username_safe = (user.username.split("_deleted_")[0] if user.is_deleted else user.username)
    username_safe = username_safe.replace(" ", "_")
    room_label = (room.room_number or "room").replace(" ", "_")
    period_label = (reading.period.name or "period").replace(" ", "_")
    filename = f"Kvitanciya_{room_label}_{username_safe}_{period_label}.pdf"
    encoded_filename = quote(filename)

    return FileResponse(
        path=pdf_path,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                f"attachment; filename*=utf-8''{encoded_filename}",
            "Cache-Control": "no-store, must-revalidate",
        },
        # Cleanup срабатывает после того как Starlette отстримит весь файл клиенту.
        background=BackgroundTask(shutil.rmtree, output_dir, ignore_errors=True),
    )


@router.get("/api/admin/export_report", summary="Скачать отчет Excel (XLSX)")
async def export_report(
        period_id: Optional[int] = Query(None, description="ID периода"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db)
):
    if current_user.role not in ("accountant", "admin"):
        raise HTTPException(status_code=403)

    target_period_id = period_id
    if not target_period_id:
        active_period = (
            await db.execute(select(BillingPeriod).where(BillingPeriod.is_active.is_(True)))).scalars().first()
        if active_period:
            target_period_id = active_period.id
        else:
            last_closed = (
                await db.execute(select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1))).scalars().first()
            if not last_closed:
                raise HTTPException(404, "Нет периодов для отчета")
            target_period_id = last_closed.id

    period = await db.get(BillingPeriod, target_period_id)
    if not period: raise HTTPException(404, "Выбранный период не найден")

    # ИСПРАВЛЕНИЕ: Используем потоковое чтение (yield_per) из БД, чтобы не грузить все в RAM
    statement = (
        select(User, MeterReading, Room)
        .join(MeterReading, User.id == MeterReading.user_id)
        .join(Room, User.room_id == Room.id)
        .where(
            MeterReading.period_id == target_period_id,
            MeterReading.is_approved.is_(True)
        )
        .order_by(Room.dormitory_name, Room.room_number, User.username)
    ).execution_options(yield_per=1000)

    # ИСПРАВЛЕНИЕ: Используем write_only=True для потоковой записи Excel
    # (openpyxl не хранит документ в памяти, а пишет напрямую в байтовый буфер)
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Сводная ведомость")

    headers = ["Общежитие/Комната", "ФИО (Логин)", "Площадь", "Жильцов", "ГВС (руб)", "ХВС (руб)", "Водоотв. (руб)",
               "Электроэнергия (руб)", "Содержание (руб)", "Наем (руб)", "ТКО (руб)", "Отопление + ОДН (руб)",
               "Счет 209 (Комм.)", "Счет 205 (Найм)", "ИТОГО (руб)"]
    worksheet.append(headers)

    total_sum, total_209, total_205 = ZERO, ZERO, ZERO

    # Потоковое получение результатов
    result = await db.stream(statement)

    async for row in result:
        user, reading, room = row
        total_cost = Decimal(reading.total_cost or 0)
        t_209 = Decimal(reading.total_209 or 0)
        t_205 = Decimal(reading.total_205 or 0)

        total_sum += total_cost
        total_209 += t_209
        total_205 += t_205

        worksheet.append([
            room.format_address,
            user.username.split("_deleted_")[0] if user.is_deleted else user.username,
            room.apartment_area,
            f"{user.residents_count}/{room.total_room_residents}",
            reading.cost_hot_water, reading.cost_cold_water, reading.cost_sewage, reading.cost_electricity,
            reading.cost_maintenance, reading.cost_social_rent, reading.cost_waste, reading.cost_fixed_part,
            t_209, t_205, total_cost
        ])

    worksheet.append([""] * 12 + ["ИТОГО:", total_209, total_205, total_sum])
    filename = f"Report_{period.name.replace(' ', '_')}.xlsx"
    encoded_filename = quote(filename)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        headers={'Content-Disposition': f"attachment; filename*=utf-8''{encoded_filename}"},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _fmt_contract_1c(rc) -> str:
    """Строка договора для 1С: «Договор от ДД.ММ.ГГГГ № N». Пусто — если нет."""
    if not rc:
        return ""
    num = (rc.number or "").strip()
    dt = rc.signed_date.strftime("%d.%m.%Y") if rc.signed_date else ""
    if num and dt:
        return f"Договор от {dt} № {num}"
    if num:
        return f"Договор № {num}"
    if dt:
        return f"Договор от {dt}"
    return ""


@router.get("/api/admin/export-1c", summary="Выгрузка в 1С (XLSX: Контрагент/Договор/Сумма 209+205)")
async def export_1c(
        period_id: Optional[int] = Query(None, description="ID периода"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db),
):
    """Выгрузка начислений за месяц в формате загрузки 1С. На каждого жильца с
    утверждённым показанием: ФИО (Контрагент), договор найма, Количество=1,
    Сумма = коммуналка за месяц (счёт 209 + счёт 205). Колонки документа-
    основания/статуса оставляем пустыми — заполняет 1С. Формат «один в один»
    с шаблоном загрузки."""
    if current_user.role not in ("accountant", "admin"):
        raise HTTPException(status_code=403)

    target_period_id = period_id
    if not target_period_id:
        active_period = (await db.execute(
            select(BillingPeriod).where(BillingPeriod.is_active.is_(True)))).scalars().first()
        if active_period:
            target_period_id = active_period.id
        else:
            last_closed = (await db.execute(
                select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1))).scalars().first()
            if not last_closed:
                raise HTTPException(404, "Нет периодов для выгрузки")
            target_period_id = last_closed.id

    period = await db.get(BillingPeriod, target_period_id)
    if not period:
        raise HTTPException(404, "Выбранный период не найден")

    rows = (await db.execute(
        select(User, MeterReading)
        .join(MeterReading, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == target_period_id,
            MeterReading.is_approved.is_(True),
            User.is_deleted.is_(False),
        )
        .order_by(User.username)
    )).all()

    # Активные договоры найма одним запросом (последний по дате подписания).
    uids = [u.id for u, _ in rows]
    contracts: dict[int, RentalContract] = {}
    if uids:
        for rc in (await db.execute(
            select(RentalContract)
            .where(RentalContract.user_id.in_(set(uids)), RentalContract.is_active.is_(True))
            .order_by(RentalContract.id.asc())
        )).scalars().all():
            contracts[rc.user_id] = rc  # asc по id → последний перезапишет = самый свежий

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист_1"
    worksheet.append([
        "N", "Контрагент", "Договор", "Количество", "Сумма",
        "Вид документа-основания", "Номер (108)", "Дата (109)",
        "Отразить в графике исполнения договора", "Статус платежа",
    ])

    n = 0
    for user, reading in rows:
        n += 1
        total = Decimal(reading.total_209 or 0) + Decimal(reading.total_205 or 0)
        worksheet.append([
            n,
            user.username,
            _fmt_contract_1c(contracts.get(user.id)),
            1,
            float(total.quantize(Decimal("0.01"))),
            "", "", "", "", "",
        ])

    filename = f"Vygruzka_1C_{period.name.replace(' ', '_')}.xlsx"
    encoded_filename = quote(filename)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        headers={'Content-Disposition': f"attachment; filename*=utf-8''{encoded_filename}"},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@router.post("/api/admin/receipts/{reading_id}/generate")
async def start_receipt_generation(reading_id: int, current_user: User = Depends(get_current_user)):
    if current_user.role not in ("accountant", "admin"): raise HTTPException(status_code=403)
    return {"task_id": generate_receipt_task.delay(reading_id).id, "status": "processing"}


@router.get("/api/admin/tasks/{task_id}")
async def get_task_status(task_id: str, current_user: User = Depends(get_current_user)):
    # Раньше любой авторизованный пользователь мог запросить статус
    # админской задачи и получить presigned-URL на PDF/Excel. Теперь
    # только admin (см. упрощение ролей — раньше было accountant/admin).
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Только для администратора")
    task_result = AsyncResult(task_id)
    if task_result.state == 'PENDING':
        return {"state": "PENDING", "status": "Pending..."}
    elif task_result.state != 'FAILURE':
        result = task_result.result
        if isinstance(result, dict) and result.get("status") in ["done", "ok"]:
            if s3_key := result.get("s3_key"):
                url = await asyncio.to_thread(s3_service.get_presigned_url, s3_key, 300)
                return {"state": task_result.state, "status": "done", "download_url": url}
        return {"state": task_result.state, "result": result}
    else:
        return {"state": "FAILURE", "error": str(task_result.info)}


@router.post("/api/admin/reports/bulk-zip", summary="Сгенерировать ZIP архива квитанций")
async def create_bulk_zip(
        period_id: Optional[int] = Query(None, description="ID периода"),
        current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)
):
    if current_user.role not in ("accountant", "admin"): raise HTTPException(status_code=403)

    target_period_id = period_id
    if not target_period_id:
        period = (await db.execute(select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1))).scalars().first()
        if not period: raise HTTPException(404, "Нет периодов")
        target_period_id = period.id

    task = start_bulk_receipt_generation.delay(target_period_id)
    return {"task_id": task.id, "status": "processing", "period_id": target_period_id}


@router.get("/api/admin/summary")
async def get_accountant_summary(
        period_id: Optional[int] = Query(None),
        current_user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)
):
    if current_user.role not in ("accountant", "admin"): raise HTTPException(status_code=403, detail="Доступ запрещен")

    # ИСПРАВЛЕНИЕ: Добавляем yield_per, чтобы не грузить весь массив в RAM разом.
    stmt = (
        select(User, MeterReading, Room)
        .join(MeterReading, User.id == MeterReading.user_id)
        .join(Room, User.room_id == Room.id)
        .where(MeterReading.is_approved.is_(True))
    ).execution_options(yield_per=1000)

    if period_id:
        stmt = stmt.where(MeterReading.period_id == period_id)
    else:
        last_period = (
            await db.execute(select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1))).scalars().first()
        if not last_period: return {}
        stmt = stmt.where(MeterReading.period_id == last_period.id)

    stmt = stmt.order_by(Room.dormitory_name, Room.room_number, User.username)

    summary = {}

    # Потоковое получение
    result = await db.stream(stmt)
    async for row in result:
        user, reading, room = row
        dorm = room.dormitory_name or "Без общежития"
        if dorm not in summary: summary[dorm] = []
        summary[dorm].append({
            "reading_id": reading.id, "user_id": user.id, "username": user.username, "area": room.apartment_area,
            "residents": user.residents_count, "hot": reading.cost_hot_water or 0, "cold": reading.cost_cold_water or 0,
            "sewage": reading.cost_sewage or 0, "electric": reading.cost_electricity or 0,
            "maintenance": reading.cost_maintenance or 0, "rent": reading.cost_social_rent or 0,
            "waste": reading.cost_waste or 0, "fixed": reading.cost_fixed_part or 0,
            "total_cost": reading.total_cost or 0, "total_209": reading.total_209 or 0,
            "total_205": reading.total_205 or 0,
            "date": reading.created_at.strftime("%Y-%m-%d %H:%M")
        })

    return summary


# =====================================================================
# SUMMARY v2 — расширенная сводка с финансовыми анализаторами
# =====================================================================
# Отличия от v1 (/api/admin/summary):
#   * Группировка по общежитиям + KPI на верхнем уровне
#   * Δ vs прошлый период для каждого жильца (рост/падение суммы)
#   * Sparkline за последние 6 периодов для каждого жильца
#   * Финансовые флаги от finance_analyzer (DEBT_GROWING и т.д.)
#   * Список MISSING_RECEIPT — жильцы без квитанции в этом периоде
#   * Поддержка фильтров: only_debtors, only_anomaly, only_overpaid, search
#
# Используется новой версткой «Финансовая отчётность» в админке.
@router.get("/api/admin/summary/v2")
async def get_accountant_summary_v2(
    period_id: Optional[int] = Query(None),
    only_debtors: bool = Query(False),
    only_overpaid: bool = Query(False),
    only_anomaly: bool = Query(False),
    only_missing: bool = Query(False),
    search: Optional[str] = Query(None),
    # housing_001/E2-C: фильтры по типу помещения. Когда админ хочет
    # увидеть финотчёт ТОЛЬКО по домам/общагам или по конкретной улице.
    # На начислении не отражается — это только сужение выборки.
    place_type: Optional[str] = Query(
        None, pattern="^(dormitory|house)$",
        description="dormitory | house — фильтр по типу помещения",
    ),
    street: Optional[str] = Query(
        None, description="Точное название улицы (для домов)",
    ),
    history_periods: int = Query(
        12,
        ge=1,
        le=12,
        description="Сколько ПРЕДЫДУЩИХ периодов показать в истории жильца "
                    "(1-12). По умолчанию 12 — год истории. Кап 12 (был 24): "
                    "загрузка 2 лет истории в память — DoS-вектор (security-аудит)."
    ),
    group_by: str = Query(
        "user", pattern="^(user|room)$",
        description="user — по жильцам (ФИО), room — по квартирам (адрес, "
                    "агрегат по всем жильцам комнаты). Финсводка следит за "
                    "помещениями, а не за людьми.",
    ),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role not in ("accountant", "admin"):
        raise HTTPException(403, "Доступ запрещён")

    from app.modules.utility.services.finance_analyzer import (
        analyze_finance, FLAG_CATALOG,
    )

    # 1) Период
    if period_id:
        period = await db.get(BillingPeriod, period_id)
    else:
        period = (await db.execute(
            select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1)
        )).scalars().first()
    if not period:
        return {"group_by": group_by, "period": None, "kpi": {}, "dormitories": [], "flag_catalog": FLAG_CATALOG}

    # 2) Тянем все утверждённые показания за этот период + жильцов с комнатой
    stmt = (
        select(User, MeterReading, Room)
        .join(MeterReading, User.id == MeterReading.user_id)
        .join(Room, User.room_id == Room.id)
        .where(
            MeterReading.is_approved.is_(True),
            MeterReading.period_id == period.id,
            User.is_deleted.is_(False),
        )
        .order_by(Room.dormitory_name, Room.room_number, User.username)
    )
    # housing_001/E2-C: фильтры по типу помещения сужают выборку
    # ДО группировки по домам.
    if place_type:
        stmt = stmt.where(Room.place_type == place_type)
    if street:
        stmt = stmt.where(Room.street == street)
    rows = (await db.execute(stmt)).all()

    # 3) История за N предыдущих периодов — для sparkline и Δ.
    # N приходит query-параметром history_periods (default 12 = год).
    # Раньше было захардкожено 6; админ хотел видеть глубже — сделали
    # настраиваемым. Берём ВСЕ utility-readings этих жильцов одним
    # запросом, фильтруем в Python: дешевле чем N запросов по жильцам.
    #
    # ВАЖНО: «прошлые» периоды определяем по БИЛЛИНГОВОЙ хронологии
    # (parsed name → year, month), а не по `BillingPeriod.id`. Иначе
    # подача задним числом (Февраль 2026 импортирован в мае) ломает
    # сортировку и считает Δ относительно «не того» предыдущего. См.
    # инцидент мая 2026 с Сорокиным С.А. и helper period_helpers.py.
    user_ids = [r[0].id for r in rows]
    history_map: dict[int, list[MeterReading]] = {uid: [] for uid in user_ids}
    # period_id → chronological_key. Хоистим дефолтом, чтобы room-режим мог
    # безопасно ссылаться даже когда история пустая.
    period_id_to_key: dict = {}
    if user_ids:
        all_periods = (await db.execute(select(BillingPeriod))).scalars().all()
        cur_key = period_chron_key(period.name)
        # Строго раньше текущего по хронологии (для Δ нужны только прошлые).
        prev_periods_sorted = sorted(
            (p for p in all_periods if period_chron_key(p.name) < cur_key),
            key=lambda p: period_chron_key(p.name),
            reverse=True,  # DESC — самый свежий первый
        )[:history_periods]
        prev_period_ids = [p.id for p in prev_periods_sorted]
        # period_id → chronological_key — для сортировки readings ниже.
        period_id_to_key = {p.id: period_chron_key(p.name) for p in prev_periods_sorted}

        if prev_period_ids:
            hist_rows = (await db.execute(
                select(MeterReading)
                .where(
                    MeterReading.user_id.in_(user_ids),
                    MeterReading.period_id.in_(prev_period_ids),
                    MeterReading.is_approved.is_(True),
                )
            )).scalars().all()
            # Сортируем по БИЛЛИНГОВОЙ хронологии ASC (старые → новые).
            # `prev_costs[-1]` будет САМЫМ СВЕЖИМ прошлым → правильная Δ.
            # Раньше сортировали по `period_id` — задним-числом импорт давал
            # «прошлым» Февральскую призрачную подачу с миллионными cost.
            tmp: dict[int, list] = {}
            for hr in hist_rows:
                tmp.setdefault(hr.user_id, []).append(hr)
            for uid in user_ids:
                history_map[uid] = sorted(
                    tmp.get(uid, []),
                    key=lambda r: period_id_to_key.get(r.period_id, (0, 0)),
                )

    # 4) MISSING_RECEIPT — жильцы с комнатой, но без MeterReading в этом периоде.
    missing_users = []
    if not only_debtors and not only_overpaid and not only_anomaly:
        # Только если фильтры не отсекают этот класс жильцов
        all_residents = (await db.execute(
            select(User, Room)
            .join(Room, User.room_id == Room.id)
            .where(
                User.is_deleted.is_(False),
                User.role == "user",
            )
        )).all()
        present = {r[0].id for r in rows}
        for user, room in all_residents:
            if user.id in present:
                continue
            if search and search.lower() not in (user.username or "").lower() \
                    and search not in (room.room_number or ""):
                continue
            missing_users.append((user, room))

    # ============================================================
    # РЕЖИМ «КВАРТИРЫ» (group_by=room): агрегируем по помещению, а не по
    # жильцу. Долг квартиры = сумма по ВСЕМ её жильцам (решение 30.05.2026).
    # Фин-фильтры (only_debtors/overpaid/anomaly/missing) применяются на
    # УРОВНЕ комнаты после суммирования, не по каждому жильцу. Внутри
    # каждой комнаты несём список жильцов — для разворота строки.
    # ============================================================
    if group_by == "room":
        from app.modules.utility.services.anomaly_flags import real_flags

        room_acc: dict[int, dict] = {}

        def _ensure_room(room) -> dict:
            if room.id not in room_acc:
                room_acc[room.id] = {
                    "room_id": room.id,
                    "dorm_name": room.dormitory_name or "Без общежития",
                    "room_number": room.room_number,
                    "address": room.format_address or (room.room_number or "—"),
                    "place_type": getattr(room, "place_type", None),
                    "area": float(room.apartment_area or 0),
                    "residents_count": 0,
                    "total_cost": Decimal("0"),
                    "total_209": Decimal("0"),
                    "total_205": Decimal("0"),
                    "debt": Decimal("0"),
                    "overpayment": Decimal("0"),
                    "anomaly_score": 0,
                    "meter_flags": set(),
                    "missing_count": 0,
                    "reading_ids": [],
                    "residents": [],
                    "_cost_by_key": {},   # chrono_key -> сумма cost жильцов
                    "_debt_by_key": {},   # chrono_key -> сумма debt жильцов
                }
            return room_acc[room.id]

        # 1) Суммируем поданные показания по комнатам.
        for user, reading, room in rows:
            if search:
                s = search.lower().strip()
                if s not in (user.username or "").lower() and s not in (room.room_number or ""):
                    continue
            debt = Decimal(str((reading.debt_209 or 0) + (reading.debt_205 or 0)))
            overpay = Decimal(str((reading.overpayment_209 or 0) + (reading.overpayment_205 or 0)))
            cur_cost = Decimal(str(reading.total_cost or 0))
            ra = _ensure_room(room)
            ra["residents_count"] += 1
            ra["total_cost"] += cur_cost
            ra["total_209"] += Decimal(str(reading.total_209 or 0))
            ra["total_205"] += Decimal(str(reading.total_205 or 0))
            ra["debt"] += debt
            ra["overpayment"] += overpay
            ra["anomaly_score"] = max(ra["anomaly_score"], int(reading.anomaly_score or 0))
            for mf in real_flags(reading.anomaly_flags):
                ra["meter_flags"].add(mf)
            if reading.id:
                ra["reading_ids"].append(reading.id)
            for h in history_map.get(user.id, []):
                k = period_id_to_key.get(h.period_id)
                if k is None:
                    continue
                ra["_cost_by_key"][k] = ra["_cost_by_key"].get(k, Decimal("0")) + Decimal(str(h.total_cost or 0))
                ra["_debt_by_key"][k] = ra["_debt_by_key"].get(k, Decimal("0")) + Decimal(str((h.debt_209 or 0) + (h.debt_205 or 0)))
            ra["residents"].append({
                "user_id": user.id,
                "username": user.username,
                "residents_count": user.residents_count or 1,
                "reading_id": reading.id,
                "total_cost": float(cur_cost),
                "debt": float(debt),
                "overpayment": float(overpay),
            })

        # 2) Жильцы без подачи (MISSING_RECEIPT) — помечаем их комнату.
        for user, room in missing_users:
            ra = _ensure_room(room)
            ra["missing_count"] += 1
            ra["residents"].append({
                "user_id": user.id,
                "username": user.username,
                "residents_count": user.residents_count or 1,
                "reading_id": None,
                "total_cost": 0.0,
                "debt": 0.0,
                "overpayment": 0.0,
            })

        # 3) Финализация: re-analyze по агрегату комнаты, флаги, фильтры.
        rooms_by_dorm: dict[str, dict] = {}
        grand_billed_r = Decimal("0")
        grand_debt_r = Decimal("0")
        grand_overpay_r = Decimal("0")
        flagged_rooms = 0
        missing_rooms = 0
        all_rooms_flat = []

        for ra in room_acc.values():
            keys_sorted = sorted(ra["_cost_by_key"].keys())
            prev_costs = [ra["_cost_by_key"][k] for k in keys_sorted]
            prev_debts = [ra["_debt_by_key"].get(k, Decimal("0")) for k in keys_sorted]
            cur_cost = ra["total_cost"]
            debt = ra["debt"]
            overpay = ra["overpayment"]
            has_reading = bool(ra["reading_ids"])

            flags, fin_score = analyze_finance(
                user_id=ra["room_id"],
                residents_count=ra["residents_count"] or 1,
                current_total_cost=cur_cost,
                current_debt=debt,
                current_overpayment=overpay,
                prev_costs=prev_costs,
                prev_debts=prev_debts,
                has_reading=has_reading,
                resident_type="family",
                billing_mode="by_meter",
            )
            if ra["missing_count"] and not has_reading and "MISSING_RECEIPT" not in flags:
                flags = list(flags) + ["MISSING_RECEIPT"]
            meter_flags = sorted(ra["meter_flags"])

            # Фин-фильтры на уровне комнаты
            if only_debtors and debt <= 0:
                continue
            if only_overpaid and overpay <= 0:
                continue
            if only_anomaly and not flags and not meter_flags:
                continue
            if only_missing and not ra["missing_count"]:
                continue

            delta_amount = None
            delta_percent = None
            if prev_costs:
                last = prev_costs[-1]
                delta_amount = float(cur_cost - last)
                if last > 0:
                    delta_percent = float((cur_cost - last) / last * 100)
            sparkline = [float(c) for c in prev_costs] + [float(cur_cost)]

            row = {
                "room_id": ra["room_id"],
                "room_number": ra["room_number"],
                "address": ra["address"],
                "place_type": ra["place_type"],
                "area": ra["area"],
                "residents_count": ra["residents_count"],
                "missing_count": ra["missing_count"],
                "total_cost": float(cur_cost),
                "total_209": float(ra["total_209"]),
                "total_205": float(ra["total_205"]),
                "debt": float(debt),
                "overpayment": float(overpay),
                "delta_amount": delta_amount,
                "delta_percent": delta_percent,
                "sparkline": sparkline,
                "finance_flags": flags,
                "finance_score": fin_score,
                "meter_flags": meter_flags,
                "anomaly_score": ra["anomaly_score"],
                "reading_ids": ra["reading_ids"],
                "residents": sorted(ra["residents"], key=lambda r: (r.get("username") or "")),
            }
            dn = ra["dorm_name"]
            if dn not in rooms_by_dorm:
                rooms_by_dorm[dn] = {
                    "name": dn, "rooms": [], "total_billed": Decimal("0"),
                    "total_debt": Decimal("0"), "total_overpay": Decimal("0"),
                    "flagged_count": 0,
                }
            dd = rooms_by_dorm[dn]
            dd["rooms"].append(row)
            dd["total_billed"] += cur_cost
            dd["total_debt"] += debt
            dd["total_overpay"] += overpay
            if flags or meter_flags:
                dd["flagged_count"] += 1
                flagged_rooms += 1
            if ra["missing_count"]:
                missing_rooms += 1
            grand_billed_r += cur_cost
            grand_debt_r += debt
            grand_overpay_r += overpay
            all_rooms_flat.append(row)

        dorms_out = []
        for name in sorted(rooms_by_dorm.keys()):
            dd = rooms_by_dorm[name]
            dd["rooms"].sort(key=lambda r: (r.get("room_number") or ""))
            dorms_out.append({
                "name": dd["name"],
                "total_billed": float(dd["total_billed"]),
                "total_debt": float(dd["total_debt"]),
                "total_overpay": float(dd["total_overpay"]),
                "flagged_count": dd["flagged_count"],
                "rooms_count": len(dd["rooms"]),
                "rooms": dd["rooms"],
            })

        top_debtor_rooms = sorted(
            [r for r in all_rooms_flat if r["debt"] > 0], key=lambda r: -r["debt"]
        )[:5]
        top_overpay_rooms = sorted(
            [r for r in all_rooms_flat if r["overpayment"] > 0], key=lambda r: -r["overpayment"]
        )[:5]

        return {
            "group_by": "room",
            "period": {"id": period.id, "name": period.name, "is_active": period.is_active},
            "kpi": {
                "total_billed": float(grand_billed_r),
                "total_debt": float(grand_debt_r),
                "total_overpay": float(grand_overpay_r),
                "flagged_count": flagged_rooms,
                "rooms_count": sum(d["rooms_count"] for d in dorms_out),
                "missing_count": missing_rooms,
            },
            "top_debtors": top_debtor_rooms,
            "top_overpayers": top_overpay_rooms,
            "dormitories": dorms_out,
            "flag_catalog": FLAG_CATALOG,
        }

    # 5) Сборка ответа
    grand_billed = Decimal("0")
    grand_debt = Decimal("0")
    grand_overpay = Decimal("0")
    flagged_count = 0

    by_dorm: dict[str, dict] = {}

    def _ensure_dorm(name: str) -> dict:
        if name not in by_dorm:
            by_dorm[name] = {
                "name": name,
                "residents": [],
                "total_billed": Decimal("0"),
                "total_debt": Decimal("0"),
                "total_overpay": Decimal("0"),
                "flagged_count": 0,
            }
        return by_dorm[name]

    for user, reading, room in rows:
        debt = (reading.debt_209 or 0) + (reading.debt_205 or 0)
        overpay = (reading.overpayment_209 or 0) + (reading.overpayment_205 or 0)
        debt = Decimal(str(debt))
        overpay = Decimal(str(overpay))
        cur_cost = Decimal(str(reading.total_cost or 0))

        # Поиск
        if search:
            s = search.lower().strip()
            if s not in (user.username or "").lower() and s not in (room.room_number or ""):
                continue

        # Фильтры
        if only_debtors and debt <= 0:
            continue
        if only_overpaid and overpay <= 0:
            continue

        history = history_map.get(user.id, [])
        prev_costs = [Decimal(str(h.total_cost or 0)) for h in history]
        prev_debts = [
            Decimal(str((h.debt_209 or 0) + (h.debt_205 or 0)))
            for h in history
        ]

        flags, fin_score = analyze_finance(
            user_id=user.id,
            residents_count=user.residents_count or 1,
            current_total_cost=cur_cost,
            current_debt=debt,
            current_overpayment=overpay,
            prev_costs=prev_costs,
            prev_debts=prev_debts,
            has_reading=True,
            resident_type=getattr(user, "resident_type", "family"),
            billing_mode=getattr(user, "billing_mode", "by_meter"),
        )

        # Только настоящие аномалии — без source-маркеров (GSHEETS_AUTO,
        # AUTO_GENERATED, DATA_OVERFLOW_RESET и т.п.). Раньше KPI «Аномалий»
        # включал все записи с anomaly_flags!=NULL — даже служебные, и
        # показывал сотни проблем при том что filter «Аномалии» находил
        # единицы. См. services/anomaly_flags.py:SOURCE_MARKERS.
        from app.modules.utility.services.anomaly_flags import real_flags
        meter_flags = real_flags(reading.anomaly_flags)

        if only_anomaly and not flags and not meter_flags:
            continue

        # Δ vs прошлый
        delta_amount = None
        delta_percent = None
        if prev_costs:
            last = prev_costs[-1]
            delta_amount = float(cur_cost - last)
            if last > 0:
                delta_percent = float((cur_cost - last) / last * 100)

        # Sparkline: 6 точек (включая текущий — последняя)
        sparkline = [float(c) for c in prev_costs] + [float(cur_cost)]

        d = _ensure_dorm(room.dormitory_name or "Без общежития")
        d["residents"].append({
            "user_id": user.id,
            "username": user.username,
            "room_number": room.room_number,
            "room_id": room.id,
            "area": float(room.apartment_area or 0),
            "residents_count": user.residents_count or 1,
            "reading_id": reading.id,
            "total_cost": float(cur_cost),
            "total_209": float(reading.total_209 or 0),
            "total_205": float(reading.total_205 or 0),
            "debt": float(debt),
            "overpayment": float(overpay),
            "delta_amount": delta_amount,
            "delta_percent": delta_percent,
            "sparkline": sparkline,
            "finance_flags": flags,
            "finance_score": fin_score,
            "meter_flags": meter_flags,
            "anomaly_score": int(reading.anomaly_score or 0),
            "created_at": reading.created_at.isoformat() if reading.created_at else None,
        })
        d["total_billed"] += cur_cost
        d["total_debt"] += debt
        d["total_overpay"] += overpay
        if flags or meter_flags:
            d["flagged_count"] += 1
            flagged_count += 1
        grand_billed += cur_cost
        grand_debt += debt
        grand_overpay += overpay

    # MISSING_RECEIPT добавляем как «жильцов без подачи»
    if missing_users and not only_anomaly and not only_debtors and not only_overpaid:
        for user, room in missing_users:
            d = _ensure_dorm(room.dormitory_name or "Без общежития")
            d["residents"].append({
                "user_id": user.id,
                "username": user.username,
                "room_number": room.room_number,
                "room_id": room.id,
                "area": float(room.apartment_area or 0),
                "residents_count": user.residents_count or 1,
                "reading_id": None,
                "total_cost": 0.0,
                "total_209": 0.0,
                "total_205": 0.0,
                "debt": 0.0,
                "overpayment": 0.0,
                "delta_amount": None,
                "delta_percent": None,
                "sparkline": [],
                "finance_flags": ["MISSING_RECEIPT"],
                "finance_score": 40,
                "meter_flags": [],
                "anomaly_score": 0,
                "created_at": None,
            })
            # flagged_count для missing-receipt НЕ инкрементим — у них есть
            # своя отдельная KPI «Без квитанции» (missing_count). Иначе KPI
            # «Аномалий» = реальные + missing, и расходится с фильтром
            # «Аномалии», который missing игнорирует.

    if only_missing:
        # отфильтровываем всё кроме MISSING_RECEIPT
        for dn in list(by_dorm.keys()):
            by_dorm[dn]["residents"] = [
                r for r in by_dorm[dn]["residents"]
                if "MISSING_RECEIPT" in (r.get("finance_flags") or [])
            ]
            if not by_dorm[dn]["residents"]:
                del by_dorm[dn]

    # Сортируем общежития по имени, жильцов внутри — по комнате
    dormitories_out = []
    for name in sorted(by_dorm.keys()):
        d = by_dorm[name]
        d["residents"].sort(key=lambda r: (r.get("room_number") or "", r.get("username") or ""))
        dormitories_out.append({
            "name": d["name"],
            "total_billed": float(d["total_billed"]),
            "total_debt": float(d["total_debt"]),
            "total_overpay": float(d["total_overpay"]),
            "flagged_count": d["flagged_count"],
            "residents_count": len(d["residents"]),
            "residents": d["residents"],
        })

    # Топ-должники / топ-плательщики (по всем общежитиям)
    all_residents = [r for d in dormitories_out for r in d["residents"]]
    top_debtors = sorted(
        [r for r in all_residents if r["debt"] > 0],
        key=lambda r: -r["debt"],
    )[:5]
    top_overpayers = sorted(
        [r for r in all_residents if r["overpayment"] > 0],
        key=lambda r: -r["overpayment"],
    )[:5]

    return {
        "group_by": "user",
        "period": {"id": period.id, "name": period.name, "is_active": period.is_active},
        "kpi": {
            "total_billed": float(grand_billed),
            "total_debt": float(grand_debt),
            "total_overpay": float(grand_overpay),
            "flagged_count": flagged_count,
            "residents_count": sum(d["residents_count"] for d in dormitories_out),
            "missing_count": len(missing_users),
        },
        "top_debtors": top_debtors,
        "top_overpayers": top_overpayers,
        "dormitories": dormitories_out,
        "flag_catalog": FLAG_CATALOG,
    }


# =========================================================================
# ДЕТАЛИ ПО ОДНОМУ ЖИЛЬЦУ (expandable-панель в «Финансовой отчётности»)
# =========================================================================
# Возвращает всё, что нужно показать в развёрнутой строке:
#   * основная квитанция за выбранный период (детализация по статьям, дельты)
#   * история показаний за 6 последних периодов (счётчики + дельты + источник)
#   * корректировки (Adjustment) за эти периоды с пояснениями
#   * активный договор найма (№/дата) — чтобы не лезть в другую вкладку
# Один запрос вместо N+1 — подтягиваем всё батчами.

def _infer_source_flag(anomaly_flags):
    """Bug AN: расширено различение auto-стратегий вместо единого «auto».
    UI показывает админу, *как именно* система начислила: норматив×коэф,
    среднее по дельтам, повтор предыдущего, или baseline."""
    if not anomaly_flags:
        return "manual"
    af = anomaly_flags.upper()
    if "GSHEETS" in af:
        return "gsheets"
    if "ONE_TIME_CHARGE" in af:
        return "one_time"
    if "AUTO_NORM_SANCTION" in af:
        return "auto_norm_sanction"
    if "AUTO_AVG_FALLBACK" in af:
        return "auto_avg_fallback"
    if "AUTO_AVG" in af:
        return "auto_avg"
    if "AUTO_NO_HISTORY" in af:
        return "auto_no_history"
    if "AUTO_GENERATED" in af:
        return "auto"
    if "MANUAL_RECEIPT" in af:
        return "manual_receipt"
    if "BASELINE" in af:
        return "baseline"
    if "INITIAL_SETUP" in af:
        return "initial"
    if "METER_CLOSED" in af or "METER_REPLACEMENT" in af:
        return "meter_op"
    return "app"


@router.get("/api/admin/rooms/{room_id}/submission-history")
async def get_room_submission_history(
    room_id: int,
    periods: int = Query(12, ge=1, le=24,
                         description="Сколько последних периодов вернуть (1-24)"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """История подач по КВАРТИРЕ: по каждому месяцу (периоду) — какие показания
    подали жильцы этой комнаты (ФИО, тип, ГВС/ХВС/электр, сумма, источник).
    Привязка по `MeterReading.room_id` — ловит и текущих, и прошлых жильцов
    помещения. Для разворота квартиры в «Финансовой отчётности» (анализ квартиры
    по месяцам подач). Периоды — в хронологическом порядке (свежие сверху)."""
    if current_user.role not in ("accountant", "admin", "financier"):
        raise HTTPException(403, "Доступ запрещён")

    room = await db.get(Room, room_id)

    rows = (await db.execute(
        select(MeterReading, User, BillingPeriod)
        .join(BillingPeriod, BillingPeriod.id == MeterReading.period_id)
        .outerjoin(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.room_id == room_id,
            MeterReading.is_approved.is_(True),
        )
    )).all()

    by_period: dict[int, dict] = {}
    for reading, user, period in rows:
        slot = by_period.get(period.id)
        if slot is None:
            slot = by_period[period.id] = {
                "period_id": period.id,
                "period_name": period.name,
                "_chrono": period_chron_key(period.name),
                "submissions": [],
                "total_cost": 0.0,
                "debt": 0.0,
                "overpayment": 0.0,
            }
        debt = float((reading.debt_209 or 0) + (reading.debt_205 or 0))
        overpay = float((reading.overpayment_209 or 0) + (reading.overpayment_205 or 0))
        slot["submissions"].append({
            "user_id": user.id if user else None,
            "username": user.username if user else "—",
            "full_name": getattr(user, "full_name", None) if user else None,
            "resident_type": (getattr(user, "resident_type", None) or "family") if user else None,
            "reading_id": reading.id,
            "is_approved": bool(reading.is_approved),
            "hot_water": float(reading.hot_water or 0),
            "cold_water": float(reading.cold_water or 0),
            "electricity": float(reading.electricity or 0),
            "total_cost": float(reading.total_cost or 0),
            "total_209": float(reading.total_209 or 0),
            "total_205": float(reading.total_205 or 0),
            "debt": debt,
            "overpayment": overpay,
            "source": _infer_source_flag(reading.anomaly_flags),
            "created_at": reading.created_at.isoformat() if reading.created_at else None,
        })
        slot["total_cost"] += float(reading.total_cost or 0)
        slot["debt"] += debt
        slot["overpayment"] += overpay

    history = sorted(by_period.values(), key=lambda s: s["_chrono"], reverse=True)[:periods]
    for s in history:
        s.pop("_chrono", None)
        s["submissions"].sort(key=lambda x: (x["username"] or ""))

    return {
        "room_id": room_id,
        "address": (room.format_address if room else None) or (room.room_number if room else "—"),
        "dormitory_name": room.dormitory_name if room else None,
        "periods_count": len(history),
        "history": history,
    }


@router.get("/api/admin/residents/{user_id}/finance-detail")
async def get_resident_finance_detail(
    user_id: int,
    period_id: Optional[int] = Query(None, description="Период квитанции. По умолчанию — последний."),
    history_periods: int = Query(6, ge=1, le=24),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Детальная финансовая справка по жильцу для разворачивающейся панели.

    Используется в «Финансовой отчётности» — админ кликает на строку жильца
    и получает всю финансовую картину: текущая квитанция + история счётчиков
    за 6 мес + корректировки + договор.
    """
    if current_user.role not in ("accountant", "admin", "financier"):
        raise HTTPException(403, "Доступ запрещён")

    user = (await db.execute(
        select(User)
        .options(selectinload(User.room))
        .where(User.id == user_id, User.is_deleted.is_(False))
    )).scalars().first()
    if not user:
        raise HTTPException(404, "Жилец не найден")

    # 1) Определяем целевой период.
    if period_id:
        period = await db.get(BillingPeriod, period_id)
    else:
        period = (await db.execute(
            select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1)
        )).scalars().first()
    if not period:
        raise HTTPException(400, "Периоды не заведены — нет данных для показа")

    # 2) N последних периодов (включая текущий) для истории — БИЛЛИНГОВО.
    # Раньше сортировали по `BillingPeriod.id.desc()`, но id отражает порядок
    # создания записи в БД, а не биллинговый месяц. Если админ задним числом
    # импортировал «Февраль 2026» в мае, у него period.id > чем у мая, и в
    # таблице февраль появлялся ВЫШЕ майских данных. Из-за этого дельты тоже
    # съезжали (prev определялось по created_at, см. _prev_for ниже).
    # Теперь сортируем по распарсенному `(year, month)` имени периода —
    # хронологически. Нестандартные имена («Начальный период», тестовые)
    # получают ключ (0, 0) и оказываются в самом начале (baseline).
    all_periods = (await db.execute(select(BillingPeriod))).scalars().all()
    cur_key = period_chron_key(period.name)
    # Только периоды у которых биллинговая хронология <= текущей.
    # Это аналог прежнего `BillingPeriod.id <= period.id`, но корректный.
    periods_filtered = [p for p in all_periods if period_chron_key(p.name) <= cur_key]
    periods = sorted(
        periods_filtered,
        key=lambda p: period_chron_key(p.name),
        reverse=True,
    )[:history_periods]
    period_ids = [p.id for p in periods]
    period_name_map = {p.id: p.name for p in periods}

    # 3) Одним запросом — все показания жильца за эти периоды (включая drafts).
    readings = (await db.execute(
        select(MeterReading)
        .where(
            MeterReading.user_id == user_id,
            MeterReading.period_id.in_(period_ids) if period_ids else False,
        )
    )).scalars().all()

    # Для дельт нужно предыдущее approved показание ЖИЛЬЦА В ЭТОЙ КОМНАТЕ
    # в БИЛЛИНГОВОЙ хронологии (не по created_at — задний-числом импорт ломает).
    # Загружаем ВСЕ его approved readings (а не только за выбранные history N),
    # потому что предыдущее показание для самого раннего из N может лежать
    # ВНЕ выборки. Затем строим словарь reading.id → prev_reading.
    prev_reading_map: dict[int, Optional[MeterReading]] = {}
    if user.room_id:
        all_user_readings = (await db.execute(
            select(MeterReading, BillingPeriod)
            .join(BillingPeriod, BillingPeriod.id == MeterReading.period_id)
            .where(
                MeterReading.user_id == user_id,
                MeterReading.room_id == user.room_id,
                MeterReading.is_approved.is_(True),
            )
        )).all()
        # Сортируем хронологически ASC: baseline → старые → новые.
        all_user_readings_chronological = sorted(
            all_user_readings,
            key=lambda row: period_chron_key(row[1].name),
        )
        # Двигаемся по цепочке: текущему reading prev = предыдущий в хронологии.
        prev_reading: Optional[MeterReading] = None
        for r, _bp in all_user_readings_chronological:
            prev_reading_map[r.id] = prev_reading
            prev_reading = r

    def _prev_for(reading):
        """Предыдущее approved показание жильца в биллинговой хронологии
        (а не по `created_at` — задний-числом импорт ломал предыдущую логику,
        см. инцидент мая 2026 с Сорокиным С.А.)."""
        return prev_reading_map.get(reading.id)

    # 4) Корректировки за эти периоды.
    adjustments = []
    if period_ids:
        adj_rows = (await db.execute(
            select(Adjustment)
            .where(Adjustment.user_id == user_id, Adjustment.period_id.in_(period_ids))
            .order_by(Adjustment.period_id.desc(), Adjustment.created_at.desc())
        )).scalars().all()
        for a in adj_rows:
            adjustments.append({
                "id": a.id,
                "period_id": a.period_id,
                "period_name": period_name_map.get(a.period_id),
                "amount": float(a.amount or 0),
                "description": a.description or "",
                "account_type": a.account_type or "209",
                "created_at": a.created_at.isoformat() if a.created_at else None,
            })

    # 5) Договор найма — активный.
    from app.modules.utility.models import RentalContract
    contract_row = (await db.execute(
        select(RentalContract)
        .where(
            RentalContract.user_id == user_id,
            RentalContract.is_active.is_(True),
        )
        .limit(1)
    )).scalars().first()
    contract_data = None
    if contract_row:
        contract_data = {
            "id": contract_row.id,
            "number": contract_row.number,
            "signed_date": contract_row.signed_date.isoformat() if contract_row.signed_date else None,
            "valid_until": contract_row.valid_until.isoformat() if contract_row.valid_until else None,
            "has_file": bool(contract_row.file_s3_key),
            "file_name": contract_row.file_name,
        }

    # 6) Сборка истории показаний (6 периодов).
    readings_by_period = {r.period_id: r for r in readings}
    history = []
    for p in periods:
        r = readings_by_period.get(p.id)
        if r is None:
            history.append({
                "period_id": p.id,
                "period_name": p.name,
                "reading_id": None,
                "is_approved": False,
                "has_pdf": False,
                "source": None,
                "flags": [],
                "hot_water": None, "cold_water": None, "electricity": None,
                "delta_hot": None, "delta_cold": None, "delta_elect": None,
                "total_cost": None, "total_209": None, "total_205": None,
            })
            continue
        prev = _prev_for(r)
        p_hot = float(prev.hot_water or 0) if prev else 0.0
        p_cold = float(prev.cold_water or 0) if prev else 0.0
        p_elect = float(prev.electricity or 0) if prev else 0.0
        cur_hot = float(r.hot_water or 0)
        cur_cold = float(r.cold_water or 0)
        cur_elect = float(r.electricity or 0)
        flags_list = [f.strip() for f in (r.anomaly_flags or "").split(",") if f.strip()]
        history.append({
            "period_id": p.id,
            "period_name": p.name,
            "reading_id": r.id,
            "is_approved": bool(r.is_approved),
            "has_pdf": bool(getattr(r, "pdf_s3_key", None)),
            "source": _infer_source_flag(r.anomaly_flags),
            "flags": flags_list,
            "hot_water": cur_hot,
            "cold_water": cur_cold,
            "electricity": cur_elect,
            "delta_hot": cur_hot - p_hot if prev else None,
            "delta_cold": cur_cold - p_cold if prev else None,
            "delta_elect": cur_elect - p_elect if prev else None,
            "total_cost": float(r.total_cost or 0),
            "total_209": float(r.total_209 or 0),
            "total_205": float(r.total_205 or 0),
        })

    # 7) Детализация текущей квитанции.
    cur = readings_by_period.get(period.id)
    current = None
    if cur:
        prev = _prev_for(cur)
        current = {
            "reading_id": cur.id,
            "is_approved": bool(cur.is_approved),
            "source": _infer_source_flag(cur.anomaly_flags),
            "anomaly_flags": cur.anomaly_flags,
            "anomaly_score": int(cur.anomaly_score or 0),
            "hot_water": float(cur.hot_water or 0),
            "cold_water": float(cur.cold_water or 0),
            "electricity": float(cur.electricity or 0),
            "prev_hot": float(prev.hot_water or 0) if prev else 0.0,
            "prev_cold": float(prev.cold_water or 0) if prev else 0.0,
            "prev_elect": float(prev.electricity or 0) if prev else 0.0,
            "delta_hot": float((cur.hot_water or 0) - (prev.hot_water or 0)) if prev else None,
            "delta_cold": float((cur.cold_water or 0) - (prev.cold_water or 0)) if prev else None,
            "delta_elect": float((cur.electricity or 0) - (prev.electricity or 0)) if prev else None,
            "total_cost": float(cur.total_cost or 0),
            "total_209": float(cur.total_209 or 0),
            "total_205": float(cur.total_205 or 0),
            "debt_209": float(cur.debt_209 or 0),
            "debt_205": float(cur.debt_205 or 0),
            "overpayment_209": float(cur.overpayment_209 or 0),
            "overpayment_205": float(cur.overpayment_205 or 0),
            "hot_correction": float(cur.hot_correction or 0),
            "cold_correction": float(cur.cold_correction or 0),
            "electricity_correction": float(cur.electricity_correction or 0),
            "sewage_correction": float(cur.sewage_correction or 0),
            "costs": {
                "cost_hot_water": float(cur.cost_hot_water or 0),
                "cost_cold_water": float(cur.cost_cold_water or 0),
                "cost_sewage": float(cur.cost_sewage or 0),
                "cost_electricity": float(cur.cost_electricity or 0),
                "cost_maintenance": float(cur.cost_maintenance or 0),
                "cost_social_rent": float(cur.cost_social_rent or 0),
                "cost_waste": float(cur.cost_waste or 0),
                "cost_fixed_part": float(cur.cost_fixed_part or 0),
            },
            "has_pdf": bool(getattr(cur, "pdf_s3_key", None)),
        }

    return {
        "user": {
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "role": user.role,
            "residents_count": user.residents_count or 1,
            "resident_type": getattr(user, "resident_type", "family"),
            "billing_mode": getattr(user, "billing_mode", "by_meter"),
            "room": (
                {
                    "id": user.room.id,
                    "dormitory_name": user.room.dormitory_name,
                    "room_number": user.room.room_number,
                    "apartment_area": float(user.room.apartment_area or 0),
                    "total_room_residents": user.room.total_room_residents or 1,
                    "hw_meter_serial": getattr(user.room, "hw_meter_serial", None),
                    "cw_meter_serial": getattr(user.room, "cw_meter_serial", None),
                    "el_meter_serial": getattr(user.room, "el_meter_serial", None),
                }
                if user.room else None
            ),
        },
        "period": {"id": period.id, "name": period.name},
        "current": current,
        "history": history,  # от свежих к старым
        "adjustments": adjustments,
        "contract": contract_data,
        "balance": await _compute_user_balance(db, user.id, user.room_id),
    }


async def _compute_user_balance(db: AsyncSession, user_id: int, room_id: Optional[int]) -> dict:
    """Текущий баланс жильца — единое число «должен/переплатил» с учётом
    ОБОИХ счетов (209 коммуналка + 205 найм).

    debt/overpayment живут в каждом MeterReading периода. Импорт 1С
    обновляет их в АКТИВНОМ периоде, но если 209-импорт прошёл в Мае,
    а 205-импорт в Январе — самые свежие сальдо лежат на РАЗНЫХ
    reading-ах. Раньше брали один свежий с любым сальдо → теряли
    второй счёт (Галко: 209-долг = 26420.92 виден, 205-долг = 3125.50
    из старого reading «пропадал»).

    Фикс: отдельный поиск САМОГО СВЕЖЕГО reading где ненулевой 209,
    и отдельный — где ненулевой 205. balance_209/205 берутся независимо.

    balance_X > 0  → жилец должен по этому счёту
    balance_X < 0  → переплата по этому счёту
    balance_X == 0 → ноль
    """
    if not room_id:
        return {
            "balance_209": 0.0, "balance_205": 0.0, "total": 0.0,
            "kind": "no_room", "source_209_reading_id": None,
            "source_205_reading_id": None,
        }

    # Свежий reading с НЕНУЛЕВЫМ 209-сальдо
    latest_209 = (await db.execute(
        select(MeterReading).where(
            MeterReading.room_id == room_id,
            (MeterReading.debt_209 > 0) | (MeterReading.overpayment_209 > 0),
        ).order_by(MeterReading.created_at.desc()).limit(1)
    )).scalars().first()

    # Свежий reading с НЕНУЛЕВЫМ 205-сальдо (может быть тот же или другой)
    latest_205 = (await db.execute(
        select(MeterReading).where(
            MeterReading.room_id == room_id,
            (MeterReading.debt_205 > 0) | (MeterReading.overpayment_205 > 0),
        ).order_by(MeterReading.created_at.desc()).limit(1)
    )).scalars().first()

    if not latest_209 and not latest_205:
        return {
            "balance_209": 0.0, "balance_205": 0.0, "total": 0.0,
            "kind": "zero", "source_209_reading_id": None,
            "source_205_reading_id": None,
        }

    debt_209 = float(latest_209.debt_209 or 0) if latest_209 else 0.0
    overpay_209 = float(latest_209.overpayment_209 or 0) if latest_209 else 0.0
    debt_205 = float(latest_205.debt_205 or 0) if latest_205 else 0.0
    overpay_205 = float(latest_205.overpayment_205 or 0) if latest_205 else 0.0

    balance_209 = debt_209 - overpay_209
    balance_205 = debt_205 - overpay_205
    total = balance_209 + balance_205

    if total > 0:
        kind = "debtor"
    elif total < 0:
        kind = "overpaid"
    else:
        kind = "even"

    return {
        "balance_209": round(balance_209, 2),
        "balance_205": round(balance_205, 2),
        "total": round(total, 2),
        "kind": kind,
        "debt_209": debt_209,
        "overpayment_209": overpay_209,
        "debt_205": debt_205,
        "overpayment_205": overpay_205,
        "source_209_reading_id": latest_209.id if latest_209 else None,
        "source_205_reading_id": latest_205.id if latest_205 else None,
    }


# =====================================================================
# EXPLAIN — детальный пересчёт одного reading с трассировкой умножений
# =====================================================================
@router.get("/api/admin/readings/{reading_id}/explain")
async def explain_reading_calculation(
    reading_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Возвращает детальный breakdown расчёта одного MeterReading.

    Цель: админ может убедиться что счёт выставлен ПРАВИЛЬНО — видит
    каждое умножение тариф×объём, какие были предыдущие показания, какие
    корректировки применены, и совпадает ли пересчитанный итог с тем,
    что лежит в БД. Если не совпадает — пересчёт прав, в БД мусор.

    Используется кнопкой «Проверить расчёт» в админ-UI рядом с reading.
    """
    if current_user.role not in ("accountant", "admin", "financier"):
        raise HTTPException(403, "Доступ запрещён")

    # Весь основной блок завернём в try/except — если что-то падает на
    # подзадаче (битый тариф, отсутствующий period, etc), вернём 200 с
    # ключом explain_error вместо 500. UI тогда покажет красную плашку
    # с объяснением вместо обвинения «ошибка загрузки».
    try:
        return await _build_explain_response(reading_id, db)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"explain failed for reading_id={reading_id}: {e}")
        return {
            "explain_error": f"{type(e).__name__}: {e}",
            "reading_id": reading_id,
        }


async def _build_explain_response(reading_id: int, db: AsyncSession) -> dict:
    # 1. Reading с user, room, period
    reading = (await db.execute(
        select(MeterReading)
        .options(
            selectinload(MeterReading.user).selectinload(User.room),
            selectinload(MeterReading.period),
        )
        .where(MeterReading.id == reading_id)
    )).scalars().first()
    if not reading:
        raise HTTPException(404, "Reading не найден")

    user = reading.user
    room = user.room if user else None
    if not user or not room:
        raise HTTPException(400, "У reading'а нет связанного жильца или комнаты")

    # 2. Тариф через тот же кеш что использовался при расчёте
    from app.modules.utility.services.tariff_cache import tariff_cache
    tariff = tariff_cache.get_effective_tariff(user=user, room=room)
    if not tariff:
        tariff = (await db.execute(
            select(Tariff).where(Tariff.is_active.is_(True))
        )).scalars().first()
    if not tariff:
        raise HTTPException(400, "Активный тариф не найден")

    # 3. Предыдущее utverждённое показание для дельт. Ищем по period_id
    # (не created_at — инцидент may 2026 с подачами заднего числа через
    # гугл-таблицу). И ПРОПУСКАЕМ synth-prev (AUTO_GENERATED, DATA_OVERFLOW_RESET,
    # MANUAL_RECEIPT, AUTO_NO_HISTORY) — их значения = 0, использование как
    # baseline даёт фантастическую дельту в следующем периоде.
    # См. инцидент Капранов 2026-05-21: prev=AUTO_GENERATED с 0/0/0, текущее
    # 1468 ГВС → формула выдавала 818 049 ₽ как «правильный пересчёт».
    # selectinload(period) — чтобы prev.period.name не дёргал lazy-load в async.
    from app.modules.utility.services.reading_calculator import is_meaningful_prev
    prev_candidates = (await db.execute(
        select(MeterReading)
        .options(selectinload(MeterReading.period))
        .where(
            MeterReading.user_id == user.id,
            MeterReading.room_id == room.id,
            MeterReading.is_approved.is_(True),
            MeterReading.period_id < reading.period_id,
        )
        .order_by(MeterReading.period_id.desc())
        .limit(20)
    )).scalars().all()
    prev = next((c for c in prev_candidates if is_meaningful_prev(c)), None)

    # 4. Корректировки за период reading'а
    adj_rows = (await db.execute(
        select(Adjustment).where(
            Adjustment.user_id == user.id,
            Adjustment.period_id == reading.period_id,
        )
    )).scalars().all()

    # 5. Считаем дельты — те же формулы что в client_readings.py:
    #    delta_hot = current - prev (или current если prev=None)
    #    delta_sewage = delta_hot + delta_cold
    #    delta_elect_share = (residents/total_room) × delta_elect
    z = Decimal("0")
    cur_hot = Decimal(str(reading.hot_water or 0))
    cur_cold = Decimal(str(reading.cold_water or 0))
    cur_elect = Decimal(str(reading.electricity or 0))
    p_hot = Decimal(str(prev.hot_water or 0)) if prev else z
    p_cold = Decimal(str(prev.cold_water or 0)) if prev else z
    p_elect = Decimal(str(prev.electricity or 0)) if prev else z

    d_hot = cur_hot - p_hot
    d_cold = cur_cold - p_cold
    d_elect = cur_elect - p_elect
    d_sewage = d_hot + d_cold

    residents = Decimal(user.residents_count or 1)
    total_room = Decimal(room.total_room_residents or 1)
    if total_room <= 0:
        total_room = Decimal("1")
    elect_share = (residents / total_room) * d_elect

    area = Decimal(str(room.apartment_area or 0))

    # 6. Пересчёт через ту же calculate_utilities (для сравнения с БД).
    # Пробуем; если падает на CalculationError — показываем причину явно.
    from app.modules.utility.services.calculations import (
        calculate_utilities, CalculationError
    )
    calc_error = None
    calc_result = None
    is_baseline = prev is None
    # Сезонные флаги — «Проверить расчёт» обязан использовать тот же набор,
    # что и реальный /api/calculate: global + per-tariff.
    from app.modules.utility.routers.settings import _load_seasonal
    _seasonal = await _load_seasonal(db)
    _heating = _seasonal.heating_season_active and tariff.is_heating_active_now()
    _hw = _seasonal.hot_water_heating_active and tariff.is_hw_heating_active_now()
    try:
        # BASELINE: потребление-зависимые статьи (вода/свет) = 0 (счётчик
        # «накручен», дельту от 0 брать нельзя), НО area-based (содержание/
        # наём/ТКО/отопление) ПЛАТЯТСЯ ВСЕГДА (Bug L — см.
        # reading_calculator.compute_reading_breakdown, инцидент Резунов 04.2026).
        # Раньше тут был ХАРДКОД всех нулей → «Проверка расчёта» показывала
        # ЛОЖНОЕ расхождение на КАЖДОМ baseline (формула 0 vs БД area-based
        # ~5000-7000 ₽). Реальный биллинг и батч-анализатор периода считают
        # baseline правильно (area-based) — теперь и эта модалка тоже:
        # calculate_utilities с volume_*=0 для baseline.
        calc_result = calculate_utilities(
            user=user, room=room, tariff=tariff,
            volume_hot=(z if is_baseline else d_hot),
            volume_cold=(z if is_baseline else d_cold),
            volume_sewage=(z if is_baseline else d_sewage),
            volume_electricity_share=(z if is_baseline else elect_share),
            heating_season_active=_heating,
            hot_water_heating_active=_hw,
        )
    except CalculationError as e:
        calc_error = str(e)

    # 7. Формируем breakdown — по компонентам, с явными формулами.
    # Null-safe: tariff-поля могут быть None для редких/тестовых тарифов,
    # без guard это даёт Decimal('None') → InvalidOperation → 500.
    def f(d):
        """Decimal/None/число → строка с 2 знаками для ₽-сумм."""
        if d is None:
            return "0.00"
        try:
            return f"{Decimal(str(d)):.2f}"
        except Exception:
            return str(d)

    def f3(d):
        """Decimal/None/число → строка с 3 знаками для м³/кВт·ч."""
        if d is None:
            return "0.000"
        try:
            return f"{Decimal(str(d)):.3f}"
        except Exception:
            return str(d)

    def _dec_or_zero(value):
        """Безопасный Decimal: None и невалидные → ZERO."""
        if value is None:
            return ZERO
        try:
            return Decimal(str(value))
        except Exception:
            return ZERO

    components = []
    if not is_baseline and calc_result:
        # Холостяцкая квартира: счётчики (ГВС/ХВС/канализация) делятся на факт.
        # число жильцов, а area-based статьи считаются от «проектного места»
        # area / max_capacity. Здесь готовим текст-объяснение; сами суммы —
        # из calc_result (= calculate_utilities). См. её docstring.
        _is_singles = bool(getattr(room, "is_singles_apartment", False))
        if _is_singles:
            _cap = room.max_capacity if (room.max_capacity and int(room.max_capacity) > 0) else (room.total_room_residents or 1)
            _area_base = area / Decimal(str(_cap or 1))
            _area_expr = f"({f(area)} / {_cap})"
            _meter_div = f" ÷ {total_room}" if (total_room and int(total_room) > 1) else ""
            _meter_note = " ÷ жильцов" if _meter_div else ""
        else:
            _area_base = area
            _area_expr = f(area)
            _meter_div = ""
            _meter_note = ""
        # ГВС (Bug AQ/AR): формула обновлена под новую бизнес-логику —
        # water_heating уже включает в себя стоимость воды + подогрева,
        # поэтому НЕ суммируем с water_supply. Летняя профилактика
        # (hot_water_heating_active=False) — переключение на water_supply.
        t_w_sup = _dec_or_zero(tariff.water_supply)
        t_w_heat = _dec_or_zero(tariff.water_heating)
        if _hw:
            hot_formula = "v_hot × water_heating" + _meter_note
            hot_calc = f"{f3(d_hot)} × {f(t_w_heat)}{_meter_div}"
        else:
            hot_formula = "v_hot × water_supply (профилактика ГВС)" + _meter_note
            hot_calc = f"{f3(d_hot)} × {f(t_w_sup)}{_meter_div}"
        components.append({
            "label": "Горячая вода",
            "kbk": "209",
            "formula": hot_formula,
            "calculation": hot_calc,
            "result": f(calc_result["cost_hot_water"]) + " ₽",
        })
        # ХВС
        components.append({
            "label": "Холодная вода",
            "kbk": "209",
            "formula": "v_cold × water_supply" + _meter_note,
            "calculation": f"{f3(d_cold)} × {f(t_w_sup)}{_meter_div}",
            "result": f(calc_result["cost_cold_water"]) + " ₽",
        })
        # Канализация
        t_sewage = _dec_or_zero(tariff.sewage)
        components.append({
            "label": "Водоотведение",
            "kbk": "209",
            "formula": "(v_hot + v_cold) × sewage_rate" + _meter_note,
            "calculation": f"{f3(d_sewage)} × {f(t_sewage)}{_meter_div}",
            "result": f(calc_result["cost_sewage"]) + " ₽",
        })
        # Электро
        t_el = _dec_or_zero(tariff.electricity_rate)
        components.append({
            "label": "Электроэнергия",
            "kbk": "209",
            "formula": "(жильцов / всех в комнате) × delta_elect × rate",
            "calculation": (
                f"({residents} / {total_room}) × {f3(d_elect)} × "
                f"{f(t_el)} = {f3(elect_share)} × {f(t_el)}"
            ),
            "result": f(calc_result["cost_electricity"]) + " ₽",
        })
        # area-based статьи: для семьи база = вся площадь, для холостяка =
        # area / max_capacity («проектное место»), НЕ делится на жильцов.
        # Содержание
        t_maint = _dec_or_zero(tariff.maintenance_repair)
        components.append({
            "label": "Содержание и ремонт",
            "kbk": "205",
            "formula": "area_base × maintenance_repair",
            "calculation": f"{_area_expr} × {f(t_maint)}",
            "result": f(calc_result["cost_maintenance"]) + " ₽",
        })
        # Наём
        t_rent = _dec_or_zero(tariff.social_rent)
        components.append({
            "label": "Социальный найм",
            "kbk": "205",
            "formula": "area_base × social_rent",
            "calculation": f"{_area_expr} × {f(t_rent)}",
            "result": f(calc_result["cost_social_rent"]) + " ₽",
        })
        # ТКО
        t_waste = _dec_or_zero(tariff.waste_disposal)
        components.append({
            "label": "Вывоз ТКО",
            "kbk": "205",
            "formula": "area_base × waste_disposal",
            "calculation": f"{_area_expr} × {f(t_waste)}",
            "result": f(calc_result["cost_waste"]) + " ₽",
        })
        # Отопление. ОДН (electricity_per_sqm) удалён из системы 29.05.2026 —
        # фиксированная часть = только отопление.
        t_h = _dec_or_zero(tariff.heating)
        components.append({
            "label": "Отопление",
            "kbk": "205",
            "formula": "area_base × heating",
            "calculation": f"{_area_expr} × {f(t_h)}",
            "result": f(calc_result["cost_fixed_part"]) + " ₽",
        })

    # 8. Сравнение пересчитанного с тем что в БД.
    #
    # ВАЖНО (инцидент may 2026 — «Капранов 818k → 0»):
    # calc_total — это только НАЧИСЛЕНИЯ за период (cost_*).
    # reading.total_cost — это total_209 + total_205, где total_209 =
    # cost_209 + debt_209 - overpayment_209 + adjustments. То есть он
    # ВКЛЮЧАЕТ долги перенесённые из прошлого периода.
    #
    # Раньше сравнивали calc_total с total_cost напрямую — у baseline
    # с долгом 7 323 ₽ это давало «РАСХОЖДЕНИЕ -7 323». Ложная тревога.
    # Теперь сравниваем calc_total с СУММОЙ cost_* полей (= чистые
    # начисления без долгов). Если они равны → БД актуальна,
    # даже если total_cost ≠ calc_total из-за переноса долга.
    def _dec(x):
        return Decimal(str(x or 0))
    stored_total = _dec(reading.total_cost)
    stored_cost_pure = (
        _dec(reading.cost_hot_water)
        + _dec(reading.cost_cold_water)
        + _dec(reading.cost_sewage)
        + _dec(reading.cost_electricity)
        + _dec(reading.cost_maintenance)
        + _dec(reading.cost_social_rent)
        + _dec(reading.cost_waste)
        + _dec(reading.cost_fixed_part)
    )
    calc_total = (
        Decimal(str(calc_result["total_cost"])) if calc_result else None
    )
    # match по чистым начислениям — это и есть «формула актуальна?»
    match = (
        calc_total is not None
        and abs(calc_total - stored_cost_pure) < Decimal("0.02")
    )
    # Разница между total_cost и stored_cost_pure = долги/переплаты/коррекции,
    # перенесённые из прошлых периодов. Не баг расчёта.
    carried_balance = stored_total - stored_cost_pure

    return {
        "reading": {
            "id": reading.id,
            "is_approved": bool(reading.is_approved),
            "anomaly_flags": reading.anomaly_flags,
            "anomaly_score": reading.anomaly_score,
            "created_at": reading.created_at.isoformat() if reading.created_at else None,
            "is_baseline": is_baseline,
        },
        "user": {
            "id": user.id,
            "username": user.username,
            "residents_count": user.residents_count or 1,
            "billing_mode": getattr(user, "billing_mode", "by_meter"),
        },
        "room": {
            "id": room.id,
            "dormitory_name": room.dormitory_name,
            "room_number": room.room_number,
            "apartment_area": f(area),
            "total_room_residents": room.total_room_residents or 1,
        },
        "period": {"id": reading.period_id, "name": reading.period.name if reading.period else None},
        "tariff": {
            "id": tariff.id,
            "name": tariff.name,
            "rates": {
                "water_supply": f(tariff.water_supply),
                "water_heating": f(tariff.water_heating),
                "sewage": f(tariff.sewage),
                "electricity_rate": f(tariff.electricity_rate),
                "maintenance_repair": f(tariff.maintenance_repair),
                "social_rent": f(tariff.social_rent),
                "waste_disposal": f(tariff.waste_disposal),
                "heating": f(tariff.heating),
            },
        },
        "previous_reading": (
            {
                "reading_id": prev.id,
                "period_name": prev.period.name if prev.period else None,
                "hot_water": f3(p_hot),
                "cold_water": f3(p_cold),
                "electricity": f3(p_elect),
                "created_at": prev.created_at.isoformat() if prev.created_at else None,
            } if prev else None
        ),
        "current_values": {
            "hot_water": f3(cur_hot),
            "cold_water": f3(cur_cold),
            "electricity": f3(cur_elect),
        },
        "deltas": {
            "hot_water": f3(d_hot),
            "cold_water": f3(d_cold),
            "electricity": f3(d_elect),
            "electricity_share": f3(elect_share),
            "sewage": f3(d_sewage),
        },
        "components": components,
        "adjustments": [
            {
                "id": a.id,
                "amount": f(a.amount),
                "description": a.description,
                "kbk": a.account_type,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            }
            for a in adj_rows
        ],
        "balances_carried_in": {
            "debt_209": f(reading.debt_209 or 0),
            "overpayment_209": f(reading.overpayment_209 or 0),
            "debt_205": f(reading.debt_205 or 0),
            "overpayment_205": f(reading.overpayment_205 or 0),
        },
        "totals": {
            "calculated_total_cost": f(calc_total) if calc_total is not None else None,
            "stored_total_cost": f(stored_total),
            "stored_cost_pure": f(stored_cost_pure),
            "stored_total_209": f(reading.total_209 or 0),
            "stored_total_205": f(reading.total_205 or 0),
            # match сравнивает calc с чистыми начислениями (без долгов).
            # Если true — формула актуальна. Если false — реально надо пересчитать.
            "match": match,
            # Разница между total_cost и чистыми cost — это переносы баланса
            # (debt_209/205, переплаты, корректировки). НЕ баг расчёта.
            "carried_balance": f(carried_balance),
            # Старое поле для обратной совместимости фронта. Не использовать
            # для match-логики — используйте diff_calc_minus_pure_cost.
            "diff_calc_minus_stored": (
                f(calc_total - stored_total) if calc_total is not None else None
            ),
            "diff_calc_minus_pure_cost": (
                f(calc_total - stored_cost_pure) if calc_total is not None else None
            ),
        },
        "sanity_warning": (
            calc_result.get("sanity_warning") if calc_result else None
        ),
        "calculation_error": calc_error,
    }
