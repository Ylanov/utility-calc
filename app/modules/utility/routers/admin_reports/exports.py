# Экспорты Excel: сводная ведомость периода и выгрузка начислений в 1С.
# Вербатим-перенос из admin_reports.py (строки 232-480), поведение 1:1.

import io
from decimal import Decimal
from typing import List, Optional
from urllib.parse import quote

from fastapi import Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from openpyxl import Workbook

from app.core.database import get_db
from app.core.dependencies import get_current_user
from app.modules.utility.models import (
    User, MeterReading, BillingPeriod, Room, RentalContract,
)

from ._shared import ZERO, _report_group, router


@router.get("/api/admin/export_report", summary="Скачать отчет Excel (XLSX)")
async def export_report(
        period_id: Optional[int] = Query(None, description="ID периода"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db)
):
    if current_user.role not in ("accountant", "admin"):
        raise HTTPException(status_code=403)

    target_period_id = period_id
    if not target_period_id:
        active_period = (
            await db.execute(select(BillingPeriod).where(BillingPeriod.is_active.is_(True)))).scalars().first()
        if active_period:
            target_period_id = active_period.id
        else:
            last_closed = (
                await db.execute(select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1))).scalars().first()
            if not last_closed:
                raise HTTPException(404, "Нет периодов для отчета")
            target_period_id = last_closed.id

    period = await db.get(BillingPeriod, target_period_id)
    if not period: raise HTTPException(404, "Выбранный период не найден")

    # ИСПРАВЛЕНИЕ: Используем потоковое чтение (yield_per) из БД, чтобы не грузить все в RAM
    statement = (
        select(User, MeterReading, Room)
        .join(MeterReading, User.id == MeterReading.user_id)
        .join(Room, User.room_id == Room.id)
        .where(
            MeterReading.period_id == target_period_id,
            MeterReading.is_approved.is_(True)
        )
        .order_by(Room.dormitory_name, Room.room_number, User.username)
    ).execution_options(yield_per=1000)

    # ИСПРАВЛЕНИЕ: Используем write_only=True для потоковой записи Excel
    # (openpyxl не хранит документ в памяти, а пишет напрямую в байтовый буфер)
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Сводная ведомость")

    headers = ["Общежитие/Комната", "ФИО (Логин)", "Площадь", "Жильцов", "ГВС (руб)", "ХВС (руб)", "Водоотв. (руб)",
               "Электроэнергия (руб)", "Содержание (руб)", "Наем (руб)", "ТКО (руб)", "Отопление + ОДН (руб)",
               "Счет 209 (Комм.)", "Счет 205 (Найм)", "ИТОГО (руб)"]
    worksheet.append(headers)

    total_sum, total_209, total_205 = ZERO, ZERO, ZERO

    # Потоковое получение результатов
    result = await db.stream(statement)

    async for row in result:
        user, reading, room = row
        total_cost = Decimal(reading.total_cost or 0)
        t_209 = Decimal(reading.total_209 or 0)
        t_205 = Decimal(reading.total_205 or 0)

        total_sum += total_cost
        total_209 += t_209
        total_205 += t_205

        worksheet.append([
            room.format_address,
            user.username.split("_deleted_")[0] if user.is_deleted else user.username,
            room.apartment_area,
            str(room.total_room_residents or 1),
            reading.cost_hot_water, reading.cost_cold_water, reading.cost_sewage, reading.cost_electricity,
            reading.cost_maintenance, reading.cost_social_rent, reading.cost_waste, reading.cost_fixed_part,
            t_209, t_205, total_cost
        ])

    worksheet.append([""] * 12 + ["ИТОГО:", total_209, total_205, total_sum])
    filename = f"Report_{period.name.replace(' ', '_')}.xlsx"
    encoded_filename = quote(filename)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        headers={'Content-Disposition': f"attachment; filename*=utf-8''{encoded_filename}"},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _fmt_contract_1c(rc) -> str:
    """Строка договора для 1С: «Договор от ДД.ММ.ГГГГ № N». Пусто — если нет."""
    if not rc:
        return ""
    num = (rc.number or "").strip()
    dt = rc.signed_date.strftime("%d.%m.%Y") if rc.signed_date else ""
    if num and dt:
        return f"Договор от {dt} № {num}"
    if num:
        return f"Договор № {num}"
    if dt:
        return f"Договор от {dt}"
    return ""


@router.get("/api/admin/export-1c/groups", summary="Список домов/общаг для выгрузки в 1С")
async def export_1c_groups(
        period_id: Optional[int] = Query(None),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db),
):
    """Список зданий (общага/дом) с числом жильцов за период — для модалки
    выбора «за какой дом/общагу выгрузить в 1С»."""
    if current_user.role not in ("accountant", "admin"):
        raise HTTPException(status_code=403)
    target_period_id = period_id
    if not target_period_id:
        ap = (await db.execute(select(BillingPeriod).where(BillingPeriod.is_active.is_(True)))).scalars().first()
        target_period_id = ap.id if ap else None
    if not target_period_id:
        return {"groups": []}
    rows = (await db.execute(
        select(Room, func.count(MeterReading.id))
        .join(MeterReading, MeterReading.room_id == Room.id)
        .join(User, User.id == MeterReading.user_id)
        .where(MeterReading.period_id == target_period_id,
               MeterReading.is_approved.is_(True), User.is_deleted.is_(False))
        .group_by(Room.id)
    )).all()
    agg: dict = {}
    for room, cnt in rows:
        g = _report_group(room)
        agg[g] = agg.get(g, 0) + int(cnt)
    groups = sorted(({"name": k, "count": v} for k, v in agg.items()), key=lambda x: x["name"])
    return {"period_id": target_period_id, "groups": groups}


@router.get("/api/admin/export-1c", summary="Выгрузка в 1С (XLSX: Контрагент/Договор/Сумма по счёту 209 или 205)")
async def export_1c(
        period_id: Optional[int] = Query(None, description="ID периода"),
        group: Optional[List[str]] = Query(None, description="Дом/общага (имя блока) — можно несколько, выгрузить только по ним"),
        account: Optional[str] = Query(None, description="Счёт: '209' (коммуналка) или '205' (наём). Пусто = 209+205 вместе (легаси)"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db),
):
    """Выгрузка начислений за месяц в формате загрузки 1С. На каждого жильца с
    утверждённым показанием: ФИО (Контрагент), договор найма, Количество=1,
    Сумма = начисление за месяц ПО ВЫБРАННОМУ СЧЁТУ (209 — коммуналка, 205 —
    наём; для бухгалтерии счета грузятся раздельно). Нулевые по выбранному счёту
    строки пропускаем. Колонки документа-основания/статуса пустые — заполняет 1С."""
    if current_user.role not in ("accountant", "admin"):
        raise HTTPException(status_code=403)

    target_period_id = period_id
    if not target_period_id:
        active_period = (await db.execute(
            select(BillingPeriod).where(BillingPeriod.is_active.is_(True)))).scalars().first()
        if active_period:
            target_period_id = active_period.id
        else:
            last_closed = (await db.execute(
                select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1))).scalars().first()
            if not last_closed:
                raise HTTPException(404, "Нет периодов для выгрузки")
            target_period_id = last_closed.id

    period = await db.get(BillingPeriod, target_period_id)
    if not period:
        raise HTTPException(404, "Выбранный период не найден")

    rows = (await db.execute(
        select(User, MeterReading, Room)
        .join(MeterReading, User.id == MeterReading.user_id)
        .join(Room, User.room_id == Room.id)
        .where(
            MeterReading.period_id == target_period_id,
            MeterReading.is_approved.is_(True),
            User.is_deleted.is_(False),
        )
        .order_by(User.username)
    )).all()

    # Фильтр по дому/общаге (можно несколько — галочки в модалке).
    if group:
        wanted = {g for g in group if g}
        if wanted:
            rows = [(u, mr, room) for (u, mr, room) in rows if _report_group(room) in wanted]

    # Активные договоры найма одним запросом (последний по дате подписания).
    uids = [u.id for u, _, _ in rows]
    contracts: dict[int, RentalContract] = {}
    if uids:
        for rc in (await db.execute(
            select(RentalContract)
            .where(RentalContract.user_id.in_(set(uids)), RentalContract.is_active.is_(True))
            .order_by(RentalContract.id.asc())
        )).scalars().all():
            contracts[rc.user_id] = rc  # asc по id → последний перезапишет = самый свежий

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист_1"
    # Колонка «N» убрана (2026-06-18) — не нужна для загрузки 1С.
    worksheet.append([
        "Контрагент", "Договор", "Количество", "Сумма",
        "Вид документа-основания", "Номер (108)", "Дата (109)",
        "Отразить в графике исполнения договора", "Статус платежа",
    ])

    _acc = (account or "").strip()
    for user, reading, _room in rows:
        t209 = Decimal(reading.total_209 or 0)
        t205 = Decimal(reading.total_205 or 0)
        if _acc == "209":
            amount = t209
        elif _acc == "205":
            amount = t205
        else:
            amount = t209 + t205
        # При выборе конкретного счёта нулевые строки пропускаем (нет начисления
        # по этому счёту — нечего грузить).
        if _acc in ("209", "205") and amount == 0:
            continue
        # Сумма — СТРОКОЙ с ТОЧКОЙ-разделителем (1С требует точку, не запятую).
        # Decimal-str всегда даёт точку независимо от локали Excel.
        sum_str = str(amount.quantize(Decimal("0.01")))
        worksheet.append([
            user.username,
            _fmt_contract_1c(contracts.get(user.id)),
            1,
            sum_str,
            "", "", "", "", "",
        ])

    _picked = [g for g in (group or []) if g]
    _acc_suffix = f"_sch{_acc}" if _acc in ("209", "205") else ""
    if len(_picked) == 1:
        _suffix = f"{_acc_suffix}_{_picked[0]}"
    elif len(_picked) > 1:
        _suffix = f"{_acc_suffix}_vyborka_{len(_picked)}"
    else:
        _suffix = _acc_suffix
    filename = f"Vygruzka_1C_{period.name.replace(' ', '_')}{_suffix}.xlsx"
    encoded_filename = quote(filename)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        headers={'Content-Disposition': f"attachment; filename*=utf-8''{encoded_filename}"},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
