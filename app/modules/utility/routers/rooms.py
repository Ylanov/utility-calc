# app/modules/utility/routers/rooms.py
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, or_, update, and_
from typing import Optional, List

from app.core.database import get_db
from app.core.time_utils import utcnow
from app.modules.utility.models import Room, User, MeterReading, BillingPeriod, Tariff
from app.modules.utility.schemas import RoomCreate, RoomUpdate, RoomResponse, PaginatedResponse, ReplaceMeterSchema, RoomMeterConfigBulk
from app.core.dependencies import get_current_user, RoleChecker
from app.modules.utility.services.calculations import calculate_utilities

router = APIRouter(prefix="/api/rooms", tags=["Housing"])
allow_management = RoleChecker(["accountant", "admin", "financier"])

ZERO = Decimal("0.00")


@router.get("/analyze", summary="Мощный анализатор Жилфонда", dependencies=[Depends(allow_management)])
async def analyze_housing(db: AsyncSession = Depends(get_db)):
    """
    Сканирует весь жилфонд и пользователей на наличие аномалий:
    - Раздельные лицевые счета в одной комнате (Холостяки)
    - Перенаселение (платят за большее кол-во человек, чем вмещает комната)
    - Недобор / Пустые места
    - Жильцы без комнаты (ошибки привязки)
    - Комнаты с нулевой площадью
    """
    # 1. Получаем все комнаты
    rooms_res = await db.execute(select(Room))
    rooms = rooms_res.scalars().all()

    # 2. Получаем всех активных жильцов
    users_res = await db.execute(select(User).where(User.is_deleted.is_(False)))
    users = users_res.scalars().all()

    # 3. Группируем жильцов по комнатам
    room_users = {}
    unattached_users = []

    for u in users:
        if not u.room_id:
            unattached_users.append(u)
        else:
            if u.room_id not in room_users:
                room_users[u.room_id] = []
            room_users[u.room_id].append(u)

    issues = {
        "shared_billing": [],  # Раздельные счета (холостяки)
        "overcrowded": [],  # Платят за больше человек, чем мест
        "underpopulated": [],  # Платят за меньше человек, чем мест (пустующие места)
        "zero_area": [],  # Нулевая площадь
        "empty_rooms": [],  # Пустые комнаты
        "unattached_users": []  # Жильцы без комнаты
    }

    for room in rooms:
        # housing_001/E2-A: универсальный адрес (общага / дом).
        r_name = room.format_address
        r_id = room.id

        # Ошибка: нулевая площадь
        if room.apartment_area <= 0:
            issues["zero_area"].append({"id": r_id, "title": r_name,
                                        "desc": "Указана нулевая площадь. Коммуналка по нормативу не будет начислена корректно."})

        occupants = room_users.get(room.id, [])

        # Инфо: Пустая комната
        if not occupants:
            issues["empty_rooms"].append({"id": r_id, "title": r_name, "desc": "Никто не прописан (нет активных Л/С)"})
            continue

        acc_count = len(occupants)
        # Число людей и вместимость берём из КОМНАТЫ (per-user residents_count
        # упразднён 2026-06-17): people = total_room_residents (для холостяцкой
        # = число Л/С, авто; для семьи = размер семьи), places = max_capacity.
        people = room.total_room_residents or 0
        places = room.max_capacity if (room.max_capacity and room.max_capacity > 0) else None
        usernames = ", ".join([u.username for u in occupants])

        # Аномалия 1: Холостяки / Раздельные счета (2 и более Л/С в одной комнате)
        if acc_count > 1:
            issues["shared_billing"].append({
                "id": r_id, "title": r_name,
                "desc": f"Раздельные счета ({acc_count} Л/С) в одном помещении: {usernames}. Убедитесь, что это не дубликаты."
            })

        # Аномалия 2: Перенаселение (людей больше, чем мест по проекту)
        if places and people > places:
            issues["overcrowded"].append({
                "id": r_id, "title": r_name,
                "desc": f"Проживает {people} чел., хотя максимальная вместимость квартиры {places}. ({usernames})"
            })

        # Аномалия 3: Недобор (есть свободные места)
        elif places and people < places:
            issues["underpopulated"].append({
                "id": r_id, "title": r_name,
                "desc": f"Проживает {people} чел., а мест {places}. Возможно, есть свободные места. ({usernames})"
            })

    # Аномалия 4: Жильцы-призраки
    for u in unattached_users:
        issues["unattached_users"].append({"id": u.id, "title": u.username,
                                           "desc": "Не привязан ни к одной комнате (Начисления по счетчикам невозможны)"})

    return issues


@router.get("/dormitories", response_model=List[str], dependencies=[Depends(get_current_user)])
async def get_dormitories(db: AsyncSession = Depends(get_db)):
    """Список уникальных названий общежитий (только place_type='dormitory')."""
    result = await db.execute(
        select(Room.dormitory_name)
        .where(
            Room.place_type == "dormitory",
            Room.dormitory_name.is_not(None),
        )
        .distinct()
        .order_by(Room.dormitory_name)
    )
    return [v for v in result.scalars().all() if v]


@router.get("/streets", response_model=List[str], dependencies=[Depends(get_current_user)])
async def get_streets(db: AsyncSession = Depends(get_db)):
    """Список уникальных улиц для домов/квартир (place_type='house').

    Используется фронтом Жилфонда для автокомплита поля «Улица» в форме
    добавления дома, аналогично /dormitories для общаг.
    """
    result = await db.execute(
        select(Room.street)
        .where(Room.place_type == "house", Room.street.is_not(None))
        .distinct()
        .order_by(Room.street)
    )
    return [v for v in result.scalars().all() if v]


@router.get("/buildings", dependencies=[Depends(allow_management)])
async def get_buildings(db: AsyncSession = Depends(get_db)):
    """Все ЗДАНИЯ для «Настройки дома»: общаги (dormitory_name) + дома
    (street + house_number) с числом помещений. Выбор здания в модалке настроек —
    тариф/счётчики ставятся на всё здание сразу (на квартире тариф read-only)."""
    dorm_rows = (await db.execute(
        select(Room.dormitory_name, func.count(Room.id))
        .where(Room.place_type == "dormitory", Room.dormitory_name.is_not(None))
        .group_by(Room.dormitory_name)
        .order_by(Room.dormitory_name)
    )).all()
    house_rows = (await db.execute(
        select(Room.street, Room.house_number, func.count(Room.id))
        .where(Room.place_type == "house",
               Room.street.is_not(None), Room.house_number.is_not(None))
        .group_by(Room.street, Room.house_number)
        .order_by(Room.street, Room.house_number)
    )).all()
    out = []
    for name, cnt in dorm_rows:
        if name:
            out.append({"type": "dormitory", "dormitory_name": name,
                        "label": f"🏢 {name}", "rooms": cnt})
    for street, house, cnt in house_rows:
        if street and house:
            out.append({"type": "house", "street": street, "house_number": house,
                        "label": f"🏠 ул. {street}, д. {house}", "rooms": cnt})
    return out


@router.get("", response_model=PaginatedResponse[RoomResponse], dependencies=[Depends(allow_management)])
async def get_rooms(
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=1000),
        search: Optional[str] = Query(None),
        dormitory: Optional[str] = Query(None),
        # housing_001 (рефакторинг Жилфонда): фильтр по типу помещения.
        # Если задан — отдаём только этот тип, иначе всё подряд.
        place_type: Optional[str] = Query(
            None, pattern="^(dormitory|house)$",
            description="Фильтр по типу помещения: dormitory|house",
        ),
        street: Optional[str] = Query(
            None, description="Точное название улицы (для домов)",
        ),
        house_number: Optional[str] = Query(
            None, description="Номер дома (для домов; вместе со street = конкретный дом)",
        ),
        # Новые фильтры — без них невозможно сделать быстрый drill-down
        # по состоянию комнаты. У админа в проекте на 2000+ комнат без
        # фильтров работать нельзя.
        occupancy: Optional[str] = Query(
            None, pattern="^(empty|partial|full|overcrowded)$",
            description="empty=0 жильцов, partial<вместимости, full=, overcrowded>",
        ),
        missing_meter: Optional[bool] = Query(
            None, description="Отсутствует хотя бы один из серийников счётчиков",
        ),
        no_meters: Optional[bool] = Query(
            None,
            description="Квартиры БЕЗ счётчиков вообще (has_hw/cw/el_meter = все false). "
                        "Им не начисляется норматив по приборам и НЕ применяется санкция ×3. "
                        "Композится с place_type (напр. dormitory).",
        ),
        db: AsyncSession = Depends(get_db)
):
    """Список комнат с пагинацией и расширенными фильтрами."""
    query = select(Room)
    count_query = select(func.count(Room.id))

    if search:
        search_filter = f"%{search}%"
        # Поиск работает по полям ОБОИХ типов адреса — общага и дом —
        # чтобы админ в едином поиске мог найти и «комн. 405» и
        # «ул. Ленина 5 кв.12». Серийники счётчиков тоже сохраняем
        # (для общаг это удобный способ найти комнату по серийнику).
        condition = or_(
            Room.room_number.ilike(search_filter),
            Room.dormitory_name.ilike(search_filter),
            Room.street.ilike(search_filter),
            Room.house_number.ilike(search_filter),
            Room.apartment_number.ilike(search_filter),
            Room.hw_meter_serial.ilike(search_filter),
            Room.cw_meter_serial.ilike(search_filter),
        )
        query = query.where(condition)
        count_query = count_query.where(condition)

    if place_type:
        query = query.where(Room.place_type == place_type)
        count_query = count_query.where(Room.place_type == place_type)

    if dormitory:
        query = query.where(Room.dormitory_name == dormitory)
        count_query = count_query.where(Room.dormitory_name == dormitory)

    if street:
        query = query.where(Room.street == street)
        count_query = count_query.where(Room.street == street)

    if house_number:
        query = query.where(Room.house_number == house_number)
        count_query = count_query.where(Room.house_number == house_number)

    # Фильтр по заполненности: коррелируем через subquery количества жильцов
    if occupancy:
        residents_subq = (
            select(func.count(User.id))
            .where(User.room_id == Room.id, User.is_deleted.is_(False), User.role == "user")
            .correlate(Room)
            .scalar_subquery()
        )
        cap = func.coalesce(Room.total_room_residents, 1)
        if occupancy == "empty":
            cond = residents_subq == 0
        elif occupancy == "partial":
            cond = (residents_subq > 0) & (residents_subq < cap)
        elif occupancy == "full":
            cond = residents_subq == cap
        else:  # overcrowded
            cond = residents_subq > cap
        query = query.where(cond)
        count_query = count_query.where(cond)

    if missing_meter is True:
        cond = (
            (Room.hw_meter_serial.is_(None)) | (Room.hw_meter_serial == "")
            | (Room.cw_meter_serial.is_(None)) | (Room.cw_meter_serial == "")
            | (Room.el_meter_serial.is_(None)) | (Room.el_meter_serial == "")
        )
        query = query.where(cond)
        count_query = count_query.where(cond)

    if no_meters is True:
        # «Квартиры без счётчиков» — учёт по образцу домов. Комната, где НЕТ
        # ни одного прибора (все has_*_meter явно false). Им норматив по
        # счётчикам не начисляется и санкция ×3 исключена (_growing_norm_volumes).
        cond = (
            Room.has_hw_meter.is_(False)
            & Room.has_cw_meter.is_(False)
            & Room.has_el_meter.is_(False)
        )
        query = query.where(cond)
        count_query = count_query.where(cond)

    total = (await db.execute(count_query)).scalar_one()

    query = query.order_by(Room.dormitory_name, Room.room_number).offset((page - 1) * limit).limit(limit)
    items = (await db.execute(query)).scalars().all()

    return {"total": total, "page": page, "size": limit, "items": items}


# =====================================================================
# STATS — сводка по Жилфонду
# =====================================================================
@router.get("/stats", dependencies=[Depends(allow_management)])
async def housing_stats(
    dormitory: Optional[str] = Query(None),
    street: Optional[str] = Query(None),
    house_number: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """KPI для шапки вкладки «Жилфонд».

    Возвращает:
      * total_rooms / empty / partial / full / overcrowded
      * total_area / avg_area
      * total_capacity (суммарные места) / total_residents (сколько жильцов)
      * missing_meters_count (комнаты с неполным набором счётчиков)
      * by_dormitory (разбивка)

    Если передан `dormitory` — все метрики считаются в рамках этого общежития.
    """
    base = select(Room)
    if dormitory:
        base = base.where(Room.dormitory_name == dormitory)
    if street:
        base = base.where(Room.street == street)
    if house_number:
        base = base.where(Room.house_number == house_number)

    # Все комнаты с кол-вом проживающих одним запросом
    residents_subq = (
        select(
            User.room_id.label("rid"),
            func.count(User.id).label("residents"),
        )
        .where(User.is_deleted.is_(False), User.role == "user", User.room_id.is_not(None))
        .group_by(User.room_id)
        .subquery()
    )
    rows_q = (
        select(
            Room.id,
            Room.dormitory_name,
            Room.apartment_area,
            Room.total_room_residents,
            Room.hw_meter_serial,
            Room.cw_meter_serial,
            Room.el_meter_serial,
            func.coalesce(residents_subq.c.residents, 0).label("residents"),
        )
        .outerjoin(residents_subq, residents_subq.c.rid == Room.id)
    )
    if dormitory:
        rows_q = rows_q.where(Room.dormitory_name == dormitory)
    if street:
        rows_q = rows_q.where(Room.street == street)
    if house_number:
        rows_q = rows_q.where(Room.house_number == house_number)

    rows = (await db.execute(rows_q)).all()

    total = len(rows)
    empty = partial = full = overcrowded = 0
    total_area = Decimal("0")
    total_capacity = 0
    total_residents = 0
    missing_meters = 0
    by_dorm: dict = {}

    for rid, dorm, area, capacity, hw, cw, el, residents in rows:
        cap = int(capacity or 1)
        r = int(residents or 0)
        if r == 0:
            empty += 1
        elif r < cap:
            partial += 1
        elif r == cap:
            full += 1
        else:
            overcrowded += 1

        a = Decimal(str(area or 0))
        total_area += a
        total_capacity += cap
        total_residents += r

        if not hw or not cw or not el:
            missing_meters += 1

        d = by_dorm.setdefault(dorm or "—", {
            "rooms": 0, "residents": 0, "capacity": 0, "area": Decimal("0"),
        })
        d["rooms"] += 1
        d["residents"] += r
        d["capacity"] += cap
        d["area"] += a

    avg_area = float(total_area / total) if total else 0.0
    occupancy_pct = (
        round(total_residents / total_capacity * 100) if total_capacity else 0
    )

    return {
        "total_rooms": total,
        "empty": empty,
        "partial": partial,
        "full": full,
        "overcrowded": overcrowded,
        "total_area": float(total_area),
        "avg_area": round(avg_area, 2),
        "total_capacity": total_capacity,
        "total_residents": total_residents,
        "occupancy_pct": occupancy_pct,
        "free_slots": max(0, total_capacity - total_residents),
        "missing_meters_count": missing_meters,
        "by_dormitory": [
            {
                "name": name, "rooms": d["rooms"], "residents": d["residents"],
                "capacity": d["capacity"], "area": float(d["area"]),
                "occupancy_pct": (
                    round(d["residents"] / d["capacity"] * 100) if d["capacity"] else 0
                ),
            }
            for name, d in sorted(by_dorm.items())
        ],
    }


# =====================================================================
# EXPORT — Excel со списком комнат (с текущими фильтрами)
# =====================================================================
@router.get("/export", dependencies=[Depends(allow_management)])
async def export_rooms(
    search: Optional[str] = Query(None),
    dormitory: Optional[str] = Query(None),
    street: Optional[str] = Query(None),
    house_number: Optional[str] = Query(None),
    occupancy: Optional[str] = Query(None, pattern="^(empty|partial|full|overcrowded)$"),
    missing_meter: Optional[bool] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Excel-выгрузка всех отфильтрованных комнат + количество жильцов."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # Собираем тот же запрос что и для списка — но без пагинации
    residents_subq = (
        select(
            User.room_id.label("rid"),
            func.count(User.id).label("residents"),
        )
        .where(User.is_deleted.is_(False), User.role == "user", User.room_id.is_not(None))
        .group_by(User.room_id)
        .subquery()
    )
    q = (
        select(
            Room,
            func.coalesce(residents_subq.c.residents, 0).label("residents"),
            Tariff.name.label("tariff_name"),
        )
        .outerjoin(residents_subq, residents_subq.c.rid == Room.id)
        .outerjoin(Tariff, Tariff.id == Room.tariff_id)
    )
    if search:
        pat = f"%{search}%"
        q = q.where(or_(
            Room.room_number.ilike(pat),
            Room.hw_meter_serial.ilike(pat),
            Room.cw_meter_serial.ilike(pat),
        ))
    if dormitory:
        q = q.where(Room.dormitory_name == dormitory)
    if street:
        q = q.where(Room.street == street)
    if house_number:
        q = q.where(Room.house_number == house_number)
    if occupancy:
        cap = func.coalesce(Room.total_room_residents, 1)
        if occupancy == "empty":
            q = q.where(residents_subq.c.residents.is_(None) | (residents_subq.c.residents == 0))
        elif occupancy == "partial":
            q = q.where((residents_subq.c.residents > 0) & (residents_subq.c.residents < cap))
        elif occupancy == "full":
            q = q.where(residents_subq.c.residents == cap)
        elif occupancy == "overcrowded":
            q = q.where(residents_subq.c.residents > cap)
    if missing_meter is True:
        q = q.where(
            (Room.hw_meter_serial.is_(None)) | (Room.hw_meter_serial == "")
            | (Room.cw_meter_serial.is_(None)) | (Room.cw_meter_serial == "")
            | (Room.el_meter_serial.is_(None)) | (Room.el_meter_serial == "")
        )
    q = q.order_by(Room.dormitory_name, Room.room_number)

    rows = (await db.execute(q)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Жилфонд"
    headers = [
        "ID", "Общежитие", "Комната", "Площадь м²", "Мест", "Жильцов",
        "Заполненность", "Тариф", "№ ГВС", "№ ХВС", "№ Электр.",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DBEAFE")
    for i, (room, residents, tariff_name) in enumerate(rows, 2):
        cap = int(room.total_room_residents or 1)
        r = int(residents or 0)
        status = (
            "Переполнена" if r > cap
            else "Полная" if r == cap
            else "Частичная" if r > 0
            else "Пустая"
        )
        ws.cell(row=i, column=1, value=room.id)
        ws.cell(row=i, column=2, value=room.dormitory_name)
        ws.cell(row=i, column=3, value=room.room_number)
        ws.cell(row=i, column=4, value=float(room.apartment_area or 0))
        ws.cell(row=i, column=5, value=cap)
        ws.cell(row=i, column=6, value=r)
        ws.cell(row=i, column=7, value=status)
        ws.cell(row=i, column=8, value=tariff_name or "")
        ws.cell(row=i, column=9, value=room.hw_meter_serial or "")
        ws.cell(row=i, column=10, value=room.cw_meter_serial or "")
        ws.cell(row=i, column=11, value=room.el_meter_serial or "")

    for col, w in [("A", 6), ("B", 28), ("C", 10), ("D", 10), ("E", 6),
                   ("F", 8), ("G", 14), ("H", 22), ("I", 14), ("J", 14), ("K", 14)]:
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"housing_{utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("", response_model=RoomResponse, dependencies=[Depends(allow_management)])
async def create_room(data: RoomCreate, db: AsyncSession = Depends(get_db)):
    """Создаёт новое помещение (общежитие или дом/квартира).

    Тип определяет схема валидации: RoomCreate@model_validator уже
    обнулил поля «не своего» типа и проверил обязательные. Здесь
    проверяем только бизнес-уникальность адреса по нужным колонкам.
    """
    if data.place_type == "dormitory":
        exist = await db.execute(select(Room).where(
            Room.place_type == "dormitory",
            Room.dormitory_name == data.dormitory_name,
            Room.room_number == data.room_number,
        ))
        if exist.scalars().first():
            raise HTTPException(
                status_code=400,
                detail="Такая комната в этом общежитии уже существует",
            )
    else:  # 'house'
        exist = await db.execute(select(Room).where(
            Room.place_type == "house",
            Room.street == data.street,
            Room.house_number == data.house_number,
            Room.apartment_number == data.apartment_number,
        ))
        if exist.scalars().first():
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Квартира {data.apartment_number} в доме "
                    f"{data.house_number} на улице {data.street} уже есть"
                ),
            )

    if getattr(data, "is_singles_apartment", False) and (
        data.max_capacity is None or int(data.max_capacity) < 1
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                "Для холостяцкой квартиры укажите «Макс. вместимость» (≥ 1) — "
                "по ней делится площадь для найма / ТКО / отопления / содержания."
            ),
        )

    room = Room(**data.model_dump())
    db.add(room)
    await db.commit()
    await db.refresh(room)
    return room


@router.get("/{room_id}/residents", dependencies=[Depends(allow_management)])
async def room_residents(
    room_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Список жильцов комнаты + последнее утверждённое показание.
    Используется раскрывающимся блоком в таблице Жилфонда."""
    from sqlalchemy.orm import selectinload

    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(404, "Комната не найдена")

    users = (await db.execute(
        select(User)
        .options(selectinload(User.tariff))
        .where(User.room_id == room_id, User.is_deleted.is_(False))
        .order_by(User.username)
    )).scalars().all()

    # Последнее утверждённое показание в комнате (общее для всех жильцов)
    last_reading = (await db.execute(
        select(MeterReading)
        .where(MeterReading.room_id == room_id, MeterReading.is_approved.is_(True))
        .order_by(MeterReading.created_at.desc())
        .limit(1)
    )).scalars().first()

    return {
        "room": {
            "id": room.id,
            "dormitory_name": room.dormitory_name,
            "room_number": room.room_number,
            "apartment_area": float(room.apartment_area or 0),
            "capacity": int(room.total_room_residents or 1),
            "tariff_id": room.tariff_id,
        },
        "residents": [
            {
                "id": u.id, "username": u.username,
                "resident_type": u.resident_type, "billing_mode": u.billing_mode,
                "tariff_id": u.tariff_id,
                "tariff_name": u.tariff.name if u.tariff else None,
                "workplace": u.workplace,
            }
            for u in users
        ],
        "last_reading": {
            "created_at": last_reading.created_at.isoformat() if last_reading and last_reading.created_at else None,
            "hot_water": float(last_reading.hot_water) if last_reading else None,
            "cold_water": float(last_reading.cold_water) if last_reading else None,
            "electricity": float(last_reading.electricity) if last_reading else None,
        } if last_reading else None,
    }


# ВАЖНО: путь-конвертер `:int` — иначе GET /{room_id} матчит статические
# GET-роуты вроде /dormitory-overview (Starlette проверяет шаблон ДО валидации,
# и «dormitory-overview» ловится как room_id → 422). С `:int` матчатся только
# числовые сегменты, статические пути проходят к своим обработчикам.
@router.get("/{room_id:int}", response_model=RoomResponse, dependencies=[Depends(allow_management)])
async def get_room(room_id: int, db: AsyncSession = Depends(get_db)):
    room = await db.get(Room, room_id)
    if not room: raise HTTPException(status_code=404, detail="Комната не найдена")
    return room


@router.put("/{room_id}", response_model=RoomResponse, dependencies=[Depends(allow_management)])
async def update_room(room_id: int, data: RoomUpdate, db: AsyncSession = Depends(get_db)):
    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(status_code=404, detail="Комната не найдена")

    update_data = data.model_dump(exclude_unset=True)

    # Эффективный тип после апдейта (если меняется — берём новый, иначе
    # текущий). Используется ниже для проверки соответствующего адреса.
    effective_pt = update_data.get("place_type", room.place_type)

    # Если в payload явно сменили place_type, RoomUpdate уже прогнал
    # нормализацию (см. _normalize_room_address_fields) и обнулил
    # «чужие» поля. Если place_type не меняется — нормализатор НЕ
    # сработал, но мы должны не дать менять «чужие» поля (например
    # для дома нельзя ставить hw_meter_serial). Делаем мягкий strip.
    if effective_pt == "house":
        for k in (
            "dormitory_name", "room_number",
            "hw_meter_serial", "cw_meter_serial", "el_meter_serial",
        ):
            if update_data.get(k) is not None:
                # Просто игнорируем — для дома эти поля бессмысленны.
                update_data[k] = None
        if update_data.get("is_singles_apartment") is True:
            raise HTTPException(
                status_code=400,
                detail="is_singles_apartment недопустим для place_type='house'",
            )
    elif effective_pt == "dormitory":
        for k in ("street", "house_number", "apartment_number"):
            if update_data.get(k) is not None:
                update_data[k] = None

    # Защита от дублей при переименовании. Логика разная для типов.
    if effective_pt == "dormitory":
        if (
            "dormitory_name" in update_data
            or "room_number" in update_data
            or "place_type" in update_data
        ):
            new_dorm = update_data.get("dormitory_name", room.dormitory_name)
            new_num = update_data.get("room_number", room.room_number)
            if new_dorm != room.dormitory_name or new_num != room.room_number \
                    or effective_pt != room.place_type:
                exist = await db.execute(select(Room).where(
                    Room.place_type == "dormitory",
                    Room.dormitory_name == new_dorm,
                    Room.room_number == new_num,
                    Room.id != room_id,
                ))
                if exist.scalars().first():
                    raise HTTPException(
                        status_code=400,
                        detail="Комната с таким номером уже есть в этом общежитии",
                    )
    elif effective_pt == "house":
        if (
            "street" in update_data
            or "house_number" in update_data
            or "apartment_number" in update_data
            or "place_type" in update_data
        ):
            new_st = update_data.get("street", room.street)
            new_hn = update_data.get("house_number", room.house_number)
            new_apt = update_data.get("apartment_number", room.apartment_number)
            if (new_st, new_hn, new_apt) != (room.street, room.house_number, room.apartment_number) \
                    or effective_pt != room.place_type:
                exist = await db.execute(select(Room).where(
                    Room.place_type == "house",
                    Room.street == new_st,
                    Room.house_number == new_hn,
                    Room.apartment_number == new_apt,
                    Room.id != room_id,
                ))
                if exist.scalars().first():
                    raise HTTPException(
                        status_code=400,
                        detail="Квартира с таким адресом уже есть",
                    )

    # Если меняется tariff_id комнаты — валидируем и инвалидируем кеш расчётов
    tariff_changed = "tariff_id" in update_data and update_data["tariff_id"] != room.tariff_id
    if tariff_changed and update_data["tariff_id"] is not None:
        t = await db.get(Tariff, update_data["tariff_id"])
        if not t or not t.is_active:
            raise HTTPException(400, "Активный тариф с таким id не найден")

    # Bug 29.05.2026 (Коммит 24): синхронизация resident_type у жильцов
    # при изменении is_singles_apartment. UI в админке (Жильцы → дашборд
    # с цифрами «Семейных/Холостяков») использует user.resident_type
    # ('family'/'single'), а флаг is_singles_apartment — атрибут комнаты.
    # Юзер ставил холостяк на комнате — в админке Жильцы показывалось
    # «Семья». Теперь sync автоматический.
    singles_changed = (
        "is_singles_apartment" in update_data
        and update_data["is_singles_apartment"] != room.is_singles_apartment
    )
    new_is_singles = update_data.get("is_singles_apartment", room.is_singles_apartment)

    # Холостяцкая квартира обязана иметь макс. вместимость — это делитель
    # площади для найма/ТКО/отопления/содержания в billing (area / max_capacity).
    # Без неё расчёт area-based статей неоднозначен.
    eff_max_cap = update_data.get("max_capacity", room.max_capacity)
    if new_is_singles and (eff_max_cap is None or int(eff_max_cap) < 1):
        raise HTTPException(
            status_code=400,
            detail=(
                "Для холостяцкой квартиры укажите «Макс. вместимость» (≥ 1) — "
                "по ней делится площадь для найма / ТКО / отопления / содержания."
            ),
        )

    # Не затираем NOT NULL поля явным null: клиент мог прислать
    # has_*_meter / is_singles_apartment = null (контракт RoomUpdate допускает
    # Optional) → setattr NULL в nullable=False колонку = IntegrityError/500.
    _non_null_cols = {"has_hw_meter", "has_cw_meter", "has_el_meter",
                      "is_singles_apartment"}
    for key, value in update_data.items():
        if key in _non_null_cols and value is None:
            continue
        setattr(room, key, value)

    if singles_changed:
        # Обновляем resident_type у всех ЖИВУЩИХ жильцов этой комнаты.
        # billing_mode оставляем by_meter (общие счётчики на квартиру,
        # cost делится на N через Bug AS этап 4). per_capita (койко-место)
        # больше не используется.
        new_resident_type = "single" if new_is_singles else "family"
        from sqlalchemy import update as _sa_update
        await db.execute(
            _sa_update(User)
            .where(User.room_id == room_id, User.is_deleted.is_(False))
            .values(resident_type=new_resident_type)
        )

    # При включённом холостяцком режиме делитель счётчиков = факт. число
    # жильцов. Проставляем сразу (и при первом включении, и при любом апдейте
    # комнаты) — чтобы billing всегда делил на актуальное число.
    if new_is_singles:
        await db.flush()
        from app.modules.utility.services.room_assignment import recount_singles_residents
        await recount_singles_residents(db, room_id)
    elif singles_changed and "total_room_residents" not in update_data:
        # Аудит #20: singles→family. Делитель счётчиков был авто-холостяцким
        # (= число Л/С). Для семьи total_room_residents = «число людей», его
        # ведёт админ вручную. Если он НЕ задал значение в этом апдейте —
        # оставляем число активных Л/С как разумную отправную точку (≥1);
        # per-user residents_count упразднён 2026-06-17.
        await db.flush()
        _cnt = (await db.execute(
            select(func.count(User.id))
            .where(User.room_id == room_id, User.is_deleted.is_(False),
                   User.role == "user")
        )).scalar_one()
        room.total_room_residents = int(_cnt) if _cnt and _cnt > 0 else 1

    # Жилфонд — единый источник числа людей семьи (2026-06-17). Если админ
    # задал total_room_residents у СЕМЕЙНОЙ комнаты — синхронизируем
    # residents_count активных Л/С (поле в форме жильца убрано). Для холостяцкой
    # комнаты этого не делаем: там residents_count=1, делёж по total.
    if not new_is_singles and "total_room_residents" in update_data:
        _trr = update_data.get("total_room_residents")
        _trr = int(_trr) if _trr and int(_trr) > 0 else 1
        await db.execute(
            update(User)
            .where(User.room_id == room_id, User.is_deleted.is_(False),
                   User.role == "user")
            .values(residents_count=_trr)
        )

    await db.commit()
    await db.refresh(room)

    if tariff_changed:
        from app.modules.utility.services.tariff_cache import tariff_cache
        tariff_cache.invalidate()
    return room


@router.delete("/{room_id}", status_code=204, dependencies=[Depends(allow_management)])
async def delete_room(room_id: int, db: AsyncSession = Depends(get_db)):
    """Удаляет комнату, если к ней не привязаны жильцы или история показаний."""
    room = await db.get(Room, room_id)
    if not room: raise HTTPException(status_code=404, detail="Комната не найдена")

    # Проверка: есть ли жильцы (даже удаленные, так как история важна)
    users_exist = await db.execute(select(User.id).where(User.room_id == room_id).limit(1))
    if users_exist.scalars().first():
        raise HTTPException(status_code=400, detail="Нельзя удалить комнату, к которой привязаны жильцы (даже бывшие)")

    # Проверка: есть ли показания счетчиков
    readings_exist = await db.execute(select(MeterReading.id).where(MeterReading.room_id == room_id).limit(1))
    if readings_exist.scalars().first():
        raise HTTPException(status_code=400, detail="Нельзя удалить комнату, по которой есть история показаний")

    await db.delete(room)
    await db.commit()
    return None


@router.post("/bulk-meter-config", dependencies=[Depends(allow_management)])
async def bulk_meter_config(
    data: RoomMeterConfigBulk,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Применить конфигурацию счётчиков ко ВСЕМ комнатам дома/общежития одним
    действием — «настроил один раз на весь дом». Квартиры статичны: жильцы
    наследуют конфиг автоматически. Цель — dormitory_name ИЛИ street+house_number.
    """
    q = update(Room).values(
        has_hw_meter=data.has_hw_meter,
        has_cw_meter=data.has_cw_meter,
        has_el_meter=data.has_el_meter,
        updated_at=utcnow(),  # bulk-UPDATE не триггерит Python onupdate
    )
    if data.dormitory_name:
        q = q.where(Room.place_type == "dormitory",
                    Room.dormitory_name == data.dormitory_name)
    elif data.street and data.house_number:
        q = q.where(Room.place_type == "house",
                    Room.street == data.street,
                    Room.house_number == data.house_number)
    else:
        raise HTTPException(
            status_code=400,
            detail="Укажите общежитие (dormitory_name) или дом (street + house_number)",
        )
    result = await db.execute(q)
    # Аудит: массовая смена конфигурации счётчиков влияет на начисления всему
    # дому/общежитию — фиксируем кто/что/сколько комнат.
    from app.modules.utility.routers.admin_dashboard import write_audit_log
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="bulk_meter_config", entity_type="room", entity_id=None,
        details={
            "target": data.dormitory_name or f"{data.street} {data.house_number}",
            "has_hw_meter": data.has_hw_meter,
            "has_cw_meter": data.has_cw_meter,
            "has_el_meter": data.has_el_meter,
            "updated_rooms": result.rowcount,
        },
    )
    await db.commit()
    return {"status": "ok", "updated_rooms": result.rowcount}


@router.get("/dormitory-overview", dependencies=[Depends(allow_management)])
async def dormitory_overview(
    dormitory_name: Optional[str] = Query(None, description="Название общежития"),
    street: Optional[str] = Query(None, description="Улица (для дома)"),
    house_number: Optional[str] = Query(None, description="Номер дома (для дома)"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сводка по зданию (дом/общага) для окна «Настройки дома» (Жилфонд):
    статистика, разбивка по тарифам и счётчикам, текущий единый тариф/набор
    счётчиков (если одинаков у всех), и список доступных тарифов с профилем
    начислений. Цель — общага (dormitory_name) ИЛИ дом (street + house_number).
    Применение: /assign-to-dormitory (тариф) и /bulk-meter-config (счётчики)."""
    from collections import Counter, defaultdict

    # Здание: общага (dormitory_name) ИЛИ дом (street + house_number).
    if dormitory_name:
        cond = and_(Room.place_type == "dormitory", Room.dormitory_name == dormitory_name)
        building_label = dormitory_name
        applicable = ["dormitory", "both"]
    elif street and house_number:
        cond = and_(Room.place_type == "house",
                    Room.street == street, Room.house_number == house_number)
        building_label = f"ул. {street}, д. {house_number}"
        applicable = ["house", "both"]
    else:
        raise HTTPException(
            400, "Укажите общежитие (dormitory_name) или дом (street + house_number)")

    rooms = (await db.execute(select(Room).where(cond))).scalars().all()
    if not rooms:
        raise HTTPException(404, f"Здание «{building_label}» не найдено")
    room_ids = [r.id for r in rooms]

    residents = (await db.execute(
        select(User).where(
            User.room_id.in_(room_ids),
            User.is_deleted.is_(False),
            User.role == "user",
        )
    )).scalars().all()
    res_by_room: dict = defaultdict(int)
    family = single = 0
    for u in residents:
        res_by_room[u.room_id] += 1
        if (u.resident_type or "family") == "single":
            single += 1
        else:
            family += 1

    empty = partial = full = overcrowded = 0
    total_capacity = 0
    for r in rooms:
        cnt = res_by_room.get(r.id, 0)
        cap = r.max_capacity or r.total_room_residents or 1
        total_capacity += cap
        if cnt == 0:
            empty += 1
        elif cnt < cap:
            partial += 1
        elif cnt == cap:
            full += 1
        else:
            overcrowded += 1

    # Разбивка по тарифам (Room.tariff_id; None = тариф жильца/дефолт).
    tariff_ids = {r.tariff_id for r in rooms if r.tariff_id}
    tnames = {}
    if tariff_ids:
        tnames = {tid: nm for tid, nm in (await db.execute(
            select(Tariff.id, Tariff.name).where(Tariff.id.in_(tariff_ids))
        )).all()}
    by_tariff_counter = Counter(r.tariff_id for r in rooms)
    by_tariff = sorted([
        {"tariff_id": tid,
         "tariff_name": (tnames.get(tid) if tid else "По умолчанию (тариф жильца)"),
         "rooms": cnt}
        for tid, cnt in by_tariff_counter.items()
    ], key=lambda x: -x["rooms"])
    distinct_tariffs = set(by_tariff_counter.keys())
    current_tariff_id = next(iter(distinct_tariffs)) if len(distinct_tariffs) == 1 else None

    # Счётчики: единый набор только если у всех комнат одинаков.
    meter_combos = {(r.has_hw_meter, r.has_cw_meter, r.has_el_meter) for r in rooms}
    current_meters = None
    if len(meter_combos) == 1:
        hwm, cwm, elm = next(iter(meter_combos))
        current_meters = {"has_hw_meter": hwm, "has_cw_meter": cwm, "has_el_meter": elm}

    # Доступные тарифы для типа здания (+ универсальные) с профилем начислений.
    # NULL-tolerant (как /api/tariffs): легаси-тариф без applicable_to всё равно
    # доступен, иначе он бы пропал из дропдауна/модалки исключений.
    from sqlalchemy import or_ as _or
    avail = (await db.execute(
        select(Tariff).where(
            Tariff.is_active.is_(True),
            _or(Tariff.applicable_to.is_(None), Tariff.applicable_to.in_(applicable)),
        ).order_by(Tariff.name)
    )).scalars().all()

    def _charges(t):
        chips = []
        if t.charge_hot_water: chips.append("ГВС")
        if t.charge_cold_water: chips.append("ХВС")
        if t.charge_sewage: chips.append("водоотв.")
        if t.charge_electricity: chips.append("электр.")
        if t.charge_maintenance: chips.append("содерж.")
        if t.charge_heating: chips.append("отопл.")
        if t.charge_waste: chips.append("ТКО")
        if t.charge_social_rent: chips.append("найм(205)")
        return chips

    return {
        "dormitory_name": building_label,
        "building_label": building_label,
        "stats": {
            "total_rooms": len(rooms), "empty": empty, "partial": partial,
            "full": full, "overcrowded": overcrowded,
            "total_residents": len(residents), "total_capacity": total_capacity,
            "occupancy_pct": round(len(residents) / total_capacity * 100) if total_capacity else 0,
            "family": family, "single": single,
            "singles_apartments": sum(1 for r in rooms if r.is_singles_apartment),
        },
        "by_tariff": by_tariff,
        "by_meter": {
            "hw": sum(1 for r in rooms if r.has_hw_meter),
            "cw": sum(1 for r in rooms if r.has_cw_meter),
            "el": sum(1 for r in rooms if r.has_el_meter),
            "none": sum(1 for r in rooms if not (r.has_hw_meter or r.has_cw_meter or r.has_el_meter)),
            "total": len(rooms),
        },
        "current_tariff_id": current_tariff_id,
        "current_meters": current_meters,
        "available_tariffs": [
            {"id": t.id, "name": t.name, "charges": _charges(t)} for t in avail
        ],
        # Список помещений здания — для окна «Исключения по квартирам»
        # (точечный тариф на отдельную квартиру).
        "rooms": [
            {
                "id": r.id,
                "room_number": r.room_number,
                "apartment_number": r.apartment_number,
                "tariff_id": r.tariff_id,
                "is_singles_apartment": bool(r.is_singles_apartment),
                "residents": res_by_room.get(r.id, 0),
            }
            for r in sorted(rooms, key=lambda x: str(x.room_number or x.apartment_number or x.id))
        ],
    }


# Нормализация серийников счётчиков к шаблону «<тип>-<дом>-<комната>»
# (напр. хвс-4.8-101). Тип берётся по полю (hw=гвс / cw=хвс / el=эл), ядро
# «дом-комната» = числовые сегменты текущего серийника; мусорный буквенный
# префикс из импорта (КВС/РРР/…) и тип-суффикс отбрасываются. Формат выбран
# владельцем. Если ядро не определить (нет цифр / пустой серийник) — не трогаем.
_SERIAL_TYPE_PREFIX = {"hw": "гвс", "cw": "хвс", "el": "эл"}


def _normalize_meter_serial(old, kind: str):
    if not old or not str(old).strip():
        return None
    segs = [s for s in str(old).strip().replace(" ", "-").split("-") if s]
    core = "-".join(s for s in segs if any(ch.isdigit() for ch in s))
    if not core:
        return None
    return f"{_SERIAL_TYPE_PREFIX[kind]}-{core}"


@router.post("/normalize-serials", dependencies=[Depends(allow_management)])
async def normalize_serials(
    dormitory_name: Optional[str] = Query(None),
    street: Optional[str] = Query(None),
    house_number: Optional[str] = Query(None),
    dry_run: bool = Query(True, description="true — только предпросмотр, не сохранять"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Привести серийники счётчиков к шаблону «<тип>-<дом>-<комната>» (напр.
    хвс-4.8-101) по всем комнатам дома/общежития одним действием. Мусорный
    префикс из импорта (КВС/РРР/…) отбрасывается, тип берётся по полю. Цель —
    dormitory_name ИЛИ street+house_number. dry_run=true — предпросмотр без
    сохранения (для модалки подтверждения)."""
    q = select(Room)
    if dormitory_name:
        q = q.where(Room.place_type == "dormitory", Room.dormitory_name == dormitory_name)
    elif street and house_number:
        q = q.where(Room.place_type == "house",
                    Room.street == street, Room.house_number == house_number)
    else:
        raise HTTPException(
            status_code=400,
            detail="Укажите общежитие (dormitory_name) или дом (street + house_number)",
        )
    rooms = (await db.execute(q.order_by(Room.room_number))).scalars().all()

    changes = []
    for room in rooms:
        fields = {}
        for kind, attr in (("hw", "hw_meter_serial"),
                           ("cw", "cw_meter_serial"),
                           ("el", "el_meter_serial")):
            old = getattr(room, attr)
            new = _normalize_meter_serial(old, kind)
            if new and new != old:
                fields[kind] = {"old": old, "new": new}
                if not dry_run:
                    setattr(room, attr, new)
        if fields:
            changes.append({
                "room_id": room.id,
                "room_number": room.room_number,
                "fields": fields,
            })

    if not dry_run and changes:
        from app.modules.utility.routers.admin_dashboard import write_audit_log
        await write_audit_log(
            db, current_user.id, current_user.username,
            action="normalize_meter_serials", entity_type="room", entity_id=None,
            details={
                "target": dormitory_name or f"{street} {house_number}",
                "changed_rooms": len(changes),
            },
        )
        await db.commit()

    return {
        "dry_run": dry_run,
        "scope": dormitory_name or f"{street} {house_number}",
        "total_rooms": len(rooms),
        "changed_rooms": len(changes),
        "changes": changes,
    }


# ═══════════════════════════════════════════════════════════════════════
# МАССОВЫЕ ОПЕРАЦИИ QR-ПОРТАЛА (2026-07-16): печать QR-кодов домом/всеми
# домами одним PDF + массовый сброс паролей личных кабинетов.
# ═══════════════════════════════════════════════════════════════════════

def _qr_scope_filter(q, dormitory: Optional[str],
                     street: Optional[str], house_number: Optional[str]):
    """Фильтр «здание»: общага (dormitory) ИЛИ дом (street+house_number).
    Ничего не задано — все объекты. ЧАСТИЧНАЯ область (улица без дома) —
    400, а не молчаливое «все объекты»: для массового сброса паролей
    fail-open означал бы сброс всего жилфонда (ревью 2026-07-16)."""
    dormitory = (dormitory or "").strip()
    street = (street or "").strip()
    house_number = (house_number or "").strip()
    if (street or house_number) and not (street and house_number):
        raise HTTPException(400, "Для дома нужны ОБА поля: street и house_number")
    if dormitory:
        return q.where(Room.dormitory_name == dormitory)
    if street:
        return q.where(Room.street == street, Room.house_number == house_number)
    return q


def _room_sort_key(room: Room):
    """Натуральная сортировка: здание → номер (числом, затем суффикс «101а»)."""
    import re
    building = room.dormitory_name or f"{room.street or ''} {room.house_number or ''}"
    num = room.room_number or room.apartment_number or ""
    m = re.match(r"(\d+)", str(num))
    return (building, int(m.group(1)) if m else 10 ** 9, str(num))


@router.get("/qr-pdf", dependencies=[Depends(allow_management)])
async def rooms_qr_pdf(
    base: str = Query(..., max_length=200,
                      description="origin фронта (https://asy-tk.ru) — уходит в QR"),
    dormitory: Optional[str] = Query(None),
    street: Optional[str] = Query(None),
    house_number: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """PDF с QR-кодами личных кабинетов: по зданию или по ВСЕМ объектам.

    8 карточек на страницу A4 (2×4), под каждым кодом — адрес (дом +
    квартира/комната), чтобы распечатать и раздать/расклеить сразу всем.
    Токены недостающим комнатам выдаются лениво (как в per-room QR)."""
    base = base.strip().rstrip("/")
    if not (base.startswith("https://") or base.startswith("http://")) or " " in base:
        raise HTTPException(400, "base должен быть http(s)-origin")

    rooms = (await db.execute(
        _qr_scope_filter(select(Room), dormitory, street, house_number)
    )).scalars().all()
    if not rooms:
        raise HTTPException(404, "Помещений по этому фильтру нет")
    rooms.sort(key=_room_sort_key)

    # Лениво выдаём токены всем, у кого ещё нет (одним коммитом, не по одному).
    from app.modules.utility.services.qr_portal import generate_qr_token
    fresh = 0
    for r in rooms:
        if not r.qr_token:
            r.qr_token = generate_qr_token()
            fresh += 1
    if fresh:
        await db.commit()

    # Данные для рендера собираем из ORM ЗАРАНЕЕ: сам рендер (400+ QR-PNG +
    # многостраничный weasyprint) — секунды/десятки секунд чистого CPU,
    # поэтому уходит в поток (to_thread), не блокируя event loop
    # (ревью 2026-07-16: иначе на время рендера замирал весь портал).
    def _building_label(r: Room) -> str:
        if r.dormitory_name:
            return r.dormitory_name
        return f"ул. {r.street}, д. {r.house_number}" if r.street else "Без адреса"

    items = [(_building_label(r), r.format_address, f"{base}/qr.html#{r.qr_token}")
             for r in rooms]

    def _render_pdf(items_: list) -> bytes:
        import base64 as _b64
        import html as _html
        from io import BytesIO
        import qrcode
        from weasyprint import HTML as _WeasyHTML

        def _qr_data_uri(text: str) -> str:
            qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M,
                               box_size=8, border=2)
            qr.add_data(text)
            qr.make(fit=True)
            buf = BytesIO()
            qr.make_image(fill_color="black", back_color="white").save(buf, format="PNG")
            return "data:image/png;base64," + _b64.b64encode(buf.getvalue()).decode()

        cards_by_building: dict[str, list[str]] = {}
        for bld, addr, url in items_:
            cards_by_building.setdefault(bld, []).append(
                f'<div class="card"><img src="{_qr_data_uri(url)}">'
                f'<div class="addr">{_html.escape(addr)}</div>'
                f'<div class="hint">Личный кабинет — подача показаний.<br>'
                f'Отсканируйте камерой телефона; пароль задаётся при первом входе.</div></div>'
            )

        sections = []
        for i, (bld, cards) in enumerate(cards_by_building.items()):
            brk = ' style="page-break-before: always;"' if i else ""
            sections.append(
                f'<div{brk}><h2>{_html.escape(bld)} — QR-коды личных кабинетов '
                f'({len(cards)} шт.)</h2>{"".join(cards)}</div>'
            )

        html_doc = (
            '<html><head><meta charset="utf-8"><style>'
            '@page { size: A4; margin: 8mm; }'
            "body { font-family: 'DejaVu Sans', sans-serif; margin: 0; }"
            'h2 { font-size: 12pt; margin: 0 0 4mm 0; }'
            '.card { display: inline-block; width: 49%; text-align: center;'
            '  padding: 4mm 0 5mm; page-break-inside: avoid; vertical-align: top; }'
            '.card img { width: 46mm; height: 46mm; }'
            '.addr { font-size: 12pt; font-weight: bold; margin-top: 1mm; }'
            '.hint { font-size: 7.5pt; color: #555; margin-top: 1mm; }'
            '</style></head><body>' + "".join(sections) + '</body></html>'
        )
        return _WeasyHTML(string=html_doc).write_pdf()

    import asyncio
    pdf_bytes = await asyncio.to_thread(_render_pdf, items)

    from urllib.parse import quote
    scope = dormitory or (f"{street}_{house_number}" if street else "все_объекты")
    scope = scope.replace("/", "_").replace("\\", "_")[:80]
    return Response(
        content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition":
                 "attachment; filename=\"qr_codes.pdf\"; "
                 f"filename*=UTF-8''{quote('QR_' + scope + '.pdf', safe='')}"},
    )


class QrBulkResetScope(BaseModel):
    dormitory: Optional[str] = None
    street: Optional[str] = None
    house_number: Optional[str] = None


@router.post("/qr/reset-passwords", dependencies=[Depends(allow_management)])
async def bulk_reset_qr_passwords(
    scope: QrBulkResetScope,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Массовый сброс паролей личных кабинетов (QR-портала): по зданию или
    по ВСЕМ объектам сразу. Пароль (4+ знака) задаёт сам жилец — сброс
    обнуляет хэш, портал попросит придумать новый при следующем входе.
    QR-токены НЕ меняются (наклейки продолжают работать)."""
    q = _qr_scope_filter(
        update(Room).values(qr_password_hash=None)
        .where(Room.qr_password_hash.isnot(None)),
        scope.dormitory, scope.street, scope.house_number,
    )
    res = await db.execute(q)
    n = res.rowcount or 0
    from app.modules.utility.routers.admin_dashboard import write_audit_log
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="bulk_reset_qr_passwords", entity_type="room", entity_id=0,
        details={"dormitory": scope.dormitory, "street": scope.street,
                 "house_number": scope.house_number, "reset_count": n},
    )
    await db.commit()
    return {"reset": n}


@router.post("/{room_id}/qr", dependencies=[Depends(allow_management)])
async def get_room_qr(
    room_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Получить (создать при отсутствии) QR-токен квартиры для портала подачи.
    Возвращает токен и относительный путь /q/<token>; полный URL и картинку
    QR строит фронт (origin + путь → /api/qr?text=...)."""
    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(404, "Помещение не найдено")
    from app.modules.utility.services.qr_portal import get_or_create_room_token
    token = await get_or_create_room_token(db, room)
    return {
        "room_id": room_id, "token": token, "portal_path": f"/qr.html#{token}",
        "password_set": bool(room.qr_password_hash),
    }


@router.post("/{room_id}/qr/regenerate", dependencies=[Depends(allow_management)])
async def regenerate_room_qr(
    room_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Перевыпустить QR-токен квартиры (отзыв): старый QR перестаёт работать."""
    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(404, "Помещение не найдено")
    from app.modules.utility.services.qr_portal import regenerate_room_token
    token = await regenerate_room_token(db, room)
    from app.modules.utility.routers.admin_dashboard import write_audit_log
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="regenerate_room_qr", entity_type="room", entity_id=room_id,
    )
    await db.commit()
    return {"room_id": room_id, "token": token, "portal_path": f"/qr.html#{token}"}


@router.post("/{room_id}/qr/reset-password", dependencies=[Depends(allow_management)])
async def reset_room_qr_password(
    room_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сбросить пароль QR-портала квартиры (жилец забыл). Портал при
    следующем входе попросит установить новый. Токен НЕ меняется."""
    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(404, "Помещение не найдено")
    room.qr_password_hash = None
    from app.modules.utility.routers.admin_dashboard import write_audit_log
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="reset_room_qr_password", entity_type="room", entity_id=room_id,
    )
    await db.commit()
    return {"status": "ok"}


@router.post("/{room_id}/make-singles", dependencies=[Depends(allow_management)])
async def make_room_singles(
    room_id: int,
    max_capacity: int = Query(..., ge=1, description="Макс. вместимость (делитель площади)"),
    apartment_area: Optional[float] = Query(None, gt=0, description="Площадь м² (делитель найма/ТКО/отопления). Если задана — обновим."),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Быстрое решение из Центра анализа («Типы квартир»): перевести квартиру в
    холостяцкую одним кликом. Ставит is_singles_apartment=True + max_capacity
    (+ площадь, если передана) и приводит ВСЕХ живущих жильцов к
    resident_type='single' (резолвит multi_family/unmarked_singles/mixed_types/
    singles_with_family). Зеркалит синк из update_room, но работает и когда флаг
    уже True. Площадь у холостяцкой критична (делится на вместимость), поэтому
    UI её спрашивает и передаёт — особенно у малых комнат."""
    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(404, "Помещение не найдено")
    if room.place_type == "house":
        raise HTTPException(400, "Холостяцкая квартира — только для общежития (place_type='dormitory')")

    room.is_singles_apartment = True
    room.max_capacity = max_capacity
    if apartment_area is not None:
        room.apartment_area = apartment_area
    room.updated_at = utcnow()

    res = await db.execute(
        update(User)
        .where(
            User.room_id == room_id,
            User.is_deleted.is_(False),
            User.role == "user",
        )
        .values(resident_type="single")
    )

    from app.modules.utility.routers.admin_dashboard import write_audit_log
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="make_room_singles", entity_type="room", entity_id=room_id,
        details={
            "room": f"{room.dormitory_name} / {room.room_number}",
            "max_capacity": max_capacity,
            "apartment_area": apartment_area,
            "residents_converted": res.rowcount,
        },
    )
    # Аудит #12: после перевода в singles делитель счётчиков обязан
    # пересчитаться (= число жильцов), иначе остаётся семейным → холостяк
    # недоплачивает за воду в N раз. Зеркалит update_room.
    await db.flush()
    from app.modules.utility.services.room_assignment import recount_singles_residents
    await recount_singles_residents(db, room_id)
    await db.commit()
    return {"status": "ok", "room_id": room_id, "residents_converted": res.rowcount}


@router.post("/{room_id}/replace-meter", dependencies=[Depends(allow_management)])
async def replace_meter(room_id: int, data: ReplaceMeterSchema, db: AsyncSession = Depends(get_db)):
    """
    Безопасная замена счетчика.
    Рассчитывает долг по старому счетчику и устанавливает нулевую базу для нового.
    """
    room = await db.get(Room, room_id)
    if not room:
        raise HTTPException(status_code=404, detail="Комната не найдена")

    # Ищем активный период
    active_period = (await db.execute(select(BillingPeriod).where(BillingPeriod.is_active))).scalars().first()
    if not active_period:
        raise HTTPException(status_code=400, detail="Нет активного расчетного периода")

    # Ищем черновики. Если есть черновик, просим сначала его утвердить или удалить
    draft = (await db.execute(select(MeterReading).where(
        MeterReading.room_id == room.id,
        MeterReading.period_id == active_period.id,
        MeterReading.is_approved == False
    ))).scalars().first()
    if draft:
        raise HTTPException(status_code=400,
                            detail="По этой комнате висит необработанный черновик показаний. Утвердите или удалите его перед заменой счетчика.")

    # Ищем жильца для расчета (чтобы взять тариф и долю)
    user = (await db.execute(
        select(User).where(User.room_id == room.id, User.is_deleted == False).limit(1))).scalars().first()

    # Ищем последнее показание для базы закрытия. Аудит (замена счётчика,
    # дефект 3): берём последнее MEANINGFUL approved показание ИМЕННО этого
    # жильца (если есть) — не AUTO/synth/METER_CLOSED/чужое, иначе дельта
    # закрытия и перенос долга считаются от неверной базы.
    from app.modules.utility.services.reading_calculator import is_meaningful_prev
    _pr_q = select(MeterReading).where(
        MeterReading.room_id == room.id, MeterReading.is_approved == True)
    if user:
        _pr_q = _pr_q.where(MeterReading.user_id == user.id)
    _pr_cands = (await db.execute(
        _pr_q.order_by(MeterReading.created_at.desc()).limit(20)
    )).scalars().all()
    prev_reading = next((r for r in _pr_cands if is_meaningful_prev(r)), None)

    p_hot = prev_reading.hot_water if prev_reading else ZERO
    p_cold = prev_reading.cold_water if prev_reading else ZERO
    p_elect = prev_reading.electricity if prev_reading else ZERO

    # 1. Расчет дельты по закрываемому счетчику
    d_hot, d_cold, d_elect = ZERO, ZERO, ZERO

    if data.meter_type == "hot":
        if data.final_old_value < p_hot: raise HTTPException(400, "Финальное показание меньше предыдущего!")
        d_hot = data.final_old_value - p_hot
        room.hw_meter_serial = data.new_serial
        c_hot, c_cold, c_elect = data.final_old_value, p_cold, p_elect
        n_hot, n_cold, n_elect = data.initial_new_value, p_cold, p_elect
    elif data.meter_type == "cold":
        if data.final_old_value < p_cold: raise HTTPException(400, "Финальное показание меньше предыдущего!")
        d_cold = data.final_old_value - p_cold
        room.cw_meter_serial = data.new_serial
        c_hot, c_cold, c_elect = p_hot, data.final_old_value, p_elect
        n_hot, n_cold, n_elect = p_hot, data.initial_new_value, p_elect
    else:  # elect
        if data.final_old_value < p_elect: raise HTTPException(400, "Финальное показание меньше предыдущего!")
        d_elect = data.final_old_value - p_elect
        room.el_meter_serial = data.new_serial
        c_hot, c_cold, c_elect = p_hot, p_cold, data.final_old_value
        n_hot, n_cold, n_elect = p_hot, p_cold, data.initial_new_value

    # Если в комнате есть жилец, начисляем ему стоимость остатка старого счетчика
    if user and (d_hot > 0 or d_cold > 0 or d_elect > 0):
        from app.modules.utility.services.tariff_cache import tariff_cache
        t = tariff_cache.get_effective_tariff(user=user, room=room)

        from app.modules.utility.services.calculations import paying_residents
        residents = Decimal(paying_residents(user, room))
        total_res = Decimal(room.total_room_residents) if room.total_room_residents > 0 else Decimal(1)
        elect_share = (residents / total_res) * d_elect

        # Замена счётчика — те же сезонные правила: global + per-tariff.
        from app.modules.utility.routers.settings import _load_seasonal
        _seasonal = await _load_seasonal(db)
        _heating = _seasonal.heating_season_active and t.is_heating_active_now()
        _hw = _seasonal.hot_water_heating_active and t.is_hw_heating_active_now()
        costs = calculate_utilities(
            user, room, t, d_hot, d_cold, d_hot + d_cold, elect_share,
            heating_season_active=_heating,
            hot_water_heating_active=_hw,
        )

        # Запись 1: Закрытие старого счетчика (с начислением суммы).
        #
        # ВАЖНО: переносим непогашенный долг (debt_209/205) и переплаты
        # (overpayment_209/205) из предыдущего утверждённого показания.
        # Раньше они обнулялись — жилец видел, что "долг исчез", но в БД
        # он висел на закрытом показании и при следующем расчёте не учитывался.
        closing_reading = MeterReading(
            user_id=user.id, room_id=room.id, period_id=active_period.id,
            hot_water=c_hot, cold_water=c_cold, electricity=c_elect,
            is_approved=True, anomaly_flags="METER_CLOSED", anomaly_score=0,
            total_209=costs['total_cost'] - costs['cost_social_rent'],
            total_205=costs['cost_social_rent'], total_cost=costs['total_cost'],
            debt_209=(prev_reading.debt_209 if prev_reading else ZERO) or ZERO,
            overpayment_209=(prev_reading.overpayment_209 if prev_reading else ZERO) or ZERO,
            debt_205=(prev_reading.debt_205 if prev_reading else ZERO) or ZERO,
            overpayment_205=(prev_reading.overpayment_205 if prev_reading else ZERO) or ZERO,
        )
        from app.modules.utility.services.calculations import costs_for_model_fields
        for k, v in costs_for_model_fields(costs).items():
            setattr(closing_reading, k, v)
        db.add(closing_reading)
        await db.flush()

    # Запись 2: Базовый старт нового счетчика (сумма 0 руб)
    # Это нужно чтобы следующий расчет считал от initial_new_value
    base_reading = MeterReading(
        user_id=user.id if user else None, room_id=room.id, period_id=active_period.id,
        hot_water=n_hot, cold_water=n_cold, electricity=n_elect,
        is_approved=True, anomaly_flags="METER_REPLACEMENT", anomaly_score=0,
        total_209=ZERO, total_205=ZERO, total_cost=ZERO
    )
    db.add(base_reading)

    # Обновляем кэш комнаты
    room.last_hot_water = n_hot
    room.last_cold_water = n_cold
    room.last_electricity = n_elect
    db.add(room)

    await db.commit()
    return {"status": "success"}
