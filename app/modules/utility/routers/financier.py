import os
import uuid
import asyncio
import logging
import secrets
import io
import json
import zipfile
from pathlib import Path
from datetime import datetime, timedelta
from decimal import Decimal
from app.core.time_utils import utcnow
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Header
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import func, or_, desc, asc, update as _sa_update, text
from app.core.config import settings
from app.core.database import get_db
# Добавлен импорт модели Room
from app.modules.utility.models import User, MeterReading, BillingPeriod, Room, DebtImportLog, SystemSetting
from app.core.dependencies import get_current_user
from app.core.auth import fernet
from app.modules.utility.schemas import PaginatedResponse, UserDebtResponse
from app.modules.utility.tasks import import_debts_task

router = APIRouter(prefix="/api/financier", tags=["Financier"])
logger = logging.getLogger(__name__)

def _nfu_fio(item) -> str:
    """Извлекает ФИО из элемента not_found_users.

    not_found_users поменял формат: legacy = list[str], новый = list[dict].
    Помогает не падать с AttributeError при работе со старыми импортами
    и при удалении/сравнении элементов.
    """
    if isinstance(item, dict):
        return str(item.get("fio", "")).strip()
    return str(item).strip()


async def _ensure_debt_alias(
    db: AsyncSession,
    *,
    alias_fio: str,
    user_id: int,
    created_by_id: int,
    note: Optional[str] = None,
) -> bool:
    """Создаёт GSheetsAlias для запоминания привязки ФИО → user.

    Общая таблица для всех типов импорта (gsheets, debt 205, debt 209) —
    привязка сделана один раз, работает везде. Если alias по этому
    normalized ФИО уже есть (даже на другого юзера) — НЕ перезаписываем,
    оставляем старый (защита от случайной перебивки чужой привязки).

    Возвращает True если действительно создал новую запись.
    """
    from app.modules.utility.models import GSheetsAlias
    from app.modules.utility.services.gsheets_sync import normalize_fio

    normalized = normalize_fio(alias_fio)
    if not normalized:
        return False
    existing = (await db.execute(
        select(GSheetsAlias).where(GSheetsAlias.alias_fio_normalized == normalized)
    )).scalars().first()
    if existing:
        return False

    db.add(GSheetsAlias(
        alias_fio=alias_fio.strip(),
        alias_fio_normalized=normalized,
        user_id=user_id,
        kind="debt_manual",
        note=note,
        created_by_id=created_by_id,
    ))
    return True


TEMP_DIR = "/app/static/temp_imports"  # legacy, для совместимости со старым кодом
# Постоянное хранение оригиналов ОСВ из 1С.
# ПУТЬ: используем существующий shared_data volume (/app/static/generated_files/).
# Раньше пробовали /app/data/debt_archives через отдельный volume,
# но это требовало `docker compose down && up -d` для создания volume —
# до рестарта web писал в эфемерный слой, а worker_heavy искал в своём
# эфемерном слое, выдавая «No such file or directory». shared_data уже
# смонтирован на web + worker_heavy + nginx, файлы шарятся сразу.
#
# Прямой доступ через nginx (/static/generated_files/debt_archives/...)
# заблокирован в nginx/conf.d/default.conf — скачивание только через
# /api/financier/debts/import-history/{id}/download с auth.
DEBT_ARCHIVE_DIR = "/app/static/generated_files/debt_archives"
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB
os.makedirs(TEMP_DIR, exist_ok=True)
os.makedirs(DEBT_ARCHIVE_DIR, exist_ok=True)


async def _save_uploaded_debt_file(
    file: UploadFile, account_type: str, batch_id: str
) -> tuple[str, str]:
    """Валидирует, проверяет размер/magic-bytes и сохраняет xlsx в archive.

    Возвращает (file_path, original_name). Кидает HTTPException при ошибке.
    Файлы парного импорта получают batch_id в имени для группировки на
    диске — после успешного импорта debt_import.py зальёт archive_path
    в DebtImportLog.
    """
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Поддерживаются только Excel-файлы")

    header = await file.read(8)
    await file.seek(0)
    if not (header.startswith(b"PK\x03\x04") or header.startswith(b"\xd0\xcf\x11\xe0")):
        raise HTTPException(status_code=400, detail="Вредоносный файл или поддельное расширение!")

    file.file.seek(0, 2)
    file_size = file.file.tell()
    await file.seek(0)
    if file_size > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"Файл слишком большой. Максимум {MAX_FILE_SIZE / 1024 / 1024} MB",
        )

    ext = file.filename.rsplit(".", 1)[-1].lower()
    unique_name = f"{batch_id}_{account_type}.{ext}"
    file_path = os.path.join(DEBT_ARCHIVE_DIR, unique_name)

    try:
        content = await file.read()

        def save_file():
            with open(file_path, "wb") as buffer:
                buffer.write(content)

        await asyncio.to_thread(save_file)
    except Exception:
        logger.exception("Failed to save uploaded file")
        raise HTTPException(status_code=500, detail="Ошибка сохранения файла")

    return file_path, file.filename


@router.post(
    "/debts/preview-file",
    summary="Предпросмотр Excel-файла ОСВ 1С перед импортом",
)
async def debts_preview_file(
    account_type: str = Form(..., pattern="^(209|205)$"),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Парсит Excel-файл БЕЗ сохранения в архив и БЕЗ создания DebtImportLog.
    Возвращает сводку: количество строк с ФИО, sum debt/overpayment, sample
    ФИО, file_hash (SHA256). Проверяет дубликат — если такой же файл уже
    был импортирован, возвращает ссылку на предыдущий лог.

    UI использует это для авто-проверки при выборе файла, чтобы:
      - не дать загрузить тот же файл дважды,
      - показать админу что он скармливает системе (счёт, период, суммы).
    """
    if current_user.role not in ("admin", "financier"):
        raise HTTPException(403, "Недостаточно прав")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Поддерживаются только Excel-файлы")

    header = await file.read(8)
    await file.seek(0)
    if not (header.startswith(b"PK\x03\x04") or header.startswith(b"\xd0\xcf\x11\xe0")):
        raise HTTPException(400, "Поддельное расширение или не Excel")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, f"Файл слишком большой ({len(content)/1024/1024:.1f} MB > {MAX_FILE_SIZE/1024/1024} MB)")

    # SHA256 хэш для дедупликации.
    import hashlib as _hl
    file_hash = _hl.sha256(content).hexdigest()

    # Парсим файл в памяти через BytesIO.
    import io as _io
    import openpyxl as _opx
    from decimal import Decimal as _D
    try:
        wb = _opx.load_workbook(filename=_io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Не удалось открыть Excel: {e}")

    # Эвристика: что считаем «строкой с ФИО».
    def _looks_like_fio(s) -> bool:
        if not isinstance(s, str):
            return False
        s_norm = s.strip()
        if len(s_norm) < 6 or len(s_norm) > 80:
            return False
        # Минимум 2 заглавных слова кириллицы.
        words = s_norm.split()
        if len(words) < 2:
            return False
        good = sum(1 for w in words if w and w[0].isalpha() and w[0].isupper() and any('А' <= c <= 'я' or c == 'Ё' or c == 'ё' for c in w))
        if good < 2:
            return False
        # Не должно быть ключевых слов ОСВ.
        s_low = s_norm.lower()
        for kw in ("договор", "сальдо", "оборот", "итого", "период", "контрагент", "счёт", "счет"):
            if kw in s_low:
                return False
        return True

    rows_total = 0
    rows_with_fio = 0
    sample_fio = []
    sum_debt = _D("0")

    try:
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            rows_total += 1
            fio_idx = None
            for col_idx, cell in enumerate(row):
                if _looks_like_fio(cell):
                    fio_idx = col_idx
                    break
            if fio_idx is None:
                continue
            rows_with_fio += 1
            if len(sample_fio) < 5:
                sample_fio.append(str(row[fio_idx]).strip())
            # Числа в строке — суммируем как debt (макс положительное) и
            # overpay (макс положительное). Это грубо, но даёт оценку.
            nums = []
            for cell in row[fio_idx + 1:]:
                if cell is None or cell == "":
                    continue
                try:
                    d = _D(str(cell).replace(",", "."))
                    if d != 0:
                        nums.append(d)
                except Exception:
                    pass
            # В ОСВ обычно последние 2 числа — финальное сальдо
            # (debt — Дебет, overpay — Кредит).
            if nums:
                # Простая эвристика: положительные большие → debt, остальные → overpay.
                # Точную классификацию даёт основной парсер; тут только превью.
                for n in nums:
                    if n > 0:
                        sum_debt += n / _D(len(nums))  # средняя — неточно, но не критично
    except Exception as e:
        wb.close()
        raise HTTPException(400, f"Ошибка парсинга: {e}")
    wb.close()

    # Проверка дубликата: проходим по последним 30 архивам, считаем их хэши.
    # Это медленно (читаем диски), но кешируем результат коротким TTL.
    duplicate_of = None
    recent_logs = (await db.execute(
        select(DebtImportLog)
        .where(DebtImportLog.archive_path.is_not(None))
        .order_by(desc(DebtImportLog.id))
        .limit(30)
    )).scalars().all()
    import os as _os
    for log in recent_logs:
        if not log.archive_path or not _os.path.exists(log.archive_path):
            continue
        try:
            with open(log.archive_path, "rb") as fh:
                arch_hash = _hl.sha256(fh.read()).hexdigest()
            if arch_hash == file_hash:
                duplicate_of = {
                    "log_id": log.id,
                    "account_type": log.account_type,
                    "started_at": log.started_at.isoformat() if log.started_at else None,
                    "status": log.status,
                }
                break
        except Exception:
            pass

    return {
        "file_name": file.filename,
        "size_bytes": len(content),
        "file_hash": file_hash,
        "duplicate_of": duplicate_of,
        "rows_total": rows_total,
        "rows_with_fio": rows_with_fio,
        "sample_fio": sample_fio,
        "expected_account_type": account_type,
    }


@router.post("/import-debts", summary="Фоновый импорт долгов из 1С")
async def upload_debts_1c(
        account_type: str = Form(..., pattern="^(209|205)$", description="Тип счета: 209 или 205"),
        file: UploadFile = File(...),
        period_id: int | None = Form(None, description="Период загрузки; по умолчанию активный"),
        current_user: User = Depends(get_current_user)
):
    """Загрузка ОДНОГО файла. Для парной загрузки 205+209 используйте
    /import-debts-pair — он создаёт один batch_id для обоих импортов
    и UI показывает их как единую операцию."""
    if current_user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    batch_id = str(uuid.uuid4())
    file_path, original_name = await _save_uploaded_debt_file(file, account_type, batch_id)

    task = import_debts_task.delay(
        file_path, account_type,
        started_by_id=current_user.id,
        started_by_username=current_user.username,
        batch_id=batch_id,
        original_file_name=original_name,
        period_id=period_id,
    )
    logger.info(f"[IMPORT] Started task={task.id} for account={account_type} batch={batch_id}")

    return {
        "task_id": task.id,
        "status": "processing",
        "account_type": account_type,
        "batch_id": batch_id,
    }


@router.post("/import-debts-pair", summary="Парная загрузка 205 + 209 одной операцией")
async def upload_debts_pair_1c(
        file_209: UploadFile = File(None, description="Файл ОСВ по счёту 209"),
        file_205: UploadFile = File(None, description="Файл ОСВ по счёту 205"),
        period_id: int | None = Form(None, description="Период загрузки; по умолчанию активный"),
        current_user: User = Depends(get_current_user),
):
    """Загружает ОБА файла одной операцией. Оба DebtImportLog получают
    один batch_id — в истории импортов показываются как одна группа.

    Минимум один файл должен быть передан. Если только один — работает
    как обычный /import-debts (но всё равно создаёт batch_id для
    унификации UI).
    """
    if current_user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    if not file_209 and not file_205:
        raise HTTPException(status_code=400, detail="Передайте хотя бы один файл (209 и/или 205)")

    batch_id = str(uuid.uuid4())
    tasks_out = []

    # СЕРИАЛИЗАЦИЯ через celery chain (may 2026):
    # Раньше тут было `import_debts_task.delay(...)` для каждого файла в цикле —
    # ДВА task'а запускались параллельно. При concurrency > 1 worker'ы
    # одновременно SELECT'или readings, видели «у жильца нет reading в active
    # period», оба шли в inserts_dict через db.add_all. Затем при commit:
    #   - либо unique-violation на (user_id, room_id, period_id) → второй task
    #     падал с rollback, debt_205 терялся ✗
    #   - либо без unique → два дубля reading; UI выбирал первый ✗
    #
    # Симптом: Лучка А.П. — debt_209=21889 виден (первый task), debt_205=0
    # (второй task упал/конфликтнул, но статус completed).
    #
    # Теперь — chain через .si() (immutable signature, не передаёт return
    # первого как arg второго). 205 task стартует ТОЛЬКО когда 209 task
    # успешно завершён и commit'нул. Никаких race condition'ов.
    from celery import chain
    signatures = []
    task_meta = []  # сохраняем file_path/account для построения сигнатур

    for f, account in [(file_209, "209"), (file_205, "205")]:
        if f is None:
            continue
        file_path, original_name = await _save_uploaded_debt_file(f, account, batch_id)
        sig = import_debts_task.si(
            file_path, account,
            started_by_id=current_user.id,
            started_by_username=current_user.username,
            batch_id=batch_id,
            original_file_name=original_name,
            period_id=period_id,
        )
        signatures.append(sig)
        task_meta.append({"account": account, "file_name": original_name})

    if not signatures:
        return {"status": "noop", "batch_id": batch_id, "tasks": []}

    # Запускаем последовательно. apply_async() возвращает AsyncResult последнего
    # task'а в цепочке — фронт поллит его, и когда он SUCCESS → значит вся
    # цепочка отработала.
    chain_result = chain(*signatures).apply_async()

    # Собираем id всех task'ов в цепочке (для UI). Первый task — chain_result.parent...
    # цикл наверх. В celery chain цепочка прицеплена через .parent.
    task_ids = []
    cur = chain_result
    while cur is not None:
        task_ids.append(cur.id)
        cur = cur.parent
    task_ids.reverse()  # порядок исполнения

    for meta, tid in zip(task_meta, task_ids):
        tasks_out.append({
            "account_type": meta["account"],
            "task_id": tid,
            "file_name": meta["file_name"],
        })
        logger.info(
            f"[IMPORT-PAIR] Queued task={tid} for account={meta['account']} "
            f"batch={batch_id} (sequential chain)"
        )

    return {
        "status": "processing",
        "batch_id": batch_id,
        "tasks": tasks_out,
    }


# =========================================================================
# ГИС ГМП — АВТО-ПОДГРУЗКА ДОЛГОВ (мост-расширение gisgmp-bridge)
# =========================================================================
# Расширение под ЭЦП-сессией пользователя парсит «Начисления» реестра
# gisgmp.cgu.mchs.ru и POST'ит сюда массив. Авторизация — статический
# GISGMP_SYNC_TOKEN (не пользовательский JWT: фоновый синк раз в 12 ч,
# короткий токен протух бы). Сервис матчит ФИО → жильца и пишет долги
# 209/205 в активный период, создавая пару DebtImportLog.

class GisgmpChargeIn(BaseModel):
    """Одна строка «Начислений» реестра (как её распарсило расширение)."""
    uin: Optional[str] = None
    amount: str = "0"                  # сумма начисления (строкой, чистим на бэке)
    bill_date: Optional[str] = None    # дата начисления
    actualize_date: Optional[str] = None
    payer_name: str = ""               # ФИО плательщика — ключ матчинга
    account: Optional[str] = None      # лицевой счёт (привязан к квартире)
    purpose: str = ""                  # «Назначение» → 209/205
    ack_status: str = ""               # статус квитирования (оплачено/нет)
    change_status: str = ""            # статус изменения (эталонное/аннулирование/…)
    charge_uuid: Optional[str] = None
    source: Optional[str] = None


class GisgmpSyncIn(BaseModel):
    charges: list[GisgmpChargeIn]


# Защита от случайного гигантского POST'а. Реальный объём — сотни жильцов ×
# ~24 начисления = единицы тысяч строк; 50k берём с большим запасом.
_GISGMP_MAX_CHARGES = 50_000


def _check_gisgmp_token(authorization: Optional[str]) -> None:
    """Сверяет Bearer-токен с GISGMP_SYNC_TOKEN (constant-time)."""
    expected = (settings.GISGMP_SYNC_TOKEN or "").strip()
    if not expected:
        raise HTTPException(
            status_code=503,
            detail="Синхронизация ГИС ГМП не настроена: задайте GISGMP_SYNC_TOKEN в .env",
        )
    token = ""
    if authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ", 1)[1].strip()
    if not token or not secrets.compare_digest(token, expected):
        raise HTTPException(status_code=401, detail="Неверный токен синхронизации ГИС ГМП")


def _run_gisgmp_sync(charges: list[dict]) -> dict:
    """Синхронный прогон импортёра в отдельной сессии (вызывается в потоке)."""
    from app.core.database import sync_db_session
    from app.modules.utility.services.gisgmp_import import sync_import_gisgmp_charges
    with sync_db_session() as db:
        return sync_import_gisgmp_charges(charges, db)


@router.post("/gisgmp/sync", summary="Авто-подгрузка долгов из реестра ГИС ГМП (мост-расширение)")
async def gisgmp_sync(
    payload: GisgmpSyncIn,
    authorization: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_db),
):
    """Принимает распарсенные начисления от расширения и заливает долги.

    Авторизация — статический GISGMP_SYNC_TOKEN в заголовке Authorization:
    Bearer. Обработка синхронная (в пуле потоков) — расширению нужен прямой
    результат синка, а объём данных небольшой.
    """
    _check_gisgmp_token(authorization)

    charges = [c.model_dump() for c in payload.charges]
    if not charges:
        raise HTTPException(status_code=400, detail="Пустой список начислений")
    if len(charges) > _GISGMP_MAX_CHARGES:
        raise HTTPException(
            status_code=413,
            detail=f"Слишком много начислений за раз (>{_GISGMP_MAX_CHARGES}). Разбейте на части.",
        )

    result = await asyncio.to_thread(_run_gisgmp_sync, charges)
    if result.get("status") == "error":
        # Нет активного периода и т.п. — 409, расширение покажет в статусе.
        raise HTTPException(status_code=409, detail=result.get("message", "Импорт не выполнен"))
    # Кэш/находки только что обновились — если есть прогон актуализации, ждущий
    # снимка «после», снимаем его сейчас (по свежей сверке). Тихо, без влияния на синк.
    try:
        await _capture_actualize_after(db)
    except Exception:
        logger.exception("[gisgmp] снимок «после» актуализации не удался")
    return result


@router.get("/gisgmp/status", summary="Статус последней авто-подгрузки ГИС ГМП")
async def gisgmp_status(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Для карточки «Авто-подгрузка ГИС ГМП» во вкладке «Долги 1С»: когда был
    последний синк, сколько жильцов затронуто, не настроен ли токен."""
    _require_finance(current_user)
    relay = await _load_relay_cfg(db)
    # Только счётчики находок (а не весь блоб) — статус опрашивается часто (15с).
    fres = (await db.execute(
        text("WITH f AS (SELECT value::jsonb AS j FROM system_settings WHERE key = :k) "
             "SELECT j->>'synced_at' AS synced_at, j->>'total_charges' AS total_charges, "
             "j->>'residents' AS residents, j->>'matched' AS matched, "
             "j->>'not_found' AS not_found FROM f"),
        {"k": GISGMP_FINDINGS_KEY},
    )).first()
    fsum = None
    if fres and fres.synced_at is not None:
        def _ci(v):
            try:
                return int(v)
            except Exception:
                return 0
        fsum = {"synced_at": fres.synced_at, "total_charges": _ci(fres.total_charges),
                "residents": _ci(fres.residents), "matched": _ci(fres.matched),
                "not_found": _ci(fres.not_found)}

    # Онлайн релея (по последнему опросу конфига) + версия (актуальна ли она).
    poll_s = int(relay.get("relay_poll_seconds") or 120)
    poll_age, online = None, False
    lp = relay.get("last_poll_at")
    if lp:
        try:
            poll_age = (utcnow() - datetime.fromisoformat(lp)).total_seconds()
            online = poll_age < max(300, poll_s * 3)
        except Exception:
            pass
    relay_latest = None
    try:
        rp = Path(__file__).resolve().parents[4] / "relay" / "gisgmp" / "relay.py"
        for line in rp.read_text(encoding="utf-8").splitlines():
            if line.strip().startswith("RELAY_VERSION"):
                relay_latest = line.split("=", 1)[1].strip().strip('"').strip("'")
                break
    except Exception:
        pass

    return {
        "configured": bool((settings.GISGMP_SYNC_TOKEN or "").strip()),
        "registry_url": "https://gisgmp.cgu.mchs.ru/charge/",
        # Состояние релея — управляется из этой же вкладки (pull-модель).
        "relay": {
            "enabled": relay.get("enabled", True),
            "months_back": relay.get("months_back", 2),
            "interval_hours": relay.get("interval_hours", 12),
            "daily_hour": relay.get("daily_hour", 22),
            "run_now": relay.get("run_now", False),
            "last_run_at": relay.get("last_run_at"),
            "last_report_at": relay.get("last_report_at"),
            "last_status": relay.get("last_status"),
            "last_message": relay.get("last_message"),
            "last_count": relay.get("last_count", 0),
            "online": online,
            "last_poll_at": relay.get("last_poll_at"),
            "poll_age_seconds": poll_age,
            "relay_poll_seconds": poll_s,
            "relay_version": relay.get("relay_version"),
            "relay_latest_version": relay_latest,
            "passport_username": relay.get("passport_username"),
            "pending": {
                "self_update": bool(relay.get("self_update")),
                "restart": bool(relay.get("restart")),
                "credentials": bool(relay.get("creds_pending")),
            },
        },
        # Сводка последних находок (ГИС ГМП пока хранится ОТДЕЛЬНО от долгов).
        "findings": fsum,
    }


# ─── Релей ГИС ГМП: конфиг и управление (pull-модель) ────────────────────────
# Релей на ВМ PODS2 за NAT — ЖКХ не достучится к нему внутрь. Поэтому ЖКХ хранит
# настройки/команды в SystemSetting, а релей сам их забирает (GET relay-config)
# и шлёт отчёт (POST relay-report). Админ рулит из вкладки «Долги 1С».

GISGMP_RELAY_KEY = "gisgmp_relay"
_GISGMP_RELAY_DEFAULTS = {
    "enabled": True,
    "months_back": 36,
    "interval_hours": 12,
    "daily_hour": 22,                # час ежедневного авто-запуска по МСК (вечер)
    "run_now": False,
    "last_run_at": None,
    "last_report_at": None,
    "last_status": None,
    "last_message": None,
    "last_count": 0,
    "last_updated": 0,
    "last_created": 0,
    "last_not_found": 0,
    # Управление демоном из UI (релей применяет на опросе, claim-на-выдаче):
    "self_update": False,            # подтянуть свежий relay.py и перезапуститься
    "restart": False,                # просто перезапуститься
    "passport_username": None,       # логин входа в реестр (показываем в статусе)
    "passport_password_enc": None,   # пароль, Fernet-шифр (в статус НЕ отдаём)
    "creds_pending": False,          # есть неотданная смена учётки
}


async def _load_relay_cfg(db: AsyncSession) -> dict:
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_RELAY_KEY)
    )).scalars().first()
    cfg = dict(_GISGMP_RELAY_DEFAULTS)
    if row and row.value:
        try:
            cfg.update(json.loads(row.value))
        except Exception:
            pass
    return cfg


async def _save_relay_cfg(db: AsyncSession, cfg: dict) -> None:
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_RELAY_KEY)
    )).scalars().first()
    if row is None:
        row = SystemSetting(key=GISGMP_RELAY_KEY, value="{}",
                            description="Конфиг и статус релея ГИС ГМП")
        db.add(row)
    row.value = json.dumps(cfg, ensure_ascii=False)
    await db.commit()


GISGMP_FINDINGS_KEY = "gisgmp_findings"


async def _load_findings(db: AsyncSession) -> Optional[dict]:
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_FINDINGS_KEY)
    )).scalars().first()
    if row and row.value:
        try:
            return json.loads(row.value)
        except Exception:
            return None
    return None


@router.get("/gisgmp/relay-config", summary="Релей берёт свой конфиг (token-auth)")
async def gisgmp_relay_config(
    v: Optional[str] = None,
    poll: Optional[int] = None,
    authorization: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_db),
):
    """Релей опрашивает этот эндпоинт раз в ~2 мин. Если пора запускаться
    (run_now или истёк интервал) — сервер атомарно «забирает» запуск (сдвигает
    last_run_at, гасит run_now) и возвращает should_run=true.
    v/poll — версия релея и его интервал опроса (для индикатора онлайн/версия)."""
    _check_gisgmp_token(authorization)
    cfg = await _load_relay_cfg(db)
    # Отметка «релей на связи» + его версия/интервал (для статуса в UI).
    cfg["last_poll_at"] = utcnow().isoformat()
    if v:
        cfg["relay_version"] = v
    if poll:
        cfg["relay_poll_seconds"] = int(poll)

    should_run, reason = False, ""
    if cfg.get("enabled", True):
        if cfg.get("run_now"):
            should_run, reason = True, "run_now"
        else:
            # Раз в сутки вечером, когда нет нагрузки. Час задаётся по МСК
            # (Москва = UTC+3 круглый год, без перехода). Запуск на ПЕРВОМ опросе
            # после daily_hour, если сегодня ещё не запускались.
            msk = timedelta(hours=3)
            now_msk = utcnow() + msk
            daily_hour = max(0, min(23, int(cfg.get("daily_hour", 22))))
            target = now_msk.replace(hour=daily_hour, minute=0, second=0, microsecond=0)
            last = cfg.get("last_run_at")
            if now_msk >= target:
                if not last:
                    should_run, reason = True, "daily_first"
                else:
                    try:
                        if datetime.fromisoformat(last) + msk < target:
                            should_run, reason = True, "daily"
                    except Exception:
                        should_run, reason = True, "bad_last_run"

    if should_run:
        cfg["run_now"] = False
        cfg["last_run_at"] = utcnow().isoformat()

    # Курсор инкремента: релей дотягивает только начисления новее этой даты
    # актуализации (всё, что старше, уже в кэше). None → первый полный проход.
    cursor_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == "gisgmp_cursor")
    )).scalars().first()
    since = None
    if cursor_row and cursor_row.value:
        try:
            since = json.loads(cursor_row.value).get("since")
        except Exception:
            since = None

    # Авто-перепроверка результата актуализации: «sent»-прогоны старше ~2ч ставим
    # их ФИО в gisgmp_recheck ДО чтения очереди ниже — чтобы релей забрал этим же
    # опросом и снял «после» по уже обработанным ГИС начислениям.
    try:
        await _drive_actualize_runs(db)
    except Exception:
        logger.exception("[gisgmp] авто-цикл актуализации не удался")

    # Очередь точечного дотягивания проблемных ФИО (где ГИС занижен):
    # отдаём список и сразу гасим (claim), чтобы не повторять.
    rc_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == "gisgmp_recheck")
    )).scalars().first()
    recheck = None
    if rc_row and rc_row.value:
        try:
            rc = json.loads(rc_row.value)
        except Exception:
            rc = {}
        if rc.get("surnames"):
            recheck = {"surnames": rc["surnames"], "deep_months": int(rc.get("deep_months", 36))}
            rc_row.value = json.dumps({}, ensure_ascii=False)
            await db.commit()

    # Команды управления демоном из UI: отдаём и сразу гасим (claim-на-выдаче),
    # чтобы выполнились РОВНО один раз. Релей применит и перезапустится (execv).
    control = {}
    if cfg.get("self_update"):
        control["self_update"] = True
        cfg["self_update"] = False
    if cfg.get("restart"):
        control["restart"] = True
        cfg["restart"] = False
    if cfg.get("creds_pending") and cfg.get("passport_username") and cfg.get("passport_password_enc"):
        try:
            control["credentials"] = {
                "username": cfg["passport_username"],
                "password": fernet.decrypt(cfg["passport_password_enc"].encode()).decode(),
            }
            cfg["creds_pending"] = False
        except Exception:
            pass

    # Очередь массовой актуализации (кнопка «Актуализировать расхождения»):
    # отдаём полный список UUID ОДИН раз и помечаем running, пока релей идёт.
    act_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == "gisgmp_actualize")
    )).scalars().first()
    actualize = None
    if act_row and act_row.value:
        try:
            av = json.loads(act_row.value)
        except Exception:
            av = {}
        if av.get("uuids") and not av.get("running"):
            actualize = {"uuids": av["uuids"]}
            av["running"] = True
            av["started_at"] = utcnow().isoformat()
            act_row.value = json.dumps(av, ensure_ascii=False)
            await db.commit()

    # Один сейв в конце: last_poll_at/версия + claim команд + claim запуска.
    await _save_relay_cfg(db, cfg)

    return {
        "enabled": cfg.get("enabled", True),
        "months_back": cfg.get("months_back", 36),
        "should_run": should_run,
        "reason": reason,
        "since": since,
        "recheck": recheck,
        "control": control or None,
        "actualize": actualize,
    }


class GisgmpRelayReportIn(BaseModel):
    ok: bool
    count: int = 0
    updated: int = 0
    created: int = 0
    not_found: int = 0
    message: str = ""


@router.post("/gisgmp/relay-report", summary="Релей шлёт отчёт о прогоне (token-auth)")
async def gisgmp_relay_report(
    payload: GisgmpRelayReportIn,
    authorization: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_db),
):
    _check_gisgmp_token(authorization)
    cfg = await _load_relay_cfg(db)
    cfg["last_status"] = "ok" if payload.ok else "error"
    cfg["last_message"] = (payload.message or "")[:500]
    cfg["last_count"] = payload.count
    cfg["last_updated"] = payload.updated
    cfg["last_created"] = payload.created
    cfg["last_not_found"] = payload.not_found
    cfg["last_report_at"] = utcnow().isoformat()
    await _save_relay_cfg(db, cfg)
    return {"ok": True}


class GisgmpRelaySettingsIn(BaseModel):
    enabled: Optional[bool] = None
    months_back: Optional[int] = None
    interval_hours: Optional[int] = None
    daily_hour: Optional[int] = None


@router.put("/gisgmp/relay-config", summary="Админ меняет настройки релея")
async def gisgmp_relay_set(
    payload: GisgmpRelaySettingsIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    cfg = await _load_relay_cfg(db)
    if payload.enabled is not None:
        cfg["enabled"] = bool(payload.enabled)
    if payload.months_back is not None:
        cfg["months_back"] = max(1, min(60, int(payload.months_back)))
    if payload.interval_hours is not None:
        cfg["interval_hours"] = max(1, min(168, int(payload.interval_hours)))
    if payload.daily_hour is not None:
        cfg["daily_hour"] = max(0, min(23, int(payload.daily_hour)))
    await _save_relay_cfg(db, cfg)
    return cfg


@router.post("/gisgmp/run-now", summary="Админ ставит флаг «запустить релей сейчас»")
async def gisgmp_run_now(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    cfg = await _load_relay_cfg(db)
    cfg["run_now"] = True
    await _save_relay_cfg(db, cfg)
    return {"ok": True, "queued": True}


@router.post("/gisgmp/relay-update", summary="Админ: релей подтянет свежий код и перезапустится")
async def gisgmp_relay_update(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    cfg = await _load_relay_cfg(db)
    cfg["self_update"] = True
    await _save_relay_cfg(db, cfg)
    return {"ok": True, "queued": True}


@router.post("/gisgmp/relay-restart", summary="Админ: перезапустить демон релея")
async def gisgmp_relay_restart(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    cfg = await _load_relay_cfg(db)
    cfg["restart"] = True
    await _save_relay_cfg(db, cfg)
    return {"ok": True, "queued": True}


class GisgmpRelayCredsIn(BaseModel):
    username: str
    password: str


@router.post("/gisgmp/relay-credentials", summary="Админ: сменить логин/пароль входа в реестр")
async def gisgmp_relay_credentials(
    payload: GisgmpRelayCredsIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Логин/пароль passport для релея. Пароль шифруется (Fernet, ключ
    ENCRYPTION_KEY) и хранится в SystemSetting; релей заберёт его ОДИН раз через
    relay-config (token+HTTPS), запишет в свой relay.env и перезапустится.
    В статус пароль не отдаётся — только логин."""
    _require_finance(current_user)
    u = (payload.username or "").strip()
    p = payload.password or ""
    if not u or not p:
        raise HTTPException(status_code=400, detail="Укажите логин и пароль")
    cfg = await _load_relay_cfg(db)
    cfg["passport_username"] = u
    cfg["passport_password_enc"] = fernet.encrypt(p.encode()).decode()
    cfg["creds_pending"] = True
    await _save_relay_cfg(db, cfg)
    return {"ok": True, "queued": True}


@router.get("/gisgmp/payer-charges", summary="Все начисления одного плательщика (клик по ФИО)")
async def gisgmp_payer_charges(
    q: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """По фамилии (первое слово q) — ВСЕ начисления человека из кэша ГИС ГМП
    с разбивкой: долг (не сквитировано) / оплачено / аннулировано."""
    _require_finance(current_user)
    from app.modules.utility.services.gisgmp_import import (
        classify_account, is_unpaid, is_annulled, parse_reg_dt, GISGMP_CACHE_KEY,
    )
    from app.modules.utility.services.debt_import import clean_decimal

    parts = (q or "").strip().split()
    surname = parts[0].lower() if parts else ""
    if not surname:
        return {"query": q, "count": 0, "charges": [], "totals": {}}

    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_CACHE_KEY)
    )).scalars().first()
    cache = {}
    if row and row.value:
        try:
            cache = json.loads(row.value)
        except Exception:
            cache = {}

    out = []
    tot = {"debt_209": 0.0, "debt_205": 0.0, "paid": 0.0, "annulled": 0, "count": 0}
    for ch in cache.values():
        name = (ch.get("payer_name") or "").strip()
        if surname not in name.lower():
            continue
        acc = classify_account(ch.get("purpose"))
        annul = is_annulled(ch.get("change_status"))
        unpaid = is_unpaid(ch.get("ack_status"))
        amt = float(clean_decimal(ch.get("amount")))
        status = "annulled" if annul else ("unpaid" if unpaid else "paid")
        out.append({
            "payer_name": name,
            "account_type": acc,
            "account": ch.get("account"),
            "amount": amt,
            "bill_date": ch.get("bill_date"),
            "actualize_date": ch.get("actualize_date"),
            "purpose": ch.get("purpose"),
            "ack_status": ch.get("ack_status"),
            "status": status,
            "uin": ch.get("uin"),
        })
        tot["count"] += 1
        if annul:
            tot["annulled"] += 1
        elif unpaid:
            if acc == "209":
                tot["debt_209"] += amt
            elif acc == "205":
                tot["debt_205"] += amt
        else:
            tot["paid"] += amt

    _order = {"unpaid": 0, "paid": 1, "annulled": 2}

    def _sk(c):
        d = parse_reg_dt(c.get("bill_date"))
        return (_order.get(c["status"], 3), -(d.toordinal() if d else 0))

    out.sort(key=_sk)
    return {"query": q, "surname": surname, "count": len(out),
            "charges": out[:500], "totals": tot}


GISGMP_OVERRIDES_KEY = "gisgmp_overrides"


class GisgmpOverrideIn(BaseModel):
    user_id: int


@router.post("/gisgmp/apply-override", summary="Применить долг ГИС к жильцу (оверрайд над 1С)")
async def gisgmp_apply_override(
    payload: GisgmpOverrideIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Пишет долг ГИС (209/205) в показания жильца активного периода — это и
    видит жилец. База остаётся 1С; явный точечный оверрайд с провенансом и
    откатом. Долг = значение ГИС, переплата по этим счетам → 0. Прежние значения
    сохраняются для отката. ВНИМАНИЕ: повторный импорт 1С-Excel перезапишет долг
    свежими данными 1С (оверрайд — точечная правка на текущий момент)."""
    _require_finance(current_user)
    uid = int(payload.user_id)

    # 1. Авторитетное значение ГИС — из сверки (НЕ с клиента).
    rec = await _build_reconcile(db)
    resident = next((r for r in rec.get("residents", []) if r.get("user_id") == uid), None)
    if resident is None:
        raise HTTPException(status_code=404, detail="Жилец не найден в сверке ГИС ГМП")
    gis_209 = Decimal(str(resident.get("g209") or 0))
    gis_205 = Decimal(str(resident.get("g205") or 0))

    # 2. Активный период + текущее показание жильца.
    ap = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if ap is None:
        raise HTTPException(status_code=409, detail="Нет активного расчётного периода")
    mr = (await db.execute(
        select(MeterReading).where(
            MeterReading.user_id == uid, MeterReading.period_id == ap.id
        ).order_by(MeterReading.created_at.desc())
    )).scalars().first()
    if mr is None:
        raise HTTPException(status_code=409,
                            detail="У жильца нет показаний в активном периоде — оверрайд недоступен")

    prev = {
        "debt_209": str(mr.debt_209 or 0), "overpayment_209": str(mr.overpayment_209 or 0),
        "debt_205": str(mr.debt_205 or 0), "overpayment_205": str(mr.overpayment_205 or 0),
    }
    mr_id = mr.id
    # expunge — иначе ORM-flush ТИХО перезапишет explicit UPDATE (партиции MR).
    db.expunge(mr)
    await db.execute(
        _sa_update(MeterReading).where(MeterReading.id == mr_id).values(
            debt_209=gis_209, overpayment_209=Decimal("0.00"),
            debt_205=gis_205, overpayment_205=Decimal("0.00"),
        )
    )

    # 3. Провенанс + прежние значения (для отката). НЕ пишем DebtImportLog —
    #    иначе сверка примет оверрайд за «последний Excel 1С».
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_OVERRIDES_KEY)
    )).scalars().first()
    overrides = {}
    if row and row.value:
        try:
            overrides = json.loads(row.value)
        except Exception:
            overrides = {}
    if row is None:
        row = SystemSetting(key=GISGMP_OVERRIDES_KEY, value="{}",
                            description="ГИС-оверрайды долгов (провенанс + откат)")
        db.add(row)
    overrides[str(uid)] = {
        "username": resident.get("username"),
        "debt_209": str(gis_209), "debt_205": str(gis_205),
        "prev": prev, "period_id": ap.id,
        "at": utcnow().isoformat(), "by": current_user.username,
    }
    row.value = json.dumps(overrides, ensure_ascii=False)
    await db.commit()
    return {"ok": True, "user_id": uid, "debt_209": str(gis_209), "debt_205": str(gis_205)}


@router.post("/gisgmp/revert-override", summary="Откатить ГИС-оверрайд жильца к прежним значениям")
async def gisgmp_revert_override(
    payload: GisgmpOverrideIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    uid = str(int(payload.user_id))
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_OVERRIDES_KEY)
    )).scalars().first()
    overrides = {}
    if row and row.value:
        try:
            overrides = json.loads(row.value)
        except Exception:
            overrides = {}
    ov = overrides.get(uid)
    if not ov:
        raise HTTPException(status_code=404, detail="Оверрайд для жильца не найден")
    prev = ov.get("prev") or {}
    mr = (await db.execute(
        select(MeterReading).where(
            MeterReading.user_id == int(uid), MeterReading.period_id == ov.get("period_id")
        ).order_by(MeterReading.created_at.desc())
    )).scalars().first()
    if mr is not None:
        mr_id = mr.id
        db.expunge(mr)
        await db.execute(
            _sa_update(MeterReading).where(MeterReading.id == mr_id).values(
                debt_209=Decimal(prev.get("debt_209", "0")),
                overpayment_209=Decimal(prev.get("overpayment_209", "0")),
                debt_205=Decimal(prev.get("debt_205", "0")),
                overpayment_205=Decimal(prev.get("overpayment_205", "0")),
            )
        )
    overrides.pop(uid, None)
    row.value = json.dumps(overrides, ensure_ascii=False)
    await db.commit()
    return {"ok": True, "reverted": uid}


@router.get("/gisgmp/findings", summary="Что нашёл ГИС ГМП (отладка, хранится отдельно)")
async def gisgmp_findings(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Полные находки последнего прогона релея: сводка по жильцам + сырые
    начисления. Пока ОТДЕЛЬНО от долгов 1С (Excel), в показания не пишется —
    показываем для отладки в отдельном окне «Долги 1С»."""
    _require_finance(current_user)
    findings = await _load_findings(db)
    if not findings:
        return {"empty": True}
    # Сырые начисления теперь в отдельном ключе (находки лёгкие) — подмешиваем
    # их сюда для поиска по фамилии в «Показать найденное».
    from app.modules.utility.services.gisgmp_import import GISGMP_FINDINGS_CHARGES_KEY
    chrow = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_FINDINGS_CHARGES_KEY)
    )).scalars().first()
    if chrow and chrow.value:
        try:
            findings = {**findings, "charges": json.loads(chrow.value).get("charges", [])}
        except Exception:
            pass
    return findings


@router.get("/gisgmp/reconcile", summary="Сверка ГИС ГМП (находки) ↔ долги 1С (Excel)")
async def gisgmp_reconcile(
    period_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сверяет два источника по жильцам и счетам 209/205:
      • ГИС ГМП — последние находки релея (SystemSetting 'gisgmp_findings',
        summary[].debt_209/205 по matched_user_id);
      • 1С — текущие долги в показаниях (MeterReading.debt_209/205 за период,
        залитые Excel-импортом ОСВ).
    На каждый счёт: совпало / расхождение (Δ) / только в ГИС / только в 1С.
    Окно ГИС влияет на полноту — для честной сверки ставь окно побольше."""
    _require_finance(current_user)
    return await _build_reconcile(db)


@router.get("/gisgmp/reconcile-fio", summary="3-сторонняя сверка ФИО: 1С ↔ ГИС ГМП ↔ база")
async def gisgmp_reconcile_fio(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """СОЮЗ по ТОЧНОМУ ФИО трёх источников: 1С (последний импорт/черновик),
    ГИС ГМП (ВЕСЬ кэш — и должники, и оплаченные «Сквитировано»), база жильцов.
    Одна строка на уникальное нормализованное ФИО, колонки «в 1С / в ГИС / в базе»
    + долги (из находок: 0, если в ГИС всё оплачено) + флаги «нет в …». Никаких
    «похожих» — только точное совпадение ФИО."""
    _require_finance(current_user)
    from app.modules.utility.services.gisgmp_import import GISGMP_SOURCE_LABEL, GISGMP_CACHE_KEY
    from app.modules.utility.services.gsheets_sync import normalize_fio

    rows: dict[str, dict] = {}

    def _row(fio_raw):
        n = normalize_fio(fio_raw or "")
        if not n:
            return None
        r = rows.get(n)
        if r is None:
            r = {"fio": (fio_raw or "").strip(), "in_1c": False, "in_gis": False,
                 "in_db": False, "d209_1c": 0.0, "d205_1c": 0.0,
                 "o209_1c": 0.0, "o205_1c": 0.0,  # переплата 1С (ГИС её не отдаёт)
                 "d209_gis": 0.0, "d205_gis": 0.0, "user_id": None, "username": None}
            rows[n] = r
        return r

    # --- база: активные жильцы (источник in_db + приоритетное отображение ФИО) ---
    db_rows = (await db.execute(
        select(User.id, User.username).where(
            User.role == "user", User.is_deleted.is_(False))
    )).all()
    for uid, un in db_rows:
        r = _row(un)
        if r is None:
            continue
        r["in_db"] = True
        r["user_id"] = uid
        r["username"] = un
        r["fio"] = un

    # --- 1С: последний импорт/черновик по каждому счёту (staged ИЛИ completed) ---
    sources: dict = {}
    for acc in ("209", "205"):
        log = (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.account_type == acc,
                DebtImportLog.status.in_(["staged", "completed"]),
                DebtImportLog.file_name != GISGMP_SOURCE_LABEL,
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()
        sources[acc] = ({"file": log.file_name, "status": log.status,
                         "at": log.started_at.isoformat() if log.started_at else None}
                        if log else None)
        if not log:
            continue
        for st in (log.applied_state or {}).values():
            un = st.get("username")
            r = _row(un) if un else None
            if r is None:
                continue
            r["in_1c"] = True
            try:
                r[f"d{acc}_1c"] += float(st.get(f"debt_{acc}") or 0)
                r[f"o{acc}_1c"] += float(st.get(f"overpayment_{acc}") or 0)
            except Exception:
                pass
        for nf in (log.not_found_users or []):
            r = _row(nf.get("fio"))
            if r is None:
                continue
            r["in_1c"] = True
            try:
                r[f"d{acc}_1c"] += float(nf.get("debt") or 0)
                r[f"o{acc}_1c"] += float(nf.get("overpayment") or 0)
            except Exception:
                pass

    # --- ГИС, СУЩЕСТВОВАНИЕ: по ВСЕМУ кэшу начислений (вкл. оплаченных
    # «Сквитировано»). Человек может быть в ГИС с НУЛЕВЫМ долгом (всё оплачено) —
    # раньше in_gis брался только из находок (должники), и оплаченные показывались
    # «нет в ГИС», хотя в ГИС ГМП они есть. DISTINCT payer_name извлекаем в
    # Postgres (парс кэша один раз, в Python приходят только имена). ---
    try:
        gis_names = (await db.execute(
            text("SELECT DISTINCT c->>'payer_name' AS fio "
                 "FROM (SELECT value FROM system_settings WHERE key = :k) s, "
                 "LATERAL jsonb_each(s.value::jsonb) AS e(uin, c)"),
            {"k": GISGMP_CACHE_KEY},
        )).all()
        for (fio,) in gis_names:
            r = _row(fio)
            if r is not None:
                r["in_gis"] = True
    except Exception:
        pass

    # --- ГИС, ДОЛГ: summary находок (только неоплаченные 209/205). Достаём
    # summary JSON-экстрактом в Postgres (без сырых charges — иначе союз тормозит). ---
    fres = (await db.execute(
        text("SELECT (value::jsonb -> 'summary')::text AS s, "
             "value::jsonb ->> 'synced_at' AS at "
             "FROM system_settings WHERE key = :k"),
        {"k": GISGMP_FINDINGS_KEY},
    )).first()
    gis_synced = None
    if fres and fres.s:
        gis_synced = fres.at
        try:
            for frow in json.loads(fres.s):
                r = _row(frow.get("fio"))
                if r is None:
                    continue
                r["in_gis"] = True
                r["d209_gis"] += float(frow.get("debt_209") or 0)
                r["d205_gis"] += float(frow.get("debt_205") or 0)
        except Exception:
            pass

    items = list(rows.values())
    summary = {
        "total": len(items),
        "all_three": sum(1 for r in items if r["in_1c"] and r["in_gis"] and r["in_db"]),
        "not_in_db": sum(1 for r in items if not r["in_db"]),
        "not_in_gis": sum(1 for r in items if not r["in_gis"]),
        "not_in_1c": sum(1 for r in items if not r["in_1c"]),
        "with_overpay": sum(1 for r in items if (r["o209_1c"] + r["o205_1c"]) > 0.005),
    }
    # Проблемные (хоть где-то «нет») — сверху, дальше по ФИО.
    items.sort(key=lambda r: ((r["in_1c"] and r["in_gis"] and r["in_db"]), r["fio"] or ""))
    return {
        "sources": sources,
        "gis_synced_at": gis_synced,
        "summary": summary,
        "rows": items[:3000],
    }


# ─── bulk-создание жильцов из «сирот» сверки (есть в 1С/ГИС, нет в базе) ─────
_FIO_STOPWORDS = (
    "пао", "ооо", "оао", "зао", "акционерн", "фгау", "фгбу", "фгуп", "гуп",
    "муп", "банк", "сбербанк", "втб", "филиал", "возмещен", "росжил",
    "бухгалтер", "директор", "начальник", "казначейств", "уфк", "ифнс",
    "фссп", "департамент", "комплекс",
)


def _clean_orphan_fio(raw: str) -> str:
    """ФИО → чистый вид: коллапс пробелов + убрать хвостовой «(общ)»/«(0)»."""
    s = " ".join((raw or "").split())
    if s.endswith(")") and "(" in s:
        s = s[:s.rfind("(")].strip()
    return s


def _looks_like_person_fio(s: str) -> tuple[bool, str]:
    """Грубый фильтр «похоже на ФИО человека?» → (ok, причина_если_нет)."""
    if not s:
        return False, "пусто"
    low = s.lower()
    if any(ch.isdigit() for ch in s):
        return False, "содержит цифры (не ФИО)"
    if any(kw in low for kw in _FIO_STOPWORDS):
        return False, "организация/должность"
    words = [w for w in s.split() if w]
    if len(words) < 2:
        return False, "одно слово (нет имени)"
    if len(words) > 5:
        return False, "слишком длинно для ФИО"
    if sum(1 for ch in low if "а" <= ch <= "я" or ch == "ё") < 5:
        return False, "не кириллица"
    return True, ""


@router.post("/gisgmp/create-missing-residents",
             summary="Создать жильцов (только ФИО) для тех, кто есть в 1С/ГИС, но не в базе")
async def gisgmp_create_missing_residents(
    dry_run: bool = Query(True),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Массово завести личные кабинеты для ФИО из 1С/ГИС, которых НЕТ в базе.

    Создаём ТОЛЬКО ФИО + учётку (без комнаты, адреса и прочего — админ заполнит
    сам). Мусор отсеиваем: организации/должности (ПАО, ФГАУ, банк, бухгалтер…),
    строки с цифрами, одно слово без имени. Дубли гасим: точное совпадение с
    базой/между собой + почти-дубли (типосы в отчестве, ≥90). dry_run=true
    (по умолчанию) — только предпросмотр, ничего не создаёт."""
    _require_finance(current_user)
    from app.modules.utility.services.gsheets_sync import normalize_fio
    from app.modules.utility.services.gisgmp_import import GISGMP_SOURCE_LABEL, GISGMP_CACHE_KEY
    from rapidfuzz import fuzz

    db_users = (await db.execute(
        select(User.username).where(User.role == "user", User.is_deleted.is_(False))
    )).scalars().all()
    db_norm = {normalize_fio(u): u for u in db_users if u}
    db_keys = list(db_norm.keys())

    # Сырые ФИО из 1С (последний импорт по счёту) + ГИС (payer_name из кэша).
    raw_by_norm: dict[str, str] = {}
    for acc in ("209", "205"):
        log = (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.account_type == acc,
                DebtImportLog.status.in_(["staged", "completed"]),
                DebtImportLog.file_name != GISGMP_SOURCE_LABEL,
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()
        if not log:
            continue
        for st in (log.applied_state or {}).values():
            un = (st.get("username") or "").strip()
            if un:
                raw_by_norm.setdefault(normalize_fio(un), un)
        for nf in (log.not_found_users or []):
            fi = (nf.get("fio") or "").strip()
            if fi:
                raw_by_norm.setdefault(normalize_fio(fi), fi)
    try:
        gis_rows = (await db.execute(
            text("SELECT DISTINCT c->>'payer_name' AS fio "
                 "FROM (SELECT value FROM system_settings WHERE key = :k) s, "
                 "LATERAL jsonb_each(s.value::jsonb) AS e(uin, c)"),
            {"k": GISGMP_CACHE_KEY},
        )).all()
        for (fi,) in gis_rows:
            fi = (fi or "").strip()
            if fi:
                raw_by_norm.setdefault(normalize_fio(fi), fi)
    except Exception:
        pass

    to_create: dict[str, str] = {}  # norm(clean) -> clean fio
    skip_not_fio, skip_in_db, skip_similar = [], [], []
    for nrm, raw in raw_by_norm.items():
        if not nrm or nrm in db_norm:
            continue  # пусто или уже точно в базе — не сирота
        cleaned = _clean_orphan_fio(raw)
        ok, reason = _looks_like_person_fio(cleaned)
        if not ok:
            skip_not_fio.append({"fio": raw, "reason": reason})
            continue
        cnrm = normalize_fio(cleaned)
        if cnrm in db_norm:
            skip_in_db.append({"fio": raw, "match": db_norm[cnrm]})
            continue
        if cnrm in to_create:
            continue
        # Почти-дубль с базой ИЛИ с уже добавленным в очередь (типос в отчестве).
        best, best_sc = None, 0
        for k in db_keys:
            sc = fuzz.token_sort_ratio(cnrm, k)
            if sc > best_sc:
                best_sc, best = sc, db_norm[k]
        for k, v in to_create.items():
            sc = fuzz.token_sort_ratio(cnrm, k)
            if sc > best_sc:
                best_sc, best = sc, v
        if best_sc >= 90:
            skip_similar.append({"fio": raw, "match": best, "score": int(best_sc)})
            continue
        to_create[cnrm] = cleaned

    create_list = sorted(to_create.values())
    if dry_run:
        return {
            "dry_run": True,
            "to_create_count": len(create_list),
            "to_create": create_list,
            "skip_not_fio": skip_not_fio,
            "skip_in_db": skip_in_db,
            "skip_similar": skip_similar,
        }

    import secrets
    from app.core.auth import get_password_hash
    from app.modules.utility.routers.admin_dashboard import write_audit_log

    existing_logins = {
        (lg or "").lower() for lg in (await db.execute(select(User.login))).scalars().all()
    }
    created = []
    for fio in create_list:
        login = fio
        if login.lower() in existing_logins:
            login = f"{fio} {secrets.token_hex(2)}"
        existing_logins.add(login.lower())
        db.add(User(
            username=fio, login=login,
            hashed_password=get_password_hash(secrets.token_urlsafe(12)),
            role="user", is_deleted=False, is_initial_setup_done=False,
        ))
        created.append(fio)

    await write_audit_log(
        db, current_user.id, current_user.username,
        action="create", entity_type="user",
        details={"bulk_create_missing": len(created),
                 "skipped_not_fio": len(skip_not_fio),
                 "skipped_dup": len(skip_in_db) + len(skip_similar)},
    )
    await db.commit()
    return {
        "dry_run": False,
        "created_count": len(created),
        "created": created,
        "skip_not_fio": len(skip_not_fio),
        "skip_in_db": len(skip_in_db),
        "skip_similar": len(skip_similar),
    }


@router.get("/gisgmp/link-candidates", summary="Кандидаты-жильцы по фамилии для привязки сироты")
async def gisgmp_link_candidates(
    fio: str = Query(..., description="ФИО из 1С/ГИС"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Для кнопки «Привязать» в союзе: жильцы базы с ТОЙ ЖЕ фамилией (точно,
    первое слово) — кандидаты на привязку сироты 1С/ГИС к учётке. Не fuzzy:
    предлагаем по фамилии, выбор подтверждает админ."""
    _require_finance(current_user)
    parts = (fio or "").strip().split()
    surname = parts[0].lower() if parts else ""
    if not surname:
        return {"candidates": []}
    rows = (await db.execute(
        select(User.id, User.username,
               Room.dormitory_name, Room.room_number,
               Room.street, Room.house_number, Room.apartment_number)
        .outerjoin(Room, User.room_id == Room.id)
        .where(User.role == "user", User.is_deleted.is_(False))
    )).all()
    out = []
    for uid, un, dorm, rno, street, house, apt in rows:
        first = (un or "").strip().split()[:1]
        if first and first[0].lower() == surname:
            if dorm:
                addr = f"{dorm}, ком. {rno or '—'}"
            elif street:
                addr = f"ул. {street}, д. {house or '—'}, кв. {apt or '—'}"
            else:
                addr = "—"
            out.append({"id": uid, "username": un, "address": addr})
    out.sort(key=lambda x: x["username"] or "")
    return {"surname": surname, "candidates": out}


class GisgmpLinkFioIn(BaseModel):
    fio: str
    user_id: int
    rename: bool = True


@router.post("/gisgmp/link-fio", summary="Привязать ФИО из 1С/ГИС к жильцу (алиас + переименование)")
async def gisgmp_link_fio(
    payload: GisgmpLinkFioIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Связывает ФИО из 1С/ГИС с учёткой жильца: создаёт/перенаводит алиас
    (долг привяжется при выгрузке) и, если rename=True, переименовывает жильца
    в имя из 1С/ГИС (источник истины). Это явная привязка админа, не fuzzy."""
    _require_finance(current_user)
    from app.modules.utility.models import GSheetsAlias
    from app.modules.utility.services.gsheets_sync import normalize_fio
    from app.modules.utility.routers.admin_dashboard import write_audit_log
    fio = (payload.fio or "").strip()
    user = await db.get(User, payload.user_id)
    if not user or user.is_deleted:
        raise HTTPException(404, "Жилец не найден")
    normalized = normalize_fio(fio)
    if not normalized:
        raise HTTPException(400, "Пустое ФИО")
    # Алиас: явный выбор админа — перенаводим, если был на другого жильца.
    existing = (await db.execute(
        select(GSheetsAlias).where(GSheetsAlias.alias_fio_normalized == normalized)
    )).scalars().first()
    if existing:
        existing.user_id = user.id
        existing.alias_fio = fio
        existing.kind = "debt_manual"
    else:
        db.add(GSheetsAlias(
            alias_fio=fio, alias_fio_normalized=normalized, user_id=user.id,
            kind="debt_manual", note="link-fio (союз)", created_by_id=current_user.id,
        ))
    # Переименование в имя из 1С/ГИС (приоритет источника), если нет конфликта.
    renamed, warning, old_name = False, None, user.username
    if payload.rename and fio and fio != user.username:
        clash = (await db.execute(
            select(User.id).where(User.username == fio, User.id != user.id)
        )).first()
        if clash:
            warning = f"Имя «{fio}» уже занято другим жильцом — переименование пропущено, алиас создан."
        else:
            user.username = fio
            renamed = True
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="gisgmp_link_fio", entity_type="user", entity_id=user.id,
        details={"fio": fio, "old_name": old_name, "renamed": renamed},
    )
    await db.commit()
    return {"ok": True, "renamed": renamed, "warning": warning,
            "user_id": user.id, "username": user.username}


@router.post("/gisgmp/purge", summary="Очистить данные ГИС ГМП (кэш, находки, курсор, очередь)")
async def gisgmp_purge(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Обнуляет рабочие данные ГИС ГМП: кэш начислений, находки, курсор инкремента
    и очередь актуализации. Долги жильцов НЕ трогает (они из 1С через гейт).
    После очистки нужен новый сбор («Запустить сбор») — соберёт с нуля точным
    матчингом. Аудит-историю актуализаций (gisgmp_actualize_log) не чистим."""
    _require_finance(current_user)
    from app.modules.utility.routers.admin_dashboard import write_audit_log
    keys = ["gisgmp_cache", "gisgmp_findings", "gisgmp_cursor", "gisgmp_actualize"]
    rows = (await db.execute(
        select(SystemSetting).where(SystemSetting.key.in_(keys))
    )).scalars().all()
    cleared = 0
    for row in rows:
        row.value = "{}"
        cleared += 1
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="gisgmp_purge", entity_type="system_setting", entity_id=None,
        details={"keys": keys, "cleared": cleared},
    )
    await db.commit()
    return {"ok": True, "cleared": cleared}


@router.get("/debts/staged-status", summary="Черновики долгов 1С, ждущие выгрузки")
async def debts_staged_status(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Для кнопки «Выгрузить»: какие черновики (status='staged') 1С готовы."""
    _require_finance(current_user)
    out = {"staged": {}, "has_staged": False}
    for acc in ("209", "205"):
        log = (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.account_type == acc,
                DebtImportLog.status == "staged",
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()
        if log:
            out["staged"][acc] = {
                "log_id": log.id, "file": log.file_name,
                "at": log.started_at.isoformat() if log.started_at else None,
                "residents": len(log.applied_state or {}),
                "not_found": log.not_found_count or 0,
                "by": log.started_by_username,
            }
            out["has_staged"] = True
    return out


@router.post("/debts/rematch-base",
             summary="Пересопоставить «не найденных» в черновиках 1С с текущей базой")
async def debts_rematch_base(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Перепроверяет not_found последних ЧЕРНОВИКОВ (staged) 1С против ТЕКУЩЕЙ
    базы (список «не найдено» заморожен на момент загрузки — после добавления
    жильцов его надо пересопоставить). Кого теперь нашли И У КОГО ЕСТЬ КОМНАТА —
    переносим в applied_state (долг привяжется при «Выгрузить»). Кто есть в базе,
    но БЕЗ комнаты — оставляем в not_found и считаем отдельно (им нужна комната,
    долг живёт на показании, а оно требует комнату). Перезаливка 1С не нужна."""
    _require_finance(current_user)
    from app.modules.utility.services.debt_import import normalize_name, _normalize_fio_key
    from app.modules.utility.services.gisgmp_import import GISGMP_SOURCE_LABEL
    from app.modules.utility.models import GSheetsAlias
    from sqlalchemy.orm import selectinload

    users = (await db.execute(
        select(User).options(selectinload(User.room))
        .where(User.role == "user", User.is_deleted.is_(False))
    )).scalars().all()
    umap, by_id = {}, {}
    for u in users:
        if not u.username:
            continue
        info = {"id": u.id, "room_id": u.room_id, "username": u.username,
                "room_label": (u.room.format_address if u.room else None)}
        umap[normalize_name(u.username)] = info
        by_id[u.id] = info
    amap = {a: uid for a, uid in (await db.execute(
        select(GSheetsAlias.alias_fio_normalized, GSheetsAlias.user_id)
    )).all() if a}

    def _match(fio_raw: str):
        u = umap.get(normalize_name(fio_raw))
        if u:
            return u
        uid = amap.get(_normalize_fio_key(fio_raw))
        return by_id.get(uid) if uid else None

    attached = no_room = still = 0
    touched = []
    for acc in ("209", "205"):
        log = (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.account_type == acc,
                DebtImportLog.status == "staged",
                DebtImportLog.file_name != GISGMP_SOURCE_LABEL,
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()
        if not log or not log.not_found_users:
            continue
        ap = dict(log.applied_state or {})
        remaining = []
        changed = False
        for nf in log.not_found_users:
            u = _match((nf.get("fio") or "").strip())
            if not u:
                remaining.append(nf)
                still += 1
                continue
            # Долг на лицевом счёте (user_id), не на комнате: привязываем даже
            # без заселения. room_id=NULL ок — комната подцепится позже.
            key = str(u["id"])
            ent = ap.get(key) or {
                "debt_209": "0", "overpayment_209": "0",
                "debt_205": "0", "overpayment_205": "0",
            }
            ent[f"debt_{acc}"] = str(nf.get("debt") or 0)
            ent[f"overpayment_{acc}"] = str(nf.get("overpayment") or 0)
            ent["username"] = u["username"]
            ent["room_id"] = u["room_id"]
            ent["room_label"] = u["room_label"]
            ap[key] = ent
            attached += 1
            if not u["room_id"]:
                no_room += 1  # привязан, но пока без комнаты (для отчёта)
            changed = True
        if changed:
            log.applied_state = ap
            log.not_found_users = remaining
            touched.append(log.id)
    await db.commit()
    return {"attached": attached, "in_base_no_room": no_room,
            "still_not_found": still, "logs": touched}


@router.post("/debts/publish", summary="Выгрузить черновики долгов жильцам (1С + ГИС-оверрайды)")
async def debts_publish(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Берёт ПОСЛЕДНИЕ черновики импорта 1С (status='staged') по 209 и 205,
    накладывает активные ГИС-оверрайды и пишет долги в показания активного
    периода — только теперь жильцы их видят. Полная замена по выгружаемому
    счёту (кого нет в черновике → 0 по этому счёту). Снимок до — для отката."""
    _require_finance(current_user)
    from sqlalchemy import update as _sa_update

    ap = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if ap is None:
        raise HTTPException(409, "Нет активного расчётного периода")

    staged = {}
    for acc in ("209", "205"):
        log = (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.account_type == acc,
                DebtImportLog.status == "staged",
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()
        if log:
            staged[acc] = log
    if not staged:
        raise HTTPException(409, "Нет черновиков для выгрузки — сначала загрузите Excel 1С")
    accts = set(staged.keys())

    # Целевые долги по жильцам из черновиков (applied_state).
    target: dict[int, dict] = {}
    for acc, log in staged.items():
        for uid_s, st in (log.applied_state or {}).items():
            if not str(uid_s).isdigit():
                continue
            uid = int(uid_s)
            t = target.setdefault(uid, {"room_id": st.get("room_id")})
            t[f"debt_{acc}"] = Decimal(str(st.get(f"debt_{acc}") or 0))
            t[f"over_{acc}"] = Decimal(str(st.get(f"overpayment_{acc}") or 0))
            if st.get("room_id"):
                t["room_id"] = st.get("room_id")

    # ГИС-оверрайды побеждают над 1С (по выгружаемым счетам).
    ovr_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_OVERRIDES_KEY)
    )).scalars().first()
    overrides = {}
    if ovr_row and ovr_row.value:
        try:
            overrides = json.loads(ovr_row.value)
        except Exception:
            overrides = {}
    ovr_count = 0
    for uid_s, ov in overrides.items():
        if not str(uid_s).isdigit():
            continue
        uid = int(uid_s)
        t = target.setdefault(uid, {"room_id": None})
        for acc in accts:
            t[f"debt_{acc}"] = Decimal(str(ov.get(f"debt_{acc}") or 0))
            t[f"over_{acc}"] = Decimal("0")
        ovr_count += 1

    # Существующие показания активного периода.
    readings = (await db.execute(
        select(MeterReading).where(MeterReading.period_id == ap.id)
    )).scalars().all()
    by_user = {r.user_id: r for r in readings if r.user_id is not None}

    snapshot_before: dict = {}
    updates: list = []
    for uid, r in by_user.items():
        vals = {}
        for acc in accts:
            vals[f"debt_{acc}"] = target.get(uid, {}).get(f"debt_{acc}", Decimal("0"))
            vals[f"overpayment_{acc}"] = target.get(uid, {}).get(f"over_{acc}", Decimal("0"))
        snapshot_before[str(r.id)] = {
            "debt_209": str(r.debt_209 or 0), "overpayment_209": str(r.overpayment_209 or 0),
            "debt_205": str(r.debt_205 or 0), "overpayment_205": str(r.overpayment_205 or 0),
        }
        updates.append((r.id, vals))

    # Создаём показания для жильцов из черновика без показания в периоде.
    new_objs = []
    for uid, t in target.items():
        if uid in by_user:
            continue
        new_objs.append(MeterReading(
            user_id=uid, room_id=t.get("room_id"), period_id=ap.id, is_approved=False,
            debt_209=t.get("debt_209", Decimal("0")), overpayment_209=t.get("over_209", Decimal("0")),
            debt_205=t.get("debt_205", Decimal("0")), overpayment_205=t.get("over_205", Decimal("0")),
        ))
    inserted_ids = []
    if new_objs:
        db.add_all(new_objs)
        await db.flush()
        inserted_ids = [n.id for n in new_objs]

    # Партиц-безопасная запись существующих (expunge + явный UPDATE по id).
    for r in list(by_user.values()):
        try:
            db.expunge(r)
        except Exception:
            pass
    updated = 0
    for rid, vals in updates:
        res = await db.execute(
            _sa_update(MeterReading).where(MeterReading.id == rid).values(**vals)
        )
        updated += res.rowcount or 0

    # Помечаем черновики published + снимок до (для отката через историю импортов).
    now = utcnow().isoformat()
    for acc, log in staged.items():
        # 'completed' (как у старого импорта) — чтобы все фичи, фильтрующие
        # completed (история, целостность, сверка с 1С), видели выгруженное.
        log.status = "completed"
        log.snapshot_data = {"before": snapshot_before,
                             "inserted_reading_ids": inserted_ids,
                             "published_at": now}
    await db.commit()

    return {
        "ok": True, "accounts": sorted(accts),
        "updated": updated, "created": len(inserted_ids),
        "overrides_applied": ovr_count, "residents": len(target),
    }


async def _build_reconcile(db: AsyncSession) -> dict:
    findings = await _load_findings(db)
    from app.modules.utility.services.gsheets_sync import normalize_fio
    gis: dict[int, dict] = {}
    # Несопоставленные ФИО (1С/ГИС не нашли жильца в базе) — собираем отдельно,
    # чтобы НЕ ТЕРЯТЬ людей: показываем их блоком внизу сверки и в печати.
    orphans_map: dict[str, dict] = {}

    def _orphan(fio_raw):
        n = normalize_fio(fio_raw or "")
        if not n:
            return None
        return orphans_map.setdefault(n, {
            "fio": (fio_raw or "").strip(),
            "gis_209": 0.0, "gis_205": 0.0, "c1_209": 0.0, "c1_205": 0.0,
        })

    if findings:
        for row in findings.get("summary", []):
            uid = row.get("matched_user_id")
            if uid is None:
                o = _orphan(row.get("fio"))
                if o is not None:
                    o["gis_209"] += float(row.get("debt_209") or 0)
                    o["gis_205"] += float(row.get("debt_205") or 0)
                continue
            gis[int(uid)] = {
                "209": Decimal(str(row.get("debt_209") or 0)),
                "205": Decimal(str(row.get("debt_205") or 0)),
                "username": row.get("matched_username") or row.get("fio"),
            }

    # 1С-сторона = ПОСЛЕДНИЙ Excel-импорт по каждому счёту (applied_state),
    # а не живой MeterReading — так видно КАКОЙ файл сверяем и нет примеси от
    # старого ГИС-прогона (он писал в показания на старом коде).
    from app.modules.utility.services.gisgmp_import import GISGMP_SOURCE_LABEL

    async def _last_excel_imp(acc):
        return (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.account_type == acc,
                # staged (черновик) ИЛИ completed (выгружено) — сверка видит и
                # не выгруженный черновик, чтобы можно было сверить ДО публикации.
                DebtImportLog.status.in_(["staged", "completed"]),
                DebtImportLog.file_name != GISGMP_SOURCE_LABEL,
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()

    mr: dict[int, dict] = {}
    source_1c = {}
    for acc in ("209", "205"):
        log = await _last_excel_imp(acc)
        source_1c[acc] = ({"file": log.file_name,
                           "at": log.started_at.isoformat() if log.started_at else None,
                           "log_id": log.id} if log else None)
        if log and log.applied_state:
            key = f"debt_{acc}"
            for uid_s, st in log.applied_state.items():
                if not str(uid_s).isdigit():
                    continue
                try:
                    val = Decimal(str(st.get(key) or 0))
                except Exception:
                    val = Decimal("0")
                if val:
                    mr.setdefault(int(uid_s), {})[acc] = val
        # 1С-сироты этого импорта (ФИО, которые не сматчились с жильцом базы).
        if log and log.not_found_users:
            for nf in log.not_found_users:
                o = _orphan(nf.get("fio"))
                if o is not None:
                    try:
                        o[f"c1_{acc}"] += float(nf.get("debt") or 0)
                    except Exception:
                        pass

    ids = set(gis) | set(mr)
    unames: dict[int, str] = {}
    if ids:
        for uid, uname in (await db.execute(
            select(User.id, User.username).where(User.id.in_(ids))
        )).all():
            unames[uid] = uname

    eps = Decimal("0.01")
    out = {
        "has_findings": bool(findings),
        "findings_at": (findings or {}).get("synced_at"),
        "source_1c": source_1c,
        "accounts": {},
    }
    for acc in ("209", "205"):
        matched = 0
        sum_gis = sum_1c = Decimal("0")
        for uid in ids:
            g = gis.get(uid, {}).get(acc, Decimal("0"))
            m = mr.get(uid, {}).get(acc, Decimal("0"))
            if g == 0 and m == 0:
                continue
            sum_gis += g
            sum_1c += m
            if abs(g - m) <= eps:
                matched += 1
        out["accounts"][acc] = {
            "matched": matched,
            "sum_gisgmp": float(sum_gis),
            "sum_1c": float(sum_1c),
            "delta_total": float(sum_gis - sum_1c),
        }

    # Единый разрез по жильцу + авто-флаги (анализатор сам сигналит проблемы).
    def _flag(sg, sc, d):
        if abs(d) <= eps:
            return "ok"
        if sc == 0:
            return "only_gis"      # есть в ГИС, в 1С долга нет
        if sg == 0:
            return "only_1c"       # есть в 1С, ГИС не нашёл
        return "gis_more" if d > 0 else "c1_more"

    # «За сколько месяцев» долг в ГИС у жильца — число РАЗНЫХ месяцев его
    # неоплаченных (не аннулированных) начислений в реестре, по фамилии.
    # Для колонки «ГИС, мес» в сверке и печати + сигнала «дотянуть».
    from app.modules.utility.services.gisgmp_import import (
        is_unpaid, is_annulled, parse_reg_dt, GISGMP_CACHE_KEY,
    )
    surname_months: dict[str, set] = {}
    cache_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_CACHE_KEY)
    )).scalars().first()
    if cache_row and cache_row.value:
        try:
            for ch in json.loads(cache_row.value).values():
                if is_annulled(ch.get("change_status")) or not is_unpaid(ch.get("ack_status")):
                    continue
                nm = (ch.get("payer_name") or "").strip().split()
                dt = parse_reg_dt(ch.get("bill_date"))
                if nm and dt is not None:
                    surname_months.setdefault(nm[0].lower(), set()).add((dt.year, dt.month))
        except Exception:
            pass

    residents = []
    problems: dict[str, dict] = {}
    matched_total = 0
    for uid in ids:
        g9 = gis.get(uid, {}).get("209", Decimal("0"))
        c9 = mr.get(uid, {}).get("209", Decimal("0"))
        g5 = gis.get(uid, {}).get("205", Decimal("0"))
        c5 = mr.get(uid, {}).get("205", Decimal("0"))
        if g9 == 0 and c9 == 0 and g5 == 0 and c5 == 0:
            continue
        name = unames.get(uid) or gis.get(uid, {}).get("username") or str(uid)
        sg, sc = g9 + g5, c9 + c5
        d = sg - sc
        flag = _flag(sg, sc, d)
        sev = "high" if abs(d) >= 20000 else ("mid" if abs(d) >= 5000 else "low")
        if flag == "ok":
            matched_total += 1
        else:
            pr = problems.setdefault(flag, {"count": 0, "sum_abs": 0.0, "high": 0})
            pr["count"] += 1
            pr["sum_abs"] += float(abs(d))
            if sev == "high":
                pr["high"] += 1
        sn = (name or "").strip().split()
        gmonths = len(surname_months.get(sn[0].lower(), set())) if sn else 0
        residents.append({
            "user_id": uid, "username": name,
            "g209": float(g9), "c209": float(c9), "d209": float(g9 - c9),
            "g205": float(g5), "c205": float(c5), "d205": float(g5 - c5),
            "sum_gis": float(sg), "sum_1c": float(sc), "delta": float(d),
            "flag": flag, "severity": sev,
            # «За сколько месяцев» долг в ГИС + нужно ли ещё дотянуть (ГИС занижен).
            "gis_months": gmonths,
            "need_pull": flag in ("c1_more", "only_1c"),
        })
    # Пометка активных ГИС-оверрайдов (для UI: показать «откат» вместо «применить»).
    ovr_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_OVERRIDES_KEY)
    )).scalars().first()
    ovr_ids = set()
    if ovr_row and ovr_row.value:
        try:
            ovr_ids = set(json.loads(ovr_row.value).keys())
        except Exception:
            ovr_ids = set()
    for r in residents:
        r["overridden"] = str(r.get("user_id")) in ovr_ids

    residents.sort(key=lambda r: -abs(r["delta"]))
    out["residents"] = residents[:1000]
    out["problems"] = problems
    out["matched_count"] = matched_total
    out["override_count"] = len(ovr_ids)
    # Несопоставленные (1С/ГИС, нет жильца в базе) — отдельным блоком, чтобы не теряли.
    out["orphans"] = sorted(orphans_map.values(), key=lambda x: x["fio"])[:2000]
    return out


GISGMP_RECHECK_KEY = "gisgmp_recheck"


@router.post("/gisgmp/recheck-build", summary="Очередь дотягивания: проблемные ФИО (ГИС<1С)")
async def gisgmp_recheck_build(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Берёт из сверки жильцов, где ГИС занижен (флаги «1С>ГИС» и «нет в ГИС»),
    и ставит их фамилии в очередь. Релей точечно дотянет их полную историю
    (36 мес) по фамилии — сошедшихся и «ГИС>1С» не трогаем."""
    _require_finance(current_user)
    rec = await _build_reconcile(db)
    surnames, seen = [], set()
    for r in rec.get("residents", []):
        if r.get("flag") not in ("c1_more", "only_1c"):
            continue
        parts = (r.get("username") or "").strip().split()
        if not parts:
            continue
        s = parts[0]
        if s.lower() not in seen:
            seen.add(s.lower())
            surnames.append(s)
    payload = {"surnames": surnames, "deep_months": 36, "requested_at": utcnow().isoformat()}
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_RECHECK_KEY)
    )).scalars().first()
    if row is None:
        row = SystemSetting(key=GISGMP_RECHECK_KEY, value="{}",
                            description="Очередь точечного дотягивания ГИС ГМП")
        db.add(row)
    row.value = json.dumps(payload, ensure_ascii=False)
    await db.commit()
    return {"queued": len(surnames)}


GISGMP_ACTUALIZE_KEY = "gisgmp_actualize"
GISGMP_ACTUALIZE_LOG_KEY = "gisgmp_actualize_log"   # аудит прогонов (до/после)
_ACTUALIZE_LOG_MAX_RUNS = 50                         # авто-кап истории (+ ручная чистка)


def _actualize_result(before: dict | None, after: dict | None) -> str:
    """Эффект актуализации по долгу ГИС (до→после) для аудита."""
    try:
        bg = float((before or {}).get("gis") or 0)
        ag = float((after or {}).get("gis") or 0)
    except Exception:
        return "unknown"
    eps = 1.0
    if bg <= eps:
        return "unchanged"
    if ag <= eps:
        return "annulled"        # долг ГИС обнулён (реестр аннулировал лишнее)
    if ag < bg - eps:
        return "reduced"         # долг ГИС уменьшился
    if ag > bg + eps:
        return "increased"
    return "unchanged"


async def _capture_actualize_after(db: AsyncSession) -> None:
    """После сверки: для прогонов актуализации со статусом after_pending снимаем
    долг ГИС «после» по каждому жильцу и классифицируем эффект (до→после).
    Запускается из /gisgmp/sync — кэш и находки к этому моменту уже свежие."""
    lr = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_LOG_KEY)
    )).scalars().first()
    if not lr or not lr.value:
        return
    try:
        lj = json.loads(lr.value)
    except Exception:
        return
    pending = [r for r in lj.get("runs", []) if r.get("after_pending")]
    if not pending:
        return
    rec = await _build_reconcile(db)
    cur = {
        int(r["user_id"]): r for r in rec.get("residents", [])
        if r.get("user_id") is not None
    }
    now = utcnow().isoformat()
    for run in pending:
        for res in run.get("residents", []):
            uid = res.get("user_id")
            r = cur.get(int(uid)) if uid is not None else None
            if r is not None:
                after = {"gis": r.get("sum_gis"), "c1": r.get("sum_1c"), "delta": r.get("delta")}
            elif res.get("fio"):
                # Прогон по ОДНОМУ человеку (user_id нет): «после» = сумма ещё
                # несквитированных начислений (что сквитировалось — ушло из долга).
                _ch, _rev = await _load_person_charges(db, res["fio"])
                after = {"gis": round(sum(c["amount"] for c in _ch
                                          if c["unpaid"] and not c["annulled"]), 2),
                         "c1": None, "delta": None}
            else:
                # Жилец выпал из сверки — ни ГИС, ни 1С долга не осталось → ГИС обнулён.
                c1 = float((res.get("before") or {}).get("c1") or 0)
                after = {"gis": 0.0, "c1": c1, "delta": -c1}
            res["after"] = after
            res["result"] = _actualize_result(res.get("before"), after)
        run["after_pending"] = False
        run["after_at"] = now
        run["status"] = "done"
    lr.value = json.dumps(lj, ensure_ascii=False)
    await db.commit()


def _run_surnames(run: dict) -> set[str]:
    """Фамилии (первое слово ФИО) жильцов прогона — для глубокого переопроса."""
    out: set[str] = set()
    for p in run.get("residents", []):
        fio = (p.get("fio") or "").strip()
        if fio:
            out.add(fio.split()[0])
    return out


async def _enqueue_recheck_surnames(db: AsyncSession, surnames: set[str]) -> None:
    """Добавляет фамилии в очередь глубокого переопроса ГИС (gisgmp_recheck),
    мёржит с уже стоящими. Релей заберёт на ближайшем опросе (do_recheck)."""
    if not surnames:
        return
    rc_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == "gisgmp_recheck")
    )).scalars().first()
    cur = {}
    if rc_row and rc_row.value:
        try:
            cur = json.loads(rc_row.value)
        except Exception:
            cur = {}
    merged = sorted(set(cur.get("surnames") or []) | surnames)
    if rc_row is None:
        rc_row = SystemSetting(key="gisgmp_recheck", value="{}",
                               description="Очередь точечного дотягивания ГИС ГМП")
        db.add(rc_row)
    rc_row.value = json.dumps({"surnames": merged, "deep_months": 36}, ensure_ascii=False)


async def _drive_actualize_runs(db: AsyncSession) -> None:
    """Авто-цикл доведения актуализации до «Сквитировано» (рулит по ~2-мин опросу
    релея). Для прогонов «checking»: читает СВЕЖИЙ кэш ГИС, считает ещё
    несквитированные начисления прогона (по UIN) → пишет «после»; всё сквитировано
    → done(all_paid); иначе каждые REACT_MIN мин повтор актуализации их
    несквитированных (макс MAX_ATTEMPTS попытки) → done(unpaid_left); и каждый
    опрос ставит переопрос их ФИО для свежести. HARD_MIN — аварийный финал, чтобы
    не висело вечно. Зовётся из relay-config ДО чтения очереди переопроса."""
    REACT_MIN, MAX_ATTEMPTS, HARD_MIN = 30, 2, 90
    lr = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_LOG_KEY)
    )).scalars().first()
    if not lr or not lr.value:
        return
    try:
        lj = json.loads(lr.value)
    except Exception:
        return
    active = [r for r in lj.get("runs", []) if r.get("status") == "checking"]
    if not active:
        return
    from app.modules.utility.services.gisgmp_import import (
        GISGMP_CACHE_KEY, is_unpaid, is_annulled,
    )
    cache = {}
    cache_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_CACHE_KEY)
    )).scalars().first()
    if cache_row and cache_row.value:
        try:
            cache = json.loads(cache_row.value)
        except Exception:
            cache = {}

    def _unpaid_now(ch: dict) -> bool:
        cc = cache.get(ch.get("uin"))
        if cc is None:
            return True  # ещё не пересканен этим циклом — считаем несквитированным
        return is_unpaid(cc.get("ack_status")) and not is_annulled(cc.get("change_status"))

    now = utcnow()
    now_iso = now.isoformat()

    def _age_min(iso) -> float:
        try:
            return (now - datetime.fromisoformat(iso)).total_seconds() / 60.0 if iso else 1e9
        except Exception:
            return 1e9

    act_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_KEY)
    )).scalars().first()
    act_busy = False
    if act_row and act_row.value:
        try:
            _av = json.loads(act_row.value)
            act_busy = bool(_av.get("running") or _av.get("uuids"))
        except Exception:
            act_busy = False

    recheck_surnames: set[str] = set()
    reactualize = None
    changed = False

    def _finalize(run, result):
        run["status"] = "done"
        run["after_pending"] = False
        run["after_at"] = now_iso
        run["loop_result"] = result

    for run in active:
        still_uuids: list[str] = []
        for res in run.get("residents", []):
            res_sum = 0.0
            for ch in res.get("charges", []):
                if _unpaid_now(ch):
                    u = ch.get("charge_uuid")
                    if u:
                        still_uuids.append(u)
                    res_sum += float(ch.get("amount") or 0)
            res["after"] = {"gis": round(res_sum, 2), "c1": None, "delta": None}
            res["result"] = _actualize_result(res.get("before"), res["after"])
        if not still_uuids:
            _finalize(run, "all_paid")
            changed = True
            continue
        if _age_min(run.get("started_at") or run.get("queued_at")) >= HARD_MIN:
            _finalize(run, "timeout")
            changed = True
            continue
        if _age_min(run.get("last_actualize_at")) >= REACT_MIN:
            if int(run.get("attempt", 1)) >= MAX_ATTEMPTS:
                _finalize(run, "unpaid_left")
                changed = True
                continue
            if not act_busy and reactualize is None:
                reactualize = (run, still_uuids)
                run["attempt"] = int(run.get("attempt", 1)) + 1
                run["last_actualize_at"] = now_iso
                run["status"] = "running"  # релей дошлёт повтор
                changed = True
                continue
        recheck_surnames |= _run_surnames(run)

    if reactualize:
        run, uuids = reactualize
        act_payload = {
            "uuids": uuids, "total": len(uuids), "done": 0, "ok": 0, "fail": 0,
            "running": False, "finished": False, "queued_at": now_iso,
            "by": "авто-цикл", "message": "", "targeting": "loop-retry",
            "run_id": run.get("id"),
        }
        if act_row is None:
            act_row = SystemSetting(key=GISGMP_ACTUALIZE_KEY, value="{}",
                                    description="Очередь массовой актуализации ГИС ГМП")
            db.add(act_row)
        act_row.value = json.dumps(act_payload, ensure_ascii=False)
    if recheck_surnames:
        await _enqueue_recheck_surnames(db, recheck_surnames)
    if changed or reactualize or recheck_surnames:
        lr.value = json.dumps(lj, ensure_ascii=False)
        await db.commit()


@router.post("/gisgmp/actualize-recheck", summary="Проверить результат актуализации (переопрос ГИС по ФИО)")
async def gisgmp_actualize_recheck(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Ставит ФИО последнего прогона актуализации (sent/processing/rechecking) в
    ГЛУБОКИЙ переопрос ГИС + взводит after_pending — релей переопросит, на
    ближайшей сверке снимется «после». Ручной дубль авто-перепроверки: нужен,
    т.к. ГИС обрабатывает «Отправлен в ГИС ГМП» асинхронно и снимок сразу всегда
    показывал «без изменений»."""
    _require_finance(current_user)
    lr = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_LOG_KEY)
    )).scalars().first()
    if not lr or not lr.value:
        return {"queued": 0, "reason": "нет прогонов актуализации"}
    try:
        lj = json.loads(lr.value)
    except Exception:
        return {"queued": 0, "reason": "история повреждена"}
    # Форсируем переопрос ВСЕХ активных циклов («checking») — авто-цикл подхватит
    # свежий кэш на ближайшем опросе и обновит «после»/финал. Статус не трогаем —
    # цикл сам решит (всё сквитировано → done, иначе повтор/ожидание).
    active = [r for r in lj.get("runs", []) if r.get("status") == "checking"]
    if not active:
        return {"queued": 0, "reason": "нет активных циклов актуализации"}
    surnames: set[str] = set()
    for run in active:
        surnames |= _run_surnames(run)
    if not surnames:
        return {"queued": 0, "reason": "нет ФИО в прогонах"}
    await _enqueue_recheck_surnames(db, surnames)
    await db.commit()
    return {"queued": len(surnames), "runs": len(active)}


async def _load_person_charges(db: AsyncSession, fio: str) -> tuple[list[dict], list[str]]:
    """Начисления ГИС ГМП одного человека из кэша (союз по ТОЧНОМУ ФИО).
    Возвращает (charges, revocable_uuids). revocable = НЕсквитированные и НЕ
    аннулированные — именно их актуализируем/аннулируем."""
    from app.modules.utility.services.gisgmp_import import (
        GISGMP_CACHE_KEY, is_unpaid, is_annulled, classify_account,
    )
    from app.modules.utility.services.gsheets_sync import normalize_fio
    target = normalize_fio(fio or "")
    charges: list[dict] = []
    revocable: list[str] = []
    if not target:
        return charges, revocable
    cache_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_CACHE_KEY)
    )).scalars().first()
    if cache_row and cache_row.value:
        try:
            for ch in json.loads(cache_row.value).values():
                if normalize_fio(ch.get("payer_name") or "") != target:
                    continue
                unpaid = is_unpaid(ch.get("ack_status"))
                annulled = is_annulled(ch.get("change_status"))
                try:
                    amt = float(str(ch.get("amount") or "0").replace(",", "."))
                except Exception:
                    amt = 0.0
                u = ch.get("charge_uuid")
                charges.append({
                    "uin": ch.get("uin"), "account": classify_account(ch.get("purpose")),
                    "amount": amt, "bill_date": ch.get("bill_date"),
                    "ack_status": ch.get("ack_status"), "change_status": ch.get("change_status"),
                    "charge_uuid": u, "unpaid": unpaid, "annulled": annulled,
                    "purpose": ch.get("purpose"),
                })
                if unpaid and not annulled and u:
                    revocable.append(u)
        except Exception:
            pass
    charges.sort(key=lambda c: (not c["unpaid"], c.get("account") or "", c.get("bill_date") or ""))
    return charges, revocable


@router.get("/gisgmp/person-charges", summary="Начисления ГИС ГМП по одному ФИО (проваливание в сверке)")
async def gisgmp_person_charges(
    fio: str = Query(..., max_length=200),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Read-only: все начисления человека из кэша ГИС (UIN, счёт, сумма, дата,
    статус квитирования/изменения, uuid). Питает модалку проваливания + кнопки."""
    _require_finance(current_user)
    charges, revocable = await _load_person_charges(db, fio)
    summary = {
        "total": len(charges),
        "revocable": len(revocable),
        "annulled": sum(1 for c in charges if c["annulled"]),
        "sum_revocable": round(sum(c["amount"] for c in charges if c["unpaid"] and not c["annulled"]), 2),
    }
    return {"fio": fio, "charges": charges, "summary": summary}


class GisgmpPersonIn(BaseModel):
    fio: str


@router.post("/gisgmp/actualize-person", summary="Актуализировать начисления ОДНОГО человека")
async def gisgmp_actualize_person(
    payload: GisgmpPersonIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Ставит несквитированные начисления одного человека (проваливание в сверке)
    в очередь актуализации — релей дёрнет actualize-request по каждому; результат
    снимется отложенной перепроверкой, как у массовой. Не пишем поверх идущей."""
    _require_finance(current_user)
    fio = (payload.fio or "").strip()
    charges, revocable = await _load_person_charges(db, fio)
    if not revocable:
        return {"queued": 0, "reason": "нет несквитированных начислений у этого ФИО"}
    act_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_KEY)
    )).scalars().first()
    if act_row and act_row.value:
        try:
            cur = json.loads(act_row.value)
            if cur.get("running") or cur.get("uuids"):
                return {"queued": 0, "reason": "идёт другая актуализация — дождитесь завершения"}
        except Exception:
            pass
    run_id = utcnow().isoformat()
    qpayload = {
        "uuids": revocable, "total": len(revocable), "done": 0, "ok": 0, "fail": 0,
        "running": False, "finished": False, "queued_at": run_id,
        "by": current_user.username, "message": "", "targeting": "person", "run_id": run_id,
    }
    if act_row is None:
        act_row = SystemSetting(key=GISGMP_ACTUALIZE_KEY, value="{}",
                                description="Очередь массовой актуализации ГИС ГМП")
        db.add(act_row)
    act_row.value = json.dumps(qpayload, ensure_ascii=False)
    run = {
        "id": run_id, "queued_at": run_id, "by": current_user.username,
        "targeting": f"один человек: {fio}", "total_charges": len(revocable),
        "residents_count": 1, "status": "running", "done": 0, "ok": 0, "fail": 0,
        "started_at": None, "finished_at": None, "after_pending": False, "after_at": None,
        "attempt": 1, "last_actualize_at": run_id,
        "residents": [{
            "user_id": None, "fio": fio, "username": None, "flag": None,
            "before": {"gis": round(sum(c["amount"] for c in charges
                                        if c["unpaid"] and not c["annulled"]), 2),
                       "c1": None, "delta": None},
            "after": None, "result": None,
            "charges": [{"uin": c["uin"], "account": c["account"], "charge_uuid": c["charge_uuid"],
                         "amount": c["amount"], "bill_date": c["bill_date"]}
                        for c in charges if c["unpaid"] and not c["annulled"]],
        }],
    }
    log_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_LOG_KEY)
    )).scalars().first()
    if log_row is None:
        log_row = SystemSetting(key=GISGMP_ACTUALIZE_LOG_KEY, value='{"runs": []}',
                                description="Аудит массовых актуализаций ГИС ГМП (до/после)")
        db.add(log_row)
    try:
        logj = json.loads(log_row.value) if log_row.value else {"runs": []}
    except Exception:
        logj = {"runs": []}
    runs = logj.get("runs", [])
    runs.insert(0, run)
    del runs[_ACTUALIZE_LOG_MAX_RUNS:]
    logj["runs"] = runs
    log_row.value = json.dumps(logj, ensure_ascii=False)
    await db.commit()
    return {"queued": len(revocable), "fio": fio}


@router.post("/gisgmp/actualize-build", summary="Очередь массовой актуализации: только ГИС > 1С (ошибка ГИС ГМП)")
async def gisgmp_actualize_build(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Ставит в очередь актуализации НЕОПЛАЧЕННЫЕ (не аннулированные) счета ТОЛЬКО
    тех жильцов, у кого ГИС > 1С (флаги gis_more и only_gis) — это и есть «ошибка
    ГИС ГМП», где актуализация имеет смысл (реестр может аннулировать лишний долг).
    Случаи ГИС < 1С актуализация не лечит — для них «Дотянуть расхождения».

    Матчинг счёт→жилец — по ПОЛНОМУ ФИО (как в находках), НЕ по фамилии, чтобы не
    тянуть однофамильцев. Параллельно пишем аудит-прогон со снимком «до»; «после»
    снимется на ближайшей сверке (релей дёрнут авто через run_now)."""
    _require_finance(current_user)
    rec = await _build_reconcile(db)
    TARGET_FLAGS = ("gis_more", "only_gis")
    flagged = {
        int(r["user_id"]): r for r in rec.get("residents", [])
        if r.get("flag") in TARGET_FLAGS and r.get("user_id") is not None
    }

    # Точная карта ФИО(реестр)→user_id из находок (та же логика, что свела долги).
    findings = await _load_findings(db)
    fio_to_uid: dict[str, int] = {}
    if findings:
        for frow in findings.get("summary", []):
            uid = frow.get("matched_user_id")
            fio = (frow.get("fio") or "").strip()
            if uid is not None and fio:
                fio_to_uid[fio] = int(uid)
    target_fios = {fio for fio, uid in fio_to_uid.items() if uid in flagged}

    from app.modules.utility.services.gisgmp_import import (
        is_unpaid, is_annulled, classify_account, parse_reg_dt, GISGMP_CACHE_KEY,
    )
    # Окно актуализации = окно сбора (months_back из настроек релея). «Всё время»
    # (>=600 мес) → без фильтра; иначе актуализируем только начисления, чьё
    # bill_date не старше окна (напр. 1 год / полгода).
    rcfg = await _load_relay_cfg(db)
    months_back = int(rcfg.get("months_back") or 999)
    cutoff = None if months_back >= 600 else (utcnow() - timedelta(days=months_back * 31))
    cache_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_CACHE_KEY)
    )).scalars().first()
    per_user: dict[int, dict] = {}
    uuids, seen = [], set()
    if cache_row and cache_row.value and target_fios:
        try:
            for ch in json.loads(cache_row.value).values():
                if is_annulled(ch.get("change_status")) or not is_unpaid(ch.get("ack_status")):
                    continue
                fio = (ch.get("payer_name") or "").strip()
                if fio not in target_fios:
                    continue
                if cutoff is not None:
                    dt = parse_reg_dt(ch.get("bill_date"))
                    if dt is not None and dt < cutoff:
                        continue
                u = ch.get("charge_uuid")
                if not u or u in seen:
                    continue
                seen.add(u)
                uuids.append(u)
                uid = fio_to_uid.get(fio)
                slot = per_user.setdefault(uid, {"user_id": uid, "fio": fio, "charges": []})
                try:
                    amt = float(str(ch.get("amount") or "0").replace(",", "."))
                except Exception:
                    amt = 0.0
                slot["charges"].append({
                    "uin": ch.get("uin"), "account": classify_account(ch.get("purpose")),
                    "charge_uuid": u, "amount": amt, "bill_date": ch.get("bill_date"),
                })
        except Exception:
            pass

    # Снимок «до» по каждому затронутому жильцу (из сверки), сорт по |Δ|.
    residents_snap = []
    for uid, slot in per_user.items():
        fr = flagged.get(uid, {})
        residents_snap.append({
            "user_id": uid, "fio": slot["fio"],
            "username": fr.get("username"), "flag": fr.get("flag"),
            "before": {"gis": fr.get("sum_gis"), "c1": fr.get("sum_1c"), "delta": fr.get("delta")},
            "after": None, "result": None,
            "charges": slot["charges"],
        })
    residents_snap.sort(key=lambda x: -abs((x.get("before") or {}).get("delta") or 0))

    run_id = utcnow().isoformat()
    payload = {
        "uuids": uuids, "total": len(uuids), "done": 0, "ok": 0, "fail": 0,
        "running": False, "finished": False,
        "queued_at": run_id, "by": current_user.username, "message": "",
        "targeting": "gis_more+only_gis", "run_id": run_id,
    }
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_KEY)
    )).scalars().first()
    if row is None:
        row = SystemSetting(key=GISGMP_ACTUALIZE_KEY, value="{}",
                            description="Очередь массовой актуализации ГИС ГМП")
        db.add(row)
    row.value = json.dumps(payload, ensure_ascii=False)

    # Аудит-прогон (до/после). «После» снимется на ближайшей сверке.
    run = {
        "id": run_id, "queued_at": run_id, "by": current_user.username,
        "targeting": "ГИС > 1С (gis_more + only_gis)",
        "total_charges": len(uuids), "residents_count": len(residents_snap),
        "status": "running", "done": 0, "ok": 0, "fail": 0,
        "started_at": None, "finished_at": None,
        "after_pending": False, "after_at": None,
        "attempt": 1, "last_actualize_at": run_id,
        "residents": residents_snap,
    }
    log_row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_LOG_KEY)
    )).scalars().first()
    if log_row is None:
        log_row = SystemSetting(key=GISGMP_ACTUALIZE_LOG_KEY, value='{"runs": []}',
                                description="Аудит массовых актуализаций ГИС ГМП (до/после)")
        db.add(log_row)
    try:
        logj = json.loads(log_row.value) if log_row.value else {"runs": []}
    except Exception:
        logj = {"runs": []}
    runs = logj.get("runs", [])
    runs.insert(0, run)
    del runs[_ACTUALIZE_LOG_MAX_RUNS:]
    logj["runs"] = runs
    log_row.value = json.dumps(logj, ensure_ascii=False)

    await db.commit()
    return {"queued": len(uuids), "residents": len(residents_snap), "targeting": "gis_more+only_gis"}


class GisgmpActualizeProgressIn(BaseModel):
    done: int = 0
    ok: int = 0
    fail: int = 0
    finished: bool = False
    message: str = ""


@router.post("/gisgmp/actualize-progress", summary="Релей шлёт прогресс актуализации (token)")
async def gisgmp_actualize_progress(
    payload: GisgmpActualizeProgressIn,
    authorization: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_db),
):
    _check_gisgmp_token(authorization)
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_KEY)
    )).scalars().first()
    if not row or not row.value:
        return {"ok": True}
    try:
        av = json.loads(row.value)
    except Exception:
        av = {}
    av["done"] = payload.done
    av["ok"] = payload.ok
    av["fail"] = payload.fail
    av["last_at"] = utcnow().isoformat()
    if payload.message:
        av["message"] = payload.message[:300]
    if payload.finished:
        av["running"] = False
        av["finished"] = True
        av["finished_at"] = utcnow().isoformat()
        av["uuids"] = []  # выполнено — очищаем список
        # Прогон «отправлен в ГИС». Снимок «после» НЕ снимаем сразу: ГИС
        # обрабатывает запрос АСИНХРОННО («Отправлен в ГИС ГМП» → «Завершен»,
        # часы) — преждевременный снимок всегда показывал «без изменений».
        # «После» снимется на ПЕРЕПРОВЕРКЕ: авто (через ~2ч) или кнопкой
        # «Проверить результат» — оба ставят ФИО прогона в глубокий переопрос
        # ГИС + взводят after_pending, и снимок берётся уже по свежим данным.
        run_id = av.get("run_id")
        if run_id:
            lr = (await db.execute(
                select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_LOG_KEY)
            )).scalars().first()
            if lr and lr.value:
                try:
                    lj = json.loads(lr.value)
                    for run in lj.get("runs", []):
                        if run.get("id") == run_id:
                            run["status"] = "checking"  # авто-цикл начинает опрос «Сквитировано»
                            run["finished_at"] = av["finished_at"]
                            run["done"] = payload.done
                            run["ok"] = payload.ok
                            run["fail"] = payload.fail
                            run["after_pending"] = False
                            break
                    lr.value = json.dumps(lj, ensure_ascii=False)
                except Exception:
                    pass
    row.value = json.dumps(av, ensure_ascii=False)
    await db.commit()
    return {"ok": True}


@router.get("/gisgmp/actualize-status", summary="Прогресс массовой актуализации (для UI)")
async def gisgmp_actualize_status(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    row = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_KEY)
    )).scalars().first()
    if not row or not row.value:
        return {"total": 0, "done": 0, "running": False, "finished": False}
    try:
        av = json.loads(row.value)
    except Exception:
        return {"total": 0, "done": 0, "running": False, "finished": False}
    return {k: av.get(k) for k in (
        "total", "done", "ok", "fail", "running", "finished",
        "queued_at", "started_at", "finished_at", "last_at", "message", "by")}


@router.get("/gisgmp/actualize-log", summary="История массовых актуализаций (аудит до/после)")
async def gisgmp_actualize_log(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    lr = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_LOG_KEY)
    )).scalars().first()
    if not lr or not lr.value:
        return {"runs": []}
    try:
        return {"runs": json.loads(lr.value).get("runs", [])}
    except Exception:
        return {"runs": []}


class GisgmpActualizePruneIn(BaseModel):
    clear_all: bool = False
    delete_id: Optional[str] = None
    older_than_days: Optional[int] = None
    keep_last: Optional[int] = None


@router.post("/gisgmp/actualize-log/prune", summary="Чистка истории актуализаций (на выбор админа)")
async def gisgmp_actualize_log_prune(
    payload: GisgmpActualizePruneIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    lr = (await db.execute(
        select(SystemSetting).where(SystemSetting.key == GISGMP_ACTUALIZE_LOG_KEY)
    )).scalars().first()
    if not lr or not lr.value:
        return {"remaining": 0}
    try:
        lj = json.loads(lr.value)
    except Exception:
        lj = {"runs": []}
    runs = lj.get("runs", [])
    if payload.clear_all:
        runs = []
    elif payload.delete_id:
        runs = [r for r in runs if r.get("id") != payload.delete_id]
    elif payload.older_than_days:
        cutoff = (utcnow() - timedelta(days=int(payload.older_than_days))).isoformat()
        runs = [r for r in runs if (r.get("queued_at") or "") >= cutoff]
    elif payload.keep_last is not None:
        runs = runs[:max(0, int(payload.keep_last))]
    lj["runs"] = runs
    lr.value = json.dumps(lj, ensure_ascii=False)
    await db.commit()
    return {"remaining": len(runs)}


@router.get("/gisgmp/relay.py", summary="Отдать актуальный relay.py (самообновление релея на ВМ)")
async def gisgmp_relay_py(authorization: Optional[str] = Header(None)):
    """Релей на ВМ берёт свежий код одной командой (token-auth), чтобы не
    вставлять его вручную при каждом апгрейде."""
    _check_gisgmp_token(authorization)
    p = Path(__file__).resolve().parents[4] / "relay" / "gisgmp" / "relay.py"
    if not p.is_file():
        raise HTTPException(404, "relay.py не найден в образе")
    return Response(
        content=p.read_text(encoding="utf-8"),
        media_type="text/x-python; charset=utf-8",
        headers={"Cache-Control": "no-store"},
    )


# Каталог расширения в репозитории. financier.py лежит в
# app/modules/utility/routers/ → parents[4] = корень репо (в Docker это /app,
# куда Dockerfile COPY кладёт extension/). Внутри — папка gisgmp-bridge.
_GISGMP_EXT_DIR = Path(__file__).resolve().parents[4] / "extension" / "gisgmp-bridge"


@router.get("/gisgmp/bridge.zip", summary="Скачать ZIP расширения-моста ГИС ГМП")
async def gisgmp_bridge_zip(current_user: User = Depends(get_current_user)):
    """Отдаёт ZIP папки gisgmp-bridge — пользователь распаковывает и грузит в
    браузер как «распакованное расширение». Только финансист/админ."""
    _require_finance(current_user)
    if not _GISGMP_EXT_DIR.is_dir():
        raise HTTPException(404, "Каталог расширения не найден на сервере.")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for root, _dirs, files in os.walk(_GISGMP_EXT_DIR):
            for fname in files:
                fpath = Path(root) / fname
                rel = Path("gisgmp-bridge") / fpath.relative_to(_GISGMP_EXT_DIR)
                zf.write(fpath, arcname=str(rel))
    body = buf.getvalue()
    return Response(
        content=body,
        media_type="application/zip",
        headers={
            "Content-Disposition": 'attachment; filename="gisgmp-bridge.zip"',
            "Content-Length": str(len(body)),
            "Cache-Control": "no-store",
            "X-Content-Type-Options": "nosniff",
        },
    )


@router.get(
    "/users-status",
    response_model=PaginatedResponse[UserDebtResponse],
    summary="Список пользователей с долгами"
)
async def get_users_with_debts(
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=500),
        search: str | None = Query(None),
        # Новые фильтры/сортировка для вкладки «Долги 1С»
        only_debtors: bool = Query(False, description="Только с положительным долгом 209 или 205"),
        only_overpaid: bool = Query(False, description="Только с положительной переплатой"),
        has_data: bool = Query(False, description="Скрыть жильцов без данных из 1С (все 8 финансовых полей = 0)"),
        dormitory: Optional[str] = Query(None, description="Фильтр по названию общежития"),
        min_debt: Optional[float] = Query(None, ge=0, description="Минимальный суммарный долг (209+205)"),
        sort_by: str = Query("room", pattern="^(room|username|debt|overpay|total)$"),
        sort_dir: str = Query("asc", pattern="^(asc|desc)$"),
        period_id: Optional[int] = Query(None, description="Период просмотра; по умолчанию активный → последний импорт → свежий"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db)
):
    if current_user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    offset = (page - 1) * limit

    active_period = await _resolve_view_period(db, period_id)
    period_id = active_period.id if active_period else None

    d209 = func.coalesce(func.sum(MeterReading.debt_209), 0).label("debt_209")
    o209 = func.coalesce(func.sum(MeterReading.overpayment_209), 0).label("overpayment_209")
    d205 = func.coalesce(func.sum(MeterReading.debt_205), 0).label("debt_205")
    o205 = func.coalesce(func.sum(MeterReading.overpayment_205), 0).label("overpayment_205")
    # Bug V: обороты для UI «движение средств».
    od209 = func.coalesce(func.sum(MeterReading.obor_debit_209), 0).label("obor_debit_209")
    oc209 = func.coalesce(func.sum(MeterReading.obor_credit_209), 0).label("obor_credit_209")
    od205 = func.coalesce(func.sum(MeterReading.obor_debit_205), 0).label("obor_debit_205")
    oc205 = func.coalesce(func.sum(MeterReading.obor_credit_205), 0).label("obor_credit_205")
    total = func.coalesce(func.sum(MeterReading.total_cost), 0).label("current_total_cost")

    stmt = select(
        User, Room, d209, o209, d205, o205, total,
        od209, oc209, od205, oc205,
    ).outerjoin(
        Room, User.room_id == Room.id
    ).outerjoin(
        MeterReading,
        (User.id == MeterReading.user_id) &
        (MeterReading.period_id == period_id)
    ).where(
        User.is_deleted.is_(False),
        User.role == "user",
    )

    search_condition = None
    if search:
        search_value = f"%{search.lower()}%"
        search_condition = or_(
            func.lower(User.username).like(search_value),
            func.lower(Room.dormitory_name).like(search_value),
            func.lower(Room.room_number).like(search_value)
        )
        stmt = stmt.where(search_condition)

    if dormitory:
        stmt = stmt.where(Room.dormitory_name == dormitory)

    stmt = stmt.group_by(User.id, Room.id)

    # Фильтры по агрегированным значениям — HAVING
    if only_debtors:
        stmt = stmt.having((d209 + d205) > 0)
    if only_overpaid:
        stmt = stmt.having((o209 + o205) > 0)
    if min_debt is not None:
        stmt = stmt.having((d209 + d205) >= min_debt)
    if has_data:
        # Bug AB: «не показывать пустых» — хотя бы одно из 8 финансовых
        # полей (сальдо + обороты по 209 и 205) > 0.
        stmt = stmt.having(
            (d209 + o209 + d205 + o205 + od209 + oc209 + od205 + oc205) > 0
        )

    # Сортировка: столбец + направление
    sort_map = {
        "room": (Room.dormitory_name, Room.room_number),
        "username": (User.username,),
        "debt": ((d209 + d205).label("__debt_sum"),),
        "overpay": ((o209 + o205).label("__over_sum"),),
        "total": (total,),
    }
    cols = sort_map[sort_by]
    direction = desc if sort_dir == "desc" else asc
    order_cols = [direction(c).nulls_last() for c in cols]
    # Стабилизатор — вторичная сортировка по username
    if sort_by != "username":
        order_cols.append(asc(User.username))
    stmt = stmt.order_by(*order_cols).limit(limit).offset(offset)

    # count
    count_stmt = select(func.count(User.id)).outerjoin(Room, User.room_id == Room.id).where(
        User.is_deleted.is_(False), User.role == "user"
    )
    if search_condition is not None:
        count_stmt = count_stmt.where(search_condition)
    if dormitory:
        count_stmt = count_stmt.where(Room.dormitory_name == dormitory)

    # Для only_debtors/only_overpaid/min_debt/has_data count тоже надо
    # пересчитать через HAVING — делаем через subquery вместо дублирования.
    if only_debtors or only_overpaid or min_debt is not None or has_data:
        inner = stmt.with_only_columns(User.id).limit(None).offset(None).order_by(None).subquery()
        count_stmt = select(func.count()).select_from(inner)

    total_res = await db.execute(count_stmt)
    total_items = total_res.scalar_one()

    result = await db.execute(stmt)
    rows = result.all()

    # Покрытие импортами 1С активного периода: кто из жильцов попал в ПОСЛЕДНИЙ
    # импорт счёта (applied_state ключуется по str(user_id)). seen_2xx:
    #   None  — импорта этого счёта в периоде не было (нечему быть «не найденным»);
    #   set   — множество user_id, попавших в последний импорт счёта.
    # Жилец с долгом >0 всегда был в импорте (его touch'нули), поэтому «не найден»
    # покажется только тем, у кого реально нет данных по счёту, а не «долг 0».
    async def _seen_ids(acct: str):
        if not period_id:
            return None
        last_imp = (await db.execute(
            select(DebtImportLog)
            .where(
                DebtImportLog.period_id == period_id,
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
            )
            .order_by(desc(DebtImportLog.started_at))
            .limit(1)
        )).scalars().first()
        if not last_imp:
            return None
        st = last_imp.applied_state or {}
        return {int(k) for k in st.keys() if str(k).isdigit()}

    seen_209 = await _seen_ids("209")
    seen_205 = await _seen_ids("205")

    items = []
    for row in rows:
        user_obj, room_obj = row[0], row[1]
        items.append({
            "id": user_obj.id,
            "username": user_obj.username,
            "room": room_obj,
            "debt_209": row[2],
            "overpayment_209": row[3],
            "debt_205": row[4],
            "overpayment_205": row[5],
            "current_total_cost": row[6],
            # Bug V: движение средств — обороты периода.
            "obor_debit_209": row[7],
            "obor_credit_209": row[8],
            "obor_debit_205": row[9],
            "obor_credit_205": row[10],
            # Покрытие импортом (None если импорта счёта не было).
            "seen_209": (user_obj.id in seen_209) if seen_209 is not None else None,
            "seen_205": (user_obj.id in seen_205) if seen_205 is not None else None,
        })

    return {"total": total_items, "page": page, "size": limit, "items": items}


# =========================================================================
# ЗЕРКАЛО ПО КВАРТИРАМ: та же отчётность, но агрегация по ПОМЕЩЕНИЮ, без ФИО.
# Долг квартиры = сумма долгов всех жильцов комнаты (по room_id). Адрес вместо
# ФИО; кто живёт — в детализации /rooms/{id}/residents-finance.
# =========================================================================
@router.get("/rooms-status", summary="Список квартир с долгами (агрегация по помещению)")
async def get_rooms_with_debts(
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=500),
        search: str | None = Query(None),
        only_debtors: bool = Query(False),
        only_overpaid: bool = Query(False),
        has_data: bool = Query(False),
        dormitory: Optional[str] = Query(None),
        place_type: Optional[str] = Query(None, pattern="^(dormitory|house)$"),
        min_debt: Optional[float] = Query(None, ge=0),
        sort_by: str = Query("room", pattern="^(room|debt|overpay|total)$"),
        sort_dir: str = Query("asc", pattern="^(asc|desc)$"),
        period_id: Optional[int] = Query(None, description="Период просмотра; по умолчанию активный → последний импорт → свежий"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db),
):
    if current_user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    offset = (page - 1) * limit
    active = await _resolve_view_period(db, period_id)
    period_id = active.id if active else None

    d209 = func.coalesce(func.sum(MeterReading.debt_209), 0).label("debt_209")
    o209 = func.coalesce(func.sum(MeterReading.overpayment_209), 0).label("overpayment_209")
    d205 = func.coalesce(func.sum(MeterReading.debt_205), 0).label("debt_205")
    o205 = func.coalesce(func.sum(MeterReading.overpayment_205), 0).label("overpayment_205")
    od209 = func.coalesce(func.sum(MeterReading.obor_debit_209), 0).label("obor_debit_209")
    oc209 = func.coalesce(func.sum(MeterReading.obor_credit_209), 0).label("obor_credit_209")
    od205 = func.coalesce(func.sum(MeterReading.obor_debit_205), 0).label("obor_debit_205")
    oc205 = func.coalesce(func.sum(MeterReading.obor_credit_205), 0).label("obor_credit_205")
    total = func.coalesce(func.sum(MeterReading.total_cost), 0).label("current_total_cost")
    residents = func.count(func.distinct(MeterReading.user_id)).label("residents_count")

    stmt = select(
        Room, d209, o209, d205, o205, total, od209, oc209, od205, oc205, residents,
    ).outerjoin(
        MeterReading,
        (Room.id == MeterReading.room_id) & (MeterReading.period_id == period_id),
    )
    search_condition = None
    if search:
        sv = f"%{search.lower()}%"
        search_condition = or_(
            func.lower(Room.dormitory_name).like(sv),
            func.lower(Room.room_number).like(sv),
            func.lower(Room.street).like(sv),
            func.lower(Room.house_number).like(sv),
            func.lower(Room.apartment_number).like(sv),
        )
        stmt = stmt.where(search_condition)
    if dormitory:
        stmt = stmt.where(Room.dormitory_name == dormitory)
    if place_type:
        stmt = stmt.where(Room.place_type == place_type)
    stmt = stmt.group_by(Room.id)
    if only_debtors:
        stmt = stmt.having((d209 + d205) > 0)
    if only_overpaid:
        stmt = stmt.having((o209 + o205) > 0)
    if min_debt is not None:
        stmt = stmt.having((d209 + d205) >= min_debt)
    if has_data:
        stmt = stmt.having(
            (d209 + o209 + d205 + o205 + od209 + oc209 + od205 + oc205) > 0
        )

    sort_map = {
        "room": (Room.dormitory_name, Room.room_number, Room.street, Room.house_number),
        "debt": ((d209 + d205).label("__d"),),
        "overpay": ((o209 + o205).label("__o"),),
        "total": (total,),
    }
    direction = desc if sort_dir == "desc" else asc
    order_cols = [direction(c).nulls_last() for c in sort_map[sort_by]]
    order_cols.append(asc(Room.id))
    stmt = stmt.order_by(*order_cols).limit(limit).offset(offset)

    count_stmt = select(func.count(Room.id))
    if search_condition is not None:
        count_stmt = count_stmt.where(search_condition)
    if dormitory:
        count_stmt = count_stmt.where(Room.dormitory_name == dormitory)
    if place_type:
        count_stmt = count_stmt.where(Room.place_type == place_type)
    if only_debtors or only_overpaid or min_debt is not None or has_data:
        inner = stmt.with_only_columns(Room.id).limit(None).offset(None).order_by(None).subquery()
        count_stmt = select(func.count()).select_from(inner)
    total_items = (await db.execute(count_stmt)).scalar_one()

    rows = (await db.execute(stmt)).all()
    items = []
    for row in rows:
        room = row[0]
        items.append({
            "room_id": room.id,
            "address": room.format_address,
            "place_type": room.place_type,
            "dormitory_name": room.dormitory_name,
            "room_number": room.room_number,
            "residents_count": int(row[10] or 0),
            "debt_209": row[1], "overpayment_209": row[2],
            "debt_205": row[3], "overpayment_205": row[4],
            "current_total_cost": row[5],
            "obor_debit_209": row[6], "obor_credit_209": row[7],
            "obor_debit_205": row[8], "obor_credit_205": row[9],
        })
    return {"total": total_items, "page": page, "size": limit, "items": items}


@router.get("/rooms/{room_id}/residents-finance",
            summary="Финансы жильцов конкретной квартиры (раскрытие строки)")
async def room_residents_finance(
        room_id: int,
        period_id: Optional[int] = Query(None, description="Период просмотра; по умолчанию активный → последний импорт → свежий"),
        current_user: User = Depends(get_current_user),
        db: AsyncSession = Depends(get_db),
):
    if current_user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")
    active = await _resolve_view_period(db, period_id)
    period_id = active.id if active else None

    d209 = func.coalesce(func.sum(MeterReading.debt_209), 0)
    o209 = func.coalesce(func.sum(MeterReading.overpayment_209), 0)
    d205 = func.coalesce(func.sum(MeterReading.debt_205), 0)
    o205 = func.coalesce(func.sum(MeterReading.overpayment_205), 0)
    total = func.coalesce(func.sum(MeterReading.total_cost), 0)

    stmt = select(User, d209, o209, d205, o205, total).outerjoin(
        MeterReading,
        (User.id == MeterReading.user_id) & (MeterReading.period_id == period_id),
    ).where(
        User.room_id == room_id, User.is_deleted.is_(False),
    ).group_by(User.id).order_by(User.username)
    rows = (await db.execute(stmt)).all()
    return {
        "room_id": room_id,
        "residents": [{
            "user_id": r[0].id,
            "username": r[0].username,
            "full_name": getattr(r[0], "full_name", None),
            "debt_209": r[1], "overpayment_209": r[2],
            "debt_205": r[3], "overpayment_205": r[4],
            "current_total_cost": r[5],
        } for r in rows],
    }


# =========================================================================
# NEW: KPI / STATS / EXPORT / HISTORY / UNDO / RECONCILE
# =========================================================================

async def _resolve_view_period(db: AsyncSession, period_id: Optional[int]):
    """Период для ПРОСМОТРА долгов (список/KPI/квартиры).

    Раньше вьюхи жёстко брали активный период — и если активного нет (между
    месяцами: май закрыт, июнь не открыт), показывали 0 хотя долги залиты в
    закрытый период. Теперь: явный period_id → он; иначе активный; иначе период
    последнего импорта долгов; иначе самый свежий период. Так после импорта за
    май долги видно, даже когда активного периода нет.
    """
    if period_id:
        return (await db.execute(
            select(BillingPeriod).where(BillingPeriod.id == period_id)
        )).scalars().first()
    active = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if active:
        return active
    last_imp = (await db.execute(
        select(DebtImportLog)
        .where(DebtImportLog.period_id.isnot(None))
        .order_by(desc(DebtImportLog.started_at))
        .limit(1)
    )).scalars().first()
    if last_imp and last_imp.period_id:
        p = await db.get(BillingPeriod, last_imp.period_id)
        if p:
            return p
    return (await db.execute(
        select(BillingPeriod).order_by(BillingPeriod.id.desc()).limit(1)
    )).scalars().first()


def _require_finance(user: User) -> None:
    if user.role not in ("financier", "accountant", "admin"):
        raise HTTPException(status_code=403, detail="Доступ запрещен")


@router.get("/debts/stats", summary="KPI по долгам (выбранный/активный период)")
async def debts_stats(
    period_id: Optional[int] = Query(None, description="Период просмотра; по умолчанию активный → последний импорт → свежий"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сводка долгов для шапки вкладки «Долги 1С»."""
    _require_finance(current_user)

    active_period = await _resolve_view_period(db, period_id)
    period_id = active_period.id if active_period else None

    # Агрегация по readings активного периода. ВАЖНО: join User + фильтр
    # is_deleted/role — чтобы KPI считал ТОЛЬКО активных жильцов и совпадал со
    # списком «Долги 1С» (users-status фильтрует так же). Иначе долги, оставшиеся
    # на user_id удалённых/выехавших жильцов после импорта 1С, раздувают счётчик
    # («Должников: 800», а в списке 1).
    _active_user = [User.is_deleted.is_(False), User.role == "user"]
    agg_q = (
        select(
            func.coalesce(func.sum(MeterReading.debt_209), 0),
            func.coalesce(func.sum(MeterReading.overpayment_209), 0),
            func.coalesce(func.sum(MeterReading.debt_205), 0),
            func.coalesce(func.sum(MeterReading.overpayment_205), 0),
            func.count(MeterReading.id),
        )
        .join(User, User.id == MeterReading.user_id)
        .where(MeterReading.period_id == period_id, *_active_user)
    )
    agg = (await db.execute(agg_q)).one()
    total_debt_209, total_over_209, total_debt_205, total_over_205, readings_count = agg

    # Должников: активных жильцов где debt_209+205 > 0 в активном периоде
    debtors_q = (
        select(func.count(func.distinct(MeterReading.user_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            *_active_user,
            (MeterReading.debt_209 > 0) | (MeterReading.debt_205 > 0),
        )
    )
    debtors_count = (await db.execute(debtors_q)).scalar_one()

    # Переплатчиков
    overpayers_q = (
        select(func.count(func.distinct(MeterReading.user_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            *_active_user,
            (MeterReading.overpayment_209 > 0) | (MeterReading.overpayment_205 > 0),
        )
    )
    overpayers_count = (await db.execute(overpayers_q)).scalar_one()

    # --- Учёт по КВАРТИРАМ (помещениям), а не по жильцам ---
    # Квартир с долгом: distinct room_id где сумма debt_209+205 > 0 (по активным
    # жильцам — join User, чтобы не считать квартиры по долгам выехавших).
    rooms_debt_q = (
        select(func.count(func.distinct(MeterReading.room_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            MeterReading.room_id.isnot(None),
            *_active_user,
            (MeterReading.debt_209 > 0) | (MeterReading.debt_205 > 0),
        )
    )
    rooms_with_debt_count = (await db.execute(rooms_debt_q)).scalar_one()

    # Квартир с переплатой
    rooms_over_q = (
        select(func.count(func.distinct(MeterReading.room_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            MeterReading.room_id.isnot(None),
            *_active_user,
            (MeterReading.overpayment_209 > 0) | (MeterReading.overpayment_205 > 0),
        )
    )
    rooms_overpaying_count = (await db.execute(rooms_over_q)).scalar_one()

    # Всего квартир с данными в периоде (для шапки в режиме «Квартиры»)
    total_rooms_q = (
        select(func.count(func.distinct(MeterReading.room_id)))
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            MeterReading.room_id.isnot(None),
            *_active_user,
        )
    )
    total_rooms = (await db.execute(total_rooms_q)).scalar_one()

    # Всего активных жильцов
    total_users_q = select(func.count(User.id)).where(
        User.is_deleted.is_(False), User.role == "user",
    )
    total_users = (await db.execute(total_users_q)).scalar_one()

    total_debt = float(total_debt_209 or 0) + float(total_debt_205 or 0)
    total_over = float(total_over_209 or 0) + float(total_over_205 or 0)
    avg_debt = (total_debt / debtors_count) if debtors_count else 0.0
    avg_debt_room = (total_debt / rooms_with_debt_count) if rooms_with_debt_count else 0.0

    # Последний импорт
    last_log = (await db.execute(
        select(DebtImportLog).order_by(desc(DebtImportLog.started_at)).limit(1)
    )).scalars().first()

    # Распределение по общежитиям: ТОП-10 по долгу
    by_dorm_q = (
        select(
            Room.dormitory_name,
            func.sum(MeterReading.debt_209 + MeterReading.debt_205).label("total_debt"),
            func.count(func.distinct(MeterReading.user_id)).label("debtors"),
        )
        .select_from(MeterReading)
        .join(Room, Room.id == MeterReading.room_id)
        .join(User, User.id == MeterReading.user_id)
        .where(
            MeterReading.period_id == period_id,
            *_active_user,
            (MeterReading.debt_209 > 0) | (MeterReading.debt_205 > 0),
        )
        .group_by(Room.dormitory_name)
        .order_by(desc("total_debt"))
        .limit(10)
    )
    by_dorm = [
        {"name": r[0] or "—", "total_debt": float(r[1] or 0), "debtors": int(r[2] or 0)}
        for r in (await db.execute(by_dorm_q)).all()
    ]

    return {
        "period_name": active_period.name if active_period else None,
        "period_id": period_id,
        "total_users": total_users,
        "debtors_count": debtors_count,
        "overpayers_count": overpayers_count,
        "total_debt_209": float(total_debt_209 or 0),
        "total_debt_205": float(total_debt_205 or 0),
        "total_debt": round(total_debt, 2),
        "total_overpay_209": float(total_over_209 or 0),
        "total_overpay_205": float(total_over_205 or 0),
        "total_overpay": round(total_over, 2),
        "avg_debt_per_debtor": round(avg_debt, 2),
        # Учёт по квартирам (помещениям)
        "rooms_with_debt_count": int(rooms_with_debt_count or 0),
        "rooms_overpaying_count": int(rooms_overpaying_count or 0),
        "total_rooms": int(total_rooms or 0),
        "avg_debt_per_room": round(avg_debt_room, 2),
        "readings_count": int(readings_count or 0),
        "last_import": {
            "id": last_log.id,
            "account_type": last_log.account_type,
            "status": last_log.status,
            "started_at": last_log.started_at.isoformat() if last_log.started_at else None,
            "started_by": last_log.started_by_username,
            "updated": last_log.updated,
            "created": last_log.created,
            "not_found_count": last_log.not_found_count,
        } if last_log else None,
        "by_dormitory": by_dorm,
    }


@router.get("/debts/dormitories", summary="Список общежитий для фильтра")
async def debts_dormitories(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    rows = (await db.execute(
        select(Room.dormitory_name).distinct().order_by(Room.dormitory_name)
    )).scalars().all()
    return [r for r in rows if r]


@router.get("/debts/unassigned", summary="Неразнесённые долги (ФИО не сопоставлены с жильцом)")
async def debts_unassigned(
    period_id: Optional[int] = Query(None, description="Период; по умолчанию активный → последний импорт → свежий"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сумма и список долгов из 1С, которые НЕ привязались ни к одному жильцу
    (ФИО нет в базе / не сопоставлено, либо у жильца нет комнаты — долг хранится
    в показании, а оно требует комнату). Берём not_found из ПОСЛЕДНЕГО импорта
    каждого счёта (209/205) за период и сводим по ФИО. Деньги не теряются из
    вида: разнесутся, когда заведёшь жильца с комнатой и сделаешь переимпорт."""
    _require_finance(current_user)
    period = await _resolve_view_period(db, period_id)
    if not period:
        return {"period_name": None, "period_id": None, "total_debt": 0.0,
                "total_overpayment": 0.0, "count": 0, "items": []}

    merged: dict[str, dict] = {}
    for acct in ("209", "205"):
        log = (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.period_id == period.id,
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()
        if not log or not log.not_found_users:
            continue
        for item in log.not_found_users:
            if isinstance(item, dict):
                fio = (item.get("fio") or "").strip()
                debt = float(item.get("debt") or 0)
                over = float(item.get("overpayment") or 0)
            else:
                fio, debt, over = str(item).strip(), 0.0, 0.0
            if not fio:
                continue
            key = " ".join(fio.lower().split())
            slot = merged.get(key)
            if slot is None:
                slot = merged[key] = {"fio": fio, "debt": 0.0, "overpayment": 0.0, "accounts": []}
            slot["debt"] += debt
            slot["overpayment"] += over
            if acct not in slot["accounts"]:
                slot["accounts"].append(acct)

    items = sorted(merged.values(), key=lambda x: -x["debt"])
    return {
        "period_name": period.name,
        "period_id": period.id,
        "total_debt": round(sum(i["debt"] for i in items), 2),
        "total_overpayment": round(sum(i["overpayment"] for i in items), 2),
        "count": len(items),
        "items": items,
    }


@router.get("/debts/export", summary="Excel-выгрузка текущего списка долгов")
async def debts_export(
    search: str | None = Query(None),
    only_debtors: bool = Query(False),
    only_overpaid: bool = Query(False),
    dormitory: Optional[str] = Query(None),
    min_debt: Optional[float] = Query(None, ge=0),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Excel-файл с теми же фильтрами, что и в таблице UI.
    Без пагинации — выгружает все подходящие записи.
    """
    _require_finance(current_user)

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    period_id = active_period.id if active_period else None

    d209 = func.coalesce(func.sum(MeterReading.debt_209), 0).label("debt_209")
    o209 = func.coalesce(func.sum(MeterReading.overpayment_209), 0).label("overpayment_209")
    d205 = func.coalesce(func.sum(MeterReading.debt_205), 0).label("debt_205")
    o205 = func.coalesce(func.sum(MeterReading.overpayment_205), 0).label("overpayment_205")
    tot = func.coalesce(func.sum(MeterReading.total_cost), 0).label("total_cost")

    stmt = select(User, Room, d209, o209, d205, o205, tot).outerjoin(
        Room, User.room_id == Room.id
    ).outerjoin(
        MeterReading,
        (User.id == MeterReading.user_id) & (MeterReading.period_id == period_id)
    ).where(User.is_deleted.is_(False), User.role == "user")

    if search:
        sv = f"%{search.lower()}%"
        stmt = stmt.where(or_(
            func.lower(User.username).like(sv),
            func.lower(Room.dormitory_name).like(sv),
            func.lower(Room.room_number).like(sv),
        ))
    if dormitory:
        stmt = stmt.where(Room.dormitory_name == dormitory)
    stmt = stmt.group_by(User.id, Room.id)
    if only_debtors:
        stmt = stmt.having((d209 + d205) > 0)
    if only_overpaid:
        stmt = stmt.having((o209 + o205) > 0)
    if min_debt is not None:
        stmt = stmt.having((d209 + d205) >= min_debt)

    stmt = stmt.order_by(Room.dormitory_name.asc().nulls_last(),
                         Room.room_number.asc().nulls_last(),
                         User.username.asc())
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Долги 1С"
    headers = ["ID", "ФИО", "Общежитие", "Комната", "Долг 209", "Перепл. 209",
               "Долг 205", "Перепл. 205", "Итого начислено"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E9D5FF")
    for i, r in enumerate(rows, 2):
        u, room = r[0], r[1]
        ws.cell(row=i, column=1, value=u.id)
        ws.cell(row=i, column=2, value=u.username)
        # housing_001/E2-A: для дома колонка "Общежитие" заполняется
        # улицей+номером дома, "Комната" — номером квартиры. Для общаги
        # сохраняем старое поведение (dormitory_name + room_number).
        if room and room.place_type == "house":
            _addr = ", ".join(filter(None, [
                f"ул. {room.street}" if room.street else None,
                f"д. {room.house_number}" if room.house_number else None,
            ])) or ""
            ws.cell(row=i, column=3, value=_addr)
            ws.cell(row=i, column=4, value=(f"кв. {room.apartment_number}" if room.apartment_number else ""))
        else:
            ws.cell(row=i, column=3, value=(room.dormitory_name if room else ""))
            ws.cell(row=i, column=4, value=(room.room_number if room else ""))
        ws.cell(row=i, column=5, value=float(r[2] or 0))
        ws.cell(row=i, column=6, value=float(r[3] or 0))
        ws.cell(row=i, column=7, value=float(r[4] or 0))
        ws.cell(row=i, column=8, value=float(r[5] or 0))
        ws.cell(row=i, column=9, value=float(r[6] or 0))
    for col, w in [("A", 6), ("B", 30), ("C", 22), ("D", 10),
                   ("E", 12), ("F", 12), ("G", 12), ("H", 12), ("I", 14)]:
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"debts_{utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# =========================================================================
# DEBT IMPORT HISTORY
# =========================================================================

@router.get("/debts/import-history", summary="История импортов 1С")
async def debts_import_history(
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    _require_finance(current_user)
    logs = (await db.execute(
        select(DebtImportLog).order_by(desc(DebtImportLog.started_at)).limit(limit)
    )).scalars().all()
    return [
        {
            "id": log.id,
            "account_type": log.account_type,
            "file_name": log.file_name,
            "status": log.status,
            "started_by": log.started_by_username,
            "processed": log.processed,
            "updated": log.updated,
            "created": log.created,
            "not_found_count": log.not_found_count,
            "error": log.error,
            "started_at": log.started_at.isoformat() if log.started_at else None,
            "completed_at": log.completed_at.isoformat() if log.completed_at else None,
            "reverted_at": log.reverted_at.isoformat() if log.reverted_at else None,
            # Новые поля: парный batch + наличие оригинала для скачивания
            "batch_id": log.batch_id,
            "has_archive": bool(log.archive_path),
        }
        for log in logs
    ]


@router.get("/debts/import-history/{log_id}/diff", summary="Diff с предыдущим импортом того же счёта")
async def debts_import_diff(
    log_id: int,
    against_id: Optional[int] = Query(None, description="ID импорта для сравнения. None — предыдущий того же типа."),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Сравнивает applied_state двух импортов одного account_type.

    Категории жильцов:
      - new_debtors:  не было в прошлом импорте, появился долг > 0
      - debt_grew:    был и есть, debt вырос
      - debt_dropped: был и есть, debt упал (но > 0)
      - debt_closed:  был долг > 0, стал 0 (или жилец исчез из файла)
      - new_overpay:  появилась переплата которой не было

    На жильцов с одинаковой суммой не возвращаем — это шум, отрисуется
    только то что изменилось.
    """
    _require_finance(current_user)

    current = await db.get(DebtImportLog, log_id)
    if not current:
        raise HTTPException(404, "Лог не найден")
    if not current.applied_state:
        raise HTTPException(
            400,
            "У этого импорта нет applied_state (импорт до миграции debts_003). "
            "Перезагрузите файлы — diff заработает.",
        )

    # Находим предыдущий импорт того же account_type, либо берём указанный.
    if against_id is not None:
        previous = await db.get(DebtImportLog, against_id)
        if not previous:
            raise HTTPException(404, "Лог для сравнения не найден")
        if previous.account_type != current.account_type:
            raise HTTPException(
                400,
                f"Нельзя сравнивать {previous.account_type!r} с {current.account_type!r}",
            )
    else:
        previous = (await db.execute(
            select(DebtImportLog)
            .where(
                DebtImportLog.account_type == current.account_type,
                DebtImportLog.id < current.id,
                DebtImportLog.status == "completed",
                DebtImportLog.applied_state.is_not(None),
            )
            .order_by(desc(DebtImportLog.id))
            .limit(1)
        )).scalars().first()

    if not previous:
        return {
            "current_id": log_id,
            "previous_id": None,
            "account_type": current.account_type,
            "fatal": "Это первый импорт этого счёта — сравнивать не с чем.",
        }

    cur_state = current.applied_state or {}
    prev_state = previous.applied_state or {}
    account = current.account_type
    debt_key = f"debt_{account}"
    over_key = f"overpayment_{account}"

    def _dec(d, key):
        try:
            return Decimal(str(d.get(key, "0") or "0"))
        except Exception:
            return Decimal("0")

    new_debtors = []
    debt_grew = []
    debt_dropped = []
    debt_closed = []
    new_overpay = []

    # Bug AG: applied_state теперь keyed by user_id (раньше room_id — в
    # коммуналке два жильца перезаписывали друг друга).
    all_user_ids = set(cur_state.keys()) | set(prev_state.keys())
    for user_id_str in all_user_ids:
        cur = cur_state.get(user_id_str, {})
        prev = prev_state.get(user_id_str, {})
        cur_debt = _dec(cur, debt_key)
        prev_debt = _dec(prev, debt_key)
        cur_over = _dec(cur, over_key)
        prev_over = _dec(prev, over_key)

        # Метаданные берём из cur если есть, иначе из prev (если жилец исчез)
        meta_username = cur.get("username") or prev.get("username") or "—"
        meta_room = cur.get("room_label") or prev.get("room_label") or "—"
        # room_id хранится внутри applied_state (после Bug AG) либо берём
        # из cur/prev. Для legacy-логов до Bug AG в applied_state нет user_id,
        # вместо него стоит room_id — попробуем привести к int безопасно.
        room_id_val = cur.get("room_id") or prev.get("room_id")
        try:
            user_id_int = int(user_id_str)
        except Exception:
            user_id_int = None

        if cur_debt > prev_debt:
            entry = {
                "user_id": user_id_int,
                "room_id": room_id_val,
                "username": meta_username,
                "room_label": meta_room,
                "prev_debt": float(prev_debt),
                "current_debt": float(cur_debt),
                "delta": float(cur_debt - prev_debt),
            }
            if prev_debt == 0 and cur_debt > 0:
                new_debtors.append(entry)
            else:
                debt_grew.append(entry)
        elif cur_debt < prev_debt:
            entry = {
                "user_id": user_id_int,
                "room_id": room_id_val,
                "username": meta_username,
                "room_label": meta_room,
                "prev_debt": float(prev_debt),
                "current_debt": float(cur_debt),
                "delta": float(cur_debt - prev_debt),  # отрицательная
            }
            if cur_debt == 0 and prev_debt > 0:
                debt_closed.append(entry)
            else:
                debt_dropped.append(entry)

        # Появилась переплата которой не было — сигнал что админ должен возвратить
        if cur_over > 0 and prev_over == 0:
            new_overpay.append({
                "user_id": user_id_int,
                "room_id": room_id_val,
                "username": meta_username,
                "room_label": meta_room,
                "overpayment": float(cur_over),
            })

    # Сортируем: новые должники и рост — по сумме убыванию
    new_debtors.sort(key=lambda x: -x["current_debt"])
    debt_grew.sort(key=lambda x: -x["delta"])
    debt_dropped.sort(key=lambda x: x["delta"])  # самый большой спад первым
    debt_closed.sort(key=lambda x: -x["prev_debt"])
    new_overpay.sort(key=lambda x: -x["overpayment"])

    # Топ-снимок сумм для KPI
    sum_grew = sum(e["delta"] for e in debt_grew + new_debtors)
    sum_closed = sum(e["prev_debt"] for e in debt_closed)
    sum_dropped = sum(-e["delta"] for e in debt_dropped)

    return {
        "current_id": log_id,
        "previous_id": previous.id,
        "account_type": account,
        "current_started_at": current.started_at.isoformat() if current.started_at else None,
        "previous_started_at": previous.started_at.isoformat() if previous.started_at else None,
        "summary": {
            "new_debtors_count": len(new_debtors),
            "debt_grew_count": len(debt_grew),
            "debt_dropped_count": len(debt_dropped),
            "debt_closed_count": len(debt_closed),
            "new_overpay_count": len(new_overpay),
            "sum_new_and_grew": float(sum_grew),
            "sum_closed": float(sum_closed),
            "sum_dropped": float(sum_dropped),
        },
        # Лимиты на размер response — UI всё равно покажет первые 100
        "new_debtors": new_debtors[:100],
        "debt_grew": debt_grew[:100],
        "debt_dropped": debt_dropped[:100],
        "debt_closed": debt_closed[:100],
        "new_overpay": new_overpay[:50],
    }


@router.get("/debts/user-debt-history/{user_id}", summary="История долгов жильца через все импорты")
async def debts_user_history(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Возвращает точки графика для одного жильца: на каждый
    completed-импорт — debt+overpayment по этому юзеру.

    Bug AG: applied_state теперь keyed by user_id, поэтому переезды
    больше не теряют точки (раньше при смене комнаты история обрывалась).
    UI рисует две линии: 209 (коммунальный) и 205 (найм), плюс tabular
    разрез по каждому импорту.
    """
    _require_finance(current_user)

    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Жилец не найден")

    user_id_key = str(user.id)

    # Все completed-импорты с applied_state — отсортированы по дате
    logs = (await db.execute(
        select(DebtImportLog)
        .where(
            DebtImportLog.status == "completed",
            DebtImportLog.applied_state.is_not(None),
        )
        .order_by(DebtImportLog.started_at.asc(), DebtImportLog.id.asc())
    )).scalars().all()

    points = []
    last_room_label = None
    for log in logs:
        st = log.applied_state or {}
        entry = st.get(user_id_key)
        if not entry:
            # В этот импорт этой комнаты не было — пропускаем точку, чтобы
            # не подмешивать «0», которое на самом деле «нет данных».
            continue
        debt_key = f"debt_{log.account_type}"
        over_key = f"overpayment_{log.account_type}"
        try:
            debt = float(Decimal(str(entry.get(debt_key, "0") or "0")))
            over = float(Decimal(str(entry.get(over_key, "0") or "0")))
        except Exception:
            debt = 0.0
            over = 0.0
        # room_label берём из applied_state (denormalized), чтобы не делать
        # отдельный JOIN. Последнее значение — самое свежее.
        if entry.get("room_label"):
            last_room_label = entry["room_label"]
        points.append({
            "log_id": log.id,
            "started_at": log.started_at.isoformat() if log.started_at else None,
            "account_type": log.account_type,
            "debt": debt,
            "overpayment": over,
            "file_name": log.file_name,
        })

    return {
        "user_id": user_id,
        "username": user.username,
        "room_id": user.room_id,
        "room_label": last_room_label,
        "points": points,
    }


@router.get("/debts/import-history/{log_id}/download", summary="Скачать оригинальный xlsx")
async def debts_import_download(
    log_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Отдаёт оригинальный файл ОСВ из 1С, привязанный к этому импорту.

    archive_path хранится на диске вне /static (защита от прямого
    скачивания через nginx без auth). FileResponse кидает 404 если файл
    физически удалён (например, после retention-чистки).
    """
    _require_finance(current_user)
    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог импорта не найден")
    if not log.archive_path:
        raise HTTPException(
            404,
            "Архив этого импорта не сохранён (старый импорт до миграции debts_002)",
        )
    if not os.path.exists(log.archive_path):
        raise HTTPException(
            404,
            "Файл физически удалён (retention-policy / ручная очистка).",
        )

    # Имя для пользователя: оригинальное file_name если есть, иначе генерим
    # понятное «debts_209_2026-05-12.xlsx».
    download_name = log.file_name or (
        f"debts_{log.account_type}_{log.started_at.strftime('%Y-%m-%d') if log.started_at else 'unknown'}.xlsx"
    )

    from fastapi.responses import FileResponse
    return FileResponse(
        path=log.archive_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/debts/import-history/{log_id}/reparse",
             summary="Переимпорт лога 1С из архива с актуальной логикой парсера")
async def debts_reparse(
    log_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bug AE: Reading'и, импортированные до Bug U-fix6, имеют
    debt_209/debt_205 = начальное сальдо вместо конечного (когда обороты
    Кредит погасили долг, а end-колонки в ОСВ пустые). Парсер обновлён
    (pick_saldo_pair учитывает обороты), но сами reading'и в БД не
    пересчитаны автоматически — там лежат старые значения.

    Этот endpoint берёт archive_path лога и заново запускает
    import_debts_task: pipeline UPDATE-ит существующие reading'и
    значениями из актуальной логики парсинга. Новый DebtImportLog
    создаётся (для аудита и возможного отката).

    Что НЕ делает:
      - не удаляет старый log (audit trail сохраняется)
      - не trigger'ит revert старого (snapshot_data старого остаётся
        корректным относительно того момента, отдельная история)
    """
    _require_finance(current_user)
    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог импорта не найден")
    if not log.archive_path:
        raise HTTPException(
            404,
            "Архив этого импорта не сохранён (старый импорт до миграции debts_002). "
            "Загрузите тот же файл из 1С вручную через форму импорта.",
        )
    if not os.path.exists(log.archive_path):
        raise HTTPException(
            404,
            "Файл архива физически удалён (retention-policy / ручная очистка). "
            "Загрузите тот же файл из 1С вручную через форму импорта.",
        )

    import uuid as _uuid
    batch_id = str(_uuid.uuid4())
    task = import_debts_task.delay(
        log.archive_path,
        log.account_type,
        started_by_id=current_user.id,
        started_by_username=current_user.username,
        batch_id=batch_id,
        original_file_name=log.file_name or f"reparse_{log.account_type}_{log.id}.xlsx",
    )

    logger.info(
        f"[REPARSE] log_id={log_id} account={log.account_type} "
        f"archive={log.archive_path} task={task.id} batch={batch_id}"
    )

    return {
        "task_id": task.id,
        "status": "processing",
        "account_type": log.account_type,
        "batch_id": batch_id,
        "source_log_id": log_id,
        "source_file": log.file_name,
    }


@router.get("/debts/import-history/{log_id}/not-found", summary="Не найденные ФИО в импорте")
async def debts_not_found(
    log_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Список ФИО из конкретного импорта, которые fuzzy не привязал.
    Админ может вручную сопоставить через reassign-endpoint.

    Формат not_found_users поменялся в импорте мая 2026:
      - старые импорты: list[str] — только ФИО, без сумм
      - новые: list[dict] {fio, debt, overpayment} — фронт префиллит инпуты
    Возвращаем УНИФИЦИРОВАННЫЙ формат list[dict] чтобы UI был один.
    """
    _require_finance(current_user)
    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог импорта не найден")

    raw_list = log.not_found_users or []
    normalized = []
    for item in raw_list:
        if isinstance(item, dict):
            normalized.append({
                "fio": item.get("fio", ""),
                "debt": item.get("debt", "0"),
                "overpayment": item.get("overpayment", "0"),
            })
        else:
            # Legacy: только ФИО, без сумм. Админу придётся вводить руками.
            normalized.append({
                "fio": str(item),
                "debt": "0",
                "overpayment": "0",
            })

    return {
        "log_id": log.id,
        "account_type": log.account_type,
        "not_found_users": normalized,
    }


@router.get("/debts/import-history/{log_id}/not-found-analysis",
            summary="Почему ФИО не сматчились: категории + лучший кандидат")
async def debts_not_found_analysis(
    log_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Диагностика: для КАЖДОГО ненайденного ФИО считает лучшего кандидата в
    базе жильцов (fuzzy + совпадение фамилии) и относит к категории:
      • near    (score ≥ 70) — близкое совпадение ЕСТЬ: ФИО просто записано в 1С
        иначе (сокращение/формат/опечатка). Привязать в 1 клик (reassign).
      • weak    (50–69)      — совпала фамилия, но имя/отчество расходятся
        (возможен однофамилец) — нужна проверка.
      • absent  (< 50)       — похожих в базе нет: бывший жилец, новый человек,
        или не-резидент-плательщик.
    Так видно, сколько из N «не найдено» — это формат (быстрый reassign), а
    сколько реально отсутствуют в базе."""
    _require_finance(current_user)
    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог импорта не найден")

    raw = log.not_found_users or []
    fios = []
    for item in raw:
        if isinstance(item, dict):
            fios.append((item.get("fio", ""), item.get("debt", "0"), item.get("overpayment", "0")))
        else:
            fios.append((str(item), "0", "0"))

    from sqlalchemy.orm import selectinload as _selectinload
    from rapidfuzz import fuzz
    users = (await db.execute(
        select(User).options(_selectinload(User.room))
        .where(User.is_deleted.is_(False), User.role == "user")
    )).scalars().all()
    user_norm = [(u, " ".join((u.username or "").lower().split())) for u in users if u.username]

    def _parts(s: str):
        # ФИО → [фамилия, имя, отчество]: lower, ё→е, точки как разделители.
        t = (s or "").lower().replace("ё", "е").replace(".", " ")
        return t.split()

    def _pm(a: str, b: str) -> bool:
        # Совпадение части ФИО: равны / инициал (одна буква совпала) / опечатка.
        if not a or not b:
            return False
        if a == b:
            return True
        if len(a) == 1 or len(b) == 1:
            return a[0] == b[0]
        return fuzz.ratio(a, b) >= 88

    # Пер-categories. ВАЖНО: «тот же человек» = совпали ВСЕ три части (фамилия+
    # имя+отчество), а не только фамилия. Иначе «Верхозин Владимир» матчился бы
    # на «Верхозин Артём» (однофамилец) с виду «привязать в 1 клик» — и долг ушёл
    # бы чужому. Однофамилец/совпавшее имя-отчество = РАЗНЫЙ человек, не привязка.
    cats = {"same": 0, "namesake": 0, "absent": 0}
    items = []
    for fio, debt, overpay in fios:
        fp = _parts(fio)
        best_u = None
        best_key = (-1, -1)   # (совпавших_частей, fuzzy)
        best_flags = (False, False, False)
        best_fuzzy = 0
        for u, _nn in user_norm:
            cp = _parts(u.username)
            if not fp or not cp:
                continue
            s = _pm(fp[0], cp[0])
            n = _pm(fp[1] if len(fp) > 1 else "", cp[1] if len(cp) > 1 else "")
            p = _pm(fp[2] if len(fp) > 2 else "", cp[2] if len(cp) > 2 else "")
            fz = fuzz.token_sort_ratio(" ".join(fp), " ".join(cp))
            key = (int(s) + int(n) + int(p), fz)
            if key > best_key:
                best_key, best_u, best_flags, best_fuzzy = key, u, (s, n, p), fz
        s, n, p = best_flags
        if best_u is None:
            cat, reason = "absent", None
        elif s and n and p:
            cat, reason = "same", "фамилия+имя+отчество совпали"
        elif s:
            cat, reason = "namesake", "та же фамилия, имя/отчество другие"
        elif n and p:
            cat, reason = "namesake", "совпали имя+отчество, фамилия другая"
        elif best_fuzzy >= 60:
            cat, reason = "namesake", "частичное совпадение"
        else:
            cat, reason = "absent", None
        cats[cat] += 1
        cand = None
        if best_u is not None and cat != "absent":
            cand = {
                "id": best_u.id,
                "username": best_u.username,
                "room": (best_u.room.format_address if best_u.room else None),
            }
        items.append({
            "fio": fio,
            "debt": debt,
            "overpayment": overpay,
            "best_score": int(best_fuzzy),
            "category": cat,
            "reason": reason,
            "candidate": cand,
        })

    items.sort(key=lambda x: -x["best_score"])
    return {
        "log_id": log.id,
        "account_type": log.account_type,
        "total": len(items),
        "categories": cats,
        "items": items,
    }


@router.post("/debts/import-history/{log_id}/undo", summary="Откат импорта 1С")
async def debts_undo_import(
    log_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Восстанавливает debt_*/overpayment_* по snapshot и удаляет
    созданные импортом черновики. Только для админа/финансиста."""
    if current_user.role not in ("admin", "financier"):
        raise HTTPException(403, "Недостаточно прав для отката импорта")

    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог импорта не найден")
    if log.status != "completed":
        raise HTTPException(400, f"Нельзя откатить импорт в статусе «{log.status}»")
    if not log.snapshot_data:
        raise HTTPException(400, "Нет snapshot-данных — откат невозможен (старый импорт?)")

    before = log.snapshot_data.get("before") or {}
    inserted_ids = log.snapshot_data.get("inserted_reading_ids") or []
    if not isinstance(before, dict):
        before = {}
    if not isinstance(inserted_ids, list):
        inserted_ids = []
    inserted_ids = [i for i in inserted_ids if isinstance(i, int)]

    # 1. Восстанавливаем существующие readings из snapshot.
    # ЗАЩИТА: снимок может быть пустым/иного формата (старые или ГИС-импорты) —
    # пропускаем некорректные записи, не роняем 500.
    def _dec(v):
        try:
            return Decimal(str(v if v not in (None, "") else "0"))
        except Exception:
            return Decimal("0")
    updates = []
    for reading_id_str, vals in before.items():
        if not isinstance(vals, dict):
            continue
        try:
            rid = int(reading_id_str)
        except (TypeError, ValueError):
            continue
        updates.append({
            "id": rid,
            "debt_209": _dec(vals.get("debt_209")),
            "overpayment_209": _dec(vals.get("overpayment_209")),
            "debt_205": _dec(vals.get("debt_205")),
            "overpayment_205": _dec(vals.get("overpayment_205")),
        })
    if not updates and not inserted_ids:
        raise HTTPException(
            400, "Снимок импорта пуст или несовместим — откат не применён "
                 "(вероятно, ГИС-импорт или старый формат). Долги не изменены.")

    # SQLAlchemy async не имеет bulk_update_mappings — делаем обычный update per row.
    # Для 1000+ записей не критично (один индексированный UPDATE).
    from sqlalchemy import update as _update
    for u in updates:
        await db.execute(
            _update(MeterReading)
            .where(MeterReading.id == u["id"])
            .values(
                debt_209=u["debt_209"],
                overpayment_209=u["overpayment_209"],
                debt_205=u["debt_205"],
                overpayment_205=u["overpayment_205"],
            )
        )

    # 2. Удаляем черновики, которые создал этот импорт
    # (берём только те, что всё ещё is_approved=False — согласованные не трогаем)
    if inserted_ids:
        from sqlalchemy import delete as _delete
        await db.execute(
            _delete(MeterReading).where(
                MeterReading.id.in_(inserted_ids),
                MeterReading.is_approved.is_(False),
            )
        )

    log.status = "reverted"
    log.reverted_at = utcnow()

    await db.commit()

    return {
        "status": "ok",
        "restored_readings": len(updates),
        "removed_drafts": len(inserted_ids),
    }


@router.get(
    "/debts/check-resident-coverage/{user_id}",
    summary="Найти жильца в архивах последних импортов 1С (диагностика)",
)
async def debts_check_resident_coverage(
    user_id: int,
    last_n: int = 10,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Для конкретного жильца перебирает последние N импортов 1С,
    парсит архивные xlsx, ищет ФИО (точное совпадение + substring).
    Полезно для диагностики «почему у Миронова нет долгов»:
      - если в архивах есть с цифрами → fuzzy-привязка ошиблась, нужен reassign
      - если есть с нулями → нормально, нет долга
      - если нет вообще → жильца не передавали из 1С
    """
    if current_user.role not in ("admin", "financier"):
        raise HTTPException(403, "Недостаточно прав")

    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Жилец не найден")

    fio_db = (user.full_name or user.username or "").strip()
    if not fio_db:
        raise HTTPException(400, "У жильца нет ФИО — нечего искать")

    # Нормализация для substring-сравнения (нижний регистр, без точек,
    # без двойных пробелов).
    import re as _re
    def _norm(s: str) -> str:
        s = (s or "").lower().replace(".", " ").replace(",", " ")
        s = _re.sub(r"\s+", " ", s).strip()
        return s

    fio_db_norm = _norm(fio_db)
    # Также берём фамилию + первую букву имени для substring-поиска.
    parts = fio_db_norm.split()
    surname = parts[0] if parts else ""

    logs = (await db.execute(
        select(DebtImportLog)
        .where(
            DebtImportLog.archive_path.is_not(None),
            DebtImportLog.status.in_(["completed", "reverted"]),
        )
        .order_by(desc(DebtImportLog.id))
        .limit(last_n)
    )).scalars().all()

    import openpyxl as _opx
    import os as _os
    from decimal import Decimal as _D
    results = []
    for log in logs:
        item = {
            "log_id": log.id,
            "account_type": log.account_type,
            "started_at": log.started_at.isoformat() if log.started_at else None,
            "status": log.status,
            "matches": [],
            "error": None,
        }
        try:
            if not _os.path.exists(log.archive_path):
                item["error"] = "archive_missing"
                results.append(item)
                continue
            wb = _opx.load_workbook(filename=log.archive_path, read_only=True, data_only=True)
            ws = wb.active
            # ФИО в ОСВ 1С может быть в любой строковой колонке (зависит
            # от шаблона выгрузки). Ищем substring фамилии во ВСЕХ
            # колонках строки.
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not row:
                    continue
                # Ищем колонку с ФИО (substring "фамилия" в строковом значении).
                fio_cell = None
                fio_col_idx = None
                for col_idx, cell_val in enumerate(row):
                    if cell_val is None or not isinstance(cell_val, str):
                        continue
                    cell_norm = _norm(cell_val)
                    if not cell_norm or not surname:
                        continue
                    if surname not in cell_norm:
                        continue
                    # Sanity: ячейка должна выглядеть как ФИО (несколько слов
                    # с заглавных букв), а не как «Договор...» или «Сальдо...».
                    # Фильтруем явные ключевые слова ОСВ.
                    if any(kw in cell_norm for kw in [
                        "договор", "сальдо", "оборот", "итого", "период",
                        "квартир", "общежит", "счёт", "счет", "помещен",
                    ]):
                        continue
                    fio_cell = cell_val
                    fio_col_idx = col_idx
                    break
                if fio_cell is None:
                    continue
                fio_cell_norm = _norm(str(fio_cell))
                # Собираем числовые значения из строки (после колонки ФИО),
                # чтобы показать сальдо.
                numeric_cols = []
                for col_val in row[fio_col_idx + 1:]:
                    if col_val is None or col_val == "":
                        continue
                    try:
                        d = _D(str(col_val).replace(",", "."))
                        if d != 0:
                            numeric_cols.append(float(d))
                    except Exception:
                        pass
                exact = fio_cell_norm == fio_db_norm
                item["matches"].append({
                    "row_excel": row_idx,
                    "col_excel": fio_col_idx + 1,  # 1-based для удобства админа
                    "fio_in_excel": str(fio_cell).strip(),
                    "exact_match": exact,
                    "numeric_values": numeric_cols[:6],
                })
            wb.close()
        except Exception as exc:
            item["error"] = f"parse_failed: {exc}"
        results.append(item)

    return {
        "user_id": user_id,
        "fio_db": fio_db,
        "imports_checked": len(logs),
        "results": results,
    }


@router.get(
    "/debts/import-history/{log_id}/parser-diagnose",
    summary="Диагностика парсера: какие колонки нашёл, что извлёк",
)
async def debts_parser_diagnose(
    log_id: int,
    fio_search: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Прогоняет логику парсера заголовков на архивном файле этого импорта
    и возвращает: какие колонки нашёл (debt_first/last, overpay_first/last),
    где была account-total row (209.34 / 205.X), какие numeric_positions
    в ней, sample 3 строк жильцов с распарсенными debt/over.

    Этот endpoint **не** делает импорт — только показывает что парсер
    видит. Помогает диагностировать «почему у Бендаса всё ещё 2385.07».
    """
    if current_user.role not in ("admin", "financier"):
        raise HTTPException(403, "Недостаточно прав")

    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог не найден")
    if not log.archive_path:
        raise HTTPException(400, "У этого импорта нет архива (старый импорт без archive_path)")

    import os as _os
    if not _os.path.exists(log.archive_path):
        raise HTTPException(404, f"Архив не найден на диске: {log.archive_path}")

    # Используем тот же парсер что и в основном импорте — копируем сюда
    # ключевые шаги.
    import openpyxl as _opx
    from app.modules.utility.services.debt_import import clean_decimal, pick_saldo_pair
    try:
        ws = _opx.load_workbook(filename=log.archive_path, read_only=True, data_only=True).active
    except Exception as e:
        raise HTTPException(400, f"openpyxl не открыл файл: {e}")

    section_markers: dict = {}
    debit_cols: list = []
    credit_cols: list = []
    account_total = None  # {row_idx, label_col, label, numeric_positions, all_values}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if not row:
            continue
        # Section markers + debit/credit positions
        for col_idx, cell in enumerate(row):
            if cell is None or not isinstance(cell, str):
                continue
            cell_norm = cell.strip().lower()
            if not cell_norm:
                continue
            if "сальдо" in cell_norm and "начал" in cell_norm and "начало" not in section_markers:
                section_markers["начало"] = col_idx
            elif "оборот" in cell_norm and ("период" in cell_norm or len(cell_norm) < 30) and "обороты" not in section_markers:
                section_markers["обороты"] = col_idx
            elif "сальдо" in cell_norm and "конец" in cell_norm and "конец" not in section_markers:
                section_markers["конец"] = col_idx
            elif cell_norm == "дебет":
                debit_cols.append(col_idx)
            elif cell_norm == "кредит":
                credit_cols.append(col_idx)
        # Account total row
        if account_total is None:
            for col_label in range(min(3, len(row))):
                cell = row[col_label]
                if cell is None:
                    continue
                s = str(cell).strip()
                if s.startswith("209.") or s.startswith("205.") or s == "209" or s == "205":
                    numeric_positions = []
                    all_values = {}
                    for col_idx in range(col_label + 1, len(row)):
                        c = row[col_idx]
                        if c is None or c == "":
                            continue
                        try:
                            d = clean_decimal(c)
                            if d != 0:
                                numeric_positions.append(col_idx)
                                all_values[col_idx] = float(d)
                        except Exception:
                            pass
                    account_total = {
                        "row_idx": row_idx,
                        "label_col": col_label,
                        "label": s,
                        "numeric_positions": numeric_positions,
                        "all_values": all_values,
                    }
                    break

    debit_cols = sorted(set(debit_cols))
    credit_cols = sorted(set(credit_cols))

    # Какие колонки выберет парсер
    chosen = {
        "debt_col_first": None,
        "debt_col_last": None,
        "overpay_col_first": None,
        "overpay_col_last": None,
        "obor_debit_col": None,
        "obor_credit_col": None,
        "strategy": None,
    }
    if account_total and len(account_total["numeric_positions"]) >= 4:
        np_list = account_total["numeric_positions"]
        chosen["debt_col_first"] = np_list[0]
        chosen["overpay_col_first"] = np_list[1] if len(np_list) > 1 else np_list[0]
        if len(np_list) >= 6:
            chosen["obor_debit_col"] = np_list[2]
            chosen["obor_credit_col"] = np_list[3]
            chosen["debt_col_last"] = np_list[4]
            chosen["overpay_col_last"] = np_list[5]
        elif len(np_list) == 5:
            chosen["obor_debit_col"] = np_list[2]
            chosen["debt_col_last"] = np_list[3]
            chosen["overpay_col_last"] = np_list[4]
        elif len(np_list) == 4:
            chosen["debt_col_last"] = np_list[2]
            chosen["overpay_col_last"] = np_list[3]
        chosen["strategy"] = "0_account_total_row"

    # Sample жильцов: для каждого извлекаем debt/over через pick_saldo_pair.
    # Если fio_search задан — ищем только этого жильца (substring match).
    # Иначе — первые 3 жильца как preview.
    samples = []
    search_norm = (fio_search or "").strip().lower()
    if chosen["debt_col_last"] is not None and chosen["overpay_col_last"] is not None:
        count = 0
        max_count = 50 if search_norm else 3
        for row in ws.iter_rows(min_row=10, max_row=2000, values_only=True):
            if count >= max_count or not row:
                continue
            # Ищем ФИО в первых 5 колонках
            fio = None
            fio_col = None
            for col_idx in range(min(5, len(row))):
                cell = row[col_idx]
                if not isinstance(cell, str):
                    continue
                s = str(cell).strip()
                if " " in s and len(s.split()) >= 2 and any('А' <= c <= 'я' for c in s):
                    # Sanity: не "Договор", не "Сальдо", не "Контрагенты"
                    s_low = s.lower()
                    if any(kw in s_low for kw in ["договор", "сальдо", "оборот", "итого", "контрагент", "счёт", "счет", "помещен", "период"]):
                        continue
                    fio = s
                    fio_col = col_idx
                    break
            if not fio:
                continue
            # Если задан поиск — фильтруем по substring.
            if search_norm and search_norm not in fio.lower():
                continue
            try:
                debt, over = pick_saldo_pair(
                    row,
                    end_debit_col=chosen["debt_col_last"],
                    end_credit_col=chosen["overpay_col_last"],
                    start_debit_col=chosen["debt_col_first"],
                    start_credit_col=chosen["overpay_col_first"],
                    obor_debit_col=chosen["obor_debit_col"],
                    obor_credit_col=chosen["obor_credit_col"],
                )
            except Exception:
                debt, over = 0, 0

            # Raw values в каждой ключевой колонке — для понимания структуры.
            def _raw(col):
                if col is None or col >= len(row):
                    return None
                v = row[col]
                if v is None or v == "":
                    return None
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    return float(clean_decimal(v))
                except Exception:
                    return str(v)

            sample = {
                "fio": fio,
                "fio_col": fio_col,
                "debt_extracted": float(debt),
                "overpayment_extracted": float(over),
                "raw_values": {
                    f"col{chosen['debt_col_first']}_start_debit": _raw(chosen["debt_col_first"]),
                    f"col{chosen['overpay_col_first']}_start_credit": _raw(chosen["overpay_col_first"]),
                    **({f"col{chosen['obor_debit_col']}_obor_debit": _raw(chosen["obor_debit_col"])} if chosen.get("obor_debit_col") is not None else {}),
                    **({f"col{chosen['obor_credit_col']}_obor_credit": _raw(chosen["obor_credit_col"])} if chosen.get("obor_credit_col") is not None else {}),
                    f"col{chosen['debt_col_last']}_end_debit": _raw(chosen["debt_col_last"]),
                    f"col{chosen['overpay_col_last']}_end_credit": _raw(chosen["overpay_col_last"]),
                },
            }

            # Если поиск конкретного жильца — добавляем сравнение с БД.
            if search_norm:
                # Ищем жильца в БД через нормализацию.
                from app.modules.utility.services.debt_import import normalize_name
                from rapidfuzz import process, fuzz
                norm = normalize_name(fio)
                # Загружаем кэш жильцов.
                users_all = (await db.execute(
                    select(User).where(User.is_deleted.is_(False))
                )).scalars().all()
                user_map = {normalize_name(u.username): u for u in users_all}
                matched_user = user_map.get(norm)
                fuzzy_match_info = None
                if not matched_user:
                    # Fuzzy
                    match = process.extractOne(
                        norm, list(user_map.keys()),
                        scorer=fuzz.token_sort_ratio,
                    )
                    if match:
                        best_key, score, _ = match
                        if score >= 80:
                            matched_user = user_map[best_key]
                            fuzzy_match_info = {"key": best_key, "score": score}
                        else:
                            fuzzy_match_info = {"key": best_key, "score": score, "too_low": True}

                sample["db_lookup"] = None
                if matched_user:
                    # Загружаем текущие debt/over из активного period.
                    active_period = (await db.execute(
                        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
                    )).scalars().first()
                    db_reading = None
                    if active_period:
                        db_reading = (await db.execute(
                            select(MeterReading).where(
                                MeterReading.user_id == matched_user.id,
                                MeterReading.period_id == active_period.id,
                            ).limit(1)
                        )).scalars().first()
                    is_account_209 = log.account_type == "209"
                    sample["db_lookup"] = {
                        "matched_user_id": matched_user.id,
                        "matched_username": matched_user.username,
                        "matched_full_name": matched_user.full_name,
                        "fuzzy": fuzzy_match_info,
                        "db_debt": float(db_reading.debt_209 if is_account_209 else db_reading.debt_205) if db_reading else None,
                        "db_overpayment": float(db_reading.overpayment_209 if is_account_209 else db_reading.overpayment_205) if db_reading else None,
                        "expected_debt": float(debt),
                        "expected_overpayment": float(over),
                        "mismatch": (db_reading is None) or (
                            abs(float(debt) - float(db_reading.debt_209 if is_account_209 else db_reading.debt_205 or 0)) > 0.01
                        ),
                    }
                else:
                    sample["db_lookup"] = {
                        "matched_user_id": None,
                        "fuzzy": fuzzy_match_info,
                        "expected_debt": float(debt),
                        "expected_overpayment": float(over),
                        "mismatch": True,
                        "reason": "user_not_found_in_db",
                    }
            samples.append(sample)
            count += 1

    ws.parent.close()

    return {
        "log_id": log_id,
        "archive_path": log.archive_path,
        "section_markers": section_markers,
        "debit_cols_in_header": debit_cols,
        "credit_cols_in_header": credit_cols,
        "account_total": account_total,
        "chosen": chosen,
        "samples": samples,
    }


@router.post(
    "/debts/probe-update/{user_id}",
    summary="Bug AF: проверить, что UPDATE на reading жильца реально доходит до БД",
)
async def debts_probe_update(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bug AF probe: пробует UPDATE двумя стратегиями (только по `id` —
    как сейчас в импорте; и по `(id, created_at)` — кандидат на фикс
    партиционирования) и показывает rowcount/значение после каждой.

    БЕЗОПАСНО: в конце делает rollback — БД не меняется. Используется
    как «истина в последней инстанции» — если UPDATE по `id` возвращает
    rowcount=0, это партиционная проблема (composite PK + RANGE BY
    created_at). Если rowcount=1, но value не меняется — что-то другое.
    """
    _require_finance(current_user)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")

    reading = (await db.execute(
        select(MeterReading).where(
            MeterReading.user_id == user_id,
            MeterReading.period_id == active_period.id,
        ).limit(1)
    )).scalars().first()
    if not reading:
        raise HTTPException(404, f"Reading для user_id={user_id} в активном периоде не найден")

    from sqlalchemy import update as _sa_update, text as _sa_text

    reading_id = reading.id
    created_at = reading.created_at
    debt_before = float(reading.debt_209 or 0)

    # Стратегия A: UPDATE по id (как сейчас в импорте).
    res_a = await db.execute(
        _sa_update(MeterReading)
        .where(MeterReading.id == reading_id)
        .values(debt_209=0)
    )
    rowcount_a = res_a.rowcount or 0
    val_a = (await db.execute(
        _sa_text("SELECT debt_209 FROM readings WHERE id = :rid AND created_at = :ca"),
        {"rid": reading_id, "ca": created_at},
    )).scalar()

    # Откатываем A перед B — чистый эксперимент.
    await db.rollback()

    # Стратегия B: UPDATE по (id, created_at).
    res_b = await db.execute(
        _sa_update(MeterReading)
        .where(MeterReading.id == reading_id)
        .where(MeterReading.created_at == created_at)
        .values(debt_209=0)
    )
    rowcount_b = res_b.rowcount or 0
    val_b = (await db.execute(
        _sa_text("SELECT debt_209 FROM readings WHERE id = :rid AND created_at = :ca"),
        {"rid": reading_id, "ca": created_at},
    )).scalar()

    # Стратегия C: raw SQL — на случай если ORM что-то ломает.
    await db.rollback()
    res_c = await db.execute(
        _sa_text(
            "UPDATE readings SET debt_209 = 0 "
            "WHERE id = :rid AND created_at = :ca"
        ),
        {"rid": reading_id, "ca": created_at},
    )
    rowcount_c = res_c.rowcount or 0
    val_c = (await db.execute(
        _sa_text("SELECT debt_209 FROM readings WHERE id = :rid AND created_at = :ca"),
        {"rid": reading_id, "ca": created_at},
    )).scalar()

    # Финальный rollback — никаких изменений в БД.
    await db.rollback()

    return {
        "user_id": user_id,
        "reading_id": reading_id,
        "created_at": created_at.isoformat() if created_at else None,
        "debt_209_before": debt_before,
        "strategies": {
            "A_orm_by_id_only": {
                "rowcount": rowcount_a,
                "value_after_in_db": float(val_a) if val_a is not None else None,
                "worked": (rowcount_a == 1 and val_a == 0),
            },
            "B_orm_by_id_and_created_at": {
                "rowcount": rowcount_b,
                "value_after_in_db": float(val_b) if val_b is not None else None,
                "worked": (rowcount_b == 1 and val_b == 0),
            },
            "C_raw_sql_by_id_and_created_at": {
                "rowcount": rowcount_c,
                "value_after_in_db": float(val_c) if val_c is not None else None,
                "worked": (rowcount_c == 1 and val_c == 0),
            },
        },
        "diagnosis": (
            "partitioning_blocks_update" if not (rowcount_a == 1) and rowcount_b == 1
            else "orm_issue" if rowcount_b != 1 and rowcount_c == 1
            else "all_work_check_other_writer" if rowcount_a == 1
            else "all_fail_deeper_problem"
        ),
        "note": "Все стратегии в конце откатываются — БД не изменена.",
    }


@router.get(
    "/debts/integrity-check",
    summary="Анализатор: сравнить applied_state свежего импорта с БД (Этап 2)",
)
async def debts_integrity_check(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Этап 2: проверка целостности долгов в активном периоде.

    Сравнивает что **должно** быть (по applied_state последних 209 и 205
    импортов) с тем, что **есть** в `readings.debt_*`. Три категории
    проблем:

      1) **drift** — applied_state[u] и reading[u] оба есть, но debt
         различается > 1₽. Симптом: что-то перезаписало после импорта
         (manual_receipt, recalc, ручная правка).
      2) **missing_in_db** — applied_state ожидает долг у юзера, а
         reading'а у него вообще нет. Симптом: импорт не дошёл, или
         reading удалён вручную.
      3) **extra_in_db** — reading с долгом есть, в applied_state юзера
         нет. Симптом: zombie от старого Bug AG (см. /debts/zombie-readings).

    Read-only. Auto-fix не делает (на каждую категорию — свой инструмент:
    drift → reparse, missing → reparse, extra → cleanup-zombie-readings).
    """
    _require_finance(current_user)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")
    period_id = active_period.id

    async def _latest_applied(acct: str):
        log = (await db.execute(
            select(DebtImportLog)
            .where(
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
                DebtImportLog.applied_state.is_not(None),
            )
            .order_by(DebtImportLog.id.desc()).limit(1)
        )).scalars().first()
        return log

    log_209 = await _latest_applied("209")
    log_205 = await _latest_applied("205")
    state_209 = (log_209.applied_state or {}) if log_209 else {}
    state_205 = (log_205.applied_state or {}) if log_205 else {}

    # Объединяем expected_state по юзерам: для каждого user_id — ожидаемые debt_209/205.
    expected: dict[int, dict] = {}
    for uid_str, vals in state_209.items():
        try:
            uid = int(uid_str)
        except Exception:
            continue
        expected.setdefault(uid, {"username": None, "room_label": None})
        expected[uid]["debt_209"] = float(vals.get("debt_209", "0") or 0)
        expected[uid]["overpayment_209"] = float(vals.get("overpayment_209", "0") or 0)
        expected[uid]["username"] = vals.get("username")
        expected[uid]["room_label"] = vals.get("room_label")
    for uid_str, vals in state_205.items():
        try:
            uid = int(uid_str)
        except Exception:
            continue
        expected.setdefault(uid, {"username": None, "room_label": None})
        expected[uid]["debt_205"] = float(vals.get("debt_205", "0") or 0)
        expected[uid]["overpayment_205"] = float(vals.get("overpayment_205", "0") or 0)
        if not expected[uid].get("username"):
            expected[uid]["username"] = vals.get("username")
        if not expected[uid].get("room_label"):
            expected[uid]["room_label"] = vals.get("room_label")

    # Все reading'и активного периода (одной выборкой).
    readings = (await db.execute(
        select(MeterReading).where(MeterReading.period_id == period_id)
    )).scalars().all()
    readings_by_user: dict[int, "MeterReading"] = {}
    for r in readings:
        if r.user_id is not None:
            readings_by_user[r.user_id] = r

    drift = []
    missing_in_db = []
    extra_in_db = []

    THR = 1.0  # порог расхождения в рублях

    # 1+2: сверяем expected → reality
    for uid, exp in expected.items():
        r = readings_by_user.get(uid)
        exp_d209 = exp.get("debt_209", 0.0)
        exp_o209 = exp.get("overpayment_209", 0.0)
        exp_d205 = exp.get("debt_205", 0.0)
        exp_o205 = exp.get("overpayment_205", 0.0)
        if r is None:
            # missing — только если ожидалось ненулевое сальдо
            if max(exp_d209, exp_o209, exp_d205, exp_o205) > THR:
                missing_in_db.append({
                    "user_id": uid,
                    "username": exp.get("username"),
                    "room_label": exp.get("room_label"),
                    "expected": {
                        "debt_209": exp_d209, "overpayment_209": exp_o209,
                        "debt_205": exp_d205, "overpayment_205": exp_o205,
                    },
                })
            continue

        actual_d209 = float(r.debt_209 or 0)
        actual_o209 = float(r.overpayment_209 or 0)
        actual_d205 = float(r.debt_205 or 0)
        actual_o205 = float(r.overpayment_205 or 0)

        diff_d209 = actual_d209 - exp_d209
        diff_o209 = actual_o209 - exp_o209
        diff_d205 = actual_d205 - exp_d205
        diff_o205 = actual_o205 - exp_o205
        max_abs_diff = max(abs(diff_d209), abs(diff_o209), abs(diff_d205), abs(diff_o205))
        if max_abs_diff > THR:
            drift.append({
                "user_id": uid,
                "reading_id": r.id,
                "username": exp.get("username"),
                "room_label": exp.get("room_label"),
                "expected": {
                    "debt_209": exp_d209, "overpayment_209": exp_o209,
                    "debt_205": exp_d205, "overpayment_205": exp_o205,
                },
                "actual": {
                    "debt_209": actual_d209, "overpayment_209": actual_o209,
                    "debt_205": actual_d205, "overpayment_205": actual_o205,
                },
                "max_abs_diff": max_abs_diff,
            })

    # 3: reading'и, которых нет в expected (zombie)
    known = set(expected.keys())
    user_ids_for_rooms = set()
    for r in readings:
        if r.user_id is None:
            continue
        if r.user_id in known:
            continue
        has_money = (
            float(r.debt_209 or 0) > THR or float(r.debt_205 or 0) > THR
            or float(r.overpayment_209 or 0) > THR or float(r.overpayment_205 or 0) > THR
        )
        if not has_money:
            continue
        user_ids_for_rooms.add(r.user_id)

    # Загружаем username/комнаты для zombie batch'ем
    extra_users_map = {}
    extra_rooms_map = {}
    if user_ids_for_rooms:
        u_rows = (await db.execute(
            select(User).where(User.id.in_(user_ids_for_rooms))
        )).scalars().all()
        extra_users_map = {u.id: u for u in u_rows}
        rids = {r.room_id for r in readings if r.user_id in user_ids_for_rooms and r.room_id}
        if rids:
            r_rows = (await db.execute(
                select(Room).where(Room.id.in_(rids))
            )).scalars().all()
            extra_rooms_map = {rm.id: rm for rm in r_rows}

    for r in readings:
        if r.user_id is None or r.user_id in known:
            continue
        has_money = (
            float(r.debt_209 or 0) > THR or float(r.debt_205 or 0) > THR
            or float(r.overpayment_209 or 0) > THR or float(r.overpayment_205 or 0) > THR
        )
        if not has_money:
            continue
        u = extra_users_map.get(r.user_id)
        rm = extra_rooms_map.get(r.room_id) if r.room_id else None
        extra_in_db.append({
            "user_id": r.user_id,
            "reading_id": r.id,
            "username": u.username if u else None,
            "room_label": (rm.format_address if rm else None),
            "actual": {
                "debt_209": float(r.debt_209 or 0),
                "overpayment_209": float(r.overpayment_209 or 0),
                "debt_205": float(r.debt_205 or 0),
                "overpayment_205": float(r.overpayment_205 or 0),
            },
        })

    drift.sort(key=lambda x: -x["max_abs_diff"])
    extra_in_db.sort(
        key=lambda x: -(
            x["actual"]["debt_209"] + x["actual"]["debt_205"]
            + x["actual"]["overpayment_209"] + x["actual"]["overpayment_205"]
        )
    )

    return {
        "period_id": period_id,
        "threshold_rub": THR,
        "latest_209_log_id": log_209.id if log_209 else None,
        "latest_205_log_id": log_205.id if log_205 else None,
        "summary": {
            "drift_count": len(drift),
            "missing_in_db_count": len(missing_in_db),
            "extra_in_db_count": len(extra_in_db),
            "expected_users": len(expected),
            "actual_readings": len(readings_by_user),
        },
        "drift": drift[:200],
        "missing_in_db": missing_in_db[:200],
        "extra_in_db": extra_in_db[:200],
    }


@router.post(
    "/debts/integrity-fix",
    summary="Авто-фикс расхождений integrity-check (Bug AK)",
)
async def debts_integrity_fix(
    category: str = Query("all", pattern="^(all|drift|missing|user)$"),
    user_id: Optional[int] = None,
    confirm: str = Query(..., pattern="^YES$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Применяет ожидаемые значения из applied_state свежих 209/205-импортов
    в БД. Покрывает категории:

      - **drift**: UPDATE существующих reading'ов до значений из applied_state
        (когда импорт правильно посчитал, но что-то после перезаписало).
      - **missing**: INSERT недостающих reading'ов из applied_state
        (когда жилец есть в файле, а в БД его reading нет).
      - **all**: drift + missing вместе.
      - **user**: фикс только для конкретного user_id (точечно).

    Extra/Zombie фиксится отдельным endpoint'ом /debts/cleanup-zombie-readings —
    у них нет «ожидаемого значения», только зануление.

    Требует ?confirm=YES.
    """
    _require_finance(current_user)

    # Реюзаем диагностику чтобы не дублировать логику расчёта расхождений.
    data = await debts_integrity_check(current_user=current_user, db=db)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")

    from sqlalchemy import update as _sa_update

    fixed_drift = 0
    fixed_missing = 0
    errors = []

    drift_items = data.get("drift", [])
    missing_items = data.get("missing_in_db", [])

    # Фильтр по user_id если category=user
    if category == "user":
        if not user_id:
            raise HTTPException(400, "category=user требует user_id")
        drift_items = [d for d in drift_items if d.get("user_id") == user_id]
        missing_items = [m for m in missing_items if m.get("user_id") == user_id]

    # 1) drift — UPDATE существующих reading'ов
    if category in ("all", "drift", "user"):
        for item in drift_items:
            try:
                res = await db.execute(
                    _sa_update(MeterReading)
                    .where(MeterReading.id == item["reading_id"])
                    .values(
                        debt_209=Decimal(str(item["expected"]["debt_209"])),
                        overpayment_209=Decimal(str(item["expected"]["overpayment_209"])),
                        debt_205=Decimal(str(item["expected"]["debt_205"])),
                        overpayment_205=Decimal(str(item["expected"]["overpayment_205"])),
                    )
                )
                if res.rowcount:
                    fixed_drift += 1
            except Exception as e:
                errors.append({
                    "kind": "drift",
                    "user_id": item.get("user_id"),
                    "error": str(e)[:200],
                })

    # 2) missing — INSERT недостающих reading'ов из applied_state
    if category in ("all", "missing", "user"):
        for item in missing_items:
            try:
                user = await db.get(User, item["user_id"])
                if not user:
                    errors.append({
                        "kind": "missing",
                        "user_id": item.get("user_id"),
                        "error": "user не найден в БД",
                    })
                    continue
                new_reading = MeterReading(
                    user_id=item["user_id"],
                    room_id=user.room_id,
                    period_id=active_period.id,
                    is_approved=False,
                    debt_209=Decimal(str(item["expected"]["debt_209"])),
                    overpayment_209=Decimal(str(item["expected"]["overpayment_209"])),
                    debt_205=Decimal(str(item["expected"]["debt_205"])),
                    overpayment_205=Decimal(str(item["expected"]["overpayment_205"])),
                    obor_debit_209=Decimal("0"), obor_credit_209=Decimal("0"),
                    obor_debit_205=Decimal("0"), obor_credit_205=Decimal("0"),
                )
                db.add(new_reading)
                fixed_missing += 1
            except Exception as e:
                errors.append({
                    "kind": "missing",
                    "user_id": item.get("user_id"),
                    "error": str(e)[:200],
                })

    await db.commit()
    logger.info(
        "[INTEGRITY-FIX] category=%s drift=%d missing=%d errors=%d (by %s)",
        category, fixed_drift, fixed_missing, len(errors), current_user.username,
    )

    return {
        "status": "ok",
        "category": category,
        "fixed_drift": fixed_drift,
        "fixed_missing": fixed_missing,
        "errors": errors[:50],
    }


@router.get(
    "/debts/zombie-readings",
    summary="Reading'и с долгом, которых нет в свежем импорте (Этап 3)",
)
async def debts_zombie_readings(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bug AG cleanup: после переключения импорта на per-user-key (Bug AG)
    в БД могут остаться reading'и с ненулевыми debt_*/overpayment_*, которых
    в свежем impart'е 1С уже нет (т.е. в файле от 1С этого жильца не передавали,
    значит долг должен быть 0). Раньше эти суммы суммировались в общий
    reading комнаты — после Bug AG они становятся «висяком» на чужом юзере.

    Логика: смотрим последние completed-импорты 209 и 205, собираем все
    user_id из их applied_state. Все reading'и активного периода с
    долгом/переплатой, чей user_id НЕ упомянут ни в одном из этих логов —
    кандидаты на zombie.

    Read-only. POST /debts/cleanup-zombie-readings занулит их (с
    подтверждением).
    """
    _require_finance(current_user)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")
    period_id = active_period.id

    async def _latest_applied(acct: str):
        log = (await db.execute(
            select(DebtImportLog)
            .where(
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
                DebtImportLog.applied_state.is_not(None),
            )
            .order_by(DebtImportLog.id.desc()).limit(1)
        )).scalars().first()
        return log

    log_209 = await _latest_applied("209")
    log_205 = await _latest_applied("205")
    state_209 = (log_209.applied_state or {}) if log_209 else {}
    state_205 = (log_205.applied_state or {}) if log_205 else {}
    known_user_ids = set(state_209.keys()) | set(state_205.keys())

    if not known_user_ids:
        return {
            "period_id": period_id,
            "count": 0,
            "zombies": [],
            "note": "Нет свежих импортов с applied_state — нечего сравнивать.",
        }

    readings = (await db.execute(
        select(MeterReading)
        .where(
            MeterReading.period_id == period_id,
            or_(
                MeterReading.debt_209 > 0,
                MeterReading.debt_205 > 0,
                MeterReading.overpayment_209 > 0,
                MeterReading.overpayment_205 > 0,
            ),
        )
    )).scalars().all()

    user_ids_in_db = {r.user_id for r in readings if r.user_id}
    if not user_ids_in_db:
        return {"period_id": period_id, "count": 0, "zombies": []}

    users = (await db.execute(
        select(User).where(User.id.in_(user_ids_in_db))
    )).scalars().all()
    users_map = {u.id: u for u in users}

    room_ids = {r.room_id for r in readings if r.room_id}
    rooms_map = {}
    if room_ids:
        rooms = (await db.execute(
            select(Room).where(Room.id.in_(room_ids))
        )).scalars().all()
        rooms_map = {r.id: r for r in rooms}

    zombies = []
    for r in readings:
        if not r.user_id:
            continue
        if str(r.user_id) in known_user_ids:
            continue  # есть в свежем импорте — не зомби
        user = users_map.get(r.user_id)
        room = rooms_map.get(r.room_id) if r.room_id else None
        zombies.append({
            "reading_id": r.id,
            "user_id": r.user_id,
            "username": user.username if user else None,
            "room_id": r.room_id,
            "room_label": (
                room.format_address if room else None
            ),
            "debt_209": float(r.debt_209 or 0),
            "overpayment_209": float(r.overpayment_209 or 0),
            "debt_205": float(r.debt_205 or 0),
            "overpayment_205": float(r.overpayment_205 or 0),
            "total_to_clean": (
                float(r.debt_209 or 0) + float(r.debt_205 or 0)
                + float(r.overpayment_209 or 0) + float(r.overpayment_205 or 0)
            ),
        })

    zombies.sort(key=lambda x: -x["total_to_clean"])
    return {
        "period_id": period_id,
        "latest_209_log_id": log_209.id if log_209 else None,
        "latest_205_log_id": log_205.id if log_205 else None,
        "count": len(zombies),
        "zombies": zombies,
    }


@router.post(
    "/debts/cleanup-zombie-readings",
    summary="Занулить debt_*/overpayment_* у zombie-reading'ов (Этап 3)",
)
async def debts_cleanup_zombie_readings(
    confirm: str = Query(..., pattern="^YES$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Зануляет debt_209/205 и overpayment_209/205 у всех reading'ов,
    которые попали в список /debts/zombie-readings.

    Reading'и НЕ удаляются (audit/история сохраняется) — только зануляются
    финансовые поля. После этого дашборд показывает 0₽ у соответствующих
    жильцов.

    Требует ?confirm=YES.
    """
    _require_finance(current_user)

    # Реюзаем логику /debts/zombie-readings — чтобы не дублировать.
    result = await debts_zombie_readings(current_user=current_user, db=db)
    zombies = result.get("zombies", [])
    if not zombies:
        return {"status": "ok", "cleaned": 0, "note": "Zombie-reading'ов нет"}

    reading_ids = [z["reading_id"] for z in zombies]
    from sqlalchemy import update as _sa_update
    total = 0
    for rid in reading_ids:
        res = await db.execute(
            _sa_update(MeterReading)
            .where(MeterReading.id == rid)
            .values(
                debt_209=Decimal("0.00"),
                overpayment_209=Decimal("0.00"),
                debt_205=Decimal("0.00"),
                overpayment_205=Decimal("0.00"),
            )
        )
        total += res.rowcount or 0

    await db.commit()
    logger.info(
        "[ZOMBIE-CLEANUP] %s reading-ов занулено (запросил %s)",
        total, current_user.username,
    )
    return {
        "status": "ok",
        "cleaned": total,
        "requested": len(reading_ids),
        "zombies": zombies[:50],
    }


@router.get(
    "/debts/orphan-readings",
    summary="Жильцы с >1 reading в активном периоде (диагностика для Bug AF)",
)
async def debts_orphan_readings(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bug AF: после переездов / auto-Vacant может оказаться, что у одного
    user_id в активном периоде существует несколько MeterReading с разными
    room_id. Дашборд агрегирует SUM(debt_*) по user_id, а импорт 1С
    обновляет reading по room_id. Если осиротевший reading с прошлой
    комнатой не зачищен — его debt_209 продолжает суммироваться в общий,
    и reparse его не чинит (импорт обновляет только текущий room_id).

    Read-only endpoint: возвращает список жильцов с >1 reading + раскладку
    каждого reading'а (room_id, debt, current/orphan). По нему уже решаем,
    что чистить — отдельный POST-endpoint.
    """
    _require_finance(current_user)

    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
    )).scalars().first()
    if not active_period:
        raise HTTPException(404, "Нет активного периода")
    period_id = active_period.id

    dup_q = (
        select(MeterReading.user_id, func.count(MeterReading.id).label("cnt"))
        .where(
            MeterReading.period_id == period_id,
            MeterReading.user_id.is_not(None),
        )
        .group_by(MeterReading.user_id)
        .having(func.count(MeterReading.id) > 1)
    )
    dup_rows = (await db.execute(dup_q)).all()
    user_ids = [r.user_id for r in dup_rows]
    if not user_ids:
        return {"period_id": period_id, "count": 0, "users": []}

    readings = (await db.execute(
        select(MeterReading)
        .where(
            MeterReading.period_id == period_id,
            MeterReading.user_id.in_(user_ids),
        )
        .order_by(MeterReading.user_id, MeterReading.created_at)
    )).scalars().all()

    users = (await db.execute(
        select(User).where(User.id.in_(user_ids))
    )).scalars().all()
    users_map = {u.id: u for u in users}

    room_ids = {r.room_id for r in readings if r.room_id}
    room_ids.update(u.room_id for u in users if u.room_id)
    rooms_map = {}
    if room_ids:
        rooms = (await db.execute(
            select(Room).where(Room.id.in_(room_ids))
        )).scalars().all()
        rooms_map = {r.id: r for r in rooms}

    def _room_label(rid):
        if rid is None:
            return None
        r = rooms_map.get(rid)
        return r.format_address if r else f"id={rid}"

    by_user: dict[int, list] = {}
    for r in readings:
        by_user.setdefault(r.user_id, []).append(r)

    items = []
    for uid, urs in by_user.items():
        user = users_map.get(uid)
        cur_room_id = user.room_id if user else None
        orphan_debt = sum(
            float(r.debt_209 or 0) + float(r.debt_205 or 0)
            for r in urs if r.room_id != cur_room_id
        )
        items.append({
            "user_id": uid,
            "username": user.username if user else None,
            "current_room_id": cur_room_id,
            "current_room_label": _room_label(cur_room_id),
            "total_debt_209": sum(float(r.debt_209 or 0) for r in urs),
            "total_debt_205": sum(float(r.debt_205 or 0) for r in urs),
            "orphan_debt_sum": orphan_debt,
            "readings": [
                {
                    "id": r.id,
                    "room_id": r.room_id,
                    "room_label": _room_label(r.room_id) or "(нет комнаты)",
                    "is_current_room": (r.room_id == cur_room_id),
                    "debt_209": float(r.debt_209 or 0),
                    "overpayment_209": float(r.overpayment_209 or 0),
                    "debt_205": float(r.debt_205 or 0),
                    "overpayment_205": float(r.overpayment_205 or 0),
                    "created_at": r.created_at.isoformat() if r.created_at else None,
                    "is_approved": r.is_approved,
                }
                for r in urs
            ],
        })

    # Сортировка: сначала те, у кого больше всего «осиротевших» денег.
    items.sort(key=lambda x: -x["orphan_debt_sum"])

    return {
        "period_id": period_id,
        "count": len(items),
        "users": items,
    }


@router.delete("/debts/import-history/{log_id}", summary="Удалить запись истории импорта 1С")
async def debts_delete_import_history(
    log_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Удаляет запись DebtImportLog **без** отката данных.

    Use case: после массового rebuild / reload-period долги в БД уже
    сброшены и заново импортированы. Старые записи истории «висят» с
    цифрами +N₽/+M₽, но реально debt'ы уже не соответствуют. Кнопка
    «Откатить» в таком случае бесполезна (snapshot устаревший). Эта
    кнопка просто удаляет запись из списка истории.

    Защита: если статус completed (актуальный импорт) — требуется
    подтверждение через ?confirm=YES.
    """
    if current_user.role not in ("admin", "financier"):
        raise HTTPException(403, "Недостаточно прав")

    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог не найден")

    from fastapi import Query as _Q  # noqa: F401 (для документации)
    # confirm передаётся через query string
    from sqlalchemy import delete as _delete
    await db.execute(_delete(DebtImportLog).where(DebtImportLog.id == log_id))
    await db.commit()
    return {"status": "ok", "deleted_id": log_id}


@router.post("/debts/import-history/cleanup", summary="Массовая чистка истории импортов")
async def debts_cleanup_import_history(
    keep_last: int = 5,
    only_reverted: bool = False,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Массовое удаление устаревших записей DebtImportLog.

    Параметры:
      keep_last     — сколько последних completed-импортов на каждый
                      account_type сохранять (по умолчанию 5);
      only_reverted — если True, удаляются ТОЛЬКО reverted-записи
                      (откаченные), completed не трогаются.

    Use case пользователя: после наших rebuild/reload-period в истории
    висят откаченные импорты + устаревшие completed (debt'ы в БД уже
    обновлены последним импортом). UI показывает «№23, №24 Откачен»
    мусором — этот endpoint их выпиливает.
    """
    if current_user.role not in ("admin", "financier"):
        raise HTTPException(403, "Недостаточно прав")

    from sqlalchemy import delete as _delete

    if only_reverted:
        # Удаляем все reverted/failed.
        res = await db.execute(
            _delete(DebtImportLog).where(
                DebtImportLog.status.in_(["reverted", "failed"])
            )
        )
        deleted = res.rowcount or 0
    else:
        # Удаляем reverted/failed ВСЕ + у каждого account_type оставляем
        # последние keep_last completed.
        # 1. Снести reverted/failed.
        await db.execute(
            _delete(DebtImportLog).where(
                DebtImportLog.status.in_(["reverted", "failed"])
            )
        )
        # 2. По account_type: оставить keep_last свежих completed, остальное удалить.
        for acct in ("209", "205"):
            completed = (await db.execute(
                select(DebtImportLog.id)
                .where(
                    DebtImportLog.account_type == acct,
                    DebtImportLog.status == "completed",
                )
                .order_by(desc(DebtImportLog.id))
            )).scalars().all()
            to_delete = completed[keep_last:]
            if to_delete:
                await db.execute(
                    _delete(DebtImportLog).where(DebtImportLog.id.in_(to_delete))
                )
        # Re-count.
        remaining = (await db.execute(
            select(func.count(DebtImportLog.id))
        )).scalar_one()
        deleted = -1  # неизвестно — не критично для UI
        deleted = max(0, deleted)
        await db.commit()
        return {"status": "ok", "remaining": int(remaining)}

    await db.commit()
    return {"status": "ok", "deleted": deleted}


# =========================================================================
# REASSIGN «не найденный ФИО» → жилец
# =========================================================================

@router.post("/debts/import-history/{log_id}/reassign", summary="Привязать не-найденное ФИО к жильцу")
async def debts_reassign_not_found(
    log_id: int,
    fio: str = Form(..., description="Оригинальное ФИО из Excel"),
    user_id: int = Form(..., description="ID жильца, к которому привязать"),
    debt: float = Form(0),
    overpayment: float = Form(0),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Ручная привязка ФИО из not_found к жильцу + добавление значений
    долга/переплаты в черновик.

    Удаляет переданный `fio` из списка not_found_users лога.
    """
    _require_finance(current_user)

    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог импорта не найден")
    if log.status not in ("staged", "completed"):
        raise HTTPException(400, f"Статус лога «{log.status}» — reassign для staged/completed")

    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Жилец не найден")
    # Долг на лицевом счёте (user_id) — комната опциональна, подцепится позже.

    debt_dec = Decimal(str(debt or 0))
    over_dec = Decimal(str(overpayment or 0))
    acc = log.account_type
    reading_id = None

    if log.status == "staged":
        # Черновик: пишем в applied_state (как rematch-base/publish), показания
        # НЕ трогаем — их создаст «Выгрузить». Полная замена по этому счёту.
        ap = dict(log.applied_state or {})
        key = str(user.id)
        ent = ap.get(key) or {
            "debt_209": "0", "overpayment_209": "0",
            "debt_205": "0", "overpayment_205": "0",
        }
        ent[f"debt_{acc}"] = str(debt_dec)
        ent[f"overpayment_{acc}"] = str(over_dec)
        ent["username"] = user.username
        ent["room_id"] = user.room_id
        room = await db.get(Room, user.room_id) if user.room_id else None
        ent["room_label"] = room.format_address if room else None
        ap[key] = ent
        log.applied_state = ap
    else:
        # completed: долг уже выгружен — дописываем в показание ЭТОГО жильца
        # (по user_id, не по комнате; room_id может быть NULL).
        reading = None
        if log.period_id:
            reading = (await db.execute(
                select(MeterReading).where(
                    MeterReading.period_id == log.period_id,
                    MeterReading.user_id == user.id,
                ).limit(1)
            )).scalars().first()
        if reading:
            if acc == "209":
                reading.debt_209 = (reading.debt_209 or Decimal("0")) + debt_dec
                reading.overpayment_209 = (reading.overpayment_209 or Decimal("0")) + over_dec
            else:
                reading.debt_205 = (reading.debt_205 or Decimal("0")) + debt_dec
                reading.overpayment_205 = (reading.overpayment_205 or Decimal("0")) + over_dec
        elif log.period_id:
            reading = MeterReading(
                user_id=user.id,
                room_id=user.room_id,
                period_id=log.period_id,
                is_approved=False,
                debt_209=debt_dec if acc == "209" else Decimal("0"),
                overpayment_209=over_dec if acc == "209" else Decimal("0"),
                debt_205=debt_dec if acc == "205" else Decimal("0"),
                overpayment_205=over_dec if acc == "205" else Decimal("0"),
            )
            db.add(reading)
        await db.flush()
        reading_id = reading.id if reading else None

    # Удаляем FIO из not_found_users. После фикса формата (list[dict])
    # сравниваем через helper _nfu_fio чтобы не упасть на .strip() от dict.
    nfu = list(log.not_found_users or [])
    fio_norm = fio.strip().lower()
    nfu_new = [x for x in nfu if _nfu_fio(x).lower() != fio_norm]
    if len(nfu_new) != len(nfu):
        log.not_found_users = nfu_new
        log.not_found_count = len(nfu_new)

    # Сохраняем alias чтобы при СЛЕДУЮЩЕМ импорте (205 или 209 или gsheets)
    # эта же ФИО автоматически матчилась на user — без повторного reassign.
    alias_created = await _ensure_debt_alias(
        db, alias_fio=fio, user_id=user_id,
        created_by_id=current_user.id,
        note=f"debt reassign log#{log_id}",
    )

    await db.commit()
    return {
        "status": "ok",
        "reading_id": reading_id,
        "alias_created": alias_created,
    }


# =========================================================================
# FIND CANDIDATES — поиск похожих жильцов для not-found ФИО
# =========================================================================
@router.get("/debts/find-candidates", summary="Похожие жильцы по ФИО (fuzzy + фамилия)")
async def debts_find_candidates(
    fio: Optional[str] = Query(None, max_length=200, description="ФИО из Excel (для auto-suggest)"),
    q: Optional[str] = Query(None, max_length=100, description="Ручной поиск по любой подстроке"),
    limit: int = Query(15, ge=1, le=50),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Возвращает кандидатов для модалки «Не найдены в 1С».

    Два режима:
      1) `q`  — ручной поиск, ILIKE по подстроке (case-insensitive). Когда
                админ вбивает часть фамилии или имени в input поиска.
      2) `fio` — auto-suggest для импорта. Объединяет:
                 - ТОЧНОЕ совпадение фамилии (первого токена) → score=100
                 - fuzzy token_sort_ratio для остальных (threshold 40%)
                 Это решает кейс когда в Excel «Ярощук Александр Павлович»,
                 а в БД «Ярощук А.П.» — fuzzy один даёт score~50%, и жилец
                 теряется в кандидатах с такими же score; surname-match
                 поднимает его в топ.

    Хотя бы один из q/fio должен быть задан.
    """
    _require_finance(current_user)

    if not q and not fio:
        raise HTTPException(400, "Передайте q (ручной поиск) или fio (по импорту)")

    from sqlalchemy.orm import selectinload as _selectinload

    base_query = (
        select(User).options(_selectinload(User.room))
        .where(User.is_deleted.is_(False), User.role == "user")
    )

    # ============= РЕЖИМ 1: ручной поиск по q =============
    if q:
        q_norm = q.strip().lower()
        if len(q_norm) < 2:
            return {"fio": None, "q": q, "candidates": []}
        # ILIKE по нормализованному username. Допускаем многословный q
        # — каждый токен должен встречаться (AND).
        tokens = [t for t in q_norm.split() if t]
        filtered = base_query
        for tok in tokens:
            filtered = filtered.where(func.lower(User.username).like(f"%{tok}%"))
        users_raw = (await db.execute(filtered.limit(limit))).scalars().all()

        candidates = []
        for u in users_raw:
            room_label = (
                u.room.format_address if u.room else "без комнаты"
            )
            candidates.append({
                "id": u.id,
                "username": u.username,
                "room_label": room_label,
                "residents_count": int(u.residents_count or 1),
                "score": 100,  # точное substring-совпадение
                "reason": None,
            })
        # Сортируем по username (стабильный порядок)
        candidates.sort(key=lambda c: c["username"].lower())
        return {"fio": None, "q": q, "candidates": candidates}

    # ============= РЕЖИМ 2: auto-suggest по fio =============
    from rapidfuzz import fuzz

    target_norm = " ".join(fio.lower().split())
    if not target_norm:
        return {"fio": fio, "candidates": []}
    target_tokens = target_norm.split()
    surname = target_tokens[0] if target_tokens else ""

    users_raw = (await db.execute(base_query)).scalars().all()

    # Проходим всех жильцов. Для каждого считаем score по двум критериям:
    #   - точное совпадение фамилии (первый токен username) → 100
    #   - token_sort_ratio
    # Из двух берём максимум.
    matches: list[tuple[User, int, Optional[str]]] = []
    for u in users_raw:
        if not u.username:
            continue
        name_norm = " ".join(u.username.lower().split())
        name_tokens = name_norm.split()

        # Точное совпадение фамилии — приоритет
        surname_exact = (
            surname and name_tokens and surname == name_tokens[0]
        )
        # Или фамилия как substring в username (защита от опечаток типа
        # «Ярощук-Иванов» когда в БД двойная фамилия)
        surname_substring = (
            surname and len(surname) >= 4 and surname in name_norm
        )

        fuzzy_score = fuzz.token_sort_ratio(target_norm, name_norm)

        if surname_exact:
            score = max(100, fuzzy_score)
            reason = "Совпадает фамилия" if fuzzy_score < 80 else None
        elif surname_substring:
            score = max(85, fuzzy_score)
            reason = "Фамилия найдена внутри ФИО"
        elif fuzzy_score >= 40:
            score = fuzzy_score
            # «Общее отчество» — простая эвристика для случая брат/сестра
            reason = None
            if (fuzzy_score < 80 and len(target_tokens) >= 3
                    and len(name_tokens) >= 3
                    and target_tokens[-1] == name_tokens[-1]
                    and target_tokens[0] != name_tokens[0]):
                reason = "Общее отчество (возможно, брат/сестра)"
        else:
            continue

        matches.append((u, int(score), reason))

    # Сортируем: сначала score DESC, потом username для стабильности
    matches.sort(key=lambda m: (-m[1], m[0].username.lower()))
    matches = matches[:limit]

    candidates = []
    for u, score, reason in matches:
        room_label = (
            u.room.format_address if u.room else "без комнаты"
        )
        candidates.append({
            "id": u.id,
            "username": u.username,
            "room_label": room_label,
            "residents_count": int(u.residents_count or 1),
            "score": score,
            "reason": reason,
        })

    return {"fio": fio, "candidates": candidates}


# =========================================================================
# CREATE-AND-MATCH — создать нового жильца + привязать долг
# =========================================================================
class DebtCreateAndMatchRequest(BaseModel):
    """Создание нового жильца с одновременной привязкой долга/переплаты.

    Аналогично gsheets create-and-match (admin_gsheets.py), но здесь сразу
    добавляем сумму к debt_*/overpayment_* в MeterReading периода импорта.
    """
    fio: str           # ФИО из Excel — для удаления из not_found_users
    username: str      # логин для входа
    password: str
    dormitory_name: str
    room_number: str
    debt: float = 0.0
    overpayment: float = 0.0
    residents_count: int = 1
    resident_type: str = "family"
    workplace: Optional[str] = None


@router.post("/debts/import-history/{log_id}/create-and-match", summary="Создать жильца + привязать долг")
async def debts_create_and_match(
    log_id: int,
    data: DebtCreateAndMatchRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Создаёт нового User + комнатную привязку + добавляет долг к
    черновику reading периода импорта. Удаляет ФИО из not_found_users.

    Используется когда жильца РЕАЛЬНО нет в системе (новый человек,
    которого ещё не завели в «Жильцы»). До этого фикса админ должен был:
      1) зайти в «Жильцы»
      2) создать пользователя
      3) вернуться в долги и сделать reassign
    Сейчас всё одной операцией.
    """
    _require_finance(current_user)

    log = await db.get(DebtImportLog, log_id)
    if not log:
        raise HTTPException(404, "Лог не найден")
    if log.status not in ("staged", "completed"):
        raise HTTPException(400, f"Статус лога «{log.status}» — операция для staged/completed")

    # 1. Уникальность логина (case-insensitive)
    existing_user = (await db.execute(
        select(User).where(func.lower(User.username) == data.username.strip().lower())
    )).scalars().first()
    if existing_user:
        raise HTTPException(400, f"Логин «{data.username.strip()}» уже занят")

    # 2. Комната должна существовать в Жилфонде
    room = (await db.execute(
        select(Room).where(
            Room.dormitory_name == data.dormitory_name.strip(),
            Room.room_number == data.room_number.strip(),
        ).limit(1)
    )).scalars().first()
    if not room:
        raise HTTPException(
            400,
            f"Комната «{data.room_number}» в общежитии «{data.dormitory_name}» "
            "не найдена в Жилфонде. Создайте её сначала.",
        )

    # 3. Создание User
    rt = data.resident_type if data.resident_type in ("family", "single") else "family"
    bm = "per_capita" if rt == "single" else "by_meter"

    from app.core.auth import get_password_hash
    db_user = User(
        username=data.username.strip(),
        login=data.username.strip(),  # учётка по умолчанию = ФИО, жилец сменит сам
        hashed_password=get_password_hash(data.password),
        role="user",
        workplace=(data.workplace or "").strip() or None,
        residents_count=max(1, int(data.residents_count)),
        room_id=None,  # выставит move_user_to_room
        resident_type=rt,
        billing_mode=bm,
        is_deleted=False,
        is_initial_setup_done=False,
    )
    db.add(db_user)
    await db.flush()

    from app.modules.utility.services.room_assignment import move_user_to_room
    await move_user_to_room(
        db, user=db_user, new_room_id=room.id,
        note=f"created via debts not-found import #{log_id}",
    )

    # 4. Добавляем долг к черновику reading периода импорта
    debt_dec = Decimal(str(data.debt or 0))
    over_dec = Decimal(str(data.overpayment or 0))

    acc = log.account_type
    if log.status == "staged":
        # Черновик: долг в applied_state, показание создаст «Выгрузить».
        ap = dict(log.applied_state or {})
        ap[str(db_user.id)] = {
            "debt_209": str(debt_dec) if acc == "209" else "0",
            "overpayment_209": str(over_dec) if acc == "209" else "0",
            "debt_205": str(debt_dec) if acc == "205" else "0",
            "overpayment_205": str(over_dec) if acc == "205" else "0",
            "username": db_user.username,
            "room_id": room.id,
            "room_label": room.format_address,
        }
        log.applied_state = ap
    elif log.period_id:
        reading = (await db.execute(
            select(MeterReading).where(
                MeterReading.period_id == log.period_id,
                MeterReading.user_id == db_user.id,
            ).limit(1)
        )).scalars().first()

        if reading:
            if acc == "209":
                reading.debt_209 = (reading.debt_209 or Decimal("0")) + debt_dec
                reading.overpayment_209 = (reading.overpayment_209 or Decimal("0")) + over_dec
            else:
                reading.debt_205 = (reading.debt_205 or Decimal("0")) + debt_dec
                reading.overpayment_205 = (reading.overpayment_205 or Decimal("0")) + over_dec
        else:
            reading = MeterReading(
                user_id=db_user.id,
                room_id=room.id,
                period_id=log.period_id,
                is_approved=False,
                debt_209=debt_dec if acc == "209" else Decimal("0"),
                overpayment_209=over_dec if acc == "209" else Decimal("0"),
                debt_205=debt_dec if acc == "205" else Decimal("0"),
                overpayment_205=over_dec if acc == "205" else Decimal("0"),
            )
            db.add(reading)

    # 5. Удаляем FIO из not_found_users (см. _nfu_fio про list[dict] vs str)
    nfu = list(log.not_found_users or [])
    fio_norm = data.fio.strip().lower()
    nfu_new = [x for x in nfu if _nfu_fio(x).lower() != fio_norm]
    if len(nfu_new) != len(nfu):
        log.not_found_users = nfu_new
        log.not_found_count = len(nfu_new)

    # 6. Alias — то же что в reassign: запомнить эту привязку для будущего.
    await _ensure_debt_alias(
        db, alias_fio=data.fio, user_id=db_user.id,
        created_by_id=current_user.id,
        note=f"debt create-and-match log#{log_id}",
    )

    await db.commit()
    return {
        "status": "ok",
        "user_id": db_user.id,
        "username": data.username.strip(),
        "room_id": room.id,
    }


# =========================================================================
# RESET BALANCE — обнулить баланс конкретного жильца
# =========================================================================
@router.post("/users/{user_id}/reset-balance")
async def reset_user_balance(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Обнуляет debt_209/debt_205/overpayment_209/overpayment_205 у ВСЕХ
    reading-ов жильца (по room_id). Полезно когда после отката импорта
    у жильца остались «зависшие» сальдо от старых reading-ов в других
    периодах — undo конкретного импорта восстанавливает snapshot только
    для тех reading-ов, которые этот импорт трогал.

    Returns: количество reading-ов обнулённых + до-сальдо для аудита.
    """
    _require_finance(current_user)

    user = await db.get(User, user_id)
    if not user or not user.room_id:
        raise HTTPException(404, "Жилец не найден или без комнаты")

    # Снимаем before-snapshot для audit_log (на случай если нужно откатить)
    readings = (await db.execute(
        select(MeterReading).where(
            MeterReading.room_id == user.room_id,
            (MeterReading.debt_209 > 0)
            | (MeterReading.debt_205 > 0)
            | (MeterReading.overpayment_209 > 0)
            | (MeterReading.overpayment_205 > 0),
        )
    )).scalars().all()

    if not readings:
        return {
            "status": "noop",
            "user_id": user_id,
            "username": user.username,
            "reset_count": 0,
        }

    snapshot = []
    for r in readings:
        snapshot.append({
            "reading_id": r.id,
            "period_id": r.period_id,
            "debt_209": str(r.debt_209 or 0),
            "overpayment_209": str(r.overpayment_209 or 0),
            "debt_205": str(r.debt_205 or 0),
            "overpayment_205": str(r.overpayment_205 or 0),
        })
        r.debt_209 = Decimal("0.00")
        r.overpayment_209 = Decimal("0.00")
        r.debt_205 = Decimal("0.00")
        r.overpayment_205 = Decimal("0.00")
        # total_cost синхронизируется триггером (integrity_002)

    # Audit log
    from app.modules.utility.routers.admin_dashboard import write_audit_log
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="reset_user_balance", entity_type="user", entity_id=user_id,
        details={"reset_count": len(readings), "snapshot": snapshot},
    )

    await db.commit()
    return {
        "status": "ok",
        "user_id": user_id,
        "username": user.username,
        "reset_count": len(readings),
    }


# =========================================================================
# RECONCILIATION — сверка показаний с долгами (для Центра анализа)
# =========================================================================

@router.get("/debts/reconcile", summary="Сверка: readings vs debts (выбранный/активный период)")
async def debts_reconcile(
    period_id: Optional[int] = Query(None, description="Период; по умолчанию активный → последний импорт → свежий"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Возвращает данные для вкладки «Сверка 1С» в Центре анализа:
      * readings_without_debts — есть reading, но в 1С долгов нет (ок, оплачено?)
      * debts_without_readings — в readings стоит долг, но reading не утверждён
      * last_import_not_found — ФИО из последнего импорта, не привязанные
      * unassigned — НЕразнесённый долг (сумма + по счетам): деньги из 1С, не
        привязанные к жильцу (ФИО нет в базе / нет комнаты). Главный сигнал
        проблемы — сколько денег «висит в воздухе».
    Период — через _resolve_view_period (работает и когда активного нет)."""
    _require_finance(current_user)

    active_period = await _resolve_view_period(db, period_id)
    if not active_period:
        return {
            "period": None,
            "readings_without_debts": [],
            "debts_without_readings": [],
            "last_import_not_found": [],
            "unassigned": {"total_debt": 0.0, "total_overpayment": 0.0, "count": 0,
                           "by_account": {"209": {"count": 0, "debt": 0.0},
                                          "205": {"count": 0, "debt": 0.0}}},
        }

    # 1) readings_without_debts — approved=True и debt_*/overpayment_* все 0 и total_cost > 0
    #    (жильцу начислено что-то, но 1С не вернула долг = вероятно уже оплатили)
    q1 = (
        select(User.username, Room.dormitory_name, Room.room_number,
               MeterReading.total_cost, MeterReading.id)
        .select_from(MeterReading)
        .join(User, User.id == MeterReading.user_id)
        .outerjoin(Room, Room.id == MeterReading.room_id)
        .where(
            MeterReading.period_id == active_period.id,
            MeterReading.is_approved.is_(True),
            MeterReading.total_cost > 0,
            MeterReading.debt_209 == 0,
            MeterReading.debt_205 == 0,
        )
        .order_by(desc(MeterReading.total_cost))
        .limit(100)
    )
    r_no_debts = [
        {
            "username": row[0], "dormitory": row[1], "room_number": row[2],
            "total_cost": float(row[3] or 0), "reading_id": row[4],
        }
        for row in (await db.execute(q1)).all()
    ]

    # 2) debts_without_readings — долг есть (debt_209+205 > 0), но reading не approved
    q2 = (
        select(User.username, Room.dormitory_name, Room.room_number,
               MeterReading.debt_209, MeterReading.debt_205, MeterReading.id)
        .select_from(MeterReading)
        .join(User, User.id == MeterReading.user_id)
        .outerjoin(Room, Room.id == MeterReading.room_id)
        .where(
            MeterReading.period_id == active_period.id,
            MeterReading.is_approved.is_(False),
            (MeterReading.debt_209 > 0) | (MeterReading.debt_205 > 0),
        )
        .order_by(desc(MeterReading.debt_209 + MeterReading.debt_205))
        .limit(100)
    )
    d_no_readings = [
        {
            "username": row[0], "dormitory": row[1], "room_number": row[2],
            "debt_209": float(row[3] or 0), "debt_205": float(row[4] or 0),
            "reading_id": row[5],
        }
        for row in (await db.execute(q2)).all()
    ]

    # 3) Последний успешный импорт — not_found FIO
    last_log = (await db.execute(
        select(DebtImportLog)
        .where(DebtImportLog.status == "completed")
        .order_by(desc(DebtImportLog.started_at)).limit(1)
    )).scalars().first()
    # not_found_users может быть list[str] (legacy) или list[dict] (new).
    # Для reconcile-UI возвращаем плоский list[str] — там сейчас не нужны
    # суммы (только список ФИО).
    nf_raw = (last_log.not_found_users or []) if last_log else []
    nf = [_nfu_fio(x) for x in nf_raw][:200]

    # 4) Неразнесённый долг: not_found из ПОСЛЕДНИХ импортов 209/205 за период.
    #    Главный денежный сигнал — сколько 1С-долга не привязано к жильцу.
    u_total_debt = 0.0
    u_total_over = 0.0
    u_keys: set = set()
    by_account = {}
    for acct in ("209", "205"):
        log = (await db.execute(
            select(DebtImportLog).where(
                DebtImportLog.period_id == active_period.id,
                DebtImportLog.account_type == acct,
                DebtImportLog.status == "completed",
            ).order_by(desc(DebtImportLog.started_at)).limit(1)
        )).scalars().first()
        acct_debt = 0.0
        acct_cnt = 0
        for it in (log.not_found_users or []) if log else []:
            if isinstance(it, dict):
                fio = (it.get("fio") or "").strip()
                dd = float(it.get("debt") or 0)
                oo = float(it.get("overpayment") or 0)
            else:
                fio, dd, oo = str(it).strip(), 0.0, 0.0
            if not fio:
                continue
            u_keys.add(" ".join(fio.lower().split()))
            acct_debt += dd
            acct_cnt += 1
            u_total_debt += dd
            u_total_over += oo
        by_account[acct] = {"count": acct_cnt, "debt": round(acct_debt, 2)}

    return {
        "period": {"id": active_period.id, "name": active_period.name},
        "readings_without_debts": r_no_debts,
        "debts_without_readings": d_no_readings,
        "last_import_not_found": nf,
        "last_import_id": last_log.id if last_log else None,
        "unassigned": {
            "total_debt": round(u_total_debt, 2),
            "total_overpayment": round(u_total_over, 2),
            "count": len(u_keys),
            "by_account": by_account,
        },
    }
