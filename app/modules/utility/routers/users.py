# app/modules/utility/routers/users.py

import io
import logging
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import or_, asc, desc, func
from sqlalchemy.orm import selectinload
from typing import Optional
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse
from fastapi_limiter.depends import RateLimiter

from app.core.database import get_db
from app.core.time_utils import utcnow
from app.modules.utility.models import (
    User, Room, BillingPeriod, MeterReading, Tariff,
)
from app.modules.utility.schemas import (
    UserCreate, UserResponse, UserUpdate, PaginatedResponse,
    RelocateUserSchema
)
from app.core.dependencies import get_current_user, RoleChecker
from app.core.auth import get_password_hash, verify_password, create_access_token
from app.modules.utility.services.excel_service import import_users_from_excel
from app.modules.utility.services.user_service import (
    delete_user_service, countable_resident_condition,
)

# ИМПОРТ ДЛЯ ЖУРНАЛА ДЕЙСТВИЙ
from app.modules.utility.routers.admin_dashboard import write_audit_log

router = APIRouter(prefix="/api/users", tags=["Users"])
logger = logging.getLogger(__name__)

allow_accountant = RoleChecker(["accountant", "admin"])
allow_fin_acc = RoleChecker(["financier", "accountant", "admin"])

ZERO = Decimal("0.00")


# =================================================================
# СХЕМЫ ДЛЯ НАСТРОЙКИ ПРОФИЛЯ
# =================================================================
class ChangeCredentials(BaseModel):
    # new_login — новый ЛОГИН (учётка). new_username оставлен для обратной
    # совместимости со старым моб-приложением (его сетап слал new_username) —
    # теперь трактуется тоже как логин: ФИО жилец менять не может.
    new_login: Optional[str] = Field(None, min_length=3, max_length=100)
    new_username: Optional[str] = Field(None, min_length=3, max_length=100)
    new_password: Optional[str] = Field(None, min_length=8, max_length=128)
    old_password: Optional[str] = None


# =================================================================
# ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ: Проверка tariff_id
# =================================================================
async def _validate_tariff_id(tariff_id: Optional[int], db: AsyncSession) -> None:
    """Проверяет что тариф с указанным ID существует и активен."""
    if tariff_id is None:
        return

    tariff = await db.get(Tariff, tariff_id)
    if not tariff:
        raise HTTPException(
            status_code=400,
            detail=f"Тариф с ID={tariff_id} не найден в системе"
        )
    if not tariff.is_active:
        raise HTTPException(
            status_code=400,
            detail=f"Тариф '{tariff.name}' (ID={tariff_id}) деактивирован. Выберите активный тариф."
        )


# =================================================================
# ЛИЧНЫЙ ПРОФИЛЬ
# =================================================================
@router.get("/me", response_model=UserResponse)
async def get_me(
        db: AsyncSession = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    result = await db.execute(
        select(User)
        .options(
            selectinload(User.room),
            selectinload(User.tariff)
        )
        .where(User.id == current_user.id)
    )
    user = result.scalars().first()
    return user


@router.post("/me/setup")
async def initial_setup(
        data: ChangeCredentials,
        db: AsyncSession = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    # Setup доступен только при первичной настройке. Раньше проверка была
    # `role=="user"` — это позволяло админу с уже сделанным сетапом снова
    # вызвать /me/setup и сменить пароль БЕЗ старого пароля. При краже
    # токена админа это закрепляло доступ навсегда. Теперь блокируется
    # для всех ролей одинаково. Для смены пароля — /me/change-password
    # с обязательным старым паролем.
    if current_user.is_initial_setup_done:
        raise HTTPException(status_code=400, detail="Первичная настройка уже пройдена.")

    # Первичная настройка меняет ЛОГИН (учётку), НЕ ФИО. ФИО (username) —
    # ключ сопоставления 1С/ГИС, его задаёт админ и жилец трогать не может.
    # new_username оставлен для обратной совместимости со старым приложением.
    new_login = (data.new_login or data.new_username or "").strip()
    if new_login and new_login.lower() != (current_user.login or "").lower():
        existing_check = await db.execute(
            select(User).where(
                func.lower(User.login) == new_login.lower(),
                User.id != current_user.id,
            )
        )
        if existing_check.scalars().first():
            raise HTTPException(status_code=400, detail="Этот логин уже занят другим пользователем")
        current_user.login = new_login

    if data.new_password:
        current_user.hashed_password = get_password_hash(data.new_password)
        # Сбрасываем счётчик неудачных попыток и блокировку. Это критично:
        # если жилец до сетапа пытался войти с временным паролем и ошибся
        # несколько раз (или его уже залочило MAX_FAILED_LOGINS=3 попытками),
        # после смены пароля он бы не смог войти даже с правильным новым.
        # Раньше из-за этого жильцы писали «поменял пароль — не могу войти».
        current_user.failed_login_count = 0
        current_user.locked_until = None
        # Отзываем все прежние токены — даже текущий (юзер должен
        # перелогиниться с новым паролем). Защита от token-replay.
        current_user.token_version = (current_user.token_version or 0) + 1

    current_user.is_initial_setup_done = True
    db.add(current_user)

    # ЗАПИСЬ В ЖУРНАЛ: Настройка профиля
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="update", entity_type="user", entity_id=current_user.id,
        details={"action": "initial_setup"}
    )

    await db.commit()
    # Если меняли пароль — token_version инкрементирован, и ТЕКУЩИЙ токен стал
    # невалидным (tv не совпадёт). Возвращаем свежий токен, чтобы приложение
    # после первичной настройки НЕ разлогинивалось и не «висело» на 401
    # (раньше отсюда «долго крутит + перезаход»). Прочие устройства отозваны.
    resp = {"status": "success", "message": "Данные успешно обновлены."}
    if data.new_password:
        resp["access_token"] = create_access_token(data={
            "sub": str(current_user.id),
            "role": current_user.role,
            "scope": "full",
            "tv": current_user.token_version or 0,
        })
    return resp


@router.post(
    "/me/change-password",
    dependencies=[Depends(RateLimiter(times=5, seconds=60))]
)
async def change_password(
        data: ChangeCredentials,
        db: AsyncSession = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    if not data.old_password or not data.new_password:
        raise HTTPException(status_code=400, detail="Необходимо указать старый и новый пароль")

    if not verify_password(data.old_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Неверный текущий пароль")

    current_user.hashed_password = get_password_hash(data.new_password)
    # Сбрасываем lockout state — после смены пароля жилец гарантированно
    # должен иметь возможность войти с новым (см. /me/setup для деталей).
    current_user.failed_login_count = 0
    current_user.locked_until = None
    # Отзываем все прежние токены — нужно перелогиниться с новым паролем.
    current_user.token_version = (current_user.token_version or 0) + 1
    db.add(current_user)

    # ЗАПИСЬ В ЖУРНАЛ: Смена пароля
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="change_password", entity_type="user", entity_id=current_user.id,
        details={}
    )

    await db.commit()
    # Возвращаем свежий токен (token_version инкрементирован) — ТЕКУЩАЯ сессия
    # продолжается без перелогина, прочие устройства разлогинены. Иначе
    # приложение оставалось со старым токеном → 401 на каждом запросе.
    access_token = create_access_token(data={
        "sub": str(current_user.id),
        "role": current_user.role,
        "scope": "full",
        "tv": current_user.token_version or 0,
    })
    return {
        "status": "success",
        "message": "Пароль успешно изменен",
        "access_token": access_token,
    }


@router.post(
    "/me/change-login",
    dependencies=[Depends(RateLimiter(times=5, seconds=60))]
)
async def change_login(
        data: ChangeCredentials,
        db: AsyncSession = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    """Жилец сам меняет ЛОГИН (учётку для входа).

    ФИО (username) менять нельзя — это ключ сопоставления 1С/ГИС, его правит
    только админ. Требуем текущий пароль (смена учётных данных). Сессия
    остаётся валидной — JWT sub = user.id (неизменяемый), перелогин не нужен.
    """
    new_login = (data.new_login or data.new_username or "").strip()
    if len(new_login) < 3:
        raise HTTPException(status_code=400, detail="Новый логин слишком короткий (минимум 3 символа)")
    if not data.old_password or not verify_password(data.old_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Неверный текущий пароль")

    if new_login.lower() != (current_user.login or "").lower():
        dup = await db.execute(
            select(User).where(
                func.lower(User.login) == new_login.lower(),
                User.id != current_user.id,
            )
        )
        if dup.scalars().first():
            raise HTTPException(status_code=400, detail="Этот логин уже занят")

    current_user.login = new_login
    db.add(current_user)
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="update", entity_type="user", entity_id=current_user.id,
        details={"action": "change_login", "new_login": new_login}
    )
    await db.commit()
    return {"status": "success", "message": "Логин изменён", "login": new_login}


# =================================================================
# CRUD ЖИЛЬЦОВ
# =================================================================
@router.post("", response_model=UserResponse)
async def create_user(
        new_user: UserCreate,
        current_user: User = Depends(allow_accountant),  # Добавили current_user
        db: AsyncSession = Depends(get_db)
):
    """Создание нового пользователя с привязкой к комнате по room_id."""
    # ЛК жильцов вычищен (2026-06-10): жильцы НЕ входят в систему, пароль им
    # не нужен — ставим случайный неизвестный hash. Для админа пароль обязателен.
    if new_user.role == "admin" and not new_user.password:
        raise HTTPException(status_code=400, detail="Для администратора задайте пароль")
    import secrets as _secrets
    _password = new_user.password or _secrets.token_urlsafe(24)
    # login (учётка) по умолчанию = ФИО (вход жильцам отключён, поле историческое).
    login_val = (new_user.login or new_user.username).strip()

    # resident_type/billing_mode (нужны и для реактивации, и для создания).
    rt = getattr(new_user, "resident_type", "family") or "family"
    bm = getattr(new_user, "billing_mode", None) or "by_meter"

    if new_user.room_id:
        room_check = await db.get(Room, new_user.room_id)
        if not room_check:
            raise HTTPException(status_code=400, detail="Комната не найдена в Жилфонде")

    # ФИО (username) уникально — это ключ сопоставления 1С/ГИС.
    existing = (await db.execute(
        select(User).where(func.lower(User.username) == func.lower(new_user.username))
    )).scalars().first()
    if existing:
        # Уже ЗАСЕЛЁН (активный + есть комната) — это настоящий дубль, отказ.
        if not existing.is_deleted and existing.room_id is not None:
            raise HTTPException(
                status_code=400,
                detail="Пользователь с таким ФИО уже существует и заселён. Найдите его в списке жильцов.")
        # Иначе запись есть, но «не в строю»: либо soft-deleted (75 из 1С-синка,
        # в списке невидимы, но UNIQUE username блокирует создание), либо
        # активный БЕЗ комнаты (есть в базе 1С, но не заселён — кейс «жилец без
        # помещения»). В обоих случаях НЕ плодим дубль, а возвращаем/заселяем
        # эту запись с указанными данными и комнатой.
        _was_deleted = existing.is_deleted
        existing.is_deleted = False
        existing.role = new_user.role or "user"
        existing.resident_type = rt
        existing.billing_mode = bm
        existing.has_hw_meter = getattr(new_user, "has_hw_meter", True)
        existing.has_cw_meter = getattr(new_user, "has_cw_meter", True)
        existing.has_el_meter = getattr(new_user, "has_el_meter", True)
        existing.is_initial_setup_done = False
        login_taken = (await db.execute(
            select(User.id).where(
                func.lower(User.login) == login_val.lower(),
                User.id != existing.id, User.is_deleted.is_(False),
            ).limit(1)
        )).scalars().first()
        if not login_taken:
            existing.login = login_val
        if new_user.password:
            existing.hashed_password = get_password_hash(new_user.password)
        db.add(existing)
        await db.flush()
        if new_user.room_id:
            from app.modules.utility.services.room_assignment import move_user_to_room
            await move_user_to_room(db, user=existing, new_room_id=new_user.room_id,
                                    note="заселение существующего из базы")
        await write_audit_log(
            db, current_user.id, current_user.username,
            action="reactivate", entity_type="user", entity_id=existing.id,
            details={"username": existing.username, "room_id": new_user.room_id,
                     "was_deleted": _was_deleted},
        )
        await db.commit()
        return (await db.execute(
            select(User).options(selectinload(User.room)).where(User.id == existing.id)
        )).scalars().first()

    # login уникален (case-insensitive) — среди активных.
    existing_login = (await db.execute(
        select(User).where(
            func.lower(User.login) == login_val.lower(),
            User.is_deleted.is_(False),
        )
    )).scalars().first()
    if existing_login:
        raise HTTPException(status_code=400, detail="Этот логин уже занят")

    db_user = User(
        username=new_user.username,
        login=login_val,
        hashed_password=get_password_hash(_password),
        role=new_user.role,
        tariff_id=None,  # тариф от дома/комнаты, не персональный
        # room_id выставится через move_user_to_room ниже — вместе с RoomAssignment.
        room_id=None,
        resident_type=rt,
        billing_mode=bm,
        # Конфигурация счётчиков (см. meters_001_per_user_config). По умолчанию
        # True (старая логика — все счётчики есть). Снимать галочку — для
        # жильцов которые принципиально не подают какой-то ресурс.
        has_hw_meter=getattr(new_user, "has_hw_meter", True),
        has_cw_meter=getattr(new_user, "has_cw_meter", True),
        has_el_meter=getattr(new_user, "has_el_meter", True),
        is_deleted=False,
        is_initial_setup_done=False
    )
    db.add(db_user)
    await db.flush()  # нужен db_user.id для room_assignments.user_id FK

    # Открываем активную RoomAssignment, если жильца сразу селят в комнату.
    if new_user.room_id:
        from app.modules.utility.services.room_assignment import move_user_to_room
        await move_user_to_room(
            db, user=db_user, new_room_id=new_user.room_id,
            note="initial assignment",
        )

    # ЗАПИСЬ В ЖУРНАЛ: Создание пользователя
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="create", entity_type="user",
        details={
            "new_username": new_user.username, "role": new_user.role,
            "resident_type": rt, "billing_mode": bm,
        }
    )

    await db.commit()
    await db.refresh(db_user)

    result = await db.execute(
        select(User).options(selectinload(User.room)).where(User.id == db_user.id)
    )
    return result.scalars().first()


@router.get("", response_model=PaginatedResponse[UserResponse], dependencies=[Depends(allow_fin_acc)])
async def get_users(
        page: int = Query(1, ge=1),
        limit: int = Query(50, ge=1, le=500),
        cursor_id: Optional[int] = Query(None, description="ID для Keyset Pagination"),
        direction: str = Query("next", pattern="^(next|prev)$", description="Направление пагинации"),
        search: Optional[str] = Query(None),
        sort_by: str = Query("id"),
        sort_dir: str = Query("asc", pattern="^(asc|desc)$"),
        # НОВОЕ: фильтры под доменную модель. Позволяет админу в UI быстро
        # отфильтровать «только семьи», «только холостяки», «за койко-место»
        # или жильцов конкретного общежития без загрузки всего списка.
        resident_type: Optional[str] = Query(None, pattern="^(family|single)$"),
        billing_mode: Optional[str] = Query(None, pattern="^(by_meter|per_capita)$"),
        dormitory: Optional[str] = Query(None),
        # housing_001/E2-C: фильтр по типу помещения (общага vs дом)
        # и точному названию улицы (для домов). Парный с rooms.py-фильтрами,
        # позволяет админу быстро вывести «жильцов всех домов» или «всех
        # на улице Ленина».
        place_type: Optional[str] = Query(
            None, pattern="^(dormitory|house)$",
            description="dormitory | house — тип помещения жильца",
        ),
        street: Optional[str] = Query(
            None, description="Точное название улицы (только для place_type=house)",
        ),
        tariff_id: Optional[int] = Query(None),
        db: AsyncSession = Depends(get_db)
):
    """
    Получение списка пользователей с поддержкой умной гибридной пагинации.
    Использует Keyset Pagination (O(1)) при сортировке по ID,
    и автоматически переходит на OFFSET при использовании фильтров.
    """
    # «Свой дом»/мусор: жилец (role='user') БЕЗ комнаты и БЕЗ единого reading
    # нигде не учитывается (решение 2026-06-07). Не-жильцов (admin/accountant)
    # показываем всегда — фильтр их не касается.
    _visible = or_(User.role != "user", countable_resident_condition())
    items_query = select(User).options(selectinload(User.room)).where(
        User.is_deleted.is_(False), _visible)
    count_query = select(func.count(User.id)).where(
        User.is_deleted.is_(False), _visible)

    # Нужен ли JOIN к Room: если ищем по dormitory / сортируем по нему / ищем текстом
    # / фильтруем по place_type / street.
    needs_room_join = bool(
        search or dormitory or place_type or street
        or sort_by in ("dormitory", "apartment_area")
    )

    if search:
        search_filter = f"%{search}%"
        search_condition = or_(
            User.username.ilike(search_filter),
            Room.dormitory_name.ilike(search_filter),
            Room.room_number.ilike(search_filter),
        )
        items_query = items_query.outerjoin(Room, User.room_id == Room.id).where(search_condition)
        count_query = count_query.outerjoin(Room, User.room_id == Room.id).where(search_condition)
    elif needs_room_join:
        items_query = items_query.outerjoin(Room, User.room_id == Room.id)
        count_query = count_query.outerjoin(Room, User.room_id == Room.id)

    if resident_type:
        items_query = items_query.where(User.resident_type == resident_type)
        count_query = count_query.where(User.resident_type == resident_type)
    if billing_mode:
        items_query = items_query.where(User.billing_mode == billing_mode)
        count_query = count_query.where(User.billing_mode == billing_mode)
    if tariff_id:
        items_query = items_query.where(User.tariff_id == tariff_id)
        count_query = count_query.where(User.tariff_id == tariff_id)
    if dormitory:
        items_query = items_query.where(Room.dormitory_name == dormitory)
        count_query = count_query.where(Room.dormitory_name == dormitory)
    if place_type:
        items_query = items_query.where(Room.place_type == place_type)
        count_query = count_query.where(Room.place_type == place_type)
    if street:
        items_query = items_query.where(Room.street == street)
        count_query = count_query.where(Room.street == street)

    total = (await db.execute(count_query)).scalar_one()

    valid_sort_fields = {
        "id": User.id,
        "username": User.username,
        "role": User.role,
        "dormitory": Room.dormitory_name,
        "apartment_area": Room.apartment_area,
    }
    sort_column = valid_sort_fields.get(sort_by, User.id)

    # JOIN к Room нужен и для сортировки по dormitory/apartment_area (если ещё не сделан)
    if sort_by in ["dormitory", "apartment_area"] and not needs_room_join:
        items_query = items_query.outerjoin(Room, User.room_id == Room.id)

    # Keyset работает только при сортировке по id И без сужающих фильтров —
    # иначе count/items дают несогласованные окна. При любом фильтре —
    # автоматически OFFSET.
    use_keyset = (sort_by == "id") and not any([
        search, resident_type, billing_mode, dormitory, tariff_id,
        place_type, street,
    ])

    if use_keyset and cursor_id is not None:
        if direction == "next":
            if sort_dir == "asc":
                items_query = items_query.where(User.id > cursor_id)
            else:
                items_query = items_query.where(User.id < cursor_id)
        else:  # prev
            if sort_dir == "asc":
                items_query = items_query.where(User.id < cursor_id)
            else:
                items_query = items_query.where(User.id > cursor_id)
    else:
        # Fallback на OFFSET (для текстовых сортировок)
        items_query = items_query.offset((page - 1) * limit)

    # Сортировка (инверсия для Prev)
    if use_keyset and direction == "prev":
        items_query = items_query.order_by(desc(sort_column) if sort_dir == "asc" else asc(sort_column))
    else:
        items_query = items_query.order_by(asc(sort_column) if sort_dir == "asc" else desc(sort_column))

    items_query = items_query.limit(limit)
    items = list((await db.execute(items_query)).scalars().all())

    # Возврат массива в правильном порядке при движении назад
    if use_keyset and direction == "prev":
        items.reverse()

    return {"total": total, "page": page, "size": limit, "items": items}


# =====================================================================
# STATS — агрегированная аналитика для вкладки «Жильцы»
# ВАЖНО: объявлен ДО @router.get("/{user_id}"). FastAPI роутит по порядку,
# и если /{user_id} окажется выше — путь /users/stats попадёт в него,
# FastAPI попробует распарсить "stats" как int и вернёт 422.
# =====================================================================
@router.get("/stats", dependencies=[Depends(allow_accountant)])
async def users_stats(
        db: AsyncSession = Depends(get_db),
):
    """KPI + распределения + топ-должники/переплатчики одним раундом к БД."""
    active_where = User.is_deleted.is_(False)
    # total = учитываемые жильцы (с комнатой ИЛИ с долгом), без «своих домов».
    total_users = (await db.execute(
        select(func.count(User.id)).where(
            active_where, User.role == "user", countable_resident_condition())
    )).scalar_one()
    with_room = (await db.execute(
        select(func.count(User.id)).where(
            active_where, User.role == "user", User.room_id.is_not(None)
        )
    )).scalar_one()

    by_type_rows = (await db.execute(
        select(User.resident_type, func.count(User.id))
        .where(active_where, User.role == "user")
        .group_by(User.resident_type)
    )).all()
    by_type = {rt or "family": int(c) for rt, c in by_type_rows}

    by_mode_rows = (await db.execute(
        select(User.billing_mode, func.count(User.id))
        .where(active_where, User.role == "user")
        .group_by(User.billing_mode)
    )).all()
    by_mode = {bm or "by_meter": int(c) for bm, c in by_mode_rows}

    dorm_rows = (await db.execute(
        select(Room.dormitory_name, func.count(User.id))
        .outerjoin(User, User.room_id == Room.id)
        .where(active_where | User.id.is_(None))
        .group_by(Room.dormitory_name)
        .order_by(func.count(User.id).desc())
    )).all()
    by_dormitory = [
        {"name": d or "— не указано —", "count": int(c)}
        for d, c in dorm_rows if c > 0
    ]

    tariff_rows = (await db.execute(
        select(Tariff.id, Tariff.name, func.count(User.id))
        .outerjoin(User, (User.tariff_id == Tariff.id) & active_where & (User.role == "user"))
        .group_by(Tariff.id, Tariff.name)
        .order_by(func.count(User.id).desc())
    )).all()
    by_tariff = [
        {"id": tid, "name": tname, "count": int(c)}
        for tid, tname, c in tariff_rows
    ]

    # Аудит #6: сальдо — снимок ОДНОГО периода. SUM по всем периодам без join
    # User множил долг на число периодов и включал удалённых. Период — как
    # financier (resolve_view_period: активный → последний импорт → свежий),
    # чтобы в межмесячном окне цифра не обнулялась (ревизия #5/#6).
    from sqlalchemy import false as _sa_false
    from app.modules.utility.services.period_resolver import resolve_view_period
    _vp = await resolve_view_period(db)
    _period_cond = (MeterReading.period_id == _vp.id) if _vp else _sa_false()

    debt_rows = (await db.execute(
        select(
            func.coalesce(func.sum(MeterReading.debt_209), 0),
            func.coalesce(func.sum(MeterReading.debt_205), 0),
            func.coalesce(func.sum(MeterReading.overpayment_209), 0),
            func.coalesce(func.sum(MeterReading.overpayment_205), 0),
        )
        .join(User, User.id == MeterReading.user_id)
        .where(active_where, User.role == "user",
               MeterReading.is_approved.is_(True), _period_cond)
    )).first()
    total_debt = float((debt_rows[0] or 0) + (debt_rows[1] or 0))
    total_overpayment = float((debt_rows[2] or 0) + (debt_rows[3] or 0))

    top_debtor_rows = (await db.execute(
        select(
            User.id, User.username,
            Room.dormitory_name, Room.room_number,
            func.coalesce(func.sum(MeterReading.debt_209 + MeterReading.debt_205), 0).label("debt"),
        )
        .join(MeterReading, MeterReading.user_id == User.id)
        .outerjoin(Room, User.room_id == Room.id)
        .where(active_where, User.role == "user",
               MeterReading.is_approved.is_(True), _period_cond)
        .group_by(User.id, User.username, Room.dormitory_name, Room.room_number)
        .having(func.coalesce(func.sum(MeterReading.debt_209 + MeterReading.debt_205), 0) > 0)
        .order_by(func.coalesce(func.sum(MeterReading.debt_209 + MeterReading.debt_205), 0).desc())
        .limit(5)
    )).all()
    top_debtors = [
        {"id": uid, "username": uname,
         "room": f"{dorm or '—'}, ком. {rnum or '—'}" if (dorm or rnum) else None,
         "amount": float(amount)}
        for uid, uname, dorm, rnum, amount in top_debtor_rows
    ]

    top_overpaid_rows = (await db.execute(
        select(
            User.id, User.username,
            Room.dormitory_name, Room.room_number,
            func.coalesce(func.sum(MeterReading.overpayment_209 + MeterReading.overpayment_205), 0).label("over"),
        )
        .join(MeterReading, MeterReading.user_id == User.id)
        .outerjoin(Room, User.room_id == Room.id)
        .where(active_where, User.role == "user",
               MeterReading.is_approved.is_(True), _period_cond)
        .group_by(User.id, User.username, Room.dormitory_name, Room.room_number)
        .having(func.coalesce(func.sum(MeterReading.overpayment_209 + MeterReading.overpayment_205), 0) > 0)
        .order_by(func.coalesce(func.sum(MeterReading.overpayment_209 + MeterReading.overpayment_205), 0).desc())
        .limit(5)
    )).all()
    top_overpaid = [
        {"id": uid, "username": uname,
         "room": f"{dorm or '—'}, ком. {rnum or '—'}" if (dorm or rnum) else None,
         "amount": float(amount)}
        for uid, uname, dorm, rnum, amount in top_overpaid_rows
    ]

    return {
        "total_users": int(total_users),
        "with_room": int(with_room),
        "without_room": int(total_users) - int(with_room),
        "by_resident_type": by_type,
        "by_billing_mode": by_mode,
        "by_dormitory": by_dormitory,
        "by_tariff": by_tariff,
        "total_debt": total_debt,
        "total_overpayment": total_overpayment,
        "top_debtors": top_debtors,
        "top_overpaid": top_overpaid,
    }


# =====================================================================
# EXPORT — Excel со списком жильцов (с учётом фильтров)
# Тоже ДО /{user_id} — ради единообразия и защиты от будущих конфликтов.
# =====================================================================
@router.get("/export/list", dependencies=[Depends(allow_accountant)])
async def export_users_list(
        search: Optional[str] = Query(None),
        resident_type: Optional[str] = Query(None, pattern="^(family|single)$"),
        billing_mode: Optional[str] = Query(None, pattern="^(by_meter|per_capita)$"),
        dormitory: Optional[str] = Query(None),
        # housing_001/E2-C: парные фильтры с get_users.
        place_type: Optional[str] = Query(None, pattern="^(dormitory|house)$"),
        street: Optional[str] = Query(None),
        tariff_id: Optional[int] = Query(None),
        db: AsyncSession = Depends(get_db),
):
    """Excel-выгрузка всех отфильтрованных жильцов (без пагинации)."""
    q = (
        select(User)
        .options(selectinload(User.room), selectinload(User.tariff))
        .where(User.is_deleted.is_(False), User.role == "user")
    )
    if search or dormitory or place_type or street:
        q = q.outerjoin(Room, User.room_id == Room.id)
    if search:
        p = f"%{search}%"
        q = q.where(or_(
            User.username.ilike(p),
            Room.dormitory_name.ilike(p),
            Room.room_number.ilike(p),
        ))
    if resident_type:
        q = q.where(User.resident_type == resident_type)
    if billing_mode:
        q = q.where(User.billing_mode == billing_mode)
    if dormitory:
        q = q.where(Room.dormitory_name == dormitory)
    if place_type:
        q = q.where(Room.place_type == place_type)
    if street:
        q = q.where(Room.street == street)
    if tariff_id:
        q = q.where(User.tariff_id == tariff_id)
    q = q.order_by(User.id)

    users = (await db.execute(q)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Жильцы"
    headers = [
        "ID", "Логин / ФИО", "Роль", "Тип жильца",
        "Тип помещения", "Общежитие / Улица", "Комната / Квартира",
        "Площадь м²", "Проживающих", "Тариф",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DBEAFE")
    for i, u in enumerate(users, 2):
        ws.cell(row=i, column=1, value=u.id)
        ws.cell(row=i, column=2, value=u.username)
        ws.cell(row=i, column=3, value=u.role)
        ws.cell(row=i, column=4, value="Семейный" if u.resident_type == "family" else "Холостяк")
        # housing_001/E2-C: колонки 5-7 — тип помещения + адрес. Для дома кладём
        # ул+дом в "Общежитие/Улица", квартиру в "Комната/Квартира".
        if u.room and u.room.place_type == "house":
            ws.cell(row=i, column=5, value="Дом / квартира")
            _addr = ", ".join(filter(None, [
                f"ул. {u.room.street}" if u.room.street else None,
                f"д. {u.room.house_number}" if u.room.house_number else None,
            ])) or ""
            ws.cell(row=i, column=6, value=_addr)
            ws.cell(row=i, column=7, value=(f"кв. {u.room.apartment_number}" if u.room.apartment_number else ""))
        else:
            ws.cell(row=i, column=5, value="Общежитие" if u.room else "")
            ws.cell(row=i, column=6, value=u.room.dormitory_name if u.room else "")
            ws.cell(row=i, column=7, value=u.room.room_number if u.room else "")
        ws.cell(row=i, column=8, value=float(u.room.apartment_area) if (u.room and u.room.apartment_area) else 0)
        ws.cell(row=i, column=9, value=(u.room.total_room_residents if u.room and u.room.total_room_residents else 1))
        ws.cell(row=i, column=10, value=u.tariff.name if u.tariff else "")
    # 10 колонок: A..J (убраны «Режим оплаты» и «Место работы»).
    for col, width in [("A", 6), ("B", 32), ("C", 10), ("D", 14),
                       ("E", 16), ("F", 22), ("G", 14), ("H", 10), ("I", 12),
                       ("J", 22)]:
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"residents_{utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{user_id}", response_model=UserResponse, dependencies=[Depends(allow_accountant)])
async def read_user(user_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(User).options(selectinload(User.room)).where(User.id == user_id)
    )
    user = result.scalars().first()
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    return user


@router.get("/{user_id}/residence-history", dependencies=[Depends(allow_accountant)])
async def get_residence_history(user_id: int, db: AsyncSession = Depends(get_db)):
    """История проживания жильца — где и когда жил.

    Используется в админке: вкладка «Жильцы» → клик по жильцу → история комнат.
    Также можно посмотреть «кто жил в комнате X в период Y» через сервис
    get_room_residents_at, но это пока не вынесено в HTTP-эндпоинт.
    """
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    from app.modules.utility.services.room_assignment import get_user_history
    history = await get_user_history(db, user_id, limit=100)

    # Подгружаем комнаты разом одним запросом (вместо N+1)
    room_ids = list({h.room_id for h in history})
    rooms_map = {}
    if room_ids:
        rooms = (await db.execute(select(Room).where(Room.id.in_(room_ids)))).scalars().all()
        rooms_map = {r.id: r for r in rooms}

    return {
        "user": {"id": user.id, "username": user.username},
        "current_room_id": user.room_id,
        "items": [
            {
                "id": h.id,
                "room_id": h.room_id,
                "room": (
                    rooms_map[h.room_id].format_address
                    if h.room_id in rooms_map else None
                ),
                "moved_in_at": h.moved_in_at.isoformat() if h.moved_in_at else None,
                "moved_out_at": h.moved_out_at.isoformat() if h.moved_out_at else None,
                "is_current": h.moved_out_at is None,
                "note": h.note,
            }
            for h in history
        ],
    }


@router.put("/{user_id}", response_model=UserResponse)
async def update_user(
        user_id: int,
        update_data: UserUpdate,
        current_user: User = Depends(allow_accountant),  # Добавили current_user
        db: AsyncSession = Depends(get_db)
):
    db_user = await db.get(User, user_id)
    if not db_user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    update_dict = update_data.dict(exclude_unset=True)

    if "room_id" in update_dict and update_dict["room_id"]:
        room_check = await db.get(Room, update_dict["room_id"])
        if not room_check:
            raise HTTPException(status_code=400, detail="Комната не найдена в Жилфонде")

    # Тариф больше не персональный — игнорируем tariff_id в апдейте жильца
    # (подтягивается от дома/комнаты, настраивается через «Настройки дома»).
    update_dict.pop("tariff_id", None)

    # ФИО (username) и login уникальны. Раньше апдейт делал setattr без проверки —
    # дубликат ронял IntegrityError (500). Проверяем тут (исключая самого себя).
    # PUT доступен только админу (allow_accountant) — поэтому ФИО правит админ.
    new_fio = (update_dict.get("username") or "").strip()
    if new_fio and new_fio.lower() != (db_user.username or "").lower():
        dup = await db.execute(select(User).where(
            func.lower(User.username) == new_fio.lower(), User.id != db_user.id))
        if dup.scalars().first():
            raise HTTPException(status_code=400, detail="Пользователь с таким ФИО уже существует")
        update_dict["username"] = new_fio
    new_login = (update_dict.get("login") or "").strip()
    if new_login and new_login.lower() != (db_user.login or "").lower():
        dup = await db.execute(select(User).where(
            func.lower(User.login) == new_login.lower(), User.id != db_user.id))
        if dup.scalars().first():
            raise HTTPException(status_code=400, detail="Этот логин уже занят")
        update_dict["login"] = new_login

    # Аудит #26: смена пароля/роли админом обязана отзывать старые сессии
    # (бамп token_version) — иначе угнанный токен живёт до exp. Роль теперь
    # ловится и проверкой token_role в get_current_user, но бамп надёжнее.
    _bump_tv = False
    if "password" in update_dict and update_dict["password"]:
        db_user.hashed_password = get_password_hash(update_dict.pop("password"))
        _bump_tv = True
    if "role" in update_dict and update_dict["role"] != db_user.role:
        _bump_tv = True

    # Переезд: если меняется room_id — пишем в историю проживания (RoomAssignment).
    # Раньше менялось одной строчкой setattr — без следов; теперь через сервис.
    if "room_id" in update_dict and update_dict["room_id"] != db_user.room_id:
        from app.modules.utility.services.room_assignment import move_user_to_room
        await move_user_to_room(
            db, user=db_user,
            new_room_id=update_dict.pop("room_id"),
            note="reassigned via update_user",
        )

    # billing_mode из resident_type больше не выводим: singles тоже by_meter
    # (счётчики на квартиру делятся в billing). per_capita — legacy.
    for key, value in update_dict.items():
        setattr(db_user, key, value)

    if _bump_tv:
        db_user.token_version = (db_user.token_version or 0) + 1

    # ЗАПИСЬ В ЖУРНАЛ: Обновление пользователя
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="update", entity_type="user", entity_id=db_user.id,
        details={"updated_fields": list(update_dict.keys())}
    )

    await db.commit()
    await db.refresh(db_user)

    result = await db.execute(
        select(User).options(selectinload(User.room)).where(User.id == db_user.id)
    )
    return result.scalars().first()


@router.delete("/{user_id}", status_code=204)
async def delete_user(
        user_id: int,
        current_user: User = Depends(allow_accountant),  # Добавили current_user
        db: AsyncSession = Depends(get_db)
):
    try:
        await delete_user_service(user_id, db)

        # ЗАПИСЬ В ЖУРНАЛ: Удаление (выселение)
        await write_audit_log(
            db, current_user.id, current_user.username,
            action="delete", entity_type="user", entity_id=user_id,
            details={"action": "soft_delete"}
        )

        await db.commit()
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=404, detail=str(e))
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=500, detail="Ошибка при удалении")
    return None


# =================================================================
# ЕДИНОЕ ОКНО: РАЗОВОЕ НАЧИСЛЕНИЕ И ПЕРЕСЕЛЕНИЕ/ВЫСЕЛЕНИЕ
# =================================================================
@router.post("/{user_id}/relocate")
async def relocate_user(
        user_id: int,
        data: RelocateUserSchema,
        current_user: User = Depends(allow_accountant),  # Добавили current_user
        db: AsyncSession = Depends(get_db)
):
    """Единый процесс: Разовое начисление по старой комнате + Переселение/Выселение"""
    active_period = (await db.execute(
        select(BillingPeriod).where(BillingPeriod.is_active)
    )).scalars().first()

    if not active_period:
        raise HTTPException(status_code=400, detail="Нет активного периода")

    user = (await db.execute(
        select(User).options(selectinload(User.room)).where(
            User.id == user_id,
            User.is_deleted.is_(False)
        )
    )).scalars().first()

    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    action = "evict" if data.is_eviction else "move"
    from app.modules.utility.services.room_assignment import move_user_to_room

    if action == "evict":
        user.is_deleted = True
        user.username = f"{user.username}_deleted_{user.id}"
        user.login = f"{user.login}_deleted_{user.id}"  # освобождаем и логин
        # Закрываем активную RoomAssignment (moved_out_at = now), новой не создаём
        await move_user_to_room(db, user=user, new_room_id=None, note="evicted")
        # Прим.: отдельная финальная квитанция за частичный месяц тут НЕ
        # формируется — начисление по старой комнате идёт обычным закрытием
        # периода (reading хранит свой room_id, считается по тарифу той комнаты).
        message = "Жилец успешно выселен."
    elif action == "move":
        new_room = await db.get(Room, data.new_room_id)
        if not new_room:
            raise HTTPException(status_code=404, detail="Новая комната не найдена")
        await move_user_to_room(db, user=user, new_room_id=new_room.id,
                                note=f"relocate to {new_room.format_address}")
        message = f"Жилец переведён в {new_room.format_address}."

    # ЗАПИСЬ В ЖУРНАЛ: Переселение/Выселение
    await write_audit_log(
        db, current_user.id, current_user.username,
        action=action, entity_type="user", entity_id=user.id,
        details={"new_room_id": data.new_room_id if action == "move" else None}
    )

    await db.commit()
    return {"status": "success", "message": message}


# =================================================================
# ИМПОРТ И ЭКСПОРТ EXCEL
# =================================================================
@router.post("/import_excel", summary="Умный импорт (Жилфонд + Жильцы)")
async def import_users(
        file: UploadFile = File(...),
        current_user: User = Depends(allow_accountant),  # Добавили current_user
        db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Только файлы Excel (.xlsx, .xls)")

    # Лимит размера файла — защита от zip-бомб и DoS через большой upload.
    # nginx ставит client_max_body_size 10M, но если кто-то ходит мимо
    # nginx напрямую на backend — этой защиты нет. Дублируем на уровне
    # приложения. 20M с запасом — реальные импорты редко больше 5M.
    MAX_EXCEL_BYTES = 20 * 1024 * 1024
    # Content-Length может отсутствовать (chunked), тогда читаем безопасно
    # через ограниченный буфер.
    declared_size = (file.size if hasattr(file, "size") else None) or 0
    if declared_size and declared_size > MAX_EXCEL_BYTES:
        raise HTTPException(status_code=413, detail="Файл слишком большой (макс. 20 МБ)")

    header = await file.read(8)
    await file.seek(0)
    if not (header.startswith(b"PK\x03\x04") or header.startswith(b"\xd0\xcf\x11\xe0")):
        raise HTTPException(status_code=400, detail="Неверная сигнатура Excel файла")

    # Читаем с ограничением (на случай если Content-Length отсутствовал).
    content = await file.read(MAX_EXCEL_BYTES + 1)
    if len(content) > MAX_EXCEL_BYTES:
        raise HTTPException(status_code=413, detail="Файл слишком большой (макс. 20 МБ)")

    # Выполняем импорт
    result = await import_users_from_excel(content, db)

    # ЗАПИСЬ В ЖУРНАЛ: Массовый импорт Excel
    await write_audit_log(
        db, current_user.id, current_user.username,
        action="import", entity_type="system",
        details={
            "added_users": result.get("added_users", 0),
            "updated_users": result.get("updated_users", 0),
            "added_rooms": result.get("added_rooms", 0),
            "updated_rooms": result.get("updated_rooms", 0)
        }
    )
    # Коммит нужен для сохранения записи в лог (сам excel сервис коммитит свои данные внутри себя)
    await db.commit()

    return result


@router.get("/export/template", summary="Скачать шаблон для импорта")
async def download_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон импорта"

    headers = [
        "Логин (Оставьте пустым для создания только комнаты)",
        "Пароль (Можно пусто)", "Общежитие (ОБЯЗАТЕЛЬНО)", "Номер комнаты (ОБЯЗАТЕЛЬНО)",
        "Площадь м2", "Макс. мест в комнате", "Кол-во жильцов на Л/С",
        "№ ГВС", "№ ХВС", "№ Электр.", "Место работы", "Тарифный профиль"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 25

    ws.append([
        "ivanov_i", "pass12345", "Общежитие №1", "101", 18.5, 2, 1,
        "HW-001", "CW-002", "EL-003", "МЧС", "Базовый тариф"
    ])
    ws.append(["", "", "Общежитие №1", "102", 20.0, 3, "", "HW-004", "CW-005", "EL-006", "", ""])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Import_Template.xlsx"}
    )
