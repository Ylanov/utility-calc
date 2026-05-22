# app/modules/utility/services/debt_import.py
import openpyxl
import logging
import os
import re
from datetime import datetime, timezone, date
from decimal import Decimal
from typing import Dict, Optional
from sqlalchemy.orm import Session
from sqlalchemy import select
from rapidfuzz import process, fuzz
from app.modules.utility.models import (
    User, MeterReading, BillingPeriod, DebtImportLog, RentalContract,
    GSheetsAlias,
)


# Регексы для парсинга строки договора из ОСВ 1С. В файле под каждым ФИО
# идут строки типа:
#   «Договор от 14.02.2017 № 1013»     (date первая)
#   «Договор № 923 от 28.12.2015»      (number первый)
#   «Договор 923 от 28.12.2015»        (без «№»)
#   «Договор от 07.02.2013 № 417-К»    (буква в номере)
_CONTRACT_DATE_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})")
_CONTRACT_NUM_AFTER_HASH_RE = re.compile(r"№\s*([^\s,;]+)", re.IGNORECASE)
_CONTRACT_NUM_BARE_RE = re.compile(r"^договор\s+(\d[^\s]*)\s+от", re.IGNORECASE)


def pick_saldo_pair(
    row,
    end_debit_col: int,
    end_credit_col: int,
    start_debit_col: int,
    start_credit_col: int,
    obor_debit_col: Optional[int] = None,
    obor_credit_col: Optional[int] = None,
) -> tuple[Decimal, Decimal]:
    """Возвращает (debt, overpayment) — актуальные сальдо из строки ОСВ.

    Корректная семантика 1С с тремя уровнями приоритета:

    1) Если в строке заполнено «Сальдо на КОНЕЦ периода» (хотя бы одна
       из колонок Дт/Кр) — берём end-значения. Пустая ячейка = 0.
       Кейс Глобы: оборот Кр=18000, конец Кр=7091 — мы видим 7091.

    2) Если obor_debit_col/obor_credit_col заданы И в них есть значения
       (т.е. у жильца БЫЛИ обороты), а end-колонки пусты — это значит
       «Сальдо конец = 0» (в ОСВ 1С нули не показываются). Вычисляем
       математически: end_d = start_d - start_c + obor_d - obor_c.

       Кейс Бендаса (Bug U-fix6):
         start_d=2385.07, obor_d=1458.86 (доначислили), obor_c=3843.93 (заплатил)
         end_d=null, end_c=null
         Старая логика: fallback на start → debt=2385.07 ❌
         Новая: 2385.07 + 1458.86 − 3843.93 = 0 → debt=0 ✓

    3) Если obor-колонки не заданы ИЛИ в них нет значений И end пусты —
       у жильца совсем не было движений в периоде, состояние = начало.
       Берём start как пару.

    4) Если всё пусто — возвращаем (0, 0).
    """
    def _read(col):
        if col is None:
            return None
        if 0 <= col < len(row):
            return row[col]
        return None

    end_d = _read(end_debit_col)
    end_c = _read(end_credit_col)
    has_end_data = (end_d is not None) or (end_c is not None)

    if has_end_data:
        debt = clean_decimal(end_d) if end_d is not None else Decimal("0")
        over = clean_decimal(end_c) if end_c is not None else Decimal("0")
        return debt, over

    # Bug U-fix6: проверяем, были ли обороты. Если да — вычисляем сальдо
    # математически (end = start + obor_d - obor_c).
    obor_d_val = _read(obor_debit_col)
    obor_c_val = _read(obor_credit_col)
    has_obor_data = (obor_d_val is not None) or (obor_c_val is not None)

    if has_obor_data:
        start_d_d = clean_decimal(_read(start_debit_col)) if _read(start_debit_col) is not None else Decimal("0")
        start_c_d = clean_decimal(_read(start_credit_col)) if _read(start_credit_col) is not None else Decimal("0")
        obor_d_d = clean_decimal(obor_d_val) if obor_d_val is not None else Decimal("0")
        obor_c_d = clean_decimal(obor_c_val) if obor_c_val is not None else Decimal("0")
        # Активный счёт 209/205: положительный остаток в Дт — долг,
        # отрицательный (Кр > Дт) — переплата.
        net = start_d_d - start_c_d + obor_d_d - obor_c_d
        if net > 0:
            return net, Decimal("0")
        elif net < 0:
            return Decimal("0"), -net
        else:
            return Decimal("0"), Decimal("0")

    # Fallback: end не показан, оборотов тоже нет → состояние = начало.
    start_d = _read(start_debit_col)
    start_c = _read(start_credit_col)
    debt = clean_decimal(start_d) if start_d is not None else Decimal("0")
    over = clean_decimal(start_c) if start_c is not None else Decimal("0")
    return debt, over


# DEPRECATED: оставлен для обратной совместимости со старыми unit-тестами
# и внешними скриптами; в продакшен-цикле используется pick_saldo_pair.
def pick_saldo_value(row, end_col: int, start_col: int) -> Decimal:
    """Старая поколоночная логика (бралa либо debt либо overpay независимо).

    Не использовать в новом коде — она ломается на парах когда у жильца
    одновременно есть оборот по противоположной стороне. См. pick_saldo_pair.
    """
    end_raw = row[end_col] if 0 <= end_col < len(row) else None
    if end_raw is not None:
        return clean_decimal(end_raw)
    if end_col == start_col:
        return Decimal("0")
    start_raw = row[start_col] if 0 <= start_col < len(row) else None
    return clean_decimal(start_raw) if start_raw is not None else Decimal("0")


def parse_contract_line(text: Optional[str]) -> Optional[dict]:
    """Парсит строку из колонки A ОСВ 1С если это договор.

    Возвращает {"number", "signed_date"} или None если строка не договор
    либо нет даты/номера.

    Поддерживаемые форматы:
      «Договор от 14.02.2017 № 1013»  → {number: "1013", date: 2017-02-14}
      «Договор № 923 от 28.12.2015»   → {number: "923", date: 2015-12-28}
      «Договор 923 от 28.12.2015»     → {number: "923", date: 2015-12-28}
      «Договор от 07.02.2013 № 417-К» → {number: "417-К", date: 2013-02-07}
    """
    if not text:
        return None
    s = str(text).strip()
    if not s or not s.lower().startswith("договор"):
        return None

    date_match = _CONTRACT_DATE_RE.search(s)
    if not date_match:
        return None
    day, month, year = date_match.groups()
    # Двузначный год — считаем что 20XX (1С обычно пишет 4 цифры, но защита)
    year_int = int("20" + year) if len(year) == 2 else int(year)
    try:
        signed_date = date(year_int, int(month), int(day))
    except ValueError:
        return None

    # Номер: «№ XXX» имеет приоритет, иначе формат «Договор NUM от ...»
    num_match = _CONTRACT_NUM_AFTER_HASH_RE.search(s)
    if num_match:
        number = num_match.group(1)
    else:
        num_bare = _CONTRACT_NUM_BARE_RE.search(s)
        number = num_bare.group(1) if num_bare else None

    if not number:
        return None
    return {"number": number.strip(), "signed_date": signed_date}

logger = logging.getLogger(__name__)

FUZZY_THRESHOLD = 88  # default; реально читается через _threshold() из конфига


def _threshold() -> int:
    from app.modules.utility.services.analyzer_config import config
    return config.get_int("debt.fuzzy_threshold", FUZZY_THRESHOLD)


def clean_decimal(value) -> Decimal:
    """Очищает и конвертирует значение из Excel в Decimal."""
    if value is None:
        return Decimal("0.00")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.replace(" ", "").replace("\xa0", "").replace(",", ".")
        try:
            return Decimal(cleaned)
        except Exception:
            return Decimal("0.00")
    return Decimal("0.00")


def normalize_name(value: str) -> str:
    """Приводит ФИО к единому нижнему регистру без лишних пробелов."""
    if not value:
        return ""
    return " ".join(str(value).lower().split())


def is_valid_name_row(cell_value: str) -> bool:
    """Проверяет, является ли строка Excel ФИО жильца."""
    if not cell_value:
        return False
    val = str(cell_value).strip()
    lower_val = val.lower()

    stop_words = [
        "договор", "закрыт", "итого", "счет", "контрагенты",
        "сальдо", "выводимые", "единица", "обороты", "дебет", "кредит"
    ]
    for word in stop_words:
        if word in lower_val:
            return False

    if len(val.split()) < 2:
        return False

    return True


def find_user_fuzzy(target_name: str, users_map: Dict[str, int]) -> Optional[int]:
    """Глобальная функция поиска (используется внешними модулями)."""
    if not target_name:
        return None

    norm_target = normalize_name(target_name)

    if norm_target in users_map:
        return users_map[norm_target]

    match = process.extractOne(norm_target, list(users_map.keys()), scorer=fuzz.token_sort_ratio)

    if match:
        best_match_name, score, _ = match
        if score >= _threshold():
            return users_map[best_match_name]

    return None


def sync_import_debts_process(
    file_path: str,
    db: Session,
    account_type: str,
    started_by_id: int | None = None,
    started_by_username: str | None = None,
    batch_id: str | None = None,
    original_file_name: str | None = None,
) -> dict:
    """
    Функция массового импорта долгов.
    Долг из 1С привязывается к КОМНАТЕ жильца. Долги соседей по комнате суммируются.

    Важно: ВСЁ делается одной транзакцией. Если на 5000-й строке случится
    ошибка — откатываются все 4999 предыдущих, файл не удаляется (пусть
    админ исправит и перезапустит).

    Дополнительно пишет запись в DebtImportLog:
      * snapshot_data — предыдущие debt_* по каждому затрагиваемому reading_id
        (для отмены импорта через endpoint /debts/import-history/{id}/undo);
      * not_found_users — список ФИО, не найденных fuzzy (для ручной привязки).
    """
    logger.info(f"Starting debts import from {file_path} for Account {account_type}")

    try:
        workbook = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
        worksheet = workbook.active
    except Exception as error:
        logger.exception("Failed to open Excel file")
        # Бросаем — Celery должен увидеть падение и ретраить.
        raise RuntimeError(f"Ошибка чтения файла: {error}") from error

    # Парсим заголовок чтобы найти колонки «Дебет» и «Кредит».
    #
    # Структура header в стандартной ОСВ 1С имеет ТРИ секции:
    #   row 8: 'Сальдо на начало периода' (E-G) | 'Обороты за период' (K-O)
    #          | 'Сальдо на конец периода' (P-R)
    #   row 9: ...'Дебет','','','Кредит',...,'Дебет','','','Кредит',...,'Дебет','','','Кредит'...
    #
    # Что нам реально нужно — «Сальдо на КОНЕЦ периода». Это актуальный
    # долг/переплата на дату формирования отчёта. Если у жильца не было
    # оборотов — конечная ячейка может быть пустой (1С не повторяет
    # неизменное значение); тогда fallback на «Сальдо на начало».
    #
    # debt_col_first/last — позиции 1-й и 3-й колонки «Дебет» в row 9
    # (Сальдо начало и Сальдо конец). Same для «Кредит».
    debt_col_first: Optional[int] = None
    debt_col_last: Optional[int] = None
    overpay_col_first: Optional[int] = None
    overpay_col_last: Optional[int] = None
    # Bug U-fix6: колонки оборотов (Дт обороты, Кр обороты) для
    # математического вычисления сальдо, когда end-колонки пусты.
    obor_debit_col: Optional[int] = None
    obor_credit_col: Optional[int] = None
    try:
        # ВАЖНО (Bug U-fix2): новый робустный парсер заголовков через
        # section markers. Работает БЕЗ unmerge (предыдущий подход не
        # сработал у Бендаса — в БД всё ещё лежит «Сальдо начало»).
        #
        # Логика: в openpyxl при read_only=True для merged-региона значение
        # возвращается ТОЛЬКО в верхней-левой ячейке. То есть для ОСВ 1С с
        # шапкой:
        #   row 8: [col 8: «Сальдо на начало периода»] [col 14: «Обороты»]
        #          [col 20: «Сальдо на конец периода»]   ← merged-headers
        #   row 9: [col 7: «Дебет»] [col 10: «Кредит»] [col 14: «Дебет»]
        #          [col 17: «Кредит»] [col 20: «Дебет»] [col 23: «Кредит»]
        # Мы можем привязать «Дебет»/«Кредит» к секциям по их col_idx.
        section_markers = {}  # 'начало'|'обороты'|'конец' -> col_idx
        debit_cols = []  # все col где найдено «Дебет»
        credit_cols = []

        header_ws_ro = openpyxl.load_workbook(
            filename=file_path, read_only=True, data_only=True,
        ).active
        for row in header_ws_ro.iter_rows(min_row=1, max_row=15, values_only=True):
            if not row:
                continue
            for col_idx, cell in enumerate(row):
                if cell is None or not isinstance(cell, str):
                    continue
                cell_norm = cell.strip().lower()
                if not cell_norm:
                    continue
                if "сальдо" in cell_norm and ("начал" in cell_norm):
                    if "начало" not in section_markers:
                        section_markers["начало"] = col_idx
                elif "оборот" in cell_norm and ("период" in cell_norm or len(cell_norm) < 30):
                    if "обороты" not in section_markers:
                        section_markers["обороты"] = col_idx
                elif "сальдо" in cell_norm and "конец" in cell_norm:
                    if "конец" not in section_markers:
                        section_markers["конец"] = col_idx
                elif cell_norm == "дебет":
                    debit_cols.append(col_idx)
                elif cell_norm == "кредит":
                    credit_cols.append(col_idx)

        # Дедуплицируем (в иных шаблонах одна и та же колонка может
        # появиться в двух строках).
        debit_cols = sorted(set(debit_cols))
        credit_cols = sorted(set(credit_cols))

        logger.info(
            "[debt_import] header scan: sections=%s, debit_cols=%s, credit_cols=%s",
            section_markers, debit_cols, credit_cols,
        )

        # ─── СТРАТЕГИЯ 0 (самая надёжная): итоговая строка счёта ────
        # В ОСВ 1С первая строка после заголовков — итог по счёту:
        # «209.34» (или «205.X»). У неё ВСЕ 6 колонок заполнены:
        # [Сальдо нач Дт, Сальдо нач Кр, Обороты Дт, Обороты Кр,
        #  Сальдо кон Дт, Сальдо кон Кр]
        # По их позициям однозначно определяем индексы. Это работает даже
        # если section-markers не нашлись (например другой шаблон 1С).
        try:
            account_total_row = None
            account_total_label_col = None
            for row in header_ws_ro.iter_rows(min_row=1, max_row=20, values_only=True):
                if not row:
                    continue
                # Ищем «209.X» или «205.X» в ЛЮБОЙ из первых 3 колонок
                # (в разных шаблонах ОСВ код счёта может быть в col 0, 1 или 2).
                for col_label in range(min(3, len(row))):
                    cell = row[col_label]
                    if cell is None:
                        continue
                    # Преобразуем в строку любой тип (Excel может хранить
                    # "209.34" как float).
                    s = str(cell).strip()
                    if (
                        s.startswith("209.") or s.startswith("205.")
                        or s == "209" or s == "205"
                    ):
                        account_total_row = row
                        account_total_label_col = col_label
                        break
                if account_total_row is not None:
                    break
            if account_total_row is None:
                logger.info("[debt_import] STRATEGY 0: account total row not found")
                row = None  # запасной placeholder
            else:
                row = account_total_row
                first_norm = str(row[account_total_label_col]).strip()
                logger.info(
                    "[debt_import] STRATEGY 0: account_total='%s' at col %d",
                    first_norm, account_total_label_col,
                )
            if row is not None:
                # Собираем колонки с числовыми значениями ПОСЛЕ label-колонки.
                numeric_positions = []
                start_search = account_total_label_col + 1
                for col_idx in range(start_search, len(row)):
                    cell = row[col_idx]
                    if cell is None or cell == "":
                        continue
                    try:
                        d = clean_decimal(cell)
                        if d != 0:
                            numeric_positions.append(col_idx)
                    except Exception:
                        pass
                logger.info(
                    "[debt_import] account_total row found: '%s' at numeric_positions=%s",
                    first_norm, numeric_positions,
                )
                # Ожидаем 6 ненулевых значений. Если меньше — у счёта нет
                # оборотов по какой-то стороне, fallback на меньшее число.
                if len(numeric_positions) >= 4:
                    # Берём по порядку: 0=начДт, 1=начКр, 2=оборДт, 3=оборКр,
                    # 4=концДт, 5=концКр. Если значений меньше 6 — берём
                    # первый и последний как best guess для нач/конец.
                    debt_col_first = numeric_positions[0]
                    overpay_col_first = numeric_positions[1] if len(numeric_positions) > 1 else numeric_positions[0]
                    if len(numeric_positions) >= 6:
                        obor_debit_col = numeric_positions[2]
                        obor_credit_col = numeric_positions[3]
                        debt_col_last = numeric_positions[4]
                        overpay_col_last = numeric_positions[5]
                    elif len(numeric_positions) == 5:
                        obor_debit_col = numeric_positions[2]
                        debt_col_last = numeric_positions[3]
                        overpay_col_last = numeric_positions[4]
                    elif len(numeric_positions) == 4:
                        debt_col_last = numeric_positions[2]
                        overpay_col_last = numeric_positions[3]
                    logger.info(
                        "[debt_import] STRATEGY 0 success: debt=%d/%d, overpay=%d/%d, obor=%s/%s",
                        debt_col_first, debt_col_last, overpay_col_first, overpay_col_last,
                        obor_debit_col, obor_credit_col,
                    )
        except Exception as exc:
            logger.warning("[debt_import] strategy 0 failed: %s", exc)

        # Стратегия 1: все 3 секции + минимум 3 «Дебет»/«Кредит».
        if (
            "начало" in section_markers
            and "обороты" in section_markers
            and "конец" in section_markers
            and len(debit_cols) >= 2
            and len(credit_cols) >= 2
        ):
            sn = section_markers["начало"]
            so = section_markers["обороты"]
            sc = section_markers["конец"]
            # Дебет «начало» — самый ранний «Дебет» ≥ sn и < so.
            cand_d_first = [c for c in debit_cols if sn <= c < so]
            cand_d_last = [c for c in debit_cols if c >= sc]
            cand_c_first = [c for c in credit_cols if sn <= c < so]
            cand_c_last = [c for c in credit_cols if c >= sc]
            if cand_d_first and cand_d_last and cand_c_first and cand_c_last:
                debt_col_first = cand_d_first[0]
                debt_col_last = cand_d_last[0]
                overpay_col_first = cand_c_first[0]
                overpay_col_last = cand_c_last[0]

        # Стратегия 2: если секции не нашли, но есть 3+ «Дебет» — first/last.
        if debt_col_last is None and len(debit_cols) >= 3:
            debt_col_first = debit_cols[0]
            debt_col_last = debit_cols[-1]
        if overpay_col_last is None and len(credit_cols) >= 3:
            overpay_col_first = credit_cols[0]
            overpay_col_last = credit_cols[-1]

        # Стратегия 3: всего 2 «Дебет» — берём first/last из них.
        if debt_col_last is None and len(debit_cols) == 2:
            debt_col_first = debit_cols[0]
            debt_col_last = debit_cols[1]
        if overpay_col_last is None and len(credit_cols) == 2:
            overpay_col_first = credit_cols[0]
            overpay_col_last = credit_cols[1]

        # Стратегия 4: всего 1 «Дебет» (упрощённый отчёт) — обе равны.
        if debt_col_last is None and len(debit_cols) == 1:
            debt_col_first = debt_col_last = debit_cols[0]
        if overpay_col_last is None and len(credit_cols) == 1:
            overpay_col_first = overpay_col_last = credit_cols[0]
    except Exception as e:
        logger.warning(f"Header parse failed, fallback to legacy indices: {e}")

    # Fallback на legacy-индексы если парсер не нашёл колонок.
    if debt_col_first is None:
        debt_col_first = 5
    if overpay_col_first is None:
        overpay_col_first = 6
    # Если в файле всего ОДНА секция «Дебет/Кредит» (упрощённый отчёт) —
    # last совпадает с first, fallback логика тоже сработает корректно.
    if debt_col_last is None:
        debt_col_last = debt_col_first
    if overpay_col_last is None:
        overpay_col_last = overpay_col_first

    logger.info(
        f"Debt import columns: debt_start=col{debt_col_first}/end=col{debt_col_last}, "
        f"overpay_start=col{overpay_col_first}/end=col{overpay_col_last} "
        f"(account={account_type})"
    )

    # Создаём запись в логе ПЕРЕД импортом — чтобы при падении остался след.
    # archive_path / batch_id заполняем сразу — если упадёт на парсинге,
    # админ всё равно сможет скачать оригинальный файл через
    # /debts/import-history/{id}/download.
    import_log = DebtImportLog(
        account_type=account_type,
        # file_name теперь хранит ОРИГИНАЛЬНОЕ имя из upload (если есть),
        # а не uuid с диска — для удобной навигации в истории.
        file_name=(original_file_name or os.path.basename(file_path)) if file_path else None,
        archive_path=file_path if "/debt_archives/" in (file_path or "") else None,
        batch_id=batch_id,
        status="pending",
        started_by_id=started_by_id,
        started_by_username=started_by_username,
    )
    db.add(import_log)
    db.flush()  # получаем id

    try:
        active_period = db.execute(
            select(BillingPeriod).where(BillingPeriod.is_active.is_(True))
        ).scalars().first()

        if not active_period:
            import_log.status = "failed"
            import_log.error = "Нет активного периода для загрузки долгов"
            import_log.completed_at = datetime.now(timezone.utc).replace(tzinfo=None)
            db.commit()
            return {"status": "error", "message": "Нет активного периода для загрузки долгов"}

        import_log.period_id = active_period.id

        # 1. Предзагрузка пользователей (сохраняем id и room_id)
        users_raw = db.execute(select(User).where(User.is_deleted.is_(False))).scalars().all()
        users_map = {normalize_name(u.username): {"id": u.id, "room_id": u.room_id} for u in users_raw}
        users_by_id = {u.id: {"id": u.id, "room_id": u.room_id} for u in users_raw}
        users_keys = list(users_map.keys())

        # 1.1. Предзагрузка алиасов ФИО — общая таблица GSheetsAlias.
        # Если админ ранее привязал «Кондрашов ГА» → user_id=5 (через
        # reassign в долгах ИЛИ в gsheets-импорте) — в этом импорте «Кондрашов ГА»
        # сразу попадёт к user_id=5 без fuzzy. Один alias работает на оба
        # счёта (205 и 209) и на gsheets — потому что таблица универсальная.
        alias_rows = db.execute(
            select(GSheetsAlias.alias_fio_normalized, GSheetsAlias.user_id)
        ).all()
        aliases_map: Dict[str, int] = {}
        for norm, uid in alias_rows:
            if norm and uid in users_by_id:
                aliases_map[norm] = uid

        # 2. Предзагрузка показаний активного периода. Ключ — user_id, не room_id.
        #
        # Bug AG (2026-05-22): в коммуналке (комната с >1 холостяком) долги
        # каждого жильца должны быть СВОИМИ — раньше мы ключевали по room_id
        # и суммировали долги всех жильцов в один reading первого встреченного.
        # Симптом: в комнате 4дв.стр.18/206 живут Муравьев (id=159, заплатил)
        # и Шелюков (id=160, должен 635.92). Импорт делал reset reading'а
        # Муравьева → ADD 0 → ADD 635.92 Шелюкова → debt_209=635.92 на reading
        # с user_id=159. Дашборд SUM'ил по user_id и показывал «Муравьев 635.92»,
        # хотя реально должник — Шелюков.
        #
        # Теперь: каждый жилец = свой reading в активном периоде.
        # Семья (несколько FamilyMember) всё равно остаётся одним user → один
        # reading (как было). Коммуналки двух холостяков = два reading'а с
        # одинаковым room_id, но разными user_id.
        readings_raw = db.execute(
            select(MeterReading).where(MeterReading.period_id == active_period.id)
        ).scalars().all()

        readings_map = {r.user_id: r for r in readings_raw if r.user_id is not None}

        stats = {
            "processed": 0, "updated": 0, "created": 0,
            "contracts_created": 0,  # новые RentalContract из строк «Договор от ...»
            "not_found_users": [], "errors": [], "account": account_type
        }

        # Для парсера договоров: запоминаем последнего сматченного жильца.
        # В ОСВ под каждым ФИО идут строки «Договор от ДД.ММ.ГГГГ № N» —
        # все они принадлежат предыдущему ФИО.
        last_matched_user_id: Optional[int] = None
        # Кеш уже существующих договоров: {user_id: set(number)} — чтобы
        # не делать SELECT для каждой строки.
        existing_contracts_cache: Dict[int, set] = {}

        updates_dict = {}  # reading_id -> reading object (для обновления)
        inserts_dict = {}  # user_id -> reading object (для вставки) — Bug AG: ключ per-user
        fuzzy_cache = {}
        # Bug AG: per-user обнуление — каждый жилец сбрасывает СВОЙ reading в 0
        # перед прибавлением своего долга из 1С. Раньше processed_rooms — после
        # первого жильца комнаты вторые ДОБАВЛЯЛИ к нему, не сбрасывали.
        processed_users = set()

        # snapshot_data: сохраняем ДО-состояние debt_*/overpayment_* по каждому
        # затронутому existing reading. Используется для отката импорта.
        # Новые (inserts) в snapshot не попадают — при undo они просто удалятся.
        snapshot_before = {}  # reading_id -> {debt_209, overpayment_209, debt_205, overpayment_205}
        inserts_reading_ids = []  # для undo (их создали — удалим при откате)

        def get_user_data_optimized(fio: str):
            norm = normalize_name(fio)
            # 1. Точное совпадение username — самое надёжное
            if norm in users_map:
                return users_map[norm]
            # 2. Алиас (запомненная админом привязка из прошлых reassign).
            # Работает для всех типов импорта (205, 209, gsheets) через
            # общую таблицу GSheetsAlias.
            if norm in aliases_map:
                uid = aliases_map[norm]
                cached = users_by_id.get(uid)
                if cached:
                    return cached
            if norm in fuzzy_cache:
                return fuzzy_cache[norm]

            match = process.extractOne(norm, users_keys, scorer=fuzz.token_sort_ratio)
            if match:
                best_match_name, score, _ = match
                if score >= _threshold():
                    found_data = users_map[best_match_name]
                    fuzzy_cache[norm] = found_data
                    return found_data

            fuzzy_cache[norm] = None
            return None

        # 3. Чтение строк Excel.
        # Минимальная длина row = max(всех колонок) + 1 чтобы индексация
        # не вышла за границу. pick_saldo_value определена на module-level
        # для покрытия unit-тестами.
        min_row_len = max(
            debt_col_first, debt_col_last, overpay_col_first, overpay_col_last
        ) + 1
        for row in worksheet.iter_rows(min_row=8, values_only=True):
            if not row or len(row) < min_row_len:
                continue

            name_cell = row[0]

            # Проверяем «Договор от ... № ...» — это под-строка для
            # ПОСЛЕДНЕГО сматченного жильца. Парсим и создаём
            # RentalContract если ещё нет.
            contract_data = parse_contract_line(name_cell)
            if contract_data and last_matched_user_id:
                # Lazy-load existing для этого юзера
                if last_matched_user_id not in existing_contracts_cache:
                    rows_db = db.execute(
                        select(RentalContract.number).where(
                            RentalContract.user_id == last_matched_user_id
                        )
                    ).all()
                    existing_contracts_cache[last_matched_user_id] = {
                        r[0] for r in rows_db if r[0]
                    }
                if contract_data["number"] not in existing_contracts_cache[last_matched_user_id]:
                    db.add(RentalContract(
                        user_id=last_matched_user_id,
                        number=contract_data["number"],
                        signed_date=contract_data["signed_date"],
                        is_active=True,
                        note=f"импортировано из 1С ОСВ (счёт {account_type})",
                    ))
                    existing_contracts_cache[last_matched_user_id].add(
                        contract_data["number"]
                    )
                    stats["contracts_created"] += 1
                continue

            if not is_valid_name_row(name_cell):
                continue

            fio_raw = str(name_cell).strip()
            stats["processed"] += 1

            # Сальдо на конец (актуальный долг/переплата на дату отчёта).
            # pick_saldo_pair — единая функция для пары (debt, overpay):
            # она корректно обрабатывает кейс «жилец заплатил больше, конец
            # Дебет пустой, конец Кредит = переплата» (Глоба) — раньше старая
            # pick_saldo_value падала в fallback на «Сальдо начало Дебет».
            debt_val, over_val = pick_saldo_pair(
                row,
                end_debit_col=debt_col_last,
                end_credit_col=overpay_col_last,
                start_debit_col=debt_col_first,
                start_credit_col=overpay_col_first,
                obor_debit_col=obor_debit_col,
                obor_credit_col=obor_credit_col,
            )

            # Bug V: извлекаем обороты для UI «движение средств».
            def _read_cell_dec(col):
                if col is None or col >= len(row):
                    return Decimal("0")
                v = row[col]
                if v is None or v == "":
                    return Decimal("0")
                try:
                    return clean_decimal(v)
                except Exception:
                    return Decimal("0")
            obor_d_val = _read_cell_dec(obor_debit_col)
            obor_c_val = _read_cell_dec(obor_credit_col)

            user_data = get_user_data_optimized(fio_raw)

            if not user_data or not user_data["room_id"]:
                # Сохраняем dict вместо плоской строки — суммы нужны фронту
                # чтобы при reassign автоматически перенести долг к жильцу.
                # Раньше: stats["not_found_users"].append(fio_raw) → суммы
                # терялись, админ должен был вводить руками (часто 0).
                stats["not_found_users"].append({
                    "fio": fio_raw,
                    "debt": str(debt_val),
                    "overpayment": str(over_val),
                })
                # Сбрасываем контекст — следующая строка «Договор» не
                # должна привязаться к ранее сматченному жильцу.
                last_matched_user_id = None
                continue

            user_id = user_data["id"]
            room_id = user_data["room_id"]
            last_matched_user_id = user_id  # для парсера договоров ниже

            # 4. Если у ЭТОГО ЖИЛЬЦА уже есть reading в активном периоде — апдейтим.
            # Bug AG: раньше ключ был room_id и долги нескольких жильцов
            # одной комнаты ВСЕ присваивались reading'у первого. Теперь
            # каждый user имеет свой reading со своим долгом.
            if user_id in readings_map:
                reading = readings_map[user_id]

                # Первый раз встречаем этого жильца в файле — снимок до + reset.
                # (Несколько строк на одного user'а в файле — редкий кейс,
                # но если случится, второй раз reset не делаем, просто прибавим.)
                if user_id not in processed_users:
                    if reading.id not in snapshot_before:
                        snapshot_before[reading.id] = {
                            "debt_209": str(reading.debt_209 or 0),
                            "overpayment_209": str(reading.overpayment_209 or 0),
                            "debt_205": str(reading.debt_205 or 0),
                            "overpayment_205": str(reading.overpayment_205 or 0),
                        }
                    if account_type == "209":
                        reading.debt_209 = Decimal("0.00")
                        reading.overpayment_209 = Decimal("0.00")
                        reading.obor_debit_209 = Decimal("0.00")
                        reading.obor_credit_209 = Decimal("0.00")
                    elif account_type == "205":
                        reading.debt_205 = Decimal("0.00")
                        reading.overpayment_205 = Decimal("0.00")
                        reading.obor_debit_205 = Decimal("0.00")
                        reading.obor_credit_205 = Decimal("0.00")
                    processed_users.add(user_id)
                    updates_dict[reading.id] = reading

                if account_type == "209":
                    reading.debt_209 += debt_val
                    reading.overpayment_209 += over_val
                    reading.obor_debit_209 = (reading.obor_debit_209 or Decimal("0")) + obor_d_val
                    reading.obor_credit_209 = (reading.obor_credit_209 or Decimal("0")) + obor_c_val
                elif account_type == "205":
                    reading.debt_205 += debt_val
                    reading.overpayment_205 += over_val
                    reading.obor_debit_205 = (reading.obor_debit_205 or Decimal("0")) + obor_d_val
                    reading.obor_credit_205 = (reading.obor_credit_205 or Decimal("0")) + obor_c_val

                stats["updated"] += 1

            # 5. Reading'а в БД нет, но мы его уже создали в памяти на этой
            # итерации (несколько строк на одного user'а в файле — редкий кейс).
            elif user_id in inserts_dict:
                reading = inserts_dict[user_id]
                if account_type == "209":
                    reading.debt_209 += debt_val
                    reading.overpayment_209 += over_val
                    reading.obor_debit_209 = (reading.obor_debit_209 or Decimal("0")) + obor_d_val
                    reading.obor_credit_209 = (reading.obor_credit_209 or Decimal("0")) + obor_c_val
                elif account_type == "205":
                    reading.debt_205 += debt_val
                    reading.overpayment_205 += over_val
                    reading.obor_debit_205 = (reading.obor_debit_205 or Decimal("0")) + obor_d_val
                    reading.obor_credit_205 = (reading.obor_credit_205 or Decimal("0")) + obor_c_val

                stats["updated"] += 1

            # 6. Reading'а нет вообще — создаём новый ИМЕННО для ЭТОГО жильца.
            else:
                new_reading = MeterReading(
                    user_id=user_id,
                    room_id=room_id,
                    period_id=active_period.id,
                    is_approved=False,
                    debt_209=Decimal("0.00"), overpayment_209=Decimal("0.00"),
                    debt_205=Decimal("0.00"), overpayment_205=Decimal("0.00"),
                    obor_debit_209=Decimal("0.00"), obor_credit_209=Decimal("0.00"),
                    obor_debit_205=Decimal("0.00"), obor_credit_205=Decimal("0.00"),
                )

                if account_type == "209":
                    new_reading.debt_209 = debt_val
                    new_reading.overpayment_209 = over_val
                    new_reading.obor_debit_209 = obor_d_val
                    new_reading.obor_credit_209 = obor_c_val
                elif account_type == "205":
                    new_reading.debt_205 = debt_val
                    new_reading.overpayment_205 = over_val
                    new_reading.obor_debit_205 = obor_d_val
                    new_reading.obor_credit_205 = obor_c_val

                inserts_dict[user_id] = new_reading
                processed_users.add(user_id)
                stats["created"] += 1

        # 7. Сохраняем в БД
        if inserts_dict:
            db.add_all(list(inserts_dict.values()))
            db.flush()  # получаем id для snapshot/undo
            inserts_reading_ids = [r.id for r in inserts_dict.values()]

        if updates_dict:
            # ИСПРАВЛЕНИЕ (may 2026): раньше использовался
            # db.bulk_update_mappings(MeterReading, updates_list). Но
            # MeterReading партиционирована по created_at, и bulk_update
            # тихо возвращает rowcount=0 даже при правильно переданном
            # составном PK (id, created_at). Импорт писал «completed»,
            # applied_state в логе показывал debt_205=10200, но в БД
            # ничего не менялось.
            #
            # Сценарий бага (Лучка А.П., прод 2026-05-21):
            #   1) 209 импорт → reading у Лучки не было → insert через
            #      db.add_all → debt_209=21889 сохранилось ✓
            #   2) 205 импорт → reading уже есть (создан в шаге 1) →
            #      updates_dict → bulk_update_mappings → молча rowcount=0
            #      → debt_205 в БД остался 0 ✗
            #
            # Now: explicit per-row update по id (SERIAL уникален без
            # created_at — partition pruning теряется, но импорт делается
            # 1-2 раза в месяц, скорость не критична). Стратегия:
            #   1) Snapshot final values из ORM объектов в dict (до
            #      любого expunge — иначе lazy-load перезаписал бы их).
            #   2) Expunge объекты из session — отвязываем от ORM, чтобы
            #      session.commit() не пытался их сам flush'ить (на
            #      партиционированной таблице ORM flush тоже тихо
            #      проваливается и могло маскировать мой execute).
            #   3) execute(update().where(id=).values()) — пишем в БД.
            #   4) Логируем rowcount=affected; если requested!=affected —
            #      WARNING (видно сразу в docker logs).
            from sqlalchemy import update as _sa_update

            # 1) snapshot
            final_values = []
            for r in updates_dict.values():
                final_values.append({
                    "id": r.id,
                    "debt_209": r.debt_209,
                    "overpayment_209": r.overpayment_209,
                    "debt_205": r.debt_205,
                    "overpayment_205": r.overpayment_205,
                })

            # 2) expunge — отвязываем от ORM-session (для безопасности
            # на партиционированной MR — чтобы ORM flush не перезаписал
            # наши explicit UPDATE'ы своим тихим no-op).
            for r in list(updates_dict.values()):
                try:
                    db.expunge(r)
                except Exception:
                    pass  # объект уже мог быть detached

            # 3) write
            total_affected = 0
            for upd in final_values:
                res = db.execute(
                    _sa_update(MeterReading)
                    .where(MeterReading.id == upd["id"])
                    .values(
                        debt_209=upd["debt_209"],
                        overpayment_209=upd["overpayment_209"],
                        debt_205=upd["debt_205"],
                        overpayment_205=upd["overpayment_205"],
                    )
                )
                total_affected += res.rowcount or 0

            # 4) lol/alert
            if total_affected != len(final_values):
                logger.warning(
                    "[DEBT-IMPORT] %s requested=%d but ONLY affected=%d. "
                    "Это значит UPDATE не записал часть строк — копать "
                    "(партиционирование, триггеры, locks). log_id=%d",
                    account_type, len(final_values), total_affected, import_log.id,
                )
            else:
                logger.info(
                    "[DEBT-IMPORT] %s updated rows OK: requested=%d affected=%d (log_id=%d)",
                    account_type, len(final_values), total_affected, import_log.id,
                )

        # 8. Финализируем DebtImportLog в той же транзакции.
        # Dedup по ФИО — set() на dict не работает, поэтому через {fio: dict}.
        # Сохраняем ПОСЛЕДНЕЕ вхождение каждого ФИО (если в xlsx несколько
        # строк с одним ФИО — обычно для семьи в одной комнате — берём
        # последнее, так делает и основной цикл).
        seen: dict[str, dict] = {}
        for item in stats["not_found_users"]:
            key = item["fio"].strip().lower()
            seen[key] = item
        stats["not_found_users"] = list(seen.values())
        import_log.status = "completed"
        import_log.processed = stats["processed"]
        import_log.updated = stats["updated"]
        import_log.created = stats["created"]
        import_log.not_found_count = len(stats["not_found_users"])
        import_log.not_found_users = stats["not_found_users"][:2000]  # защита от гигантских файлов
        import_log.snapshot_data = {
            "before": snapshot_before,
            "inserted_reading_ids": inserts_reading_ids,
        }

        # applied_state — state ПОСЛЕ применения импорта, для последующего
        # diff. Собираем denormalized {user_id: {долги, username, room_label, room_id}}
        # по всем затронутым reading'ам (updates + inserts).
        #
        # Bug AG: ключ — user_id (раньше room_id, в коммуналке два жильца
        # перезаписывали друг друга). Diff/undo тоже ходят через user_id.
        applied_state: dict[str, dict] = {}
        all_touched_readings = list(updates_dict.values()) + list(inserts_dict.values())
        if all_touched_readings:
            room_ids = list({r.room_id for r in all_touched_readings if r.room_id})
            user_ids = list({r.user_id for r in all_touched_readings if r.user_id})
            rooms_map = {}
            users_map_id = {}
            if room_ids:
                from app.modules.utility.models import Room as _Room
                rooms_rows = db.execute(
                    select(_Room).where(_Room.id.in_(room_ids))
                ).scalars().all()
                rooms_map = {r.id: r for r in rooms_rows}
            if user_ids:
                users_rows = db.execute(
                    select(User).where(User.id.in_(user_ids))
                ).scalars().all()
                users_map_id = {u.id: u for u in users_rows}

            for r in all_touched_readings:
                if not r.user_id:
                    continue
                room = rooms_map.get(r.room_id) if r.room_id else None
                user = users_map_id.get(r.user_id)
                applied_state[str(r.user_id)] = {
                    "debt_209": str(r.debt_209 or 0),
                    "overpayment_209": str(r.overpayment_209 or 0),
                    "debt_205": str(r.debt_205 or 0),
                    "overpayment_205": str(r.overpayment_205 or 0),
                    "username": user.username if user else None,
                    "room_id": r.room_id,
                    "room_label": (
                        f"{room.dormitory_name} / {room.room_number}"
                        if room else None
                    ),
                }
        import_log.applied_state = applied_state

        import_log.completed_at = datetime.now(timezone.utc).replace(tzinfo=None)
        import_log.id  # нужен id для stats

        db.commit()

        stats["log_id"] = import_log.id
        logger.info(
            "Import finished. Log=%s Processed: %s, Updated: %s, Created: %s",
            import_log.id, stats["processed"], stats["updated"], stats["created"],
        )
        return stats

    except Exception as error:
        # Полный откат: ни одна строка не должна остаться полузакоммиченной.
        db.rollback()
        logger.exception("Import failed — full rollback applied")
        # Пробуем отметить лог как failed в новой транзакции
        try:
            failed_log = DebtImportLog(
                account_type=account_type,
                file_name=(original_file_name or os.path.basename(file_path)) if file_path else None,
                archive_path=file_path if "/debt_archives/" in (file_path or "") else None,
                batch_id=batch_id,
                status="failed",
                started_by_id=started_by_id,
                started_by_username=started_by_username,
                error=str(error)[:2000],
                completed_at=datetime.now(timezone.utc).replace(tzinfo=None),
            )
            db.add(failed_log)
            db.commit()
        except Exception:
            db.rollback()
        # Пробрасываем — Celery отправит в retry (см. retry_kwargs у задачи).
        raise RuntimeError(f"Ошибка во время импорта: {error}") from error

    finally:
        # workbook.close() — даже при исключении. Иначе файл-дескриптор висит.
        try:
            workbook.close()
        except Exception:
            pass
