"""Импорт показаний из Excel с полным анализом и прямым утверждением.

Сценарий (2026-06-15): админ загружает Excel с парой показаний на каждого
жильца — «предыдущий месяц» (база) и «текущий месяц». Система сопоставляет
ФИО→жилец, прогоняет ВСЕ анализаторы на каждого, показывает повердиктную
таблицу, админ разбирает поштучно и жмёт «Утвердить» → создаются
УТВЕРЖДЁННЫЕ MeterReading сразу в финотчётность (минуя подачу жильца).

Формат Excel: листы по ресурсам («горячая»/«холодная»/«электричество»),
3 колонки: ФИО | Предыдущий месяц | Текущий месяц. Значения — целые м³/кВт.

БАЗА РАСЧЁТА — «предыдущий месяц» ИЗ EXCEL (а не из БД): расход = текущий −
предыдущий из колонок. Это решение пользователя: Excel самодостаточен.

НЕ ПОДАЛИ (текущий пуст) — начисление по нормативу (_growing_norm_volumes,
та же проверенная математика, что у авто-добивки; норма на квартиру, ×1 без
санкции). Электричество в этом импорте не начисляется (его вносят электрики
отдельно через админку) — пока в Excel нет листа «электричество».
"""
from __future__ import annotations

import asyncio
import calendar
import io
import logging
from datetime import datetime
from decimal import Decimal
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.modules.utility.models import (
    BillingPeriod, GSheetsImportRow, MeterReading, Room, Tariff, User, Adjustment,
)
from app.modules.utility.services.calculations import (
    calculate_utilities, costs_for_model_fields, CalculationError, D,
)

logger = logging.getLogger(__name__)
ZERO = Decimal("0")

# Маркер источника — чтобы реестр/аналитика видели, что это импорт Excel.
EXCEL_FLAG = "EXCEL_IMPORT"

# Классификация листов по нормализованному названию.
_RES_KEYS = ("hot", "cold", "elect")


def _sheet_kind(title: str) -> Optional[str]:
    t = (title or "").strip().lower()
    if "горяч" in t or "гвс" in t:
        return "hot"
    if "холод" in t or "хвс" in t:
        return "cold"
    if "электр" in t or "свет" in t or "квт" in t:
        return "elect"
    return None


def _num(v) -> Optional[Decimal]:
    """Ячейка → Decimal или None (пусто/мусор)."""
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return D(v)
    s = str(v).strip().replace(",", ".").replace(" ", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _is_junk_fio(v) -> bool:
    """Строки-разделители/итоги/пустые — пропускаем."""
    if v is None:
        return True
    s = str(v).strip()
    if not s or s == "0":
        return True
    low = s.lower()
    if low.startswith("итого") or low.startswith("этаж") or "общежити" in low:
        return True
    if low in ("ф.и.о.", "фио", "ф.и.о"):
        return True
    # Чисто числовая/символьная строка (не ФИО).
    if all(not c.isalpha() for c in s):
        return True
    return False


def parse_readings_workbook(content: bytes) -> dict:
    """Парсит Excel → {key: {fio, hot:{prev,cur}, cold:{...}, elect:{...}}}.

    Ключ — нормализованное ФИО (объединяет одного человека из разных листов).
    Возвращает также список ресурсов-листов и счётчик пропущенных строк.
    """
    from app.modules.utility.services.gsheets_sync import normalize_fio

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    people: dict[str, dict] = {}
    meters_present: list[str] = []
    skipped = 0
    sheet_map: dict[str, str] = {}

    for ws in wb.worksheets:
        kind = _sheet_kind(ws.title)
        if not kind:
            continue
        sheet_map[ws.title] = kind
        if kind not in meters_present:
            meters_present.append(kind)
        for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
            if not row:
                continue
            fio_raw = row[0] if len(row) > 0 else None
            if _is_junk_fio(fio_raw):
                skipped += 1
                continue
            fio = str(fio_raw).strip()
            key = normalize_fio(fio)
            if not key:
                skipped += 1
                continue
            prev = _num(row[1] if len(row) > 1 else None)
            cur = _num(row[2] if len(row) > 2 else None)
            rec = people.setdefault(key, {"fio": fio, "hot": {}, "cold": {}, "elect": {}})
            # Первое читаемое написание ФИО оставляем как отображаемое.
            rec[kind] = {"prev": prev, "cur": cur}

    try:
        wb.close()
    except Exception:
        pass

    return {
        "people": people,
        "meters_present": meters_present,
        "sheets": sheet_map,
        "skipped_rows": skipped,
    }


# =====================================================================
# Расчёт стоимости (единый для preview и commit — не дублируем формулу)
# =====================================================================

def _seasonal_for_tariff(seasonal, tariff: Tariff) -> tuple[bool, bool]:
    """Эффективные сезонные флаги: глобальный override И per-tariff (даты)."""
    heating = seasonal.heating_season_active and tariff.is_heating_active_now()
    hw = seasonal.hot_water_heating_active and tariff.is_hw_heating_active_now()
    return heating, hw


# =====================================================================
# Сверка с Google Sheets (буфер GSheetsImportRow) за окно вокруг периода
# =====================================================================

def _gsheets_window(period_name: Optional[str]) -> Optional[tuple[datetime, datetime]]:
    """Окно сверки по выбранному месяцу: с 16-го ПОЗАПРОШЛОГО месяца по конец
    выбранного. Май → 16 марта … 31 мая (май+апрель целиком + 2-я половина
    марта). Без парсящегося имени периода — None."""
    from app.modules.utility.services.period_helpers import parse_period_name
    parsed = parse_period_name(period_name or "")
    if not parsed:
        return None
    y, m = parsed
    mm2, yy2 = m - 2, y
    while mm2 <= 0:
        mm2 += 12
        yy2 -= 1
    start = datetime(yy2, mm2, 16, 0, 0, 0)
    end = datetime(y, m, calendar.monthrange(y, m)[1], 23, 59, 59)
    return start, end


async def _load_gsheets_lookup(
    db: AsyncSession, window: Optional[tuple[datetime, datetime]]
) -> tuple[dict[str, dict], dict[int, dict]]:
    """Буфер GSheets за окно → два индекса: по нормализованному ФИО и по
    matched_user_id. На каждого — последняя подача (по sheet_timestamp) +
    счётчик. Пусто, если окна нет."""
    if not window:
        return {}, {}
    from app.modules.utility.services.gsheets_sync import normalize_fio
    start, end = window
    rows = (await db.execute(
        select(GSheetsImportRow).where(
            GSheetsImportRow.sheet_timestamp.is_not(None),
            GSheetsImportRow.sheet_timestamp >= start,
            GSheetsImportRow.sheet_timestamp <= end,
        ).order_by(GSheetsImportRow.sheet_timestamp.asc())
    )).scalars().all()
    by_fio: dict[str, dict] = {}
    by_uid: dict[int, dict] = {}
    for r in rows:
        rec = {
            "hot": (float(r.hot_water) if r.hot_water is not None else None),
            "cold": (float(r.cold_water) if r.cold_water is not None else None),
            "date": r.sheet_timestamp.strftime("%d.%m.%Y") if r.sheet_timestamp else None,
            "raw_fio": r.raw_fio,
        }
        key = normalize_fio(r.raw_fio or "")
        # asc по дате → последняя подача перезаписывает; копим count.
        if key:
            rec["count"] = by_fio.get(key, {}).get("count", 0) + 1
            by_fio[key] = rec
        if r.matched_user_id:
            rec2 = dict(rec)
            rec2["count"] = by_uid.get(r.matched_user_id, {}).get("count", 0) + 1
            by_uid[r.matched_user_id] = rec2
    return by_fio, by_uid


def _gsheets_for_row(row: dict, by_fio: dict, by_uid: dict) -> dict:
    """Сводка GSheets для строки Excel + пометка расхождений с Excel-текущим.
    Связь по ФИО — нормализуем сам ФИО строки (не ключ — он может быть
    синтетическим при пересчёте)."""
    from app.modules.utility.services.gsheets_sync import normalize_fio
    m = row.get("matched") or {}
    fio_key = normalize_fio(row.get("fio") or "")
    gs = None
    if m.get("user_id") and by_uid.get(m["user_id"]):
        gs = by_uid[m["user_id"]]
    elif fio_key and by_fio.get(fio_key):
        gs = by_fio[fio_key]
    if not gs:
        return {"present": False}
    # Расхождение с Excel-текущим (целые м³ — сравниваем точно).
    cur_hot = (row.get("hot") or {}).get("cur")
    cur_cold = (row.get("cold") or {}).get("cur")
    mis_hot = (gs["hot"] is not None and cur_hot is not None
               and abs(float(gs["hot"]) - float(cur_hot)) > 0.001)
    mis_cold = (gs["cold"] is not None and cur_cold is not None
                and abs(float(gs["cold"]) - float(cur_cold)) > 0.001)
    return {
        "present": True, "hot": gs["hot"], "cold": gs["cold"],
        "date": gs["date"], "count": gs.get("count", 1),
        "mismatch": bool(mis_hot or mis_cold),
        "mismatch_hot": bool(mis_hot), "mismatch_cold": bool(mis_cold),
    }


async def _load_adjustments(
    db: AsyncSession, user_ids: list[int], period_id: Optional[int]
) -> dict[int, dict[str, Decimal]]:
    """Корректировки 209/205 по жильцам за период — одним запросом."""
    out: dict[int, dict[str, Decimal]] = {}
    if not user_ids or not period_id:
        return out
    res = await db.execute(
        select(Adjustment.user_id, Adjustment.account_type, func.sum(Adjustment.amount))
        .where(Adjustment.user_id.in_(set(user_ids)), Adjustment.period_id == period_id)
        .group_by(Adjustment.user_id, Adjustment.account_type)
    )
    for uid, acc, amount in res.all():
        out.setdefault(uid, {})[acc] = amount or ZERO
    return out


def _norm_volumes(tariff: Tariff, user: User, room: Room) -> tuple[Decimal, Decimal, Decimal]:
    """Объёмы по нормативу для не подавших — через ПРОВЕРЕННУЮ функцию
    авто-добивки (_growing_norm_volumes, miss_count=0 → без санкции ×1)."""
    from app.modules.utility.services.billing import _growing_norm_volumes
    residents = D(user.residents_count or 1)
    vol_hot, vol_cold, vol_el, _coef = _growing_norm_volumes(
        tariff, residents, miss_count=0, room=room,
    )
    total = D(room.total_room_residents if room and room.total_room_residents else 1)
    share_el = max(ZERO, (residents / total) * vol_el) if total > ZERO else ZERO
    return vol_hot, vol_cold, share_el


def _consumption_volumes(
    user: User, room: Room, cur: dict, prev: dict
) -> tuple[Decimal, Decimal, Decimal]:
    """Объёмы по показаниям: дельта = max(0, текущий − предыдущий из Excel).
    Электро — доля (residents/total). Отсутствующий ресурс → 0."""
    def _delta(res: str) -> Decimal:
        c = cur.get(res)
        p = prev.get(res)
        if c is None or p is None:
            return ZERO
        return max(ZERO, D(c) - D(p))

    d_hot = _delta("hot")
    d_cold = _delta("cold")
    d_el_raw = _delta("elect")
    residents = D(user.residents_count or 1)
    total = D(room.total_room_residents if room and room.total_room_residents else 1)
    share_el = max(ZERO, (residents / total) * d_el_raw) if total > ZERO else ZERO
    return d_hot, d_cold, share_el


class _RoomMeterProxy:
    """Прокси комнаты с принудительными has_*_meter. Для поданных строк
    импорта: начисляем СТРОГО по показаниям из Excel (has_*=True → берётся
    переданный объём; ресурс не в Excel → объём 0 → стоимость 0), не давая
    calculate_utilities подменить норматив при has_*_meter=False у комнаты."""
    __slots__ = ("_room",)

    def __init__(self, room):
        object.__setattr__(self, "_room", room)

    def __getattr__(self, name):
        if name in ("has_hw_meter", "has_cw_meter", "has_el_meter"):
            return True
        return getattr(self._room, name)


def _compute_costs(
    user: User, room: Room, tariff: Tariff,
    vol_hot: Decimal, vol_cold: Decimal, share_el: Decimal,
    seasonal, adj: dict[str, Decimal], force_meters: bool = False,
) -> dict:
    """calculate_utilities + total_209/205 (с корректировками). Чистый —
    сезонные/корректировки переданы (загружены батчем выше).

    force_meters=True (поданные строки): начисляем строго по переданным
    объёмам — электричество/вода из Excel считаются, даже если у комнаты
    флаг счётчика снят; ресурс без данных → объём 0 → 0₽ (без норматива)."""
    heating, hw = _seasonal_for_tariff(seasonal, tariff)
    calc_room = _RoomMeterProxy(room) if force_meters else room
    costs = calculate_utilities(
        user=user, room=calc_room, tariff=tariff,
        volume_hot=vol_hot, volume_cold=vol_cold,
        volume_sewage=vol_hot + vol_cold, volume_electricity_share=share_el,
        heating_season_active=heating, hot_water_heating_active=hw,
    )
    cost_rent = costs["cost_social_rent"]
    total_209 = (costs["total_cost"] - cost_rent) + (adj.get("209") or ZERO)
    total_205 = cost_rent + (adj.get("205") or ZERO)
    costs["total_209"] = total_209
    costs["total_205"] = total_205
    costs["grand_total"] = total_209 + total_205
    return costs


# =====================================================================
# PREVIEW — сопоставление + анализ каждого человека (без записи в БД)
# =====================================================================

def _build_match_indexes_sync() -> tuple[dict, list, dict, dict]:
    """Индексы матчинга (plain-dict'ы, без привязки к сессии) — строим в
    СИНХРОННОЙ сессии: build_users_index/build_aliases_index работают с sync
    Session (как gisgmp/gsheets). Зовётся через asyncio.to_thread из async."""
    from app.core.database import sync_db_session
    from app.modules.utility.services.gsheets_sync import (
        build_users_index, build_aliases_index,
    )
    with sync_db_session() as s:
        umap, ukeys, ubyid = build_users_index(s)
        amap = build_aliases_index(s)
    return umap, ukeys, ubyid, amap


async def build_preview(
    db: AsyncSession, parsed: dict, period_id: Optional[int],
    forced_match: Optional[dict] = None,
) -> dict:
    """На каждого человека из Excel: матч ФИО→жилец, прогон анализаторов,
    предварительная сумма, агрегированный вердикт. Ничего не пишет.

    forced_match: {key → user_id} — для строк, где админ уже назначил жильца
    (переназначение/создание/правка ФИО при пересчёте) — fuzzy пропускаем."""
    from app.modules.utility.services.gsheets_sync import match_user, _fuzzy_threshold
    from app.modules.utility.services.reading_validators import validate_meter_reading
    from app.modules.utility.services.tariff_cache import tariff_cache

    forced_match = forced_match or {}
    # Матчер синхронный (sync Session) — строим индексы в потоке. match_user
    # чистый (без БД), зовём прямо в async-цикле.
    users_map, users_keys, users_by_id, aliases_map = await asyncio.to_thread(
        _build_match_indexes_sync
    )
    fuzzy_thr = _fuzzy_threshold()

    # ── Фаза 1: матч всех ФИО (чисто), сбор найденных user_id ──
    matched_rows: list[tuple[dict, dict, int, bool]] = []  # (row, info, score, conflict)
    items: list[dict] = []
    counts = {"ok": 0, "warning": 0, "error": 0, "unmatched": 0, "norm": 0}
    matched_ids: set[int] = set()

    for key, rec in parsed["people"].items():
        fio = rec["fio"]
        forced_uid = forced_match.get(key)
        if forced_uid:
            info = users_by_id.get(forced_uid) or {"id": forced_uid, "username": fio}
            score, conflict = 100, None
        else:
            info, score, conflict = match_user(
                fio, None, users_map, users_keys, users_by_id, aliases_map, fuzzy=True,
            )
        row: dict = {
            "key": key, "fio": fio, "score": score,
            "hot": rec.get("hot") or {}, "cold": rec.get("cold") or {},
            "elect": rec.get("elect") or {},
            "reasons": [], "matched": None, "verdict": "unmatched",
            "status": "skip", "preview_total": None,
        }
        if not info or score < fuzzy_thr:
            row["reasons"].append(
                "Жилец не найден" + (f" (лучшее совпадение {score}%)" if score else "")
            )
            counts["unmatched"] += 1
            items.append(row)
            continue
        if conflict:
            row["reasons"].append(conflict)
        matched_ids.add(info["id"])
        matched_rows.append((row, info, score, bool(conflict)))

    # ── Батч-загрузка: жильцы с комнатами, сезонные настройки, корректировки ──
    from app.modules.utility.routers.settings import _load_seasonal
    from app.modules.utility.services.reading_validators import validate_total_cost
    seasonal = await _load_seasonal(db)
    # Сверка с буфером Google Sheets за окно вокруг выбранного месяца.
    period = await db.get(BillingPeriod, period_id) if period_id else None
    window = _gsheets_window(period.name if period else None)
    gs_by_fio, gs_by_uid = await _load_gsheets_lookup(db, window)
    users_by_id_orm: dict[int, User] = {}
    if matched_ids:
        for u in (await db.execute(
            select(User).options(selectinload(User.room)).where(User.id.in_(matched_ids))
        )).scalars().all():
            users_by_id_orm[u.id] = u
    adj_by_user = await _load_adjustments(db, list(matched_ids), period_id)
    fallback_tariff = (await db.execute(
        select(Tariff).where(Tariff.is_active)
    )).scalars().first()

    # ── Фаза 2: анализ каждого сматченного (без запросов в цикле) ──
    for row, info, score, conflict in matched_rows:
        rec = parsed["people"][row["key"]]
        user = users_by_id_orm.get(info["id"])
        room = user.room if user else None
        if not user or not room:
            row["reasons"].append("Жилец без помещения — нельзя начислить")
            row["matched"] = {"user_id": info.get("id"), "username": info.get("username"), "score": score}
            row["verdict"] = "unmatched"
            counts["unmatched"] += 1
            items.append(row)
            continue

        tariff = tariff_cache.get_effective_tariff(user=user, room=room) or fallback_tariff
        row["matched"] = {
            "user_id": user.id, "username": user.username,
            "room": (room.room_number or room.apartment_number),
            "dormitory": room.dormitory_name,
            "tariff": tariff.name if tariff else None,
            "score": score, "conflict": bool(conflict),
            "residents": user.residents_count or 1,
        }

        # Какие ресурсы реально считаем (есть счётчик у комнаты).
        def _has(attr: str) -> bool:
            v = getattr(room, attr, None)
            return bool(v) if v is not None else True
        has = {"hot": _has("has_hw_meter"), "cold": _has("has_cw_meter"), "elect": _has("has_el_meter")}

        cur = {r: (rec.get(r) or {}).get("cur") for r in _RES_KEYS}
        prev = {r: (rec.get(r) or {}).get("prev") for r in _RES_KEYS}

        # Подал ли хоть один метрируемый ресурс (есть текущее показание)?
        submitted = any(cur[r] is not None for r in _RES_KEYS if has[r])
        verdict = "ok"

        # Откат счётчика (текущий < предыдущего) — флаг на каждый ресурс.
        for r in _RES_KEYS:
            if has[r] and cur[r] is not None and prev[r] is not None and D(cur[r]) < D(prev[r]):
                row["reasons"].append(
                    f"{_res_label(r)}: счётчик уменьшился ({prev[r]}→{cur[r]}) — расход не начислим"
                )
                verdict = "warning"

        # Главный sanity-валидатор (потолок/дельта) — там, где оба значения есть.
        if submitted:
            vr = validate_meter_reading(
                hot=cur["hot"] if (has["hot"] and cur["hot"] is not None) else None,
                cold=cur["cold"] if (has["cold"] and cur["cold"] is not None) else None,
                elect=cur["elect"] if (has["elect"] and cur["elect"] is not None) else None,
                prev_hot=prev["hot"], prev_cold=prev["cold"], prev_elect=prev["elect"],
                is_baseline=False,
            )
            for e in vr.errors:
                # Монотонность уже показали выше своим текстом; не дублируем.
                if "не может уменьш" in e or "убыва" in e:
                    continue
                if "не задан" in e:
                    continue  # частичная подача (только вода/только свет) — норм
                # Электричество — счётчик КУМУЛЯТИВНЫЙ (десятки тысяч кВт),
                # абсолютный потолок (рассчитан на воду в м³) для него ложный.
                # Реальная защита — дельта (расход), она проверяется отдельно.
                if e.startswith("electricity=") and "превышает максимум" in e:
                    continue
                row["reasons"].append(e)
                verdict = "error"

        try:
            if submitted:
                vol_hot, vol_cold, share_el = _consumption_volumes(user, room, cur, prev)
                row["status"] = "submitted"
            else:
                vol_hot, vol_cold, share_el = _norm_volumes(tariff, user, room)
                row["status"] = "norm"
                row["reasons"].append("Не подал показания — начислим по нормативу")
                if verdict == "ok":
                    verdict = "warning"
                counts["norm"] += 1
            costs = _compute_costs(user, room, tariff, vol_hot, vol_cold, share_el,
                                   seasonal, adj_by_user.get(user.id, {}),
                                   force_meters=submitted)
            row["preview_total"] = float(costs["grand_total"])
            row["preview_209"] = float(costs["total_209"])
            row["preview_205"] = float(costs["total_205"])
            # Sanity на итог.
            tc = validate_total_cost(costs["total_cost"])
            if not tc.ok:
                row["reasons"].extend(tc.errors)
                verdict = "error"
        except CalculationError as ce:
            row["reasons"].append(f"Тариф не настроен: {ce}")
            verdict = "error"
        except Exception as ex:  # noqa: BLE001
            logger.warning("[EXCEL-IMPORT] preview calc failed fio=%s: %s", fio, ex)
            row["reasons"].append("Ошибка расчёта — проверьте тариф/комнату")
            verdict = "error"

        row["verdict"] = verdict
        counts[verdict] = counts.get(verdict, 0) + 1
        items.append(row)

    # Сверка с Google Sheets — на каждую строку (вкл. ненайденных, по ФИО).
    gs_present = gs_mismatch = 0
    for row in items:
        g = _gsheets_for_row(row, gs_by_fio, gs_by_uid)
        row["gsheets"] = g
        if g.get("present"):
            gs_present += 1
            if g.get("mismatch"):
                gs_mismatch += 1
                if row["verdict"] == "ok":
                    row["verdict"] = "warning"
                    counts["ok"] = max(0, counts.get("ok", 0) - 1)
                    counts["warning"] = counts.get("warning", 0) + 1
                row["reasons"].append("Расходится с Google Sheets — проверьте показания")

    items.sort(key=lambda x: ({"error": 0, "unmatched": 1, "warning": 2, "ok": 3}.get(x["verdict"], 9),
                              x["fio"]))
    return {
        "items": items,
        "counts": counts,
        "total_people": len(items),
        "meters_present": parsed.get("meters_present", []),
        "skipped_rows": parsed.get("skipped_rows", 0),
        "gsheets": {
            "checked": bool(window),
            "window": (
                {"start": window[0].strftime("%d.%m.%Y"), "end": window[1].strftime("%d.%m.%Y")}
                if window else None
            ),
            "present": gs_present, "mismatch": gs_mismatch,
        },
    }


def _res_label(r: str) -> str:
    return {"hot": "ГВС", "cold": "ХВС", "elect": "Электр."}.get(r, r)


# =====================================================================
# Черновик импорта — сохранить/продолжить (один на админку, в SystemSetting)
# =====================================================================
DRAFT_KEY = "excel_readings_draft"


async def save_draft(db: AsyncSession, payload: dict, actor: Optional[User]) -> dict:
    """Сохранить черновик (period_id + строки) — чтобы продолжить позже
    (например, после создания недостающей квартиры в Жилфонде)."""
    import json
    from app.modules.utility.models import SystemSetting
    from app.core.time_utils import utcnow
    payload = dict(payload)
    payload["saved_at"] = utcnow().isoformat()
    payload["saved_by"] = actor.username if actor else None
    blob = json.dumps(payload, ensure_ascii=False, default=str)
    row = await db.get(SystemSetting, DRAFT_KEY)
    if row:
        row.value = blob
    else:
        db.add(SystemSetting(key=DRAFT_KEY, value=blob,
                             description="Черновик импорта показаний из Excel"))
    await db.commit()
    return {"status": "ok", "rows": len(payload.get("rows", [])), "saved_at": payload["saved_at"]}


async def load_draft(db: AsyncSession) -> Optional[dict]:
    import json
    from app.modules.utility.models import SystemSetting
    row = await db.get(SystemSetting, DRAFT_KEY)
    if not row or not row.value:
        return None
    try:
        return json.loads(row.value)
    except Exception:
        return None


async def clear_draft(db: AsyncSession) -> dict:
    from app.modules.utility.models import SystemSetting
    row = await db.get(SystemSetting, DRAFT_KEY)
    if row:
        await db.delete(row)
        await db.commit()
    return {"status": "ok"}


async def recompute_preview(
    db: AsyncSession, period_id: Optional[int], rows: list[dict]
) -> dict:
    """Пересчёт превью по отредактированным строкам (правка показаний/ФИО,
    переназначение) — без файла. rows: [{fio, user_id?, hot:{prev,cur}, ...}].
    user_id (если задан) фиксирует жильца — fuzzy пропускается."""
    people: dict[str, dict] = {}
    forced: dict[str, int] = {}
    meters: list[str] = []
    for i, r in enumerate(rows):
        key = f"row{i}"
        rec = {"fio": (r.get("fio") or "").strip(), "hot": {}, "cold": {}, "elect": {}}
        for res in _RES_KEYS:
            d = r.get(res) or {}
            p, c = _num(d.get("prev")), _num(d.get("cur"))
            if p is not None or c is not None:
                rec[res] = {"prev": p, "cur": c}
                if res not in meters:
                    meters.append(res)
        people[key] = rec
        if r.get("user_id"):
            try:
                forced[key] = int(r["user_id"])
            except (TypeError, ValueError):
                pass
    parsed = {
        "people": people,
        "meters_present": [m for m in _RES_KEYS if m in meters] or ["hot", "cold"],
        "skipped_rows": 0,
    }
    return await build_preview(db, parsed, period_id, forced_match=forced)


# =====================================================================
# COMMIT — создание утверждённых MeterReading прямо в финотчётность
# =====================================================================

async def commit_import(
    db: AsyncSession, period_id: int, decisions: list[dict], actor: Optional[User]
) -> dict:
    """decisions: [{user_id, hot:{prev,cur}, cold:{...}, elect:{...},
    status: submitted|norm}]. Создаёт is_approved=True MeterReading на период.
    Уже утверждённых за период пропускает (анти-дубль). Один commit на всё."""
    from app.modules.utility.services.tariff_cache import tariff_cache
    from app.modules.utility.routers.admin_dashboard import write_audit_log

    period = await db.get(BillingPeriod, period_id)
    if not period:
        from fastapi import HTTPException
        raise HTTPException(404, "Период не найден")

    from app.modules.utility.routers.settings import _load_seasonal
    seasonal = await _load_seasonal(db)
    dec_ids = [d.get("user_id") for d in decisions if d.get("user_id")]
    adj_by_user = await _load_adjustments(db, dec_ids, period_id)
    fallback_tariff = (await db.execute(
        select(Tariff).where(Tariff.is_active)
    )).scalars().first()

    created = skipped_existing = failed = 0
    errors: list[dict] = []

    for dec in decisions:
        uid = dec.get("user_id")
        if not uid:
            continue
        try:
            user = (await db.execute(
                select(User).options(selectinload(User.room)).where(User.id == uid)
            )).scalars().first()
            if not user or not user.room:
                failed += 1
                errors.append({"user_id": uid, "reason": "нет помещения"})
                continue
            room = user.room

            # Анти-дубль: уже есть утверждённый reading за этот период.
            exists = (await db.execute(
                select(MeterReading.id).where(
                    MeterReading.user_id == uid,
                    MeterReading.room_id == room.id,
                    MeterReading.period_id == period_id,
                    MeterReading.is_approved.is_(True),
                ).limit(1)
            )).scalars().first()
            if exists:
                skipped_existing += 1
                continue

            tariff = tariff_cache.get_effective_tariff(user=user, room=room) or fallback_tariff
            if not tariff:
                failed += 1
                errors.append({"user_id": uid, "reason": "нет тарифа"})
                continue

            status = dec.get("status", "submitted")
            cur = {r: _num((dec.get(r) or {}).get("cur")) for r in _RES_KEYS}
            prev = {r: _num((dec.get(r) or {}).get("prev")) for r in _RES_KEYS}

            if status == "norm":
                vol_hot, vol_cold, share_el = _norm_volumes(tariff, user, room)
                flags = f"AUTO_NORM,{EXCEL_FLAG}"
                # Показания не меняем — берём последние известные (prev из Excel
                # или текущие в комнате), счётчик не «крутим».
                read_hot = prev["hot"] if prev["hot"] is not None else D(room.last_hot_water or 0)
                read_cold = prev["cold"] if prev["cold"] is not None else D(room.last_cold_water or 0)
                read_el = prev["elect"] if prev["elect"] is not None else D(room.last_electricity or 0)
            else:
                vol_hot, vol_cold, share_el = _consumption_volumes(
                    user, room, {k: cur[k] for k in _RES_KEYS}, {k: prev[k] for k in _RES_KEYS}
                )
                flags = EXCEL_FLAG
                read_hot = cur["hot"] if cur["hot"] is not None else (prev["hot"] if prev["hot"] is not None else D(room.last_hot_water or 0))
                read_cold = cur["cold"] if cur["cold"] is not None else (prev["cold"] if prev["cold"] is not None else D(room.last_cold_water or 0))
                read_el = cur["elect"] if cur["elect"] is not None else (prev["elect"] if prev["elect"] is not None else D(room.last_electricity or 0))

            costs = _compute_costs(user, room, tariff, vol_hot, vol_cold, share_el,
                                   seasonal, adj_by_user.get(uid, {}),
                                   force_meters=(status != "norm"))

            reading = MeterReading(
                room_id=room.id, user_id=user.id, period_id=period_id,
                hot_water=read_hot, cold_water=read_cold, electricity=read_el,
                total_209=costs["total_209"], total_205=costs["total_205"],
                total_cost=costs["grand_total"],
                is_approved=True,
                anomaly_flags=flags, anomaly_score=0,
            )
            for k, v in costs_for_model_fields(costs).items():
                setattr(reading, k, v)
            db.add(reading)

            room.last_hot_water = read_hot
            room.last_cold_water = read_cold
            room.last_electricity = read_el
            db.add(room)
            created += 1
        except Exception as ex:  # noqa: BLE001
            failed += 1
            errors.append({"user_id": uid, "reason": str(ex)[:160]})
            logger.warning("[EXCEL-IMPORT] commit failed user=%s: %s", uid, ex)

    if actor is not None:
        await write_audit_log(
            db, actor.id, actor.username,
            action="excel_readings_import", entity_type="period", entity_id=period_id,
            details={"created": created, "skipped_existing": skipped_existing,
                     "failed": failed, "period": period.name},
        )
    await db.commit()
    return {
        "status": "ok", "period": period.name,
        "created": created, "skipped_existing": skipped_existing,
        "failed": failed, "errors": errors[:50],
    }
