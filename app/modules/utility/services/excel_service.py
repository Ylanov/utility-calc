import io
import asyncio
from typing import Dict, List, Tuple
from decimal import Decimal
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload

from app.modules.utility.models import User, MeterReading, BillingPeriod, Room
from app.core.auth import get_password_hash
from app.modules.utility.services.debt_import import normalize_name

ZERO = Decimal("0.00")


def _load_workbook_sync(file_content: bytes):
    return load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)


# ======================================================
# ИНТЕЛЛЕКТУАЛЬНЫЙ ИМПОРТ: ЖИЛФОНД + ЖИЛЬЦЫ
# ======================================================
async def import_users_from_excel(file_content: bytes, db: AsyncSession) -> dict:
    try:
        workbook = await asyncio.to_thread(_load_workbook_sync, file_content)
        worksheet = workbook.active

        # Загружаем существующих пользователей
        users_result = await db.execute(select(User).where(User.is_deleted.is_(False)))
        existing_users: Dict[str, User] = {normalize_name(u.username): u for u in users_result.scalars().all()}

        # Загружаем существующие комнаты
        rooms_result = await db.execute(select(Room))
        existing_rooms: Dict[str, Room] = {f"{r.dormitory_name}_{r.room_number}": r for r in
                                           rooms_result.scalars().all()}

        added_users = 0
        updated_users = 0
        added_rooms = 0
        updated_rooms = 0
        skipped_count = 0
        errors: List[str] = []
        password_cache: Dict[str, str] = {}

        # Читаем со второй строки (пропуская заголовки)
        rows = worksheet.iter_rows(min_row=2, values_only=True)

        for row_index, row in enumerate(rows, start=2):
            try:
                # Если строка пустая
                if not row or not any(row):
                    continue

                # Колонки Excel:
                # 0: Логин, 1: Пароль, 2: Общежитие, 3: № Комнаты, 4: Площадь, 5: Мест в комнате
                # 6: Жильцов (платит за), 7: № ГВС, 8: № ХВС, 9: № Электр., 10: Место работы

                username = str(row[0]).strip() if row[0] else None
                password = str(row[1]).strip() if len(row) > 1 and row[1] else username
                dormitory = str(row[2]).strip() if len(row) > 2 and row[2] else None
                room_number = str(row[3]).strip() if len(row) > 3 and row[3] else None

                # Если нет ни общежития, ни номера комнаты - пропускаем (не к чему привязывать)
                if not dormitory or not room_number:
                    errors.append(f"Строка {row_index}: Не указано 'Общежитие' или 'Номер комнаты'")
                    skipped_count += 1
                    continue

                # Безопасное чтение числовых значений
                try:
                    apartment_area = Decimal(str(row[4]).replace(',', '.')) if len(row) > 4 and row[4] else ZERO
                    total_room_residents = int(row[5]) if len(row) > 5 and row[5] else 1
                    residents_count = int(row[6]) if len(row) > 6 and row[6] else 1
                except Exception:
                    apartment_area, total_room_residents, residents_count = ZERO, 1, 1
                    errors.append(f"Строка {row_index}: Ошибка в числовых данных (Площадь/Места). Установлены нули.")

                # Счетчики и работа
                hw_serial = str(row[7]).strip() if len(row) > 7 and row[7] else None
                cw_serial = str(row[8]).strip() if len(row) > 8 and row[8] else None
                el_serial = str(row[9]).strip() if len(row) > 9 and row[9] else None
                workplace = str(row[10]).strip() if len(row) > 10 and row[10] else None

                # ======================================================
                # 1. ОБРАБАТЫВАЕМ КОМНАТУ (ЖИЛФОНД)
                # ======================================================
                room_key = f"{dormitory}_{room_number}"
                if room_key in existing_rooms:
                    room = existing_rooms[room_key]
                    # Обновляем данные комнаты
                    if apartment_area > 0: room.apartment_area = apartment_area
                    if total_room_residents > 0: room.total_room_residents = total_room_residents
                    if hw_serial: room.hw_meter_serial = hw_serial
                    if cw_serial: room.cw_meter_serial = cw_serial
                    if el_serial: room.el_meter_serial = el_serial
                    updated_rooms += 1
                else:
                    # Создаем новую комнату
                    room = Room(
                        dormitory_name=dormitory,
                        room_number=room_number,
                        apartment_area=apartment_area,
                        total_room_residents=total_room_residents,
                        hw_meter_serial=hw_serial,
                        cw_meter_serial=cw_serial,
                        el_meter_serial=el_serial
                    )
                    db.add(room)
                    await db.flush()  # Чтобы получить room.id для пользователя
                    existing_rooms[room_key] = room
                    added_rooms += 1

                # ======================================================
                # 2. ОБРАБАТЫВАЕМ ПОЛЬЗОВАТЕЛЯ (Если логин заполнен)
                # ======================================================
                if username:
                    normalized_username = normalize_name(username)

                    if password in password_cache:
                        hashed_password = password_cache[password]
                    else:
                        hashed_password = get_password_hash(password)
                        password_cache[password] = hashed_password

                    if normalized_username in existing_users:
                        user = existing_users[normalized_username]
                        # Обновляем пользователя
                        user.workplace = workplace
                        user.residents_count = residents_count
                        user.room_id = room.id
                        if password and password != username:  # Если пароль был явно передан
                            user.hashed_password = hashed_password
                        updated_users += 1
                    else:
                        # Создаем пользователя
                        new_user = User(
                            username=username,
                            hashed_password=hashed_password,
                            role="user",
                            workplace=workplace,
                            residents_count=residents_count,
                            room_id=room.id,
                            is_deleted=False
                        )
                        db.add(new_user)
                        existing_users[normalized_username] = new_user
                        added_users += 1

            except Exception as error:
                skipped_count += 1
                errors.append(f"Строка {row_index}: Непредвиденная ошибка - {str(error)}")

        await db.commit()
        await asyncio.to_thread(workbook.close)

        return {
            "status": "success",
            "added_users": added_users,
            "updated_users": updated_users,
            "added_rooms": added_rooms,
            "updated_rooms": updated_rooms,
            "skipped": skipped_count,
            "errors": errors
        }

    except Exception as error:
        await db.rollback()
        return {"status": "error", "message": f"Ошибка файла: {str(error)}", "errors": [str(error)]}


# ======================================================
# EXPORT REPORT (СВОДКА ДЛЯ БУХГАЛТЕРА)
# ======================================================
async def generate_billing_report_xlsx(db: AsyncSession, period_id: int) -> Tuple[io.BytesIO, str]:
    period_result = await db.execute(select(BillingPeriod).where(BillingPeriod.id == period_id))
    period = period_result.scalars().first()
    if not period: raise ValueError("Период не найден")

    statement = (select(User, MeterReading).join(MeterReading, User.id == MeterReading.user_id)
                 .options(selectinload(User.room)).where(MeterReading.period_id == period_id,
                                                         MeterReading.is_approved.is_(True))
                 .order_by(Room.dormitory_name, User.username))

    result = await db.execute(statement)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Сводная ведомость"
    headers = ["Общежитие/Комната", "ФИО", "Площадь", "Жильцов", "ГВС", "ХВС", "Водоотв.", "Электроэнергия",
               "Содержание", "Наем", "ТКО", "Отопление", "209", "205", "ИТОГО"]
    worksheet.append(headers)

    total_sum, total_209_sum, total_205_sum = ZERO, ZERO, ZERO

    for user, reading in result:
        room = user.room
        total_cost = Decimal(reading.total_cost or 0)
        t_209 = Decimal(reading.total_209 or 0)
        t_205 = Decimal(reading.total_205 or 0)
        total_sum += total_cost
        total_209_sum += t_209
        total_205_sum += t_205

        room_display = f"{room.dormitory_name} / {room.room_number}" if room else "-"
        area = room.apartment_area if room else "-"
        residents = f"{user.residents_count}/{room.total_room_residents}" if room else "-"

        worksheet.append([room_display, user.username, area, residents, reading.cost_hot_water, reading.cost_cold_water,
                          reading.cost_sewage, reading.cost_electricity, reading.cost_maintenance,
                          reading.cost_social_rent, reading.cost_waste, reading.cost_fixed_part, t_209, t_205,
                          total_cost])

    worksheet.append([""] * 11 + ["ИТОГО:", total_209_sum, total_205_sum, total_sum])
    output = await asyncio.to_thread(lambda: _save_workbook_sync(workbook))
    filename = f"Report_{period.name}".replace(" ", "_") + ".xlsx"

    return output, filename


def _save_workbook_sync(workbook) -> io.BytesIO:
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output