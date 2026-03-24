import io
import re
import logging
import secrets
import string
from typing import Dict
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.dialects.postgresql import insert
from passlib.context import CryptContext

from app.modules.arsenal.models import AccountingObject, Nomenclature, WeaponRegistry, ArsenalUser

logger = logging.getLogger(__name__)
ZERO = Decimal("0.00")
BATCH_SIZE = 1000

pwd_context = CryptContext(schemes=["argon2"], deprecated="auto")


def parse_price(value) -> Decimal:
    """
    Парсит сумму/цену из Excel, очищая от пробелов и лишних символов.
    """
    if value is None:
        return ZERO
    if isinstance(value, (int, float)):
        return Decimal(str(round(float(value), 2)))
    if isinstance(value, Decimal):
        return value
    clean = str(value).replace(' ', '').replace(',', '.').replace('\xa0', '').strip()
    try:
        return Decimal(clean)
    except Exception:
        return ZERO


async def import_arsenal_from_excel(file_content: bytes, db: AsyncSession) -> dict:
    """
    Импорт данных из Excel с оптимизацией памяти и исправленной логикой количества.
    """
    workbook = None
    try:
        workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        worksheet = workbook.active

        objects_cache: Dict[str, AccountingObject] = {}
        nom_cache: Dict[str, Nomenclature] = {}

        obj_res = await db.execute(select(AccountingObject))
        for obj in obj_res.scalars().all():
            key = f"{obj.name.lower()}_{str(obj.mol_name).lower() if obj.mol_name else ''}"
            objects_cache[key] = obj

        nom_res = await db.execute(select(Nomenclature))
        for nom in nom_res.scalars().all():
            nom_cache[nom.name.lower()] = nom

        batch = []
        added_count = 0
        skipped_count = 0
        created_users_creds = []

        rows = worksheet.iter_rows(values_only=True)

        for row_index, row in enumerate(rows, start=1):
            try:
                if not row or row[0] is None:
                    continue
                first_col = str(row[0]).strip().replace('.', '')
                if not first_col.isdigit():
                    skipped_count += 1
                    continue

                raw_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if not raw_name:
                    skipped_count += 1
                    continue

                account = str(row[2]).strip() if len(row) > 2 and row[2] else None
                kbk = str(row[3]).strip() if len(row) > 3 and row[3] else None
                storage_raw = str(row[4]).strip() if len(row) > 4 and row[4] else "Главный склад"

                qty_val = row[5] if len(row) > 5 else 1
                try:
                    qty = int(float(str(qty_val).replace(" ", "").replace("\xa0", "")))
                    if qty < 1:
                        qty = 1
                except Exception:
                    qty = 1

                total_sum_val = row[6] if len(row) > 6 else 0
                total_sum = parse_price(total_sum_val)

                if qty > 0 and total_sum > 0:
                    unit_price = total_sum / Decimal(qty)
                    unit_price = unit_price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                else:
                    unit_price = total_sum

                inv_number = str(row[7]).strip() if len(row) > 7 and row[7] else None

                # --- 1. Обработка Склада и МОЛ (без изменений) ---
                obj_name = storage_raw
                mol_name = None
                if " - " in storage_raw:
                    parts = storage_raw.rsplit(" - ", 1)
                    obj_name = parts[0].strip()
                    mol_name = parts[1].strip()

                obj_key = f"{obj_name.lower()}_{str(mol_name).lower() if mol_name else ''}"
                target_object = objects_cache.get(obj_key)

                if not target_object:
                    target_object = AccountingObject(name=obj_name, obj_type="Склад", mol_name=mol_name)
                    db.add(target_object)
                    await db.flush()
                    objects_cache[obj_key] = target_object
                    # ... (создание пользователя) ...
                    new_username = f"unit_{target_object.id}"
                    alphabet = string.ascii_letters + string.digits
                    plain_password = ''.join(secrets.choice(alphabet) for _ in range(8))
                    new_user = ArsenalUser(
                        username=new_username,
                        hashed_password=pwd_context.hash(plain_password),
                        role="unit_head",
                        object_id=target_object.id
                    )
                    db.add(new_user)
                    created_users_creds.append({
                        "object": obj_name, "mol": mol_name or "-",
                        "username": new_username, "password": plain_password
                    })

                # --- 2. 🔥 ИСПРАВЛЕННАЯ ЛОГИКА ОБРАБОТКИ НОМЕНКЛАТУРЫ ---
                serial = "Б/Н"
                nom_name = raw_name

                match = re.search(r'\(([^)]+)\)$', raw_name)
                if match:
                    serial = match.group(1).strip()
                    nom_name = raw_name[:match.start()].strip()

                nom_key = nom_name.lower()
                nomenclature = nom_cache.get(nom_key)

                is_numbered = True  # Значение по-умолчанию

                if nomenclature:
                    # Если номенклатура уже есть в базе, доверяем ЕЙ
                    is_numbered = nomenclature.is_numbered
                else:
                    # Если это новый товар, пытаемся угадать его тип
                    batch_keywords = ["патрон", "граната", "мина", "снаряд", "зип", "масло", "смазка", "гвоздодер"]
                    nom_name_lower = nom_name.lower()
                    if (account and str(account).startswith("105")) or any(
                            kw in nom_name_lower for kw in batch_keywords):
                        is_numbered = False

                    # Создаем новую номенклатуру с правильным флагом
                    nomenclature = Nomenclature(name=nom_name, default_account=account, is_numbered=is_numbered)
                    db.add(nomenclature)
                    await db.flush()
                    nom_cache[nom_key] = nomenclature

                # Определяем серийник, если он не был указан в скобках
                if serial == "Б/Н":
                    if is_numbered:
                        serial = f"Инв.{inv_number}" if inv_number else f"Б/Н-{row_index}"
                    else:
                        serial = "Партия 1"

                # --- 3. 🔥 ИСПРАВЛЕННАЯ ЛОГИКА ФОРМИРОВАНИЯ КОЛИЧЕСТВА ---
                final_quantity = 1 if is_numbered else qty

                batch.append({
                    "nomenclature_id": nomenclature.id,
                    "serial_number": serial,
                    "current_object_id": target_object.id,
                    "status": 1,
                    "quantity": final_quantity,  # <-- Используем правильное количество
                    "inventory_number": inv_number,
                    "price": unit_price,
                    "account_code": account,
                    "kbk": kbk
                })
                # ... (далее без изменений) ...
                if len(batch) >= BATCH_SIZE:
                    await upsert_batch(db, batch)
                    added_count += len(batch)
                    batch.clear()

            except Exception as e:
                logger.error(f"Error parsing row {row_index}: {e}")
                skipped_count += 1

        if batch:
            await upsert_batch(db, batch)
            added_count += len(batch)

        await db.commit()

        return {
            "status": "success", "added": added_count,
            "skipped": skipped_count, "new_users": created_users_creds
        }
    except Exception as e:
        await db.rollback()
        return {"status": "error", "message": str(e)}
    finally:
        if workbook:
            workbook.close()


async def upsert_batch(db: AsyncSession, rows: list):
    """
    Массовая вставка (Upsert). (без изменений)
    """
    if not rows:
        return
    aggregated = {}
    for r in rows:
        key = (r["nomenclature_id"], r["serial_number"], r["current_object_id"])
        if key not in aggregated:
            aggregated[key] = r.copy()
        else:
            aggregated[key]["quantity"] += r["quantity"]
            if r["price"] and r["price"] > ZERO:
                aggregated[key]["price"] = r["price"]
            if r.get("inventory_number"):
                aggregated[key]["inventory_number"] = r["inventory_number"]
            if r.get("kbk"):
                aggregated[key]["kbk"] = r["kbk"]
            if r.get("account_code"):
                aggregated[key]["account_code"] = r["account_code"]
    final_rows = list(aggregated.values())
    stmt = insert(WeaponRegistry).values(final_rows)
    stmt = stmt.on_conflict_do_update(
        constraint="uix_nom_serial_obj",
        set_={
            "quantity": stmt.excluded.quantity,
            "price": stmt.excluded.price,
            "inventory_number": stmt.excluded.inventory_number,
            "kbk": stmt.excluded.kbk,
            "account_code": stmt.excluded.account_code,
            "status": 1
        }
    )
    await db.execute(stmt)
