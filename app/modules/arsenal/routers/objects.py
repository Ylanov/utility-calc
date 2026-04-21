import secrets
import string
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import or_, func

from app.core.database import get_arsenal_db
from app.modules.arsenal.models import AccountingObject, ArsenalUser, WeaponRegistry, Nomenclature
from app.modules.arsenal.schemas import ObjCreate
from app.modules.arsenal.deps import get_current_arsenal_user, pwd_context
from app.modules.arsenal.services.audit import write_arsenal_audit

router = APIRouter(tags=["Arsenal Objects"])


@router.get("/objects")
async def get_objects(
        db: AsyncSession = Depends(get_arsenal_db),
        current_user: ArsenalUser = Depends(get_current_arsenal_user)
):
    result = await db.execute(select(AccountingObject).order_by(AccountingObject.name))
    return result.scalars().all()


@router.get("/objects/with-stats")
async def get_objects_with_stats(
        db: AsyncSession = Depends(get_arsenal_db),
        current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Список объектов + для каждого: сколько единиц / общая стоимость / сколько
    прикреплённых unit_head'ов / сколько дочерних объектов (иерархия).
    Используется на странице «Объекты» для карточек и дерева."""
    objs = (await db.execute(
        select(AccountingObject).order_by(AccountingObject.name)
    )).scalars().all()

    # Агрегаты за один запрос
    stats = (await db.execute(
        select(
            WeaponRegistry.current_object_id,
            func.count(WeaponRegistry.id).label("units"),
            func.coalesce(func.sum(WeaponRegistry.quantity), 0).label("qty"),
            func.coalesce(
                func.sum(WeaponRegistry.quantity * WeaponRegistry.price), 0
            ).label("cost"),
        )
        .where(WeaponRegistry.status == 1, WeaponRegistry.current_object_id.is_not(None))
        .group_by(WeaponRegistry.current_object_id)
    )).all()
    stats_map = {oid: (int(u), int(q), float(c)) for oid, u, q, c in stats}

    # Количество пользователей, привязанных к объекту
    users_rows = (await db.execute(
        select(ArsenalUser.object_id, func.count(ArsenalUser.id))
        .where(ArsenalUser.object_id.is_not(None))
        .group_by(ArsenalUser.object_id)
    )).all()
    users_map = {oid: int(c) for oid, c in users_rows}

    # Количество дочерних объектов
    children_rows = (await db.execute(
        select(AccountingObject.parent_id, func.count(AccountingObject.id))
        .where(AccountingObject.parent_id.is_not(None))
        .group_by(AccountingObject.parent_id)
    )).all()
    children_map = {oid: int(c) for oid, c in children_rows}

    result = []
    for o in objs:
        units, qty, cost = stats_map.get(o.id, (0, 0, 0.0))
        result.append({
            "id": o.id,
            "name": o.name,
            "obj_type": o.obj_type,
            "parent_id": o.parent_id,
            "mol_name": o.mol_name,
            "units_count": units,
            "total_quantity": qty,
            "total_cost": cost,
            "users_count": users_map.get(o.id, 0),
            "children_count": children_map.get(o.id, 0),
        })
    return result


@router.get("/objects/export")
async def export_objects(
        db: AsyncSession = Depends(get_arsenal_db),
        current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Excel-выгрузка структуры объектов с количествами и стоимостью.
    Для офлайн-аудита и печати реестра подразделений."""
    if current_user.role != "admin":
        raise HTTPException(403, "Только администратор")
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    data = await get_objects_with_stats(db, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Объекты"
    headers = ["ID", "Название", "Тип", "Родитель", "МОЛ", "Ед.", "Кол-во", "Стоимость", "Подразделений"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DBEAFE")
    for i, o in enumerate(data, 2):
        parent_name = next((x["name"] for x in data if x["id"] == o["parent_id"]), "")
        ws.cell(row=i, column=1, value=o["id"])
        ws.cell(row=i, column=2, value=o["name"])
        ws.cell(row=i, column=3, value=o["obj_type"])
        ws.cell(row=i, column=4, value=parent_name)
        ws.cell(row=i, column=5, value=o["mol_name"])
        ws.cell(row=i, column=6, value=o["units_count"])
        ws.cell(row=i, column=7, value=o["total_quantity"])
        ws.cell(row=i, column=8, value=o["total_cost"])
        ws.cell(row=i, column=9, value=o["children_count"])
    for col_letter, width in [("A", 6), ("B", 34), ("C", 18), ("D", 24), ("E", 28),
                              ("F", 10), ("G", 10), ("H", 16), ("I", 14)]:
        ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="arsenal_objects.xlsx"'},
    )


class ObjPatch(BaseModel):
    name: Optional[str] = None
    obj_type: Optional[str] = None
    parent_id: Optional[int] = None
    mol_name: Optional[str] = None


@router.patch("/objects/{obj_id}")
async def update_object(
        obj_id: int,
        data: ObjPatch,
        db: AsyncSession = Depends(get_arsenal_db),
        current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Редактирование объекта: можно поменять имя, МОЛ, тип, родителя
    (для переподчинения подразделения). Имя должно быть уникальным.
    Запрещено делать объект собственным предком (циклическая иерархия)."""
    if current_user.role != "admin":
        raise HTTPException(403, "Только администратор")

    obj = await db.get(AccountingObject, obj_id)
    if not obj:
        raise HTTPException(404, "Объект не найден")

    changes: dict = {}
    if data.name and data.name != obj.name:
        dup = (await db.execute(
            select(AccountingObject).where(
                AccountingObject.name == data.name,
                AccountingObject.id != obj_id,
            )
        )).scalars().first()
        if dup:
            raise HTTPException(409, "Объект с таким именем уже существует")
        changes["name"] = {"old": obj.name, "new": data.name}
        obj.name = data.name

    if data.obj_type and data.obj_type != obj.obj_type:
        changes["obj_type"] = {"old": obj.obj_type, "new": data.obj_type}
        obj.obj_type = data.obj_type

    if data.mol_name is not None and data.mol_name != obj.mol_name:
        changes["mol_name"] = {"old": obj.mol_name, "new": data.mol_name}
        obj.mol_name = data.mol_name

    if data.parent_id is not None and data.parent_id != obj.parent_id:
        if data.parent_id == obj_id:
            raise HTTPException(400, "Объект не может быть родителем самого себя")
        # Проверка цикла: новый parent не должен быть нашим потомком
        if data.parent_id:
            cur_id = data.parent_id
            visited = {obj_id}
            for _ in range(100):  # ограничение на глубину
                if cur_id in visited:
                    raise HTTPException(400, "Циклическая иерархия недопустима")
                visited.add(cur_id)
                parent = await db.get(AccountingObject, cur_id)
                if not parent or parent.parent_id is None:
                    break
                cur_id = parent.parent_id
        changes["parent_id"] = {"old": obj.parent_id, "new": data.parent_id}
        obj.parent_id = data.parent_id

    if not changes:
        return {"status": "noop"}

    await write_arsenal_audit(
        db, user_id=current_user.id, username=current_user.username,
        action="update_object", entity_type="object", entity_id=obj_id,
        details={"changes": changes},
    )
    await db.commit()
    await db.refresh(obj)
    return {"status": "ok", "changes": changes}


@router.post("/objects")
async def create_object(
        data: ObjCreate,
        db: AsyncSession = Depends(get_arsenal_db),
        current_user: ArsenalUser = Depends(get_current_arsenal_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Только администратор может создавать новые объекты")

    existing = await db.execute(select(AccountingObject).where(AccountingObject.name == data.name))
    if existing.scalars().first():
        raise HTTPException(status_code=400, detail="Объект с таким именем уже существует")

    obj = AccountingObject(**data.dict())
    db.add(obj)
    await db.flush()

    new_username = f"unit_{obj.id}"
    alphabet = string.ascii_letters + string.digits
    new_password = ''.join(secrets.choice(alphabet) for _ in range(8))
    hashed_pw = pwd_context.hash(new_password)

    new_user = ArsenalUser(
        username=new_username,
        hashed_password=hashed_pw,
        role="unit_head",
        object_id=obj.id
    )
    db.add(new_user)

    await write_arsenal_audit(
        db, user_id=current_user.id, username=current_user.username,
        action="create_object", entity_type="object", entity_id=obj.id,
        details={"name": obj.name, "obj_type": obj.obj_type, "mol": obj.mol_name},
    )

    await db.commit()
    await db.refresh(obj)

    return {
        "id": obj.id,
        "name": obj.name,
        "obj_type": obj.obj_type,
        "mol_name": obj.mol_name,
        "credentials": {
            "username": new_username,
            "password": new_password
        }
    }


@router.delete("/objects/{obj_id}")
async def delete_object(
        obj_id: int,
        db: AsyncSession = Depends(get_arsenal_db),
        current_user: ArsenalUser = Depends(get_current_arsenal_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Только администратор может удалять объекты")

    obj = await db.get(AccountingObject, obj_id)
    if not obj:
        raise HTTPException(status_code=404, detail="Объект не найден")

    # Защита: нельзя удалить объект с остатками / дочерними объектами
    units = (await db.execute(
        select(func.count(WeaponRegistry.id)).where(
            WeaponRegistry.current_object_id == obj_id,
            WeaponRegistry.status == 1,
        )
    )).scalar_one()
    if units:
        raise HTTPException(409, f"На объекте {units} активных единиц. Сначала переместите их.")

    children = (await db.execute(
        select(func.count(AccountingObject.id)).where(AccountingObject.parent_id == obj_id)
    )).scalar_one()
    if children:
        raise HTTPException(409, f"У объекта {children} подчинённых — удалите/переподчините их сначала.")

    await write_arsenal_audit(
        db, user_id=current_user.id, username=current_user.username,
        action="delete_object", entity_type="object", entity_id=obj_id,
        details={"name": obj.name, "obj_type": obj.obj_type},
    )
    await db.delete(obj)
    await db.commit()
    return {"status": "deleted"}


@router.get("/balance/{obj_id}")
async def get_object_balance(
        obj_id: int,
        db: AsyncSession = Depends(get_arsenal_db),
        current_user: ArsenalUser = Depends(get_current_arsenal_user),
        skip: int = Query(0, ge=0, description="Смещение (для пагинации)"),
        limit: int = Query(1000, ge=1, le=5000, description="Максимальное количество возвращаемых строк"),
        q: Optional[str] = Query(None, description="Поиск по названию изделия, инвентарному или заводскому номеру")
):
    if current_user.role == "unit_head" and obj_id != current_user.object_id:
        raise HTTPException(status_code=403, detail="Вы можете просматривать остатки только своего подразделения")

    # 🔥 ОПТИМИЗАЦИЯ: Плоский запрос без ORM overhead
    stmt = (
        select(
            WeaponRegistry.nomenclature_id,
            WeaponRegistry.serial_number,
            WeaponRegistry.inventory_number,
            WeaponRegistry.price,
            WeaponRegistry.quantity,
            WeaponRegistry.account_code,
            WeaponRegistry.kbk,
            Nomenclature.name.label("nom_name"),
            Nomenclature.code.label("nom_code"),
            Nomenclature.is_numbered,
            Nomenclature.default_account
        )
        .join(Nomenclature, WeaponRegistry.nomenclature_id == Nomenclature.id)
        .where(
            WeaponRegistry.current_object_id == obj_id,
            WeaponRegistry.status == 1
        )
    )

    # 🔥 ОПТИМИЗАЦИЯ: Делегируем поиск базе данных (LIKE)
    if q:
        search_term = f"%{q}%"
        stmt = stmt.where(
            or_(
                Nomenclature.name.ilike(search_term),
                WeaponRegistry.serial_number.ilike(search_term),
                WeaponRegistry.inventory_number.ilike(search_term)
            )
        )

    # 🔥 ОПТИМИЗАЦИЯ: Пагинация. Защищаем сервер от выгрузки миллиона строк
    stmt = stmt.order_by(Nomenclature.name, WeaponRegistry.serial_number).offset(skip).limit(limit)

    result = await db.execute(stmt)
    rows = result.all()

    balance = []
    for row in rows:
        display_serial = row.serial_number if row.is_numbered else f"Партия {row.serial_number}"
        account = row.account_code or row.default_account or "Не указан"

        balance.append({
            "nomenclature_id": row.nomenclature_id,
            "nomenclature": row.nom_name,
            "code": row.nom_code,
            "serial_number": display_serial,
            "inventory_number": row.inventory_number or "Б/Н",
            "price": float(row.price) if row.price else 0.0,
            "account": account,
            "kbk": row.kbk or "---",
            "quantity": row.quantity,
            "is_numbered": row.is_numbered
        })

    return balance
