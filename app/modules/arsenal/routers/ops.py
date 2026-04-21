"""Arsenal ops — объединённые endpoints для:

  * причины списания (DisposalReason) — справочник
  * низкие остатки (GET /alerts/low-stock)
  * инвентаризация — открытие, сканирование, отчёт расхождений, закрытие
  * безопасный сброс пароля (одноразовый токен вместо plaintext в JSON)

Всё это — critical/high улучшения по аудиту модуля.
"""
from __future__ import annotations

import hashlib
import secrets
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field
from sqlalchemy import and_, or_, select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_arsenal_db
from app.core.security import get_password_hash
from app.modules.arsenal.deps import get_current_arsenal_user
from app.modules.arsenal.models import (
    AccountingObject,
    ArsenalPasswordResetToken,
    ArsenalUser,
    DisposalReason,
    Document,
    DocumentItem,
    Inventory,
    InventoryItem,
    Nomenclature,
    WeaponRegistry,
)

router = APIRouter(tags=["Arsenal Ops"])


def _require_admin(user: ArsenalUser) -> None:
    if user.role != "admin":
        raise HTTPException(403, "Только для администратора")


# =====================================================================
# DISPOSAL REASONS
# =====================================================================
@router.get("/disposal-reasons")
async def list_disposal_reasons(
    active_only: bool = Query(True),
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Справочник причин списания. Используется в формах «Списание» для
    обязательного выбора причины (иначе списание выглядит анонимно)."""
    q = select(DisposalReason)
    if active_only:
        q = q.where(DisposalReason.is_active.is_(True))
    rows = (await db.execute(q.order_by(DisposalReason.id))).scalars().all()
    return [
        {"id": r.id, "code": r.code, "name": r.name, "kind": r.kind,
         "is_active": r.is_active}
        for r in rows
    ]


class DisposalReasonCreate(BaseModel):
    code: str = Field(..., max_length=32)
    name: str
    kind: str = Field("disposal", pattern="^(disposal|external|lost|other)$")


@router.post("/disposal-reasons")
async def create_disposal_reason(
    data: DisposalReasonCreate,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    _require_admin(current_user)
    # Уникальный код — если такой уже есть, 409
    exists = (await db.execute(
        select(DisposalReason).where(DisposalReason.code == data.code)
    )).scalars().first()
    if exists:
        raise HTTPException(409, f"Причина с кодом {data.code!r} уже существует")
    obj = DisposalReason(code=data.code, name=data.name, kind=data.kind, is_active=True)
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return {"id": obj.id, "code": obj.code, "name": obj.name, "kind": obj.kind}


# =====================================================================
# LOW STOCK ALERTS
# =====================================================================
@router.get("/alerts/low-stock")
async def low_stock(
    object_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Список партий где текущий остаток (SUM qty по всем объектам или по
    конкретному объекту) ниже Nomenclature.min_quantity.

    unit_head видит только свой склад (object_id=current_user.object_id).
    """
    # Для unit_head ограничиваем
    target_object_id = object_id
    if current_user.role != "admin":
        target_object_id = current_user.object_id
        if target_object_id is None:
            return {"items": [], "total": 0}

    # Берём номенклатуры, у которых задан порог (партионные, т.к. min_quantity
    # имеет смысл только для is_numbered=False).
    base_q = (
        select(
            Nomenclature.id,
            Nomenclature.name,
            Nomenclature.code,
            Nomenclature.min_quantity,
            func.coalesce(func.sum(WeaponRegistry.quantity), 0).label("current_qty"),
        )
        .select_from(Nomenclature)
        .outerjoin(
            WeaponRegistry,
            and_(
                WeaponRegistry.nomenclature_id == Nomenclature.id,
                WeaponRegistry.status == 1,
                WeaponRegistry.current_object_id == target_object_id if target_object_id else True,
            ),
        )
        .where(
            Nomenclature.is_numbered.is_(False),
            Nomenclature.min_quantity > 0,
        )
        .group_by(Nomenclature.id, Nomenclature.name, Nomenclature.code, Nomenclature.min_quantity)
    )
    rows = (await db.execute(base_q)).all()

    alerts = [
        {
            "nomenclature_id": r[0],
            "name": r[1],
            "code": r[2],
            "min_quantity": r[3],
            "current_quantity": int(r[4] or 0),
            "deficit": max(0, r[3] - int(r[4] or 0)),
            "severity": "critical" if (r[4] or 0) == 0 else "warning",
        }
        for r in rows
        if (r[4] or 0) < r[3]
    ]
    # Сортируем: сначала полный ноль, потом по дефициту.
    alerts.sort(key=lambda x: (-x["deficit"]))
    return {"items": alerts, "total": len(alerts), "object_id": target_object_id}


# =====================================================================
# INVENTORY
# =====================================================================
class InventoryCreate(BaseModel):
    object_id: int
    note: Optional[str] = None


@router.post("/inventory")
async def start_inventory(
    data: InventoryCreate,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Начать инвентаризацию склада. Только одна активная инвентаризация
    на объект за раз — иначе расхождения смешаются."""
    if current_user.role != "admin" and current_user.object_id != data.object_id:
        raise HTTPException(403, "Нет прав на инвентаризацию этого объекта")

    obj = await db.get(AccountingObject, data.object_id)
    if not obj:
        raise HTTPException(404, "Объект не найден")

    active = (await db.execute(
        select(Inventory).where(
            Inventory.object_id == data.object_id,
            Inventory.status == "open",
        )
    )).scalars().first()
    if active:
        raise HTTPException(
            409,
            f"На объекте уже идёт инвентаризация #{active.id}. Закройте или отмените её.",
        )

    inv = Inventory(
        object_id=data.object_id,
        status="open",
        started_by_id=current_user.id,
        note=data.note,
    )
    db.add(inv)
    await db.commit()
    await db.refresh(inv)
    return {"id": inv.id, "object_id": inv.object_id, "status": inv.status,
            "started_at": inv.started_at}


class InventoryScan(BaseModel):
    nomenclature_id: Optional[int] = None  # можно не указывать — выведем по серийнику
    serial_number: Optional[str] = None
    found_quantity: int = Field(1, ge=1)
    note: Optional[str] = None


class InventoryQuickScan(BaseModel):
    """Быстрый скан: только серийник (без указания номенклатуры).
    Сервер ищет по серийнику в WeaponRegistry и сам подставляет nomenclature_id.
    Это ускоряет работу — пользователь просто сканирует QR/вводит номер."""
    serial_number: str


@router.post("/inventory/{inventory_id}/quick-scan")
async def quick_scan(
    inventory_id: int,
    data: InventoryQuickScan,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Быстрый скан по серийнику. Работает только для номерного учёта.

    Алгоритм:
      1. Ищем единицу в WeaponRegistry по serial_number.
      2. Если найдена на ЭТОМ складе — обычный scan.
      3. Если на ДРУГОМ складе — возвращаем warning (возможно «чужой» предмет).
      4. Если не найдена — warning (излишек/не числится).

    Возвращает статус-код, позволяющий UI показать цветное уведомление.
    """
    inv = await db.get(Inventory, inventory_id)
    if not inv:
        raise HTTPException(404, "Инвентаризация не найдена")
    if inv.status != "open":
        raise HTTPException(409, "Инвентаризация уже закрыта / отменена")

    serial = (data.serial_number or "").strip()
    if not serial:
        raise HTTPException(400, "Пустой серийник")

    # Ищем в реестре — только активные
    found = (await db.execute(
        select(WeaponRegistry)
        .options(selectinload(WeaponRegistry.nomenclature))
        .where(
            WeaponRegistry.serial_number == serial,
            WeaponRegistry.status == 1,
        )
        .limit(1)
    )).scalars().first()

    warning = None
    nomenclature_id = None
    nomenclature_name = None
    if found:
        nomenclature_id = found.nomenclature_id
        nomenclature_name = found.nomenclature.name if found.nomenclature else None
        # Проверка, что этот серийник числится ИМЕННО на этом складе
        if found.current_object_id != inv.object_id:
            warning = (
                f"Серийник {serial!r} числится на другом объекте (id={found.current_object_id}). "
                "Возможно чужой предмет или не было оформлено перемещение."
            )
    else:
        warning = f"Серийник {serial!r} не найден в реестре — зафиксирован как излишек."

    # Проверка дубликата внутри этой инвентаризации
    if nomenclature_id:
        dup = (await db.execute(
            select(InventoryItem).where(
                InventoryItem.inventory_id == inventory_id,
                InventoryItem.nomenclature_id == nomenclature_id,
                InventoryItem.serial_number == serial,
            )
        )).scalars().first()
        if dup:
            raise HTTPException(409, f"Серийник {serial!r} уже сканирован в этой инвентаризации")

    item = InventoryItem(
        inventory_id=inventory_id,
        nomenclature_id=nomenclature_id or 0,  # 0 не пройдёт FK — уложит в warning ниже
        serial_number=serial,
        found_quantity=1,
        scanned_by_id=current_user.id,
        note=warning,
    )
    # Если nomenclature_id неизвестен — не сохраняем сканирование в БД, но
    # возвращаем warning, чтобы UI уведомил оператора о находке.
    if not nomenclature_id:
        return {
            "saved": False,
            "warning": warning,
            "serial_number": serial,
            "suggestion": "Найдите номенклатуру вручную через форму /scan",
        }

    db.add(item)
    await db.commit()
    await db.refresh(item)
    return {
        "saved": True,
        "item_id": item.id,
        "nomenclature_id": nomenclature_id,
        "nomenclature_name": nomenclature_name,
        "warning": warning,
    }


@router.post("/inventory/{inventory_id}/scan")
async def scan_inventory_item(
    inventory_id: int,
    data: InventoryScan,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Зафиксировать найденный предмет в ходе инвентаризации.

    Для номерного учёта: добавляется одна запись на серийник. Повторное
    сканирование того же серийника — 409.
    Для партионного: ищется существующая строка по номенклатуре, если есть —
    увеличивается found_quantity, иначе создаётся новая.
    """
    inv = await db.get(Inventory, inventory_id)
    if not inv:
        raise HTTPException(404, "Инвентаризация не найдена")
    if inv.status != "open":
        raise HTTPException(409, "Инвентаризация уже закрыта / отменена")

    if not data.nomenclature_id:
        raise HTTPException(400, "nomenclature_id обязателен (или используйте /quick-scan для быстрого скана по серийнику)")
    nom = await db.get(Nomenclature, data.nomenclature_id)
    if not nom:
        raise HTTPException(400, "Номенклатура не найдена")

    if nom.is_numbered:
        # Серийник обязателен, дубли запрещены
        if not data.serial_number:
            raise HTTPException(400, "Для номерного учёта нужен серийный номер")
        dup = (await db.execute(
            select(InventoryItem).where(
                InventoryItem.inventory_id == inventory_id,
                InventoryItem.nomenclature_id == data.nomenclature_id,
                InventoryItem.serial_number == data.serial_number,
            )
        )).scalars().first()
        if dup:
            raise HTTPException(
                409, f"Серийник {data.serial_number!r} уже сканирован в этой инвентаризации",
            )
        item = InventoryItem(
            inventory_id=inventory_id,
            nomenclature_id=data.nomenclature_id,
            serial_number=data.serial_number,
            found_quantity=1,
            note=data.note,
            scanned_by_id=current_user.id,
        )
        db.add(item)
    else:
        # Партионный — один ряд на номенклатуру (+ опционально номер партии)
        existing = (await db.execute(
            select(InventoryItem).where(
                InventoryItem.inventory_id == inventory_id,
                InventoryItem.nomenclature_id == data.nomenclature_id,
                InventoryItem.serial_number.is_(data.serial_number)
                if data.serial_number is None else
                InventoryItem.serial_number == data.serial_number,
            )
        )).scalars().first()
        if existing:
            existing.found_quantity += data.found_quantity
            if data.note:
                existing.note = (existing.note or "") + " | " + data.note
            item = existing
        else:
            item = InventoryItem(
                inventory_id=inventory_id,
                nomenclature_id=data.nomenclature_id,
                serial_number=data.serial_number,
                found_quantity=data.found_quantity,
                note=data.note,
                scanned_by_id=current_user.id,
            )
            db.add(item)
    await db.commit()
    await db.refresh(item)
    return {"id": item.id, "found_quantity": item.found_quantity}


@router.get("/inventory/{inventory_id}/report")
async def inventory_report(
    inventory_id: int,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Отчёт расхождений: что ожидается (по WeaponRegistry) vs что найдено.

    Структура:
      * missing — в БД есть, но не найдено в ходе инвентаризации (недостача)
      * surplus — найдено, но в БД нет (излишек, чужое или ошибка)
      * matched — совпадает
    Для партионного учёта сравниваются количества.
    """
    inv = (await db.execute(
        select(Inventory)
        .options(selectinload(Inventory.items).selectinload(InventoryItem.nomenclature))
        .where(Inventory.id == inventory_id)
    )).scalars().first()
    if not inv:
        raise HTTPException(404, "Инвентаризация не найдена")

    # Ожидаемые остатки на складе
    expected_rows = (await db.execute(
        select(
            WeaponRegistry.nomenclature_id,
            WeaponRegistry.serial_number,
            WeaponRegistry.quantity,
            Nomenclature.name,
            Nomenclature.is_numbered,
        )
        .join(Nomenclature, Nomenclature.id == WeaponRegistry.nomenclature_id)
        .where(
            WeaponRegistry.current_object_id == inv.object_id,
            WeaponRegistry.status == 1,
        )
    )).all()

    # Ключи: для номерного — (nom_id, serial), для партионного — (nom_id, serial or None)
    expected_map: dict[tuple, dict] = {}
    for nom_id, serial, qty, name, is_numbered in expected_rows:
        key = (nom_id, serial if is_numbered else (serial or None))
        expected_map[key] = {
            "nomenclature_id": nom_id,
            "serial_number": serial,
            "name": name,
            "is_numbered": bool(is_numbered),
            "expected_quantity": int(qty or 0),
        }

    found_map: dict[tuple, dict] = {}
    for it in inv.items:
        is_num = bool(it.nomenclature.is_numbered) if it.nomenclature else False
        key = (it.nomenclature_id, it.serial_number if is_num else (it.serial_number or None))
        if key not in found_map:
            found_map[key] = {
                "nomenclature_id": it.nomenclature_id,
                "serial_number": it.serial_number,
                "name": it.nomenclature.name if it.nomenclature else "—",
                "is_numbered": is_num,
                "found_quantity": 0,
            }
        found_map[key]["found_quantity"] += it.found_quantity

    matched, missing, surplus = [], [], []
    all_keys = set(expected_map.keys()) | set(found_map.keys())
    for k in all_keys:
        exp = expected_map.get(k)
        fnd = found_map.get(k)
        if exp and fnd:
            diff = fnd["found_quantity"] - exp["expected_quantity"]
            if diff == 0:
                matched.append({**exp, "found_quantity": fnd["found_quantity"]})
            elif diff < 0:
                missing.append({**exp, "found_quantity": fnd["found_quantity"], "deficit": -diff})
            else:
                surplus.append({**fnd, "expected_quantity": exp["expected_quantity"], "excess": diff})
        elif exp:
            missing.append({**exp, "found_quantity": 0, "deficit": exp["expected_quantity"]})
        else:
            surplus.append({**fnd, "expected_quantity": 0, "excess": fnd["found_quantity"]})

    # Прогресс: сколько % ожидаемого уже сканировано.
    total_expected_units = sum(e["expected_quantity"] for e in expected_map.values())
    total_found_units = sum(f["found_quantity"] for f in found_map.values())
    progress_pct = (
        round(min(total_found_units / total_expected_units * 100, 100))
        if total_expected_units > 0 else 0
    )

    # Скорость: scans/мин (для оценки времени завершения оператором).
    elapsed_min = None
    scans_per_min = None
    if inv.started_at:
        elapsed_seconds = (datetime.utcnow() - inv.started_at).total_seconds()
        if elapsed_seconds > 0:
            elapsed_min = round(elapsed_seconds / 60, 1)
            if len(inv.items) > 0:
                scans_per_min = round(len(inv.items) / max(elapsed_min, 0.1), 1)

    return {
        "inventory": {
            "id": inv.id, "object_id": inv.object_id, "status": inv.status,
            "started_at": inv.started_at, "closed_at": inv.closed_at,
            "note": inv.note,
            "correction_document_id": inv.correction_document_id,
        },
        "matched": matched,
        "missing": missing,
        "surplus": surplus,
        "summary": {
            "matched_count": len(matched),
            "missing_count": len(missing),
            "surplus_count": len(surplus),
            "total_expected_units": total_expected_units,
            "total_found_units": total_found_units,
            "progress_pct": progress_pct,
            "elapsed_minutes": elapsed_min,
            "scans_per_minute": scans_per_min,
        },
    }


@router.get("/inventory/{inventory_id}/pending")
async def inventory_pending(
    inventory_id: int,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Что ещё НЕ отсканировано — для ускорения процесса. Оператор видит
    оставшийся список и может целенаправленно искать пропавшие предметы,
    а не бегать по складу в поисках всех подряд."""
    inv = (await db.execute(
        select(Inventory)
        .options(selectinload(Inventory.items).selectinload(InventoryItem.nomenclature))
        .where(Inventory.id == inventory_id)
    )).scalars().first()
    if not inv:
        raise HTTPException(404, "Инвентаризация не найдена")

    # Ожидаемые единицы
    expected = (await db.execute(
        select(
            WeaponRegistry.id,
            WeaponRegistry.nomenclature_id,
            WeaponRegistry.serial_number,
            WeaponRegistry.quantity,
            Nomenclature.name,
            Nomenclature.is_numbered,
        )
        .join(Nomenclature, Nomenclature.id == WeaponRegistry.nomenclature_id)
        .where(
            WeaponRegistry.current_object_id == inv.object_id,
            WeaponRegistry.status == 1,
        )
    )).all()

    # Для номерного: ключ = (nom_id, serial)
    # Для партионного: (nom_id, None) и сравнение по количеству
    scanned_keys = set()
    scanned_qty_by_nom: dict[int, int] = {}
    for it in inv.items:
        is_num = bool(it.nomenclature.is_numbered) if it.nomenclature else False
        if is_num:
            scanned_keys.add((it.nomenclature_id, it.serial_number))
        else:
            scanned_qty_by_nom[it.nomenclature_id] = (
                scanned_qty_by_nom.get(it.nomenclature_id, 0) + it.found_quantity
            )

    pending = []
    for _, nom_id, serial, qty, name, is_numbered in expected:
        if is_numbered:
            if (nom_id, serial) not in scanned_keys:
                pending.append({
                    "nomenclature_id": nom_id, "name": name,
                    "serial_number": serial, "is_numbered": True,
                    "expected_quantity": int(qty or 0),
                })
        else:
            scanned = scanned_qty_by_nom.get(nom_id, 0)
            remaining = int(qty or 0) - scanned
            if remaining > 0:
                pending.append({
                    "nomenclature_id": nom_id, "name": name,
                    "serial_number": serial, "is_numbered": False,
                    "expected_quantity": int(qty or 0),
                    "remaining_quantity": remaining,
                })
    return {
        "inventory_id": inv.id,
        "object_id": inv.object_id,
        "pending": pending,
        "pending_count": len(pending),
    }


class InventoryCloseRequest(BaseModel):
    """Параметры закрытия инвентаризации."""
    # Если True — создаётся корректирующий документ списания для недостачи
    # + первичного ввода для излишка. Если False — просто закрывается.
    auto_correct: bool = False
    # Причина списания для недостачи (используется при auto_correct=True).
    # Если не задан — берём код "LOST" из сида.
    disposal_reason_id: Optional[int] = None
    note: Optional[str] = None


@router.post("/inventory/{inventory_id}/close")
async def close_inventory(
    inventory_id: int,
    params: Optional[InventoryCloseRequest] = None,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Закрывает инвентаризацию. Если `auto_correct=True`, автоматически
    создаёт корректирующие документы:
      * недостача → документ «Списание» с причиной (default=LOST)
      * излишек   → документ «Первичный ввод»

    Без `auto_correct` — инвентаризация просто помечается closed, а админ
    решает что делать с расхождениями через обычные документы.
    Только admin может закрыть с auto_correct — безопасность.
    """
    inv = await db.get(Inventory, inventory_id)
    if not inv:
        raise HTTPException(404, "Инвентаризация не найдена")
    if inv.status != "open":
        raise HTTPException(409, f"Инвентаризация уже {inv.status!r}")

    p = params or InventoryCloseRequest()
    if p.auto_correct and current_user.role != "admin":
        raise HTTPException(
            403, "Автоматическая корректировка доступна только администратору",
        )

    correction_doc_id = None
    correction_summary = None

    if p.auto_correct:
        # Вычисляем отчёт ещё раз внутри транзакции
        report = await inventory_report(inventory_id, db, current_user)
        missing, surplus = report["missing"], report["surplus"]

        if missing or surplus:
            # Причина для недостачи. По умолчанию "LOST" из сида миграции.
            reason_id = p.disposal_reason_id
            if not reason_id:
                lost = (await db.execute(
                    select(DisposalReason).where(DisposalReason.code == "LOST")
                )).scalars().first()
                if lost:
                    reason_id = lost.id

            from app.modules.arsenal.services.weapon_service import WeaponService

            # Недостача → Списание
            if missing:
                class _DocShort:
                    doc_number = "АВТО"
                    operation_type = "Списание"
                    source_id = inv.object_id
                    target_id = None
                    operation_date = datetime.utcnow()
                    comment = f"Инвентаризация #{inv.id}: недостача"
                    disposal_reason_id = reason_id

                class _ItemShort:
                    def __init__(self, r):
                        self.nomenclature_id = r["nomenclature_id"]
                        self.serial_number = r["serial_number"]
                        self.quantity = r.get("deficit") or 1
                        self.price = None
                        self.inventory_number = None

                miss_items = [_ItemShort(m) for m in missing]
                try:
                    doc_miss = await WeaponService.process_document(
                        db, _DocShort, miss_items,
                        attached_file_path=None,
                        author_id=current_user.id,
                    )
                    correction_doc_id = doc_miss.id
                except HTTPException as e:
                    raise HTTPException(
                        400,
                        f"Не удалось автоматически списать недостачу: {e.detail}. "
                        "Закройте инвентаризацию без auto_correct и оформите документы вручную.",
                    )

            # Излишек → Первичный ввод
            if surplus:
                class _DocIn:
                    doc_number = "АВТО"
                    operation_type = "Первичный ввод"
                    source_id = None
                    target_id = inv.object_id
                    operation_date = datetime.utcnow()
                    comment = f"Инвентаризация #{inv.id}: излишек / оприходование"
                    disposal_reason_id = None

                class _ItemIn:
                    def __init__(self, r):
                        self.nomenclature_id = r["nomenclature_id"]
                        self.serial_number = r["serial_number"]
                        self.quantity = r.get("excess") or 1
                        self.price = None
                        self.inventory_number = None

                surp_items = [_ItemIn(s) for s in surplus]
                try:
                    doc_surp = await WeaponService.process_document(
                        db, _DocIn, surp_items,
                        attached_file_path=None,
                        author_id=current_user.id,
                    )
                    # Если doc_miss не создавался — ссылкой будет doc_surp.
                    if correction_doc_id is None:
                        correction_doc_id = doc_surp.id
                except HTTPException as e:
                    raise HTTPException(
                        400,
                        f"Не удалось оприходовать излишек: {e.detail}. "
                        "Закройте без auto_correct и создайте документ вручную.",
                    )

            correction_summary = {
                "missing_resolved": len(missing),
                "surplus_resolved": len(surplus),
                "correction_document_id": correction_doc_id,
            }

    inv.status = "closed"
    inv.closed_at = datetime.utcnow()
    inv.closed_by_id = current_user.id
    if p.note:
        inv.note = (inv.note or "") + ("\n" if inv.note else "") + p.note
    if correction_doc_id:
        inv.correction_document_id = correction_doc_id
    await db.commit()

    return {
        "status": "closed",
        "closed_at": inv.closed_at,
        "correction": correction_summary,
    }


@router.post("/inventory/{inventory_id}/cancel")
async def cancel_inventory(
    inventory_id: int,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Отменить инвентаризацию без сохранения расхождений. Используется
    когда оператор ошибся при запуске или решил начать заново."""
    inv = await db.get(Inventory, inventory_id)
    if not inv:
        raise HTTPException(404, "Инвентаризация не найдена")
    if inv.status != "open":
        raise HTTPException(409, f"Инвентаризация уже {inv.status!r}")
    inv.status = "cancelled"
    inv.closed_at = datetime.utcnow()
    inv.closed_by_id = current_user.id
    await db.commit()
    return {"status": "cancelled"}


@router.get("/inventory/{inventory_id}/export")
async def export_inventory_report(
    inventory_id: int,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Excel-отчёт расхождений для печати/подписи комиссией. Три листа:
    «Совпадения», «Недостача», «Излишек» + сводка."""
    import io

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    report = await inventory_report(inventory_id, db, current_user)
    inv = report["inventory"]
    obj = await db.get(AccountingObject, inv["object_id"])
    obj_name = obj.name if obj else f"id={inv['object_id']}"

    wb = Workbook(write_only=False)

    # ------ Лист «Сводка» ------
    ws = wb.active
    ws.title = "Сводка"
    bold = Font(bold=True)
    gray = PatternFill("solid", fgColor="F3F4F6")
    ws["A1"] = f"Инвентаризационная ведомость — {obj_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    summary_rows = [
        ("ID инвентаризации", inv["id"]),
        ("Объект", obj_name),
        ("Статус", inv["status"]),
        ("Начата", inv["started_at"]),
        ("Закрыта", inv["closed_at"] or "—"),
        ("Найдено ед. (по факту)", report["summary"]["total_found_units"]),
        ("Ожидалось ед. (по учёту)", report["summary"]["total_expected_units"]),
        ("Прогресс", f"{report['summary']['progress_pct']}%"),
        ("Совпадений", report["summary"]["matched_count"]),
        ("Недостача", report["summary"]["missing_count"]),
        ("Излишек", report["summary"]["surplus_count"]),
    ]
    for i, (k, v) in enumerate(summary_rows, start=3):
        ws.cell(row=i, column=1, value=k).font = bold
        ws.cell(row=i, column=2, value=str(v) if v is not None else "—")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40

    # ------ Лист «Недостача» ------
    ws2 = wb.create_sheet("Недостача")
    header = ["Номенклатура", "Серийник", "Ожидалось", "Найдено", "Дефицит"]
    for c, v in enumerate(header, start=1):
        cell = ws2.cell(row=1, column=c, value=v)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="FECACA")
    for i, r in enumerate(report["missing"], start=2):
        ws2.cell(row=i, column=1, value=r.get("name"))
        ws2.cell(row=i, column=2, value=r.get("serial_number") or "—")
        ws2.cell(row=i, column=3, value=r.get("expected_quantity"))
        ws2.cell(row=i, column=4, value=r.get("found_quantity"))
        ws2.cell(row=i, column=5, value=r.get("deficit"))
    for col in range(1, 6):
        ws2.column_dimensions[chr(64 + col)].width = 20

    # ------ Лист «Излишек» ------
    ws3 = wb.create_sheet("Излишек")
    header = ["Номенклатура", "Серийник", "Ожидалось", "Найдено", "Избыток"]
    for c, v in enumerate(header, start=1):
        cell = ws3.cell(row=1, column=c, value=v)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="FEF3C7")
    for i, r in enumerate(report["surplus"], start=2):
        ws3.cell(row=i, column=1, value=r.get("name"))
        ws3.cell(row=i, column=2, value=r.get("serial_number") or "—")
        ws3.cell(row=i, column=3, value=r.get("expected_quantity"))
        ws3.cell(row=i, column=4, value=r.get("found_quantity"))
        ws3.cell(row=i, column=5, value=r.get("excess"))
    for col in range(1, 6):
        ws3.column_dimensions[chr(64 + col)].width = 20

    # ------ Лист «Совпадения» ------
    ws4 = wb.create_sheet("Совпадения")
    header = ["Номенклатура", "Серийник", "Количество"]
    for c, v in enumerate(header, start=1):
        cell = ws4.cell(row=1, column=c, value=v)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="D1FAE5")
    for i, r in enumerate(report["matched"], start=2):
        ws4.cell(row=i, column=1, value=r.get("name"))
        ws4.cell(row=i, column=2, value=r.get("serial_number") or "—")
        ws4.cell(row=i, column=3, value=r.get("found_quantity"))
    for col in range(1, 4):
        ws4.column_dimensions[chr(64 + col)].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"inventory_{inventory_id}_{obj_name}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/inventory")
async def list_inventories(
    object_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    q = select(Inventory).order_by(Inventory.started_at.desc()).limit(100)
    if current_user.role != "admin":
        q = q.where(Inventory.object_id == current_user.object_id)
    if object_id:
        q = q.where(Inventory.object_id == object_id)
    if status:
        q = q.where(Inventory.status == status)
    rows = (await db.execute(q)).scalars().all()
    return [
        {"id": i.id, "object_id": i.object_id, "status": i.status,
         "started_at": i.started_at, "closed_at": i.closed_at}
        for i in rows
    ]


# =====================================================================
# SECURE PASSWORD RESET (одноразовый токен)
# =====================================================================
@router.post("/users/{user_id}/reset-password-link")
async def create_password_reset_link(
    user_id: int,
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Создаёт одноразовый токен сброса пароля и возвращает URL-путь.
    В отличие от старого /reset-password, новый пароль НЕ возвращается в JSON
    и не попадает в логи. Админ передаёт ссылку пользователю; пользователь
    по ней сам вводит новый пароль."""
    _require_admin(current_user)
    user = await db.get(ArsenalUser, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")

    # Генерируем 32 байта случайности, хешируем для хранения
    token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(token.encode()).hexdigest()

    # Старые активные токены пользователя инвалидируем
    old = (await db.execute(
        select(ArsenalPasswordResetToken).where(
            ArsenalPasswordResetToken.user_id == user_id,
            ArsenalPasswordResetToken.used_at.is_(None),
            ArsenalPasswordResetToken.expires_at > datetime.utcnow(),
        )
    )).scalars().all()
    for t in old:
        t.used_at = datetime.utcnow()

    record = ArsenalPasswordResetToken(
        user_id=user_id,
        token_hash=token_hash,
        expires_at=datetime.utcnow() + timedelta(hours=24),
        created_by_id=current_user.id,
    )
    db.add(record)
    await db.commit()
    # Отдаём токен ТОЛЬКО ОДИН РАЗ в ответе — админ должен сразу передать.
    return {
        "reset_url": f"/arsenal_reset_password.html?token={token}",
        "expires_at": record.expires_at.isoformat(),
        "username": user.username,
    }


class PasswordResetSubmit(BaseModel):
    token: str
    new_password: str = Field(..., min_length=8, max_length=128)


@router.post("/reset-password")
async def submit_password_reset(
    data: PasswordResetSubmit,
    db: AsyncSession = Depends(get_arsenal_db),
):
    """Публичный endpoint: пользователь приходит по ссылке и вводит новый
    пароль. Токен одноразовый, живёт 24 часа."""
    token_hash = hashlib.sha256(data.token.encode()).hexdigest()
    record = (await db.execute(
        select(ArsenalPasswordResetToken).where(
            ArsenalPasswordResetToken.token_hash == token_hash,
        )
    )).scalars().first()
    if not record:
        raise HTTPException(404, "Недействительная ссылка сброса")
    if record.used_at is not None:
        raise HTTPException(410, "Эта ссылка уже использована")
    if record.expires_at < datetime.utcnow():
        raise HTTPException(410, "Срок действия ссылки истёк")

    user = await db.get(ArsenalUser, record.user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")

    user.hashed_password = get_password_hash(data.new_password)
    record.used_at = datetime.utcnow()
    await db.commit()
    return {"status": "ok", "username": user.username}
