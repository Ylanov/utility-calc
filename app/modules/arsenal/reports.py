from datetime import datetime, timedelta
from typing import Annotated, Optional
from fastapi import APIRouter, Depends, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import or_, func

from app.core.database import get_arsenal_db
from app.modules.arsenal.models import (
    AccountingObject, Document, DocumentItem, Nomenclature,
    WeaponRegistry, ArsenalUser,
)
from app.modules.arsenal.deps import get_current_arsenal_user

router = APIRouter(tags=["Arsenal Reports"])


@router.get("/reports/search-weapon")
async def search_weapon(
        q: Annotated[str, Query(min_length=2, description="Поиск по серийнику, инв. номеру или названию")],
        db: Annotated[AsyncSession, Depends(get_arsenal_db)],
        current_user: Annotated[ArsenalUser, Depends(get_current_arsenal_user)]
):
    """
    Поиск истории конкретного изделия/партии.
    Ищет по всем документам (сохраненная история).
    """
    # Ищем в истории документов (DocumentItem)
    stmt = (
        select(
            DocumentItem.serial_number,
            DocumentItem.inventory_number,
            Nomenclature.id.label("nom_id"),
            Nomenclature.name,
            Nomenclature.code
        )
        .join(Nomenclature, DocumentItem.nomenclature_id == Nomenclature.id)
        .where(
            or_(
                DocumentItem.serial_number.ilike(f"%{q}%"),
                DocumentItem.inventory_number.ilike(f"%{q}%"),
                Nomenclature.name.ilike(f"%{q}%")
            )
        )
        .distinct()
        .limit(20)
    )

    result = await db.execute(stmt)

    items = []
    for row in result.all():
        items.append({
            "serial": row.serial_number or "Б/Н",
            "inventory": row.inventory_number or "Б/Н",
            "nom_id": row.nom_id,
            "name": row.name,
            "code": row.code
        })
    return items


@router.get("/reports/timeline")
async def get_weapon_timeline(
        serial: str,
        nom_id: int,
        db: Annotated[AsyncSession, Depends(get_arsenal_db)],
        current_user: Annotated[ArsenalUser, Depends(get_current_arsenal_user)],
        skip: Annotated[int, Query(ge=0)] = 0,
        limit: Annotated[int, Query(ge=1, le=1000)] = 50
):
    """
    Таймлайн (История движения) конкретного ствола/партии.
    """
    # 1. Запрашиваем историю из документов
    # Сортировка по дате операции (от новых к старым)
    stmt = (
        select(DocumentItem)
        .join(Document, DocumentItem.document_id == Document.id)
        .options(
            selectinload(DocumentItem.document).selectinload(Document.source),
            selectinload(DocumentItem.document).selectinload(Document.target)
            # УДАЛЕНО: selectinload(DocumentItem.document).selectinload(Document.author)
            # Причина: в модели Document пока нет relationship("author")
        )
        .where(
            # Обработка Б/Н (поиск может передавать строку "Б/Н")
            DocumentItem.serial_number == (None if serial == "Б/Н" else serial),
            DocumentItem.nomenclature_id == nom_id
        )
        .order_by(Document.operation_date.desc(), Document.created_at.desc())
        .offset(skip)
        .limit(limit)
    )

    result = await db.execute(stmt)
    history = result.scalars().all()

    timeline = []
    for item in history:
        doc = item.document
        timeline.append({
            "doc_id": doc.id,
            "date": doc.operation_date.strftime("%d.%m.%Y"),
            "doc_number": doc.doc_number,
            "op_type": doc.operation_type,
            "source": doc.source.name if doc.source else "Внешний источник / Списание",
            "target": doc.target.name if doc.target else "Списание / Вне учета",
            "quantity": item.quantity,
            "price": float(item.price) if item.price else 0.0,
            "inventory_number": item.inventory_number or "Б/Н"
        })

    # 2. Получаем текущий статус (Где это изделие лежит сейчас?)
    reg_stmt = select(WeaponRegistry).options(selectinload(WeaponRegistry.current_object)).where(
        WeaponRegistry.serial_number == (None if serial == "Б/Н" else serial),
        WeaponRegistry.nomenclature_id == nom_id
    )
    reg_res = await db.execute(reg_stmt)
    current_state = reg_res.scalars().first()

    status_text = "Списано / Вне баланса"
    if current_state:
        if current_state.status == 1:
            loc = current_state.current_object.name if current_state.current_object else "Неизвестно"
            status_text = f"В наличии: {loc}"
        elif current_state.status == 2:
            status_text = "В ремонте / В пути"

    return {
        "status": status_text,
        "history": timeline
    }


# =====================================================================
# БАЛАНСОВАЯ ВЕДОМОСТЬ — остатки по счетам / категориям / объектам
# =====================================================================
@router.get("/api/arsenal/reports/balance-summary")
async def balance_summary(
    object_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Агрегированная балансовая ведомость: группировка по счёту учёта +
    категории номенклатуры + объекту. Для бухгалтерии и годового инвентаря.
    unit_head видит только свой склад."""
    stmt = (
        select(
            WeaponRegistry.account_code,
            Nomenclature.category,
            AccountingObject.name.label("object_name"),
            func.count(WeaponRegistry.id).label("units"),
            func.coalesce(func.sum(WeaponRegistry.quantity), 0).label("qty"),
            func.coalesce(
                func.sum(WeaponRegistry.quantity * WeaponRegistry.price), 0
            ).label("total_cost"),
        )
        .join(Nomenclature, Nomenclature.id == WeaponRegistry.nomenclature_id)
        .outerjoin(AccountingObject, AccountingObject.id == WeaponRegistry.current_object_id)
        .where(WeaponRegistry.status == 1)
        .group_by(
            WeaponRegistry.account_code,
            Nomenclature.category,
            AccountingObject.name,
        )
        .order_by(WeaponRegistry.account_code, Nomenclature.category)
    )
    if object_id:
        stmt = stmt.where(WeaponRegistry.current_object_id == object_id)
    if current_user.role != "admin" and current_user.object_id:
        stmt = stmt.where(WeaponRegistry.current_object_id == current_user.object_id)

    rows = (await db.execute(stmt)).all()
    by_account: dict = {}
    grand_total_cost = 0.0
    grand_total_units = 0
    for acc, cat, obj_name, units, qty, cost in rows:
        acc_key = acc or "—"
        d = by_account.setdefault(acc_key, {"items": [], "total_cost": 0.0, "total_units": 0})
        d["items"].append({
            "category": cat or "Без категории",
            "object_name": obj_name or "—",
            "units": int(units), "quantity": int(qty), "cost": float(cost),
        })
        d["total_cost"] += float(cost)
        d["total_units"] += int(units)
        grand_total_cost += float(cost)
        grand_total_units += int(units)

    return {
        "by_account": by_account,
        "grand_total_cost": grand_total_cost,
        "grand_total_units": grand_total_units,
    }


@router.get("/api/arsenal/reports/balance-summary/export")
async def export_balance_summary(
    object_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Excel-версия балансовой ведомости. 2 листа: Детализация + Итоги по счетам."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    data = await balance_summary(object_id, db, current_user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Детализация"
    headers = ["Счёт", "Категория", "Объект", "Ед.", "Кол-во", "Стоимость, ₽"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DBEAFE")
    row = 2
    for acc, block in data["by_account"].items():
        for it in block["items"]:
            ws.cell(row=row, column=1, value=acc)
            ws.cell(row=row, column=2, value=it["category"])
            ws.cell(row=row, column=3, value=it["object_name"])
            ws.cell(row=row, column=4, value=it["units"])
            ws.cell(row=row, column=5, value=it["quantity"])
            ws.cell(row=row, column=6, value=it["cost"])
            row += 1
    for col_letter, width in [("A", 14), ("B", 28), ("C", 28), ("D", 8), ("E", 10), ("F", 14)]:
        ws.column_dimensions[col_letter].width = width

    ws2 = wb.create_sheet("Итоги по счетам")
    ws2.append(["Счёт", "Ед.", "Стоимость, ₽"])
    for acc, block in data["by_account"].items():
        ws2.append([acc, block["total_units"], block["total_cost"]])
    ws2.append([])
    ws2.append(["ИТОГО", data["grand_total_units"], data["grand_total_cost"]])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="balance_summary.xlsx"'},
    )


# =====================================================================
# ОТЧЁТ ПО МОЛ
# =====================================================================
@router.get("/api/arsenal/reports/by-mol")
async def report_by_mol(
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Что числится за каждым МОЛ — для годовых актов материальной ответственности."""
    rows = (await db.execute(
        select(
            AccountingObject.mol_name,
            AccountingObject.id.label("obj_id"),
            AccountingObject.name.label("obj_name"),
            func.coalesce(func.count(WeaponRegistry.id), 0).label("units"),
            func.coalesce(func.sum(WeaponRegistry.quantity), 0).label("qty"),
            func.coalesce(
                func.sum(WeaponRegistry.quantity * WeaponRegistry.price), 0
            ).label("cost"),
        )
        .outerjoin(
            WeaponRegistry,
            (WeaponRegistry.current_object_id == AccountingObject.id) &
            (WeaponRegistry.status == 1),
        )
        .group_by(AccountingObject.mol_name, AccountingObject.id, AccountingObject.name)
        .order_by(AccountingObject.mol_name, AccountingObject.name)
    )).all()

    by_mol: dict = {}
    for mol, oid, oname, units, qty, cost in rows:
        k = mol or "— МОЛ не назначен —"
        d = by_mol.setdefault(k, {"objects": [], "total_units": 0, "total_cost": 0.0})
        d["objects"].append({
            "object_id": oid, "object_name": oname,
            "units": int(units), "quantity": int(qty), "cost": float(cost),
        })
        d["total_units"] += int(units)
        d["total_cost"] += float(cost)
    return {"by_mol": by_mol}


# =====================================================================
# ОБОРОТ ЗА ПЕРИОД
# =====================================================================
@router.get("/api/arsenal/reports/turnover")
async def turnover_report(
    date_from: datetime,
    date_to: datetime,
    object_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Оборот за период: сколько поступило/выбыло/перемещено, в шт. и в рублях."""
    stmt = (
        select(
            Document.operation_type,
            func.count(Document.id).label("docs"),
            func.coalesce(func.sum(DocumentItem.quantity), 0).label("qty"),
            func.coalesce(
                func.sum(DocumentItem.quantity * DocumentItem.price), 0
            ).label("cost"),
        )
        .join(DocumentItem, DocumentItem.document_id == Document.id)
        .where(
            Document.operation_date >= date_from,
            Document.operation_date <= date_to,
            Document.is_reversed.is_(False),
        )
        .group_by(Document.operation_type)
    )
    if object_id:
        stmt = stmt.where(
            (Document.source_id == object_id) | (Document.target_id == object_id)
        )
    if current_user.role != "admin" and current_user.object_id:
        stmt = stmt.where(
            (Document.source_id == current_user.object_id) |
            (Document.target_id == current_user.object_id)
        )

    rows = (await db.execute(stmt)).all()
    by_type = {
        op: {"docs": int(d), "quantity": int(q), "cost": float(c)}
        for op, d, q, c in rows
    }
    inbound_types = ("Первичный ввод", "Прием")
    outbound_types = ("Списание", "Утилизация", "Отправка")
    inbound_cost = sum(by_type.get(t, {}).get("cost", 0) for t in inbound_types)
    outbound_cost = sum(by_type.get(t, {}).get("cost", 0) for t in outbound_types)

    return {
        "period": {
            "from": date_from.isoformat(),
            "to": date_to.isoformat(),
            "object_id": object_id,
        },
        "by_operation": by_type,
        "summary": {
            "total_docs": sum(v["docs"] for v in by_type.values()),
            "inbound_cost": inbound_cost,
            "outbound_cost": outbound_cost,
            "net_cost": inbound_cost - outbound_cost,
        },
    }


# =====================================================================
# ТОП ПЕРЕМЕЩАЮЩИХСЯ
# =====================================================================
@router.get("/api/arsenal/reports/top-moving")
async def top_moving(
    days: int = Query(30, ge=1, le=365),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_arsenal_db),
    current_user: ArsenalUser = Depends(get_current_arsenal_user),
):
    """Топ-N наиболее активных позиций (по кол-ву движений) за последние N дней."""
    cutoff = datetime.utcnow() - timedelta(days=days)
    rows = (await db.execute(
        select(
            Nomenclature.id,
            Nomenclature.name,
            Nomenclature.category,
            func.count(DocumentItem.id).label("movements"),
            func.coalesce(func.sum(DocumentItem.quantity), 0).label("qty"),
        )
        .join(DocumentItem, DocumentItem.nomenclature_id == Nomenclature.id)
        .join(Document, Document.id == DocumentItem.document_id)
        .where(
            Document.operation_date >= cutoff,
            Document.is_reversed.is_(False),
        )
        .group_by(Nomenclature.id, Nomenclature.name, Nomenclature.category)
        .order_by(func.count(DocumentItem.id).desc())
        .limit(limit)
    )).all()

    return {
        "days": days,
        "items": [
            {
                "nomenclature_id": nid, "name": name, "category": cat,
                "movements": int(mv), "total_quantity": int(qty),
            }
            for nid, name, cat, mv, qty in rows
        ],
    }
